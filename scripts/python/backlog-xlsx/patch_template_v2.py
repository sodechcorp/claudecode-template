# -*- coding: utf-8 -*-
"""patch_template_v2.py
対応記録テンプレート.xlsx に以下の修正を一度だけ適用する:
  1. サマリー・経緯 r11-r15 の枠線色を既存と同じ 00B4C6E7（ライトブルー）に統一
  2. サマリー・経緯 r18-r21 の枠線色統一 + 列構造変更:
     旧: A=No / B=内容 / C=影響範囲 / D=期待する判断者 / E=関連ファイル / F=空
     新: A=No / B:C マージ=内容 / D:E マージ=影響範囲 / F=関連ファイル

Usage:
    python patch_template_v2.py
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles.borders import Border, Side
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"
COLOR = "00B4C6E7"

_thin = Side(style="thin", color=COLOR)
_FULL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_TB   = Border(top=_thin, bottom=_thin)
_LTB  = Border(left=_thin, top=_thin, bottom=_thin)
_RTB  = Border(right=_thin, top=_thin, bottom=_thin)


def _apply_summary_row_border(ws, row):
    """対応サマリー行（A 単独 + B:F マージ）に色付き thin border を適用。"""
    ws.cell(row, 1).border = _FULL
    ws.cell(row, 2).border = _FULL
    ws.cell(row, 3).border = _TB
    ws.cell(row, 4).border = _TB
    ws.cell(row, 5).border = _TB
    ws.cell(row, 6).border = _RTB


def _apply_pending_row_border(ws, row):
    """判断保留事項行（A 単独 + B:C マージ + D:E マージ + F 単独）に色付き thin border を適用。"""
    ws.cell(row, 1).border = _FULL
    ws.cell(row, 2).border = _LTB
    ws.cell(row, 3).border = _RTB
    ws.cell(row, 4).border = _LTB
    ws.cell(row, 5).border = _RTB
    ws.cell(row, 6).border = _FULL


def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)
    ws = wb["サマリー・経緯"]

    # ── 1. r11-r15 枠線色を 00B4C6E7 に塗り直し ──────────────────────────────
    for r in range(11, 16):
        _apply_summary_row_border(ws, r)
    print("[OK] r11-r15 枠線色を 00B4C6E7 に更新")

    # ── 2. r18-r21 判断保留事項: 列構造変更 ──────────────────────────────────
    # 2-a. r18 ヘッダー値を新構造に書き換え
    ws.cell(18, 1).value = "No"
    ws.cell(18, 2).value = "内容"
    ws.cell(18, 3).value = None
    ws.cell(18, 4).value = "影響範囲"
    ws.cell(18, 5).value = None
    ws.cell(18, 6).value = "関連ファイル"

    # 2-b. r19-r21 の D 列（旧「期待する判断者」データ）をクリア
    for r in range(19, 22):
        ws.cell(r, 4).value = None
        ws.cell(r, 5).value = None

    # 2-c. 既存マージを解除（r18-r21 の B-F 間に残骸がある場合）
    to_discard = [
        mc for mc in list(ws.merged_cells.ranges)
        if mc.min_row >= 18 and mc.max_row <= 21
    ]
    for mc in to_discard:
        ws.merged_cells.ranges.discard(mc)

    # 2-d. 新マージを設定: B:C / D:E
    for r in range(18, 22):
        ws.merge_cells(start_row=r, end_row=r, start_column=2, end_column=3)
        ws.merge_cells(start_row=r, end_row=r, start_column=4, end_column=5)

    # 2-e. 枠線を新構造に合わせて適用
    for r in range(18, 22):
        _apply_pending_row_border(ws, r)

    print("[OK] r18-r21 列構造変更（B:C / D:E マージ + F=関連ファイル）と枠線色更新")

    # ── 保存 ──────────────────────────────────────────────────────────────────
    try:
        wb.save(TEMPLATE)
        print(f"\n[完了] テンプレート更新: {TEMPLATE}")
    except PermissionError as e:
        print(f"[ERROR] 保存失敗（Excel でファイルが開かれている可能性）: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
