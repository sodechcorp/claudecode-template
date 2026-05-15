# -*- coding: utf-8 -*-
"""backlog-xlsx / create_evidence_v2.py
エビデンス_v2.xlsx を生成する（v1 に画像が貼付された後にテストケースが追加された場合）

v1 は読み取り専用で参照し、増分テストケース（v1 にない No）のみを含む
v2 を新規ファイルとして発行する。v1 の画像は保護される。

Usage:
    python create_evidence_v2.py \\
      --v1 PATH_TO_V1 --folder FOLDER --issue-id ID \\
      --implementation-plan PATH
"""

import argparse
import os
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl がインストールされていません。`pip install openpyxl` を実行してください。")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "エビデンステンプレート.xlsx"
WRAP = Alignment(wrap_text=True, vertical="top")
STRIPE_A = PatternFill("solid", fgColor="FFFFFF")
STRIPE_B = PatternFill("solid", fgColor="F2F7FB")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F7")
CHECKLIST_BG = PatternFill("solid", fgColor="E8F0FE")
EVIDENCE_BG  = PatternFill("solid", fgColor="FFF3CD")
WHITE = PatternFill("solid", fgColor="FFFFFF")


# ── パースユーティリティ（create_evidence.py と共通） ─────────────────────────

def read_md(path):
    if path and Path(path).exists():
        try:
            return Path(path).read_text(encoding="utf-8")
        except UnicodeDecodeError as e:
            print(f"[ERROR] ファイルのエンコーディングが UTF-8 ではありません: {path}\n{e}")
            sys.exit(1)
    return ""


def extract_section(md, *headings):
    for h in headings:
        pat = r"^#{1,3}\s+" + re.escape(h) + r"\s*$"
        m = re.search(pat, md, re.MULTILINE)
        if m:
            start = m.end()
            rest = md[start:]
            end_m = re.search(r"^#{1,3}\s", rest, re.MULTILINE)
            body = rest[: end_m.start()] if end_m else rest
            return body.strip()
    return ""


def parse_md_table(section_text):
    rows = []
    headers = []
    for line in section_text.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if all(re.match(r"^[-: ]+$", c) for c in cells):
            continue
        if not headers:
            headers = cells
        else:
            rows.append(dict(zip(headers, cells)))
    return rows


# ── v1 のテストケース No を読み取る ──────────────────────────────────────────

def read_v1_nos(v1_path):
    """v1 のテスト仕様シートから既存の No セットを返す。"""
    nos = set()
    try:
        wb = load_workbook(v1_path, data_only=True)
        if "テスト仕様" not in wb.sheetnames:
            return nos
        ws = wb["テスト仕様"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is not None:
                nos.add(str(row[0]).strip())
    except Exception as e:
        print(f"[WARN] v1 読み取りエラー: {e}")
    return nos


# ── テスト仕様シート ───────────────────────────────────────────────────────

def fill_test_spec(ws, test_cases):
    for i, tc in enumerate(test_cases):
        r = 3 + i
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        vals = [
            tc.get("No", str(i + 1)),
            tc.get("確認観点", ""),
            tc.get("タイミング", ""),
            tc.get("実行種別", "UI手動"),
            tc.get("確認手順", ""),
            tc.get("期待結果", ""),
            tc.get("貼付先シート", "実装前エビデンス" if tc.get("タイミング") == "実装前" else "実装後エビデンス"),
        ]
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.alignment = WRAP
            cell.fill = fill


# ── エビデンスシート ──────────────────────────────────────────────────────────

def fill_evidence_sheet(ws, cases, sheet_label):
    # テンプレートは3件固定想定のマージ構造を持つため、データエリア (r5以降) のマージを
    # 事前に全解除してから動的書き込みする（r1-r4のタイトル/ヘッダー行マージは保持）
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 5:
            ws.unmerge_cells(str(mr))

    checklist_start = 6
    for i, tc in enumerate(cases):
        r = checklist_start + i
        cell_cb = ws.cell(row=r, column=1, value="□")
        cell_cb.alignment = WRAP
        cell_cb.fill = CHECKLIST_BG

        label = f"No.{tc.get('No', i + 1)}: {tc.get('確認観点', '')}"
        steps = tc.get("確認手順", "")
        if steps and "\n" in steps:
            first_step = steps.split("\n")[0].strip().lstrip("0123456789. ")
            label += f"（{first_step[:30]}）"
        cell_obs = ws.cell(row=r, column=2, value=label)
        cell_obs.alignment = WRAP
        cell_obs.fill = CHECKLIST_BG

        ws.cell(row=r, column=3).fill = WHITE
        ws.cell(row=r, column=4).fill = WHITE

    paste_start = checklist_start + len(cases) + 1

    # セクション見出し行: A:D 横幅マージ
    ws.merge_cells(start_row=paste_start, start_column=1, end_row=paste_start, end_column=4)
    hdr_cell = ws.cell(row=paste_start, column=1, value="■ エビデンス貼付欄")
    hdr_cell.fill = LIGHT_BLUE
    hdr_cell.alignment = WRAP

    row_ptr = paste_start + 1
    for i, tc in enumerate(cases):
        # ラベル行: A:D 横幅マージ
        label = f"エビデンス{chr(0x2460 + i)}: No.{tc.get('No', i + 1)} {tc.get('確認観点', '')}"
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=4)
        cell = ws.cell(row=row_ptr, column=1, value=label)
        cell.fill = EVIDENCE_BG
        cell.alignment = WRAP
        row_ptr += 1

        # 貼付エリア: A:D × 10行をまとめてマージ
        paste_top = row_ptr
        paste_bottom = paste_top + 9
        for r in range(paste_top, paste_bottom + 1):
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = WHITE
        ws.cell(row=paste_top, column=1, value="ここにスクリーンショットを貼り付けてください")
        ws.cell(row=paste_top, column=1).alignment = WRAP
        ws.merge_cells(start_row=paste_top, start_column=1, end_row=paste_bottom, end_column=4)

        row_ptr = paste_bottom + 2


# ── main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="エビデンス_v2.xlsx を生成する（増分テストケースのみ）")
    parser.add_argument("--v1",                  required=True,
                        help="既存の v1 xlsx パス（画像保護のため読み取り専用参照）")
    parser.add_argument("--folder",              required=True)
    parser.add_argument("--issue-id",            required=True, dest="issue_id")
    parser.add_argument("--implementation-plan", required=True, dest="implementation_plan",
                        help="docs/logs/{issueID}/implementation-plan.md のパス")
    args = parser.parse_args()

    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    if not Path(args.v1).exists():
        print(f"[ERROR] v1 ファイルが見つかりません: {args.v1}")
        sys.exit(1)

    impl_md = read_md(args.implementation_plan)
    if not impl_md:
        print(f"[ERROR] ファイルが見つかりません: {args.implementation_plan}")
        sys.exit(1)

    # 全テストケース取得
    spec_text = extract_section(impl_md, "テスト仕様", "テストケース", "テスト仕様テーブル")
    all_cases = parse_md_table(spec_text)

    # v1 に存在する No を取得し、増分のみ抽出
    v1_nos = read_v1_nos(args.v1)
    new_cases = [tc for tc in all_cases if str(tc.get("No", "")).strip() not in v1_nos]

    if not new_cases:
        print("[INFO] 増分テストケースはありません。v2 の生成をスキップします。")
        sys.exit(0)

    # 実行種別=UI手動 の行のみエビデンスファイルに記録（後方互換: 実行種別列なしは全行対象）
    has_shubetsu = any(tc.get("実行種別", "").strip() for tc in new_cases)
    if has_shubetsu:
        new_cases = [tc for tc in new_cases if tc.get("実行種別", "").strip() == "UI手動"]

    if not new_cases:
        print("[INFO] 増分 UI 手動テストケースはありません。v2 の生成をスキップします。")
        sys.exit(0)

    print(f"増分 UI 手動テストケース: {len(new_cases)} 件（v1 既存: {len(v1_nos)} 件）")

    before_cases = [tc for tc in new_cases if tc.get("タイミング", "") == "実装前"]
    after_cases  = [tc for tc in new_cases if tc.get("タイミング", "") != "実装前"]

    os.makedirs(args.folder, exist_ok=True)
    try:
        wb = load_workbook(TEMPLATE)
    except Exception as e:
        print(f"[ERROR] テンプレートファイルの読み込みに失敗しました: {TEMPLATE}\n{e}")
        sys.exit(1)

    fill_test_spec(wb["テスト仕様"], new_cases)
    fill_evidence_sheet(wb["実装前エビデンス"], before_cases, "実装前")
    fill_evidence_sheet(wb["実装後エビデンス"], after_cases, "実装後")

    path = os.path.join(args.folder, f"{args.issue_id}_エビデンス_v2.xlsx")
    try:
        wb.save(path)
    except PermissionError as e:
        print(f"[ERROR] xlsx の保存に失敗しました（ファイルが開かれている可能性があります）: {path}\n{e}")
        sys.exit(1)
    print(f"生成完了: {path}")
    print(f"  テスト仕様: {len(new_cases)} 件（実装前: {len(before_cases)} 件 / 実装後: {len(after_cases)} 件）")


if __name__ == "__main__":
    main()
