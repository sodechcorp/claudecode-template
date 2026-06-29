# -*- coding: utf-8 -*-
"""
backlog-xlsx / update_records.py
対応記録.xlsx を更新するスクリプト

Usage (タイムライン行を追加):
    python update_records.py --folder FOLDER --issue-id ID timeline \\
      --phase "調査" --source "Claude" --content "○○を調査: 原因は△△"

Usage (セルを直接更新):
    python update_records.py --folder FOLDER --issue-id ID cell \\
      --sheet "課題と対応方針" --label "ステータス" --col 2 --value "完了" --force

    python update_records.py --folder FOLDER --issue-id ID cell \\
      --sheet "対応内容" --row 2 --col 1 --value "○○を実施した" --force

Usage (変更した資材を1行追加):
    python update_records.py --folder FOLDER --issue-id ID content-list \\
      --label "preCheck 画面（preCheck）" --kind "変更" --detail "ラジオボタン追加"

Usage (Before/After 追記):
    python update_records.py --folder FOLDER --issue-id ID before-after \\
      --file "force-app/.../X.cls" --before "変更前コード" --after "変更後コード"

Usage (NG対応履歴に1行追加):
    python update_records.py --folder FOLDER --issue-id ID ng-history \\
      --round "R1" --tc "TC-003" --reason "件数NG: 期待3件だが1件" --fix "SOQL WHERE 条件を修正"
Usage (対応内容シートを implementation-summary.md から一括記入):
    python update_records.py --folder FOLDER --issue-id ID content-from-md \\
      --summary docs/logs/{issueID}/implementation-summary.md

Usage (xlsx 全枠充足確認):
    python update_records.py --folder FOLDER --issue-id ID verify \\
      --stage pre-release
    python update_records.py --folder FOLDER --issue-id ID verify \\
      --stage final --status-expected 完了
"""

import argparse
import copy
import datetime
import os
import re
import sys
from pathlib import Path

from _common import validate_folder, _stripe_fill

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("[ERROR] openpyxl がインストールされていません。`pip install openpyxl` を実行してください。")
    sys.exit(1)

WRAP = Alignment(wrap_text=True, vertical="top")


def find_next_empty_row(ws, col=1, start_row=1):
    """指定列で最初の空行を返す（start_row から下方向に検索）"""
    r = start_row
    while ws.cell(row=r, column=col).value is not None:
        r += 1
    return r


def find_header_row(ws, keywords, start_row=1):
    """キーワードをいずれか含む最初の行番号を返す。見つからなければ None。"""
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.value and any(kw in str(cell.value) for kw in keywords):
                return cell.row
    return None


def find_label_row(ws, label, col=1, start_row=1):
    """col 列で label に部分一致する行を返す。"""
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v and label in str(v).strip():
            return r
    return None


def _copy_font(src_font):
    """フォントオブジェクトを独立コピーして返す。"""
    return Font(
        name=src_font.name or "游ゴシック",
        size=src_font.size or 10,
        bold=src_font.bold,
        italic=src_font.italic,
        underline=src_font.underline,
        color=copy.copy(src_font.color) if src_font.color else None,
    )


def cmd_timeline(args, wb):
    """課題と対応方針 シートのタイムラインに1行追加する"""
    sheet_name = "課題と対応方針"
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] シート '{sheet_name}' が見つかりません。")
        sys.exit(1)
    ws = wb[sheet_name]

    # タイムラインセクションヘッダーを検索
    tl_section_row = find_header_row(ws, ("■ 対応経緯タイムライン",))
    if tl_section_row is None:
        print("[ERROR] タイムラインセクション（■ 対応経緯タイムライン）が見つかりません。")
        sys.exit(1)

    # 列ヘッダー行（No / 日時 / 発生元 / フェーズ / 内容 / 理由）
    col_header_row = tl_section_row + 1
    data_start = col_header_row + 1

    # 既存の最大 No 値を取得（位置ベースではなく実 No ベース）
    max_no = 0
    for r in range(data_start, ws.max_row + 2):
        v = ws.cell(r, 1).value
        if isinstance(v, int) and v > 0:
            max_no = max(max_no, v)
        elif isinstance(v, str) and v.strip().isdigit():
            max_no = max(max_no, int(v.strip()))
    no = max_no + 1

    # 次の空行を取得
    next_row = find_next_empty_row(ws, col=1, start_row=data_start)

    # 重複検出: 直前の非空行と phase + content が一致する場合はスキップ
    if not getattr(args, "force", False):
        for r in range(next_row - 1, data_start - 1, -1):
            if ws.cell(r, 1).value is not None:
                existing_phase = str(ws.cell(r, 4).value or "").strip()
                existing_content = str(ws.cell(r, 5).value or "").strip()
                if existing_phase == str(args.phase).strip() and existing_content == str(args.content).strip():
                    print(f"[SKIP] 重複: phase={args.phase}, content={args.content[:30]}... （--force で強制追記）")
                    return
                break

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    fill = _stripe_fill(no - 1)

    # データ開始行からフォントを継承
    src_font = ws.cell(data_start, 1).font
    row_font = _copy_font(src_font)

    for col, value in enumerate([no, now, args.source, args.phase, args.content, args.reason or ""], start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = WRAP
        cell.fill = fill
        cell.font = row_font

    print(f"タイムライン追加: No={no} / 行{next_row} / {now} / {args.phase} / {args.content[:30]}...")


def cmd_cell(args, wb):
    """指定したシート・行（または行ラベル）・列のセルを更新する"""
    if args.sheet not in wb.sheetnames:
        print(f"[ERROR] シート '{args.sheet}' が見つかりません。利用可能: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[args.sheet]

    # --label 指定時はラベル検索で行番号を決定（--row より優先）
    row = getattr(args, "row", None)
    label = getattr(args, "label", "") or ""
    if label:
        found = find_label_row(ws, label)
        if found is None:
            print(f"[ERROR] ラベル '{label}' が見つかりません（シート: {args.sheet}）")
            sys.exit(1)
        row = found

    if row is None:
        print("[ERROR] --row または --label のいずれかを指定してください。")
        sys.exit(1)

    cell = ws.cell(row=row, column=args.col)
    if cell.value and not getattr(args, "force", False):
        print(f"[WARN] 既存値あり: {cell.value!r} → --force を指定しないと上書きしません")
        return
    cell.value = args.value
    cell.alignment = WRAP
    print(f"セル更新: {args.sheet}!(行{row},{args.col}) = {str(args.value)[:40]}...")


def cmd_before_after(args, wb):
    """対応内容シートの Before/After セクションにファイルの変更前後を追記する"""
    sheet_name = "対応内容"
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] シート '{sheet_name}' が見つかりません。")
        sys.exit(1)
    ws = wb[sheet_name]

    ba_header = find_header_row(ws, ("■ Before / After", "Before / After"))
    if ba_header is None:
        print("[ERROR] ■ Before / After セクションが見つかりません。")
        sys.exit(1)

    # Before/After セクションの次の空行を探す
    next_row = find_next_empty_row(ws, col=1, start_row=ba_header + 1)

    def _write_cell_1(row, text, bold=False):
        """テンプレートの A:D マージ行にも対応した col 1 書き込み。"""
        c = ws.cell(row, 1, value=text)
        c.alignment = WRAP
        c.font = Font(name="游ゴシック", size=10, bold=bold)

    _write_cell_1(next_row,     f"【{args.file}】",        bold=True)
    _write_cell_1(next_row + 1, f"Before: {args.before}",  bold=False)
    _write_cell_1(next_row + 2, f"After:  {args.after}",   bold=False)

    print(f"[OK] Before/After 追記: 行{next_row}〜{next_row + 2} / {args.file}")


def cmd_content_list(args, wb):
    """対応内容シートの「変更を加えた資材一覧」に1行追加する。

    引数:
      --label : 資材名（日本語表示名優先・API名は括弧補助）
                例: 「preCheck 画面（preCheck）」「犯罪歴確認フラグ（CriminalHistory__c）」
      --kind  : 変更種別（例: 新規追加 / 変更 / 削除）
      --detail: 変更内容（人間が読める日本語で1〜2行）
    """
    sheet_name = "対応内容"
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] シート '{sheet_name}' が見つかりません。")
        sys.exit(1)
    ws = wb[sheet_name]

    # 「変更を加えた資材一覧」セクションヘッダーを検索
    section_row = find_header_row(ws, ("■ 変更を加えた資材一覧",))
    if section_row is None:
        print("[ERROR] ■ 変更を加えた資材一覧 セクションが見つかりません。")
        sys.exit(1)

    # 列ヘッダー行（No / 資材名 / 変更種別 / 変更内容）
    col_header_row = section_row + 1
    data_start = col_header_row + 1

    # 既存の最大 No と次の空行を取得
    max_no = 0
    for r in range(data_start, ws.max_row + 2):
        v = ws.cell(r, 1).value
        if isinstance(v, int) and v > 0:
            max_no = max(max_no, v)
        elif isinstance(v, str) and str(v).strip().isdigit():
            max_no = max(max_no, int(str(v).strip()))

    next_row = find_next_empty_row(ws, col=1, start_row=data_start)
    new_no = max_no + 1
    fill = _stripe_fill(new_no - 1)

    # データ書き込み: No(A), 資材名(B), 変更種別(C), 変更内容(D)
    for col, value in enumerate([new_no, args.label, args.kind, args.detail], start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = WRAP
        cell.fill = fill

    print(f"[OK] 変更資材追加: No={new_no} / 行{next_row} / {args.label[:30]} / {args.kind}")


def cmd_ng_history(args, wb):
    """対応内容シートの「NG対応履歴」セクションに1行追加する。

    引数:
      --round  : 回次（例: R1 / R2）
      --tc     : TC番号（例: TC-003）
      --reason : NG原因（機械判定の reason / 目視確認内容）
      --fix    : 修正内容（何をどう変えたか）
    """
    sheet_name = "対応内容"
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] シート '{sheet_name}' が見つかりません。")
        sys.exit(1)
    ws = wb[sheet_name]

    section_row = find_header_row(ws, ("■ NG対応履歴",))
    if section_row is None:
        print("[ERROR] ■ NG対応履歴 セクションが見つかりません。テンプレートを再生成してください（build_template.py を実行）。")
        sys.exit(1)

    col_header_row = section_row + 1
    data_start = col_header_row + 1

    next_row = find_next_empty_row(ws, col=1, start_row=data_start)

    # 重複検出: 回次(A) + TC(B) の組み合わせ一致でスキップ
    if not getattr(args, "force", False):
        for r in range(data_start, next_row):
            existing_round = str(ws.cell(r, 1).value or "").strip()
            existing_tc    = str(ws.cell(r, 2).value or "").strip()
            if existing_round == str(args.round).strip() and existing_tc == str(args.tc).strip():
                print(f"[SKIP] 重複: round={args.round}, tc={args.tc} （--force で強制追記）")
                return

    # 空行が足りない場合は1行追加して書き込む
    fill = _stripe_fill(next_row - data_start)
    for col, value in enumerate([args.round, args.tc, args.reason, args.fix], start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = WRAP
        cell.fill = fill

    print(f"[OK] NG対応履歴追加: 行{next_row} / {args.round} / {args.tc} / {args.reason[:30]}")


def _read_md(path):
    """MD ファイルを UTF-8 で読む。見つからなければ空文字。"""
    p = Path(path)
    if p.exists():
        try:
            return p.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            print(f"[ERROR] UTF-8 で読めません: {path}")
            sys.exit(1)
    return ""


def _extract_section_md(md, *headings):
    """## または ### 見出しのセクション本文を返す（最初にマッチしたもの）。"""
    for h in headings:
        parts = re.split(r" +", h)
        kw = r"\s*".join(re.escape(p) for p in parts)
        pat = (r"^#{1,3}[\s　]+"
               r"(?:[■●▶◆]\s*)?" + kw +
               r"(?:[・/／]\S+?)?(?:テーブル|一覧|[:：])?(?:\s*[（(][^)）]*[)）])?\s*$")
        m = re.search(pat, md, re.MULTILINE)
        if m:
            start = m.end()
            rest = md[start:]
            level = len(re.match(r"^#+", md[m.start():]).group(0))
            end_pat = rf"^#{{1,{level}}}\s"
            end_m = re.search(end_pat, rest, re.MULTILINE)
            body = rest[: end_m.start()] if end_m else rest
            stripped = body.strip()
            if stripped:
                return stripped
    return ""


def _parse_md_table(md, heading):
    """MD 中の指定見出しの直下にある Markdown テーブルを解析してリストを返す。

    Returns:
        list[dict]: 各行をヘッダ列名→値の dict で返す。テーブルなしなら空リスト。
    """
    body = _extract_section_md(md, heading)
    if not body:
        return []

    rows = []
    headers = None
    for line in body.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if headers is None:
            headers = cells
            continue
        # セパレータ行をスキップ（--- のみ）
        if all(re.fullmatch(r":?-+:?", c) for c in cells if c):
            continue
        if len(cells) >= len(headers):
            rows.append({headers[i]: cells[i] for i in range(len(headers))})
    return rows


def cmd_content_from_md(args, wb):
    """対応内容シートを implementation-summary.md から一括記入する。

    implementation-summary.md の固定見出し:
      ## 実施した対応      → 対応内容シート 行2 col1 (実施した対応 大テキスト)
      ## 変更を加えた資材一覧 → 対応内容シート 資材一覧テーブル (content-list 形式)
      ## Before / After    → 対応内容シート Before/After セクション (before-after 形式)
    """
    md = _read_md(args.summary)
    if not md:
        print(f"[ERROR] implementation-summary.md が見つかりません: {args.summary}")
        sys.exit(1)

    sheet_name = "対応内容"
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] シート '{sheet_name}' が見つかりません。")
        sys.exit(1)
    ws = wb[sheet_name]

    # ① 実施した対応（行2 col1 大テキスト）
    text_content = _extract_section_md(md, "実施した対応")
    if text_content:
        cell = ws.cell(row=2, column=1)
        if cell.value and not getattr(args, "force", False):
            print("[SKIP] 実施した対応: 既存値あり（--force で上書き可）")
        else:
            cell.value = text_content
            cell.alignment = WRAP
            print(f"[OK] 実施した対応: 行2 に記入 ({len(text_content)}文字)")
    else:
        print("[WARN] implementation-summary.md に '## 実施した対応' が見つかりませんでした。")

    # ② 変更を加えた資材一覧（テーブル行を content-list 形式で追記）
    material_rows = _parse_md_table(md, "変更を加えた資材一覧")
    if material_rows:
        section_row = find_header_row(ws, ("■ 変更を加えた資材一覧",))
        if section_row is None:
            print("[ERROR] ■ 変更を加えた資材一覧 セクションが見つかりません。")
            sys.exit(1)
        col_header_row = section_row + 1
        data_start = col_header_row + 1

        max_no = 0
        for r in range(data_start, ws.max_row + 2):
            v = ws.cell(r, 1).value
            if isinstance(v, int) and v > 0:
                max_no = max(max_no, v)
            elif isinstance(v, str) and str(v).strip().isdigit():
                max_no = max(max_no, int(str(v).strip()))

        for row_data in material_rows:
            # ヘッダ名揺れを吸収（「資材名（表示名）」「資材名」等）
            label  = next((v for k, v in row_data.items() if "資材名" in k), "")
            kind   = row_data.get("変更種別", "")
            detail = row_data.get("変更内容", "")
            if not label:
                continue
            next_row = find_next_empty_row(ws, col=1, start_row=data_start)
            max_no += 1
            fill = _stripe_fill(max_no - 1)
            for col, value in enumerate([max_no, label, kind, detail], start=1):
                cell = ws.cell(row=next_row, column=col, value=value)
                cell.alignment = WRAP
                cell.fill = fill
            print(f"[OK] 変更資材追加: No={max_no} / {label[:30]} / {kind}")
    else:
        print("[WARN] implementation-summary.md に '## 変更を加えた資材一覧' テーブルが見つかりませんでした。")

    # ③ Before / After（テーブル行を before-after 形式で追記）
    ba_rows = _parse_md_table(md, "Before / After")
    if ba_rows:
        ba_header = find_header_row(ws, ("■ Before / After", "Before / After"))
        if ba_header is None:
            print("[ERROR] ■ Before / After セクションが見つかりません。")
            sys.exit(1)
        next_row = find_next_empty_row(ws, col=1, start_row=ba_header + 1)
        for row_data in ba_rows:
            file_  = row_data.get("ファイル", "")
            before = row_data.get("Before", "")
            after  = row_data.get("After", "")
            if not file_:
                continue
            from openpyxl.styles import Font as _Font
            ws.cell(next_row,     1, value=f"【{file_}】").alignment = WRAP
            ws.cell(next_row,     1).font = _Font(name="游ゴシック", size=10, bold=True)
            ws.cell(next_row + 1, 1, value=f"Before: {before}").alignment = WRAP
            ws.cell(next_row + 2, 1, value=f"After:  {after}").alignment = WRAP
            print(f"[OK] Before/After 追記: {file_}")
            next_row += 3
    # Before/After は任意のため WARN なし


def cmd_verify(args, wb):
    """xlsx の全必須枠が充足されているか確認する（read-back 検証）。

    --stage pre-release : Phase5 前の事前検証。対応内容の資材一覧1行以上を必須チェック。
    --stage final       : Phase6 末の最終確認。ステータスが --status-expected と一致するか確認。

    充足: exit 0 / 不足: NG 列挙して exit 2
    """
    PLACEHOLDER_PREFIX = "（未記入："
    issues = []

    # ─── シート①「課題と対応方針」検証 ──────────────────────────────────────
    plan_sheet = "課題と対応方針"
    if plan_sheet not in wb.sheetnames:
        print(f"[ERROR] シート '{plan_sheet}' が見つかりません。")
        sys.exit(1)
    ws1 = wb[plan_sheet]

    # メタ情報（B2–B6）
    for label in ["課題ID", "件名", "ステータス"]:
        row = find_label_row(ws1, label)
        if row is None:
            issues.append(f"①シート: ラベル '{label}' が見つかりません")
            continue
        val = ws1.cell(row, 2).value
        if not val or str(val).strip() == "":
            issues.append(f"①シート: '{label}' が空欄")

    # 本文4枠（プレースホルダも NG 扱い）
    for label in ["課題の内容・詳細", "原因・現状", "対応方針（結論）", "方針決定の経緯・根拠"]:
        row = find_label_row(ws1, label)
        if row is None:
            issues.append(f"①シート: ラベル '{label}' が見つかりません")
            continue
        val = ws1.cell(row, 2).value
        if not val or str(val).strip() == "":
            issues.append(f"①シート: '{label}' が空欄")
        elif str(val).startswith(PLACEHOLDER_PREFIX):
            issues.append(f"①シート: '{label}' がプレースホルダのまま（要追記）")

    # ステータス期待値チェック（--stage final のみ）
    if getattr(args, "stage", "") == "final":
        expected = getattr(args, "status_expected", None)
        if expected:
            row = find_label_row(ws1, "ステータス")
            if row is not None:
                val = str(ws1.cell(row, 2).value or "").strip()
                if val != expected:
                    issues.append(f"①シート: ステータスが '{val}' のまま（期待値: '{expected}'）")

    # ─── シート②「対応内容」検証 ──────────────────────────────────────────
    content_sheet = "対応内容"
    if content_sheet not in wb.sheetnames:
        print(f"[ERROR] シート '{content_sheet}' が見つかりません。")
        sys.exit(1)
    ws2 = wb[content_sheet]

    # 実施した対応（行2 col1）
    val = ws2.cell(2, 1).value
    if not val or str(val).strip() == "":
        issues.append("②シート: '実施した対応' が空欄（行2）")
    elif str(val).startswith(PLACEHOLDER_PREFIX):
        issues.append("②シート: '実施した対応' がプレースホルダのまま")

    # 変更を加えた資材一覧（少なくとも1行）
    section_row = find_header_row(ws2, ("■ 変更を加えた資材一覧",))
    if section_row is not None:
        data_start = section_row + 2  # ヘッダ行の次
        has_data = any(ws2.cell(r, 1).value for r in range(data_start, data_start + 10))
        if not has_data:
            issues.append("②シート: '変更を加えた資材一覧' にデータ行がありません（最低1行必要）")

    # ─── 結果出力 ─────────────────────────────────────────────────────────
    if issues:
        print(f"[VERIFY NG] {len(issues)} 件の未充足:")
        for i, msg in enumerate(issues, 1):
            print(f"  {i}. {msg}")
        sys.exit(2)
    else:
        print("[VERIFY OK] 全必須枠が充足されています。")


TIMELINE_PHASES = ["調査", "対応方針", "実装方針", "実装前検証", "実装", "テスト", "最終検証", "リリース", "お客様確認"]


def main():
    parser = argparse.ArgumentParser(description="対応記録.xlsx を更新する")
    parser.add_argument("--folder",   required=True, help="保存先フォルダパス")
    parser.add_argument("--issue-id", required=True, dest="issue_id", help="課題ID (例: GF-327)")

    sub = parser.add_subparsers(dest="command", required=True)

    # タイムライン追加
    p_tl = sub.add_parser("timeline", help="タイムラインに行を追加する")
    p_tl.add_argument("--phase",   required=True, choices=TIMELINE_PHASES,
                      help="フェーズ名（固定値: " + "/".join(TIMELINE_PHASES) + "）")
    p_tl.add_argument("--source",  default="Claude", help="発生元 (例: Claude, ユーザ)")
    p_tl.add_argument("--content", required=True, help="内容・決定事項")
    p_tl.add_argument("--reason",  default="", help="変更・判断の理由（任意）")
    p_tl.add_argument("--force",   action="store_true", help="重複・既存値があっても上書きする")

    # セル直接更新（--label または --row で行を指定）
    p_cell = sub.add_parser("cell", help="特定セルを直接更新する")
    p_cell.add_argument("--sheet", required=True, help="シート名")
    p_cell.add_argument("--row",   type=int, default=None,
                        help="行番号（--label の代わりに使用）")
    p_cell.add_argument("--label", default="",
                        help="行ラベル（A列の部分一致で行を検索。--row より優先）")
    p_cell.add_argument("--col",   required=True, type=int, help="列番号")
    p_cell.add_argument("--value", required=True, help="書き込む値")
    p_cell.add_argument("--force", action="store_true", help="既存値があっても上書きする")

    # Before/After 追記
    p_ba = sub.add_parser("before-after", help="対応内容シートの Before/After セクションに変更前後を追記する")
    p_ba.add_argument("--file",   required=True, help="ファイルパス")
    p_ba.add_argument("--before", required=True, help="変更前コード / 設定値")
    p_ba.add_argument("--after",  required=True, help="変更後コード / 設定値")
    p_ba.add_argument("--force",  action="store_true")

    # 変更した資材一覧に1行追加
    p_cl = sub.add_parser("content-list", help="対応内容シートの変更を加えた資材一覧に1行追加する")
    p_cl.add_argument("--label",  required=True,
                      help="資材名（日本語表示名優先・API名括弧補助。例: 「preCheck 画面（preCheck）」）")
    p_cl.add_argument("--kind",   required=True,
                      help="変更種別（例: 新規追加 / 変更 / 削除）")
    p_cl.add_argument("--detail", required=True,
                      help="変更内容（人間が読める日本語で1〜2行）")

    # NG対応履歴に1行追加
    p_ng = sub.add_parser("ng-history", help="対応内容シートのNG対応履歴に1行追加する（/test NG 修正ループ記録）")
    p_ng.add_argument("--round",  required=True, help="回次（例: R1 / R2）")
    p_ng.add_argument("--tc",     required=True, help="TC番号（例: TC-003）")
    p_ng.add_argument("--reason", required=True, help="NG原因（機械判定のreason / 目視確認内容）")
    p_ng.add_argument("--fix",    required=True, help="修正内容（何をどう変えたか）")
    p_ng.add_argument("--force",  action="store_true", help="同一回次+TC が既にあっても上書きする")

    # implementation-summary.md から対応内容シートを一括記入
    p_cfm = sub.add_parser("content-from-md",
                            help="対応内容シートを implementation-summary.md から一括記入する（Phase 4 ハーネス直実行）")
    p_cfm.add_argument("--summary", required=True,
                       help="docs/logs/{issueID}/implementation-summary.md のパス")
    p_cfm.add_argument("--force", action="store_true", help="既存値があっても上書きする")

    # xlsx 全枠充足確認（read-back 検証）
    p_vfy = sub.add_parser("verify", help="xlsx の全必須枠が充足されているか確認する（exit 0=OK / exit 2=NG）")
    p_vfy.add_argument("--stage", required=True, choices=["pre-release", "final"],
                       help="pre-release: Phase5 前 / final: Phase6 末")
    p_vfy.add_argument("--status-expected", default="", dest="status_expected",
                       help="final ステージで期待するステータス値（例: 完了）")

    args = parser.parse_args()
    args.folder = validate_folder(args.folder)

    xlsx_path = os.path.join(args.folder, f"{args.issue_id}_対応記録.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"[ERROR] ファイルが見つかりません: {xlsx_path}")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        print(f"[ERROR] xlsx の読み込みに失敗しました: {xlsx_path}\n{e}")
        sys.exit(1)

    if args.command == "timeline":
        cmd_timeline(args, wb)
    elif args.command == "cell":
        cmd_cell(args, wb)
    elif args.command == "before-after":
        cmd_before_after(args, wb)
    elif args.command == "content-list":
        cmd_content_list(args, wb)
    elif args.command == "ng-history":
        cmd_ng_history(args, wb)
    elif args.command == "content-from-md":
        cmd_content_from_md(args, wb)
    elif args.command == "verify":
        # verify は wb を読み取るだけで保存不要。sys.exit は cmd_verify 内で行う
        cmd_verify(args, wb)
        return  # exit 0 相当（上で sys.exit(2) されなかった場合）

    try:
        wb.save(xlsx_path)
    except PermissionError as e:
        print(f"[ERROR] xlsx の保存に失敗しました（ファイルが開かれている可能性があります）: {xlsx_path}\n{e}")
        sys.exit(1)
    print(f"保存完了: {xlsx_path}")


if __name__ == "__main__":
    main()
