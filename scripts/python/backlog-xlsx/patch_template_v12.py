# -*- coding: utf-8 -*-
"""patch_template_v12.py
対応記録テンプレート.xlsx を Group P 仕様に移行する。

変更内容:
  1. テスト・検証シート: C 列「実行種別」のヘッダ・データ行フォントを「游ゴシック」に統一
     - v11 のコピー方式はコピー元が MS P ゴシックの場合に連鎖するため、明示設定に変更

冪等性: C 列ヘッダ・データ行のフォント名が既に「游ゴシック」なら skip する。

Usage:
    python patch_template_v12.py
"""

import sys
import copy
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"

_YAGOTHIC = "游ゴシック"
_HDR_FILL_RGB = "002E75B6"
_DATA_FILL_ODD  = "00FFFFFF"
_DATA_FILL_EVEN = "00EAF2F8"
_THIN_SIDE = Side(border_style="thin", color="00B4C6E7")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)
_HEADER_FONT  = Font(name=_YAGOTHIC, bold=True, color="FFFFFFFF", size=11)
_DATA_FONT    = Font(name=_YAGOTHIC, size=11)
_HEADER_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")
_DATA_ALIGN   = Alignment(wrap_text=True, vertical="top")


def _find_header_row(ws):
    """■ テストテーブル が入っている行を返す（見つからない場合は None）。"""
    for r in range(1, min(20, ws.max_row + 1)):
        v = ws.cell(r, 1).value
        if v and "テストテーブル" in str(v):
            return r
    return None


def _get_font_name(cell):
    return (cell.font.name or "") if cell.font else ""


def patch_test_sheet_font(ws):
    """テスト・検証シート C 列のフォントを游ゴシックに統一する。"""
    hdr_row = _find_header_row(ws)
    col_hdr_row = (hdr_row + 1) if hdr_row else 6

    # 冪等チェック: ヘッダ行 + データ行の C 列が全て游ゴシックなら skip
    hdr_ok = _get_font_name(ws.cell(col_hdr_row, 3)) == _YAGOTHIC
    data_ok = all(
        _get_font_name(ws.cell(r, 3)) == _YAGOTHIC
        for r in range(col_hdr_row + 1, ws.max_row + 1)
        if ws.cell(r, 3).value is not None
    )
    if hdr_ok and data_ok:
        print("[SKIP] テスト・検証: C 列フォントは既に「游ゴシック」 (ヘッダ・データ行とも) → スキップ")
        return

    if str(ws.cell(col_hdr_row, 3).value or "").strip() != "実行種別":
        print("[WARN] テスト・検証: C 列ヘッダが「実行種別」でない → patch_template_v11 未適用の可能性")
        return

    # 列ヘッダ行 C 列
    hdr_cell = ws.cell(col_hdr_row, 3)
    hdr_cell.font = _HEADER_FONT

    # データ行（col_hdr_row + 1 以降、値があるか余白行は 8 行分）
    data_start = col_hdr_row + 1
    data_end = ws.max_row
    for r in range(data_start, data_end + 1):
        cell = ws.cell(r, 3)
        cell.font = _DATA_FONT

    print(f"[OK  ] テスト・検証: C 列フォントを「游ゴシック」に設定"
          f"（列ヘッダ r{col_hdr_row}・データ r{data_start}〜r{data_end}）")


def _v(ok, msg):
    prefix = "[OK  ]" if ok else "[FAIL]"
    print(f"  {prefix} {msg}")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)
    print("=== patch_template_v12: Group P C列フォント游ゴシック統一 ===")
    print(f"現在のシート: {wb.sheetnames}\n")

    ws_name = "テスト・検証"
    if ws_name in wb.sheetnames:
        patch_test_sheet_font(wb[ws_name])
    else:
        print(f"[WARN] {ws_name} シートが見つかりません")
    print()

    try:
        wb.save(TEMPLATE)
        print(f"[OK  ] テンプレート保存完了: {TEMPLATE}")
    except Exception as e:
        print(f"[ERROR] 保存失敗: {e}")
        sys.exit(1)

    # ── 自動検証 ──────────────────────────────────────────────────────────────
    print("\n=== 自動検証 ===")
    wb2 = load_workbook(TEMPLATE)
    ok = True

    if ws_name in wb2.sheetnames:
        ws2 = wb2[ws_name]
        hdr_row = _find_header_row(ws2)
        col_hdr_row = (hdr_row + 1) if hdr_row else 6

        font_name = _get_font_name(ws2.cell(col_hdr_row, 3))
        ok1 = (font_name == _YAGOTHIC)
        _v(ok1, f"テスト・検証: r{col_hdr_row} C 列フォント = 「{font_name}」（期待: 游ゴシック）")
        if not ok1:
            ok = False

    print()
    print("[OK  ] 全検証 PASS" if ok else "[FAIL] 検証に失敗した項目があります")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
