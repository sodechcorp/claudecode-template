# -*- coding: utf-8 -*-
"""backlog-xlsx / judge_results.py
test-spec.md の期待結果と証跡ファイルの実際の結果を突き合わせ、
OK/NG を判定して対応記録.xlsx の H 列（実際の結果）を更新する。

Usage:
    python judge_results.py \\
      --folder /path/to/xlsx_folder \\
      --issue-id GF-350 \\
      --spec /path/to/test-spec.md \\
      --evidence-dir /path/to/evidence/after \\
      --out /path/to/judgment-result.json
"""

import argparse
import json
import os
import re
import subprocess
import sys
from pathlib import Path

from _common import validate_folder


# ── test-spec.md パーサ ───────────────────────────────────────────────────────

def parse_test_spec(spec_path: str) -> list:
    """test-spec.md のテーブルを dict リストとして返す。"""
    text = Path(spec_path).read_text(encoding="utf-8")
    headers = []
    rows = []
    in_table = False

    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            if in_table:
                break
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if all(re.match(r"^[-: ]+$", c) for c in cells):
            in_table = True
            continue
        if not headers:
            headers = cells
            in_table = True
        else:
            rows.append(dict(zip(headers, cells)))

    return rows


# ── 証跡ファイル探索 ────────────────────────────────────────────────────────

def find_evidence_files(evidence_dir: str, tc_no: str, shubetsu: str) -> list:
    """証跡ディレクトリから TC-001 に対応する全ファイルを返す（複数証跡・分岐ラベル対応）。"""
    # 種別別サブディレクトリ（複合種別 "AnonApex + SOQL" 等にも対応）
    subdir_map = {
        "SOQL": "soql",
        "ApexTest": "apex",
        "AnonApex": "apex",
        "UI": "screen",
        "メタ確認": "meta",
        "ファイル確認": "meta",
    }
    # " + " で分割して各サブディレクトリを収集（重複なし・順序維持）
    subdirs_ordered = []
    seen_subdirs: set = set()
    for part in re.split(r'\s*\+\s*', shubetsu):
        sd = subdir_map.get(part.strip(), "")
        if sd and sd not in seen_subdirs:
            seen_subdirs.add(sd)
            subdirs_ordered.append(sd)

    search_dirs = [os.path.join(evidence_dir, sd) for sd in subdirs_ordered]
    search_dirs.append(evidence_dir)

    found = []
    seen: set = set()
    for d in search_dirs:
        if not os.path.isdir(d):
            continue
        for fname in sorted(os.listdir(d)):
            # before / リサイズ済みサムネイルは対象外
            if "_before." in fname or "_resized." in fname:
                continue
            if fname.startswith(tc_no) or fname.startswith(tc_no.replace("TC-", "tc-")):
                fpath = os.path.join(d, fname)
                if fpath not in seen:
                    seen.add(fpath)
                    found.append(fpath)

    # after/ で見つからない場合は sibling の before/ も検索（Before 証跡ケース: TC-016 等）
    if not found:
        before_dir = os.path.join(os.path.dirname(os.path.abspath(evidence_dir)), "before")
        if os.path.isdir(before_dir):
            for fname in sorted(os.listdir(before_dir)):
                if "_resized." in fname:          # サムネイルは除外
                    continue
                if fname.startswith(tc_no) or fname.startswith(tc_no.replace("TC-", "tc-")):
                    fpath = os.path.join(before_dir, fname)
                    if fpath not in seen:
                        seen.add(fpath)
                        found.append(fpath)

    return found


def find_evidence_file(evidence_dir: str, tc_no: str, shubetsu: str) -> str:
    """後方互換: 最初の1ファイルのみ返す。"""
    files = find_evidence_files(evidence_dir, tc_no, shubetsu)
    return files[0] if files else ""


# ── 判定ロジック ─────────────────────────────────────────────────────────────

def _read_text_evidence(path: str) -> str:
    """UTF-16 LE 自動検出してテキストを読む。"""
    try:
        raw = Path(path).read_bytes()
        if raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
            return raw.decode("utf-16", errors="replace")
        elif len(raw) > 1 and raw[1] == 0x00:
            return raw.decode("utf-16-le", errors="replace")
        else:
            return raw.decode("utf-8", errors="replace")
    except Exception:
        return ""


def judge_single_evidence(evidence_path: str, kiki: str, judge_method: str, no: str) -> dict:
    """1証跡ファイルを判定し {"ok": bool|None, "actual": str, "reason": str} を返す。"""

    # スクショ（PNG）: DOM スナップショット (.txt) があれば内容照合、なければ存在判定
    if evidence_path.lower().endswith(".png"):
        size = os.path.getsize(evidence_path)
        if size < 1000:
            return {"ok": False, "actual": f"スクショあり ({size}B・小さすぎる)", "reason": "PNG が不正に小さい"}
        # 同名の .txt（DOMスナップショット）を探す
        snap_path = re.sub(r'\.png$', '.txt', evidence_path, flags=re.IGNORECASE)
        if os.path.exists(snap_path):
            snap = _read_text_evidence(snap_path)
            # 構造化証跡「判定: OK/NG」を最優先で参照
            m_verdict = re.search(r"^判定\s*:\s*(OK|NG)", snap, re.MULTILINE)
            if m_verdict:
                ok = m_verdict.group(1) == "OK"
                m_reason = re.search(r"^判定\s*:.+?[-—]\s*(.+)$", snap, re.MULTILINE)
                reason = m_reason.group(1).strip() if m_reason else ""
                # F-5: 期待結果がある場合は「実際の値:」セクション内で追加照合（採取側OK行を盲信しない）
                if ok and kiki:
                    m_snap_section = re.search(r"実際の値\s*[:：](.+?)(?=判定\s*[:：]|\Z)", snap, re.DOTALL)
                    if m_snap_section and kiki.lower() not in m_snap_section.group(1).lower():
                        ok = False
                        reason = f"採取側はOKとしているが期待値「{kiki[:30]}」が「実際の値:」セクションに見つかりません"
                actual_str = f"画面表示{'OK' if ok else 'NG'}（DOM照合済）" + (" — " + reason[:40] if reason else "")
                return {"ok": ok, "actual": actual_str, "reason": "" if ok else reason}
            # F-2/F-3: kiki による DOM 照合（「実際の値:」セクション優先スコープ）
            if kiki or "含まない" in judge_method or "非表示" in judge_method or "なし確認" in judge_method:
                m_snap_section = re.search(r"実際の値\s*[:：](.+?)(?=判定\s*[:：]|\Z)", snap, re.DOTALL)
                search_scope = m_snap_section.group(1) if m_snap_section else snap
                # F-3: 否定確認（PNG+DOM の場合も適用）
                if "含まない" in judge_method or "非表示" in judge_method or "なし確認" in judge_method:
                    ok = kiki.lower() not in search_scope.lower() if kiki else True
                    actual_str = f"画面表示{'OK' if ok else 'NG'}（DOM照合済）— 「{kiki[:20]}」{'なし(OK)' if ok else 'あり(NG)'}"
                    reason = "" if ok else f"「{kiki[:30]}」が DOM に残存（非表示のはずが表示されている）"
                    return {"ok": ok, "actual": actual_str, "reason": reason}
                ok = kiki.lower() in search_scope.lower() if kiki else True
                actual_str = f"画面表示{'OK' if ok else 'NG'}（DOM照合済）— 「{kiki[:20]}」{'あり' if ok else 'なし'}"
                reason = "" if ok else f"DOM に「{kiki[:30]}」が含まれない（DOM照合失敗）"
                return {"ok": ok, "actual": actual_str, "reason": reason}
            return {"ok": True, "actual": "画面表示OK（DOM照合済）", "reason": ""}
        # F-1: DOM スナップショットなし → 観点の自動確認不可（要目視）。ok: None = SKIP 扱い
        return {"ok": None, "actual": "スクショ取得済（DOM未取得・要目視確認）",
                "reason": "DOM スナップショット（.txt）が採取されていません。ui-evidence-runner の return 値が .txt に Write されているか確認してください"}

    # テキスト証跡（SOQL/Apex ログ / DOM スナップショット .txt）
    content = _read_text_evidence(evidence_path)

    # 構造化証跡（auto-evidence-runner 生成）: 「判定: OK/NG —」行を最優先参照
    m_verdict = re.search(r"^判定\s*:\s*(OK|NG)", content, re.MULTILINE)
    if m_verdict:
        ok = m_verdict.group(1) == "OK"
        m_reason = re.search(r"^判定\s*:.+?—\s*(.+)$", content, re.MULTILINE)
        reason = m_reason.group(1).strip() if m_reason else ""
        # F-5: 期待結果がある場合は「実際の値:」セクション内で追加照合（採取側OK行を盲信しない）
        if ok and kiki:
            m_actual_section = re.search(r"実際の値\s*[:：](.+?)(?=判定\s*[:：]|\Z)", content, re.DOTALL)
            if m_actual_section and kiki.lower() not in m_actual_section.group(1).lower():
                ok = False
                reason = f"採取側はOKとしているが期待値「{kiki[:30]}」が「実際の値:」セクションに見つかりません"
        actual_str = "OK — " + reason if ok else "NG — " + reason
        return {"ok": ok, "actual": actual_str, "reason": "" if ok else reason}

    # 件数一致判定 (期待結果が "N 件" 形式)
    m_expected_count = re.search(r"(\d+)\s*件", kiki)
    m_actual_count = re.search(r"件数\s*:\s*(\d+)\s*件", content)
    # sf CLI の "Total number of records retrieved: N." 形式にも対応
    if not m_actual_count:
        m_actual_count = re.search(r"Total number of records retrieved:\s*(\d+)", content, re.IGNORECASE)
    # kiki に "N件" がなくても "完全一致" の judge_method で1件以上取得できていれば OK とみなす
    if not m_expected_count and m_actual_count and ("完全一致" in judge_method or "件数一致" in judge_method):
        act = int(m_actual_count.group(1))
        if act > 0:
            return {"ok": True, "actual": f"SOQL {act} 件取得", "reason": ""}
        return {"ok": False, "actual": "SOQL 0件", "reason": "対象レコードが見つかりません"}
    if m_expected_count and m_actual_count:
        exp = int(m_expected_count.group(1))
        act = int(m_actual_count.group(1))
        # F-4: 件数デフォルトは完全一致。spec に「以上」と明記した場合のみ >= に緩和
        ok = (act >= exp) if "以上" in judge_method else (exp == act)
        actual_str = f"{act} 件"
        reason = "" if ok else f"期待 {exp} 件 / 実際 {act} 件"
        return {"ok": ok, "actual": actual_str, "reason": reason}

    # F-3: 否定確認（含まない/非表示/なし確認）: 期待文字列が証跡に存在しないことを確認
    if "含まない" in judge_method or "非表示" in judge_method or "なし確認" in judge_method:
        m_actual_section = re.search(r"実際の値\s*[:：](.+?)(?=判定\s*[:：]|\Z)", content, re.DOTALL)
        search_scope = m_actual_section.group(1) if m_actual_section else content
        ok = kiki.lower() not in search_scope.lower() if kiki else True
        actual_str = f"「{kiki[:30]}」{'あり（NG）' if not ok else 'なし（OK）'}"
        return {"ok": ok, "actual": actual_str,
                "reason": "" if ok else f"「{kiki[:30]}」が証跡に残存（非表示のはずが表示されている）"}

    # 含む判定 (期待結果に含まれるべき文字列): 「実際の値:」行以降のみを検索し期待値行の誤ヒットを防ぐ
    if "含む" in judge_method or "存在" in judge_method:
        m_actual_section = re.search(r"実際の値\s*[:：](.+?)(?=判定\s*[:：]|\Z)", content, re.DOTALL)
        search_scope = m_actual_section.group(1) if m_actual_section else content
        ok = kiki.lower() in search_scope.lower() if kiki else True
        actual_str = f"「{kiki[:30]}」{'あり' if ok else 'なし'}"
        return {"ok": ok, "actual": actual_str, "reason": "" if ok else f"「{kiki[:30]}」が証跡に含まれない"}

    # ApexTest 成功確認: "Pass Rate" / "Outcome: Passed" を正として "FAIL" の誤ヒットを避ける
    if re.search(r"Pass\s+Rate\s*[:\|]\s*100%", content) or re.search(r"Outcome\s*:\s*Passed", content):
        return {"ok": True, "actual": "Apex テスト PASS", "reason": ""}
    if re.search(r"Pass\s+Rate\s*[:\|]\s*0%", content) or re.search(r"Outcome\s*:\s*Failed", content):
        m_fail = re.search(r"(Failures:.+)", content)
        reason = m_fail.group(1) if m_fail else "Apex テスト FAIL"
        return {"ok": False, "actual": "Apex テスト FAIL", "reason": reason}
    # sf apex run test の表形式（TEST NAME / OUTCOME 列）: 全件 Pass 判定
    if re.search(r"\bTEST\s+NAME\b", content) and re.search(r"\bOUTCOME\b", content):
        outcomes = re.findall(r'\.\w+\s+(Pass|Fail|Error|Skip)\b', content, re.IGNORECASE)
        if outcomes and all(o.lower() == "pass" for o in outcomes):
            return {"ok": True, "actual": f"Apex テスト PASS ({len(outcomes)} 件全件)", "reason": ""}
        if outcomes and any(o.lower() in ("fail", "error") for o in outcomes):
            fail_n = sum(1 for o in outcomes if o.lower() in ("fail", "error"))
            return {"ok": False, "actual": f"Apex テスト FAIL ({fail_n}/{len(outcomes)} 件)", "reason": "テスト失敗あり"}
    # Anonymous Apex 実行成功: "Executed successfully." を正として判定
    if re.search(r"Executed successfully\.", content, re.IGNORECASE):
        if not re.search(r"(Error:|FATAL_ERROR|System\.\w+Exception)", content):
            return {"ok": True, "actual": "AnonApex 実行成功", "reason": ""}
    if re.search(r"(FATAL_ERROR|System\.\w+Exception)", content):
        m_err = re.search(r"((?:FATAL_ERROR|System\.\w+Exception).{0,80})", content)
        reason = m_err.group(1)[:80] if m_err else "AnonApex 実行エラー"
        return {"ok": False, "actual": "AnonApex 実行エラー", "reason": reason}
    # Fallback: 旧パターン（sf CLI raw 出力等）
    if "success" in content.lower() and re.search(r"\bPASS\b|\bPassed\b", content):
        return {"ok": True, "actual": "Apex テスト PASS", "reason": ""}
    if re.search(r"\bFailed\b", content) or re.search(r"Fail\s+Rate\s*[:\|]\s*(?!0%)\d", content):
        m_fail = re.search(r"(Failures:.+)", content)
        reason = m_fail.group(1) if m_fail else "Apex テスト FAIL"
        return {"ok": False, "actual": "Apex テスト FAIL", "reason": reason}

    # デフォルト: 判定パターン未一致は「要確認」（NG扱い）— 証跡があるだけで OK にしない
    return {"ok": False, "actual": "証跡あり（判定パターン未一致）",
            "reason": "判定方法を機械可読な値（含む/件数一致/完全一致/含まない/前後比較等）にしてください",
            "ng_type": "要確認"}


def _judge_transition(tc: dict, after_txts: list, evidence_dir: str) -> dict:
    """F-7: 状態遷移前後比較判定（判定方法に「前後比較」を含むケース専用）。
    before/{No}_*.txt と after/{soql|apex|screen}/{No}_*.txt を突き合わせる。
    期待結果の形式: 「before:初期値 / after:変更後値」（例: before:未送信 / after:送信済）"""
    no = tc.get("No", "")
    kiki = tc.get("期待結果", "").strip()

    # 期待結果を before: / after: で分解
    m_before_exp = re.search(r"(?:before|変更前)\s*[:：]\s*(.+?)(?:\s*/\s*(?:after|変更後)\s*[:：]|$)",
                             kiki, re.IGNORECASE)
    m_after_exp  = re.search(r"(?:after|変更後)\s*[:：]\s*(.+)", kiki, re.IGNORECASE)
    exp_before = m_before_exp.group(1).strip() if m_before_exp else ""
    exp_after  = m_after_exp.group(1).strip()  if m_after_exp  else kiki

    # before .txt を収集（before/ ディレクトリ）
    before_dir = os.path.join(os.path.dirname(os.path.abspath(evidence_dir)), "before")
    before_txts = []
    if os.path.isdir(before_dir):
        for fname in sorted(os.listdir(before_dir)):
            if "_resized." in fname:
                continue
            if (fname.startswith(no) or fname.startswith(no.replace("TC-", "tc-"))) \
                    and fname.lower().endswith(".txt"):
                before_txts.append(os.path.join(before_dir, fname))

    # after DOM .txt を確認
    after_ok = False
    after_actual = ""
    for fpath in after_txts:
        content = _read_text_evidence(fpath)
        m_sec = re.search(r"実際の値\s*[:：](.+?)(?=判定\s*[:：]|\Z)", content, re.DOTALL)
        scope = m_sec.group(1) if m_sec else content
        if exp_after and exp_after.lower() in scope.lower():
            after_ok = True
            after_actual = f"after:「{exp_after[:20]}」あり"
            break
    if not after_txts:
        after_actual = "after DOM 証跡なし"
    elif not after_ok:
        after_actual = f"after:「{exp_after[:20]}」なし"

    # before DOM .txt を確認（.txt が無い場合は PNG のみ＝参考のみ・判定は after のみで行う）
    before_ok = True
    before_actual = ""
    if before_txts and exp_before:
        before_ok = False
        for fpath in before_txts:
            content = _read_text_evidence(fpath)
            if exp_before.lower() in content.lower():
                before_ok = True
                before_actual = f"before:「{exp_before[:20]}」確認済"
                break
        if not before_ok:
            before_actual = f"before:「{exp_before[:20]}」なし"
    elif not before_txts:
        before_actual = "before DOM 未取得（PNG のみ・参考）"

    ok = after_ok and before_ok
    actual_parts = [p for p in [before_actual, after_actual] if p]
    actual = " / ".join(actual_parts) if actual_parts else "証跡不足"
    reason = ""
    if not after_ok:
        reason = f"after に「{exp_after[:30]}」が見つかりません（状態遷移が未確認）"
    elif not before_ok:
        reason = f"before に「{exp_before[:30]}」が見つかりません（初期状態が未確認）"
    return {"ok": ok, "actual": actual, "reason": reason}


def judge_case(tc: dict, evidence_path: str, evidence_dir: str = "") -> dict:
    """1テストケースを判定し {"ok": bool, "actual": str, "reason": str} を返す。
    複数証跡（分岐ラベル付き）がある場合は全証跡を AND 評価する。"""
    no = tc.get("No", "")
    kiki = tc.get("期待結果", "").strip()
    judge_method = tc.get("判定方法", "").strip()
    auto = tc.get("自動化可否", "自動").strip()
    shubetsu = tc.get("種別", tc.get("実行種別", "")).strip()

    # 要手動ケースは判定スキップ
    if "要手動" in auto:
        return {"ok": None, "actual": "要手動確認", "reason": "自動化不可・ユーザー手動確認"}

    # 複数証跡を収集（evidence_dir が渡されていれば全ファイルを探す）
    if evidence_dir:
        all_files = find_evidence_files(evidence_dir, no, shubetsu)
    elif evidence_path and os.path.exists(evidence_path):
        all_files = [evidence_path]
    else:
        all_files = []

    # .txt DOMスナップショットは PNG 判定の補助として使うため、単独では PNG のサブ証跡扱い
    # PNG の判定内で snap.txt を読むため、ここでは PNG のみを判定対象とし txt 単独は除外しない
    if not all_files:
        return {"ok": False, "actual": "", "reason": f"証跡ファイルが見つかりません (No: {no})", "ng_type": "未実行"}

    # F-7: 状態遷移前後比較（before/after DOM テキストを突き合わせる）
    if "前後比較" in judge_method and evidence_dir:
        after_txts = [f for f in all_files if f.lower().endswith(".txt")]
        return _judge_transition(tc, after_txts, evidence_dir)

    # PNG と txt を分ける: PNG がある場合は PNG で判定（内部で snap.txt を参照）
    # PNG が無く txt のみの場合は txt で判定
    png_files = [f for f in all_files if f.lower().endswith(".png")]
    txt_files = [f for f in all_files if f.lower().endswith(".txt")]

    # PNG + DOM ペア判定（UI 証跡の場合）
    judge_targets = png_files if png_files else txt_files

    results = []
    for fpath in judge_targets:
        r = judge_single_evidence(fpath, kiki, judge_method, no)
        results.append(r)

    if not results:
        return {"ok": False, "actual": "", "reason": f"判定可能な証跡ファイルがありません (No: {no})"}

    # AND 評価: 全分岐 OK で OK（NG > SKIP > OK の優先順位）
    ng_results   = [r for r in results if r.get("ok") is False]
    skip_results = [r for r in results if r.get("ok") is None]
    ok_results   = [r for r in results if r.get("ok") is True]

    if ng_results:
        # 最初の NG の理由を採用（ng_type も伝播）
        ng = ng_results[0]
        actuals = " / ".join(r["actual"] for r in results)
        return {"ok": False, "actual": actuals, "reason": ng["reason"], "ng_type": ng.get("ng_type", "")}

    if skip_results:
        # ok: None（DOM未取得・要目視等）は SKIP として伝播
        actuals = " / ".join(r["actual"] for r in skip_results)
        return {"ok": None, "actual": actuals, "reason": skip_results[0].get("reason", "")}

    # 全件 OK
    actuals = " / ".join(r["actual"] for r in ok_results)
    return {"ok": True, "actual": actuals, "reason": ""}


# ── 対応記録.xlsx H 列の追記 ─────────────────────────────────────────────────

def update_xlsx_h_col(folder: str, issue_id: str, tc_no: str, value: str) -> bool:
    """update_records.py cell コマンドで H 列を更新する。"""
    # 対象行番号を特定（実装後・UI手動以外）
    xlsx_path = os.path.join(folder, f"{issue_id}_対応記録.xlsx")
    if not os.path.exists(xlsx_path):
        return False

    find_row_code = f"""
import openpyxl, os
wb = openpyxl.load_workbook(r'{xlsx_path}')
ws = wb['テスト・検証']
for r in range(1, ws.max_row + 1):
    v = [ws.cell(r, c).value for c in range(1, 9)]
    if (any(v) and str(v[1] or '').strip() == '実装後'
            and str(v[2] or '').strip() != 'UI手動'
            and str(v[3] or '').strip().startswith('{tc_no}')):
        print(r)
        break
"""
    result = subprocess.run(
        ["python", "-c", find_row_code],
        capture_output=True, text=True
    )
    row_str = result.stdout.strip()
    if not row_str.isdigit():
        return False

    row_num = int(row_str)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    update_script = os.path.join(script_dir, "update_records.py")

    result = subprocess.run(
        ["python", update_script,
         "--folder", folder, "--issue-id", issue_id,
         "cell", "--sheet", "テスト・検証",
         "--row", str(row_num), "--col", "8",
         "--value", value],
        capture_output=True, text=True
    )
    return result.returncode == 0


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="テストケースの OK/NG 判定と xlsx H 列更新")
    parser.add_argument("--folder", required=True, help="xlsx 出力フォルダ")
    parser.add_argument("--issue-id", required=True, dest="issue_id")
    parser.add_argument("--spec", required=True, help="test-spec.md のパス")
    parser.add_argument("--evidence-dir", required=True, dest="evidence_dir",
                        help="証跡ファイルの after/ ディレクトリ")
    parser.add_argument("--out", default="", help="判定結果 JSON の出力パス（省略時は stdout）")
    parser.add_argument("--no-xlsx", action="store_true", dest="no_xlsx",
                        help="対応記録.xlsx の H 列更新をスキップする")
    parser.add_argument("--prev", default="", dest="prev_json",
                        help="前回判定 JSON（差分再実行時に前回 OK を流用する）")
    args = parser.parse_args()

    args.folder = validate_folder(args.folder)
    test_cases = parse_test_spec(args.spec)
    if not test_cases:
        print("[WARN] test-spec.md にテストケースが見つかりませんでした。")
        sys.exit(0)

    # 前回判定の読み込み（差分再実行用）
    prev_results = {}
    if args.prev_json and os.path.exists(args.prev_json):
        try:
            prev_data = json.loads(Path(args.prev_json).read_text(encoding="utf-8"))
            for r in prev_data.get("results", []):
                if r.get("status") == "OK":
                    prev_results[r["no"]] = r
            print(f"[INFO] 前回判定を読み込み: OK={len(prev_results)} 件を流用")
        except Exception as e:
            print(f"[WARN] 前回判定 JSON の読み込み失敗: {e}（全件再実行）")

    results = []
    ng_list = []
    skip_list = []

    for tc in test_cases:
        no = tc.get("No", "")
        shubetsu = tc.get("種別", tc.get("実行種別", "")).strip()

        # 差分再実行: 前回 OK の TC は流用
        if no in prev_results:
            prev = prev_results[no]
            results.append(prev)
            print(f"[REUSE] {no}: {tc.get('観点', '')} → 前回OK流用 ({prev.get('actual', '')})")
            continue

        evidence_path = find_evidence_file(args.evidence_dir, no, shubetsu)
        judgment = judge_case(tc, evidence_path, evidence_dir=args.evidence_dir)

        ok = judgment["ok"]
        actual = judgment["actual"]
        reason = judgment["reason"]
        ng_type = judgment.get("ng_type", "")

        if ok is None:
            status = "SKIP"
            skip_list.append(no)
            xlsx_value = "要手動確認"
        elif ok:
            status = "OK"
            xlsx_value = f"OK"
        else:
            status = "NG"
            ng_list.append({"no": no, "label": tc.get("観点", ""), "reason": reason, "ng_type": ng_type})
            xlsx_value = f"NG: {reason}" if reason else "NG"

        results.append({
            "no": no,
            "label": tc.get("観点", ""),
            "status": status,
            "actual": actual,
            "reason": reason,
            "ng_type": ng_type if status == "NG" else "",
            "evidence": evidence_path,
        })

        # xlsx H 列更新: 対応記録.xlsx のテスト・検証シートは廃止済みのためスキップ
        # （テストエビデンスはエビデンス.xlsx に集約。--no-xlsx フラグは互換維持のため残す）

        icon = {"OK": "[OK]", "NG": "[NG]", "SKIP": "[--]"}[status]
        print(f"{icon} {no}: {tc.get('観点', '')} → {actual}" + (f" ({reason})" if reason else ""))

    # サマリー
    ok_count = sum(1 for r in results if r["status"] == "OK")
    ng_count = len(ng_list)
    skip_count = len(skip_list)
    print(f"\n判定サマリー: OK={ok_count} / NG={ng_count} / 要手動={skip_count} / 合計={len(results)}")

    output = {
        "ok": ok_count,
        "ng": ng_count,
        "skip": skip_count,
        "total": len(results),
        "ng_list": ng_list,
        "skip_list": skip_list,
        "results": results,
    }

    if args.out:
        os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
        Path(args.out).write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[OK] 判定結果を保存: {args.out}")
    else:
        print(json.dumps(output, ensure_ascii=False, indent=2))

    # NG があれば exit 1
    if ng_count > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
