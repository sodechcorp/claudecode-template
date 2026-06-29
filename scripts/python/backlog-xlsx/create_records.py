# -*- coding: utf-8 -*-
"""backlog-xlsx / create_records.py
対応記録.xlsx を生成する（Phase 3 完了直後・メタ情報と4本文セクションを埋める）

Usage:
    python create_records.py \\
      --folder FOLDER --issue-id ID \\
      --investigation PATH \\
      [--approach-plan PATH]

  --approach-plan は任意。指定した場合、approach-plan.md の4固定見出し
  （課題の内容・詳細 / 原因・現状 / 対応方針（結論）/ 方針決定の経緯・根拠）
  の内容を LLM が書いた人間向け日本語のまま ①シートに転記する。

残りのセクションはハーネスが update_records.py で記入する:
  ② 対応内容: Phase 4 でハーネスが content-from-md で一括記入
              （implementer が implementation-summary.md を書き出し後、ハーネスが直接実行）
  タイムライン: 各Phaseエージェントが timeline で追記
  ステータス→完了: Phase 6 末でハーネスが cell で更新
"""

import argparse
import copy
import datetime
import os
import re
import sys
from pathlib import Path

from _common import validate_folder, _stripe_fill

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl がインストールされていません。`pip install openpyxl` を実行してください。")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"
WRAP = Alignment(wrap_text=True, vertical="top")


# ── MD パースユーティリティ ─────────────────────────────────────────────────

def read_md(path):
    if path and Path(path).exists():
        try:
            return Path(path).read_text(encoding="utf-8")
        except UnicodeDecodeError as e:
            print(f"[ERROR] ファイルのエンコーディングが UTF-8 ではありません: {path}\n{e}")
            sys.exit(1)
    return ""


def _flex_keyword(h):
    """見出しキーワード内の半角スペースを \\s* で吸収した正規表現パターンを返す。"""
    parts = re.split(r" +", h)
    return r"\s*".join(re.escape(p) for p in parts)


def extract_section(md, *headings):
    """指定見出し（## または ###）のセクション本文を返す。
    複数見出しは先にマッチしたものを使用。本文が空のセクションはスキップ。
    見出し揺れ吸収: 先頭「■ 」等の記号・末尾の「:」「テーブル」等・括弧付記を許容。
    """
    for h in headings:
        pat = (
            r"^#{1,3}[\s　]+"
            r"(?:[■●▶◆]\s*)?"
            + _flex_keyword(h) +
            r"(?:[・/／]\S+?)?"
            r"(?:テーブル|一覧|[:：])?"
            r"(?:\s*[（(][^)）]*[)）])?\s*$"
        )
        m = re.search(pat, md, re.MULTILINE)
        if m:
            start = m.end()
            rest = md[start:]
            level = len(re.match(r"^#+", md[m.start():]).group(0))
            end_pat = rf"^#{{1,{level}}}\s"
            end_m = re.search(end_pat, rest, re.MULTILINE)
            body = rest[: end_m.start()] if end_m else rest
            stripped = body.strip()
            if stripped:
                return stripped
    return ""


def extract_section_fallback(approach_md, inv_md, primary, *inv_headings):
    """approach-plan.md から一次抽出、なければ investigation.md から二次抽出する。

    Args:
        approach_md: approach-plan.md の本文（空文字 or None 可）
        inv_md: investigation.md の本文
        primary: approach-plan.md で検索する見出しキーワード
        *inv_headings: investigation.md で検索するフォールバック見出し（複数可）
    Returns:
        見つかった本文（strip 済み）、どこにも無ければ空文字
    """
    text = extract_section(approach_md, primary) if approach_md else ""
    if not text and inv_md and inv_headings:
        text = extract_section(inv_md, *inv_headings)
    return text


def extract_metadata(md, key):
    """key: value 形式の値を取る。半角・全角コロン両対応。"""
    pat = rf"^\s*(?:[-*+]\s+)?(?:\*\*)?{re.escape(key)}(?:\*\*)?\s*[:|：]\s*(.+?)\s*$"
    m = re.search(pat, md, re.MULTILINE)
    if m:
        return re.sub(r"\*\*\s*$", "", m.group(1)).strip()
    return ""


# ── 行検索ユーティリティ ──────────────────────────────────────────────────────

def find_header_row(ws, candidates):
    """A 列を走査して candidates のいずれかに一致する行番号を返す。見つからなければ None。"""
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and any(str(c) in str(cell.value) for c in candidates):
            return cell.row
    return None


def find_label_row(ws, label, col=1, start_row=1):
    """col 列で label に部分一致する行を返す。"""
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v and label in str(v).strip():
            return r
    return None


# ── 行スタイルコピー / 挿入ユーティリティ ────────────────────────────────────

def copy_row_style(ws, src_row, dst_row, max_col=6):
    """src_row の書式を dst_row にコピーする（insert_rows 後のスタイル継承用）。"""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst._style = copy.copy(src._style)
            dst.alignment = WRAP


def insert_rows_with_format(ws, insert_at, count, source_row, max_col):
    """insert_rows + 行高継承 + マージ補修を一括で行う (openpyxl の既知バグを回避)。"""
    all_merges = [(m.min_row, m.max_row, m.min_col, m.max_col)
                  for m in list(ws.merged_cells.ranges)]
    row_heights = {r: ws.row_dimensions[r].height
                   for r in ws.row_dimensions
                   if ws.row_dimensions[r].height is not None}
    src_h = row_heights.get(source_row)

    for mcr in list(ws.merged_cells.ranges):
        ws.merged_cells.ranges.discard(mcr)

    ws.insert_rows(insert_at, amount=count)

    for r in row_heights:
        if r >= insert_at:
            ws.row_dimensions[r].height = None
    for r, h in row_heights.items():
        new_r = r + count if r >= insert_at else r
        ws.row_dimensions[new_r].height = h

    for r in range(insert_at, insert_at + count):
        if src_h:
            ws.row_dimensions[r].height = src_h
        copy_row_style(ws, source_row, r, max_col=max_col)

    for (min_r, max_r, min_c, max_c) in all_merges:
        if min_r >= insert_at:
            ws.merge_cells(start_row=min_r + count, end_row=max_r + count,
                           start_column=min_c, end_column=max_c)
        elif max_r >= insert_at:
            ws.merge_cells(start_row=min_r, end_row=max_r + count,
                           start_column=min_c, end_column=max_c)
        else:
            ws.merge_cells(start_row=min_r, end_row=max_r,
                           start_column=min_c, end_column=max_c)


# ── セル書き込みユーティリティ ──────────────────────────────────────────────

def wset(ws, row, col, value, stripe=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = WRAP
    if stripe is not None:
        cell.fill = stripe
    return cell


# ── 行高自動調整ヘルパー ─────────────────────────────────────────────────────

def _get_merged_width(ws, row, col):
    """row,col が属する merge 範囲の全列幅合計を返す。merge なしなら単独列幅。"""
    for m in ws.merged_cells.ranges:
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
            return sum(
                ws.column_dimensions[get_column_letter(c)].width or 8
                for c in range(m.min_col, m.max_col + 1)
            )
    return ws.column_dimensions[get_column_letter(col)].width or 8


def _calc_row_height(text, width_chars, line_height=20, padding=4,
                     min_height=28, max_height=None):
    """テキストの折り返し行数を概算して row.height を返す。"""
    if not text or not str(text).strip():
        return min_height
    text = str(text)
    chars_per_row = max(int(width_chars * 2), 4)
    total_lines = 0
    for line in text.split("\n"):
        if not line.strip():
            total_lines += 1
            continue
        visual_width = sum(2 if ord(c) > 127 else 1 for c in line)
        total_lines += max(1, (visual_width + chars_per_row - 1) // chars_per_row)
    h = max(min_height, total_lines * line_height + padding)
    return min(h, max_height) if max_height else h


def auto_fit_row(ws, row_idx, target_cols=None, min_height=28, max_height=None):
    """行 row_idx の各セル値から折り返し行数を概算し row.height を設定。"""
    if target_cols is None:
        target_cols = list(range(1, ws.max_column + 1))
    max_h = min_height
    for col in target_cols:
        cell = ws.cell(row=row_idx, column=col)
        if cell.value:
            width = _get_merged_width(ws, row_idx, col)
            h = _calc_row_height(cell.value, width, min_height=min_height,
                                  max_height=max_height)
            if h > max_h:
                max_h = h
    ws.row_dimensions[row_idx].height = min(max_h, max_height) if max_height else max_h


# ── border 補完ユーティリティ ────────────────────────────────────────────────

_THIN_SIDE = Side(border_style="thin", color="00B4C6E7")


def _ensure_borders(ws):
    """値ありセル / マージ範囲のセルに枠線が無い箇所を thin border で補う。
    既存の太線・色付き border は保持し、欠けている辺だけ thin で埋める。
    """
    merged_coords = set()
    for mcr in ws.merged_cells.ranges:
        for r in range(mcr.min_row, mcr.max_row + 1):
            for c in range(mcr.min_col, mcr.max_col + 1):
                merged_coords.add((r, c))

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if cell.value is None and (r, c) not in merged_coords:
                continue
            b = cell.border
            need_left   = not (b.left   and b.left.style)
            need_right  = not (b.right  and b.right.style)
            need_top    = not (b.top    and b.top.style)
            need_bottom = not (b.bottom and b.bottom.style)
            if not (need_left or need_right or need_top or need_bottom):
                continue
            cell.border = Border(
                left=  (_THIN_SIDE if need_left   else b.left),
                right= (_THIN_SIDE if need_right  else b.right),
                top=   (_THIN_SIDE if need_top    else b.top),
                bottom=(_THIN_SIDE if need_bottom else b.bottom),
            )


# ── ① 課題と対応方針 シート 埋め ────────────────────────────────────────────

def fill_header(ws, args, inv_md, approach_md):
    """① 課題と対応方針 シートのメタ情報と4本文セクションを埋める。

    埋めるセル:
      メタ情報: 課題ID / 件名 / 優先度・期限 / 種別 / ステータス=対応中
      4本文: 課題の内容・詳細 / 原因・現状 / 対応方針（結論）/ 方針決定の経緯・根拠
             一次: approach-plan.md の固定見出しから転記
             二次: approach-plan.md が薄い/未生成なら investigation.md からフォールバック
             どちらにも無い場合はプレースホルダを書く（黙った空欄を作らない）

    Returns:
        list[str]: プレースホルダになった枠のラベル一覧（空なら全枠記入済み）

    残りはハーネスが update_records.py で記入する（Phase 4 content-from-md, Phase 6 cell）。
    """
    issue_id   = args.issue_id
    title      = extract_metadata(inv_md, "件名") or extract_metadata(inv_md, "タイトル") or ""
    priority   = extract_metadata(inv_md, "優先度") or ""
    deadline   = extract_metadata(inv_md, "期限") or ""
    issue_type = (extract_metadata(inv_md, "種別") or
                  extract_metadata(inv_md, "課題種別") or "")

    priority_str = ""
    if priority or deadline:
        priority_str = f"優先度: {priority}" + (f" / 期限: {deadline}" if deadline else "")

    # メタ情報をラベル検索で書き込む
    meta_map = [
        ("課題ID",    issue_id),
        ("件名",      title),
        ("優先度・期限", priority_str),
        ("種別",      issue_type),
        ("ステータス", "対応中"),
    ]
    for label, value in meta_map:
        row = find_label_row(ws, label)
        if row is not None:
            wset(ws, row, 2, value)
            auto_fit_row(ws, row, max_height=40)

    # 4本文: approach-plan.md から一次抽出、薄い/未生成なら investigation.md からフォールバック
    # planner A-3 規約の4固定見出しを一次として探し、なければ investigation.md 相当セクションで補う。
    # どちらにも無い場合はプレースホルダを書いて「黙った空欄」を作らない。
    PLACEHOLDER = "（未記入：approach-plan.md 未生成のため要追記）"
    # (ラベル, approach 見出し, inv フォールバック見出し...)
    text_sections = [
        ("課題の内容・詳細",     "課題の内容・詳細",     "課題の概要", "課題サマリー", "要件理解"),
        ("原因・現状",           "原因・現状",           "原因仮説", "前提・背景"),
        ("対応方針（結論）",     "対応方針（結論）"),
        ("方針決定の経緯・根拠", "方針決定の経緯・根拠"),
    ]
    unfilled = []
    for entry in text_sections:
        label         = entry[0]
        appr_heading  = entry[1]
        inv_fallbacks = entry[2:]
        text = extract_section_fallback(approach_md, inv_md, appr_heading, *inv_fallbacks)
        if not text:
            text = PLACEHOLDER
            unfilled.append(label)
        row = find_label_row(ws, label)
        if row is not None:
            wset(ws, row, 2, text)
            auto_fit_row(ws, row, max_height=300)

    # 後処理: 全セル行高・枠線補完
    for r in range(1, ws.max_row + 1):
        h = ws.row_dimensions[r].height
        has_content = any(ws.cell(r, c).value for c in range(1, ws.max_column + 1))
        if has_content and (h is None or h < 18):
            auto_fit_row(ws, r, min_height=20)

    return unfilled


# ── main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="対応記録.xlsx を生成する（Phase 3 完了後・メタ情報＋4本文のみ埋め版）"
    )
    parser.add_argument("--folder",         required=True)
    parser.add_argument("--issue-id",       required=True, dest="issue_id")
    parser.add_argument("--investigation",  required=True, dest="investigation",
                        help="docs/logs/{issueID}/investigation.md のパス")
    parser.add_argument("--approach-plan",  default="", dest="approach_plan",
                        help="docs/logs/{issueID}/approach-plan.md のパス（任意）")
    args = parser.parse_args()
    args.folder = validate_folder(args.folder)

    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        print("       python build_template.py を実行してテンプレートを生成してください。")
        sys.exit(1)

    inv_md = read_md(args.investigation)
    if not inv_md:
        print(f"[ERROR] investigation.md が見つかりません: {args.investigation}")
        sys.exit(1)

    approach_md = read_md(args.approach_plan) if args.approach_plan else ""

    os.makedirs(args.folder, exist_ok=True)
    try:
        wb = load_workbook(TEMPLATE)
    except Exception as e:
        print(f"[ERROR] テンプレートファイルの読み込みに失敗しました: {TEMPLATE}\n{e}")
        sys.exit(1)

    sheet_name = "課題と対応方針"
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] テンプレートにシート '{sheet_name}' がありません。"
              f"利用可能: {wb.sheetnames}")
        print("       python build_template.py を実行してテンプレートを再生成してください。")
        sys.exit(1)

    unfilled = fill_header(wb[sheet_name], args, inv_md, approach_md)

    # 後処理: 枠線補完
    for ws in wb.worksheets:
        _ensure_borders(ws)

    path = os.path.join(args.folder, f"{args.issue_id}_対応記録.xlsx")
    try:
        wb.save(path)
        print(f"[OK] 生成完了: {path}")
    except PermissionError as e:
        print(f"[ERROR] xlsx の保存に失敗しました（ファイルが開かれている可能性があります）: {path}\n{e}")
        sys.exit(1)

    if unfilled:
        print(f"  シート ①「課題と対応方針」: メタ情報＋4本文"
              f"（プレースホルダ {len(unfilled)} 枠: {', '.join(unfilled)}）")
        print(f"  ※ プレースホルダ枠は approach-plan.md 生成後に create_records.py を再実行するか、xlsx を直接更新してください。")
    else:
        print(f"  シート ①「課題と対応方針」: メタ情報＋4本文（全枠記入済み）")
    print(f"  シート ②「対応内容」: implementation-summary.md 書き出し後に Phase 4 ハーネスが update_records.py で記入")


if __name__ == "__main__":
    main()
