# -*- coding: utf-8 -*-
"""patch_template_v11.py
対応記録テンプレート.xlsx を Group O 仕様に移行する。

変更内容:
  1. テスト・検証シート: C 列に「実行種別」を挿入（7 列 → 8 列）
     - 全横幅マージ（A:G）を A:H に拡張
     - 列ヘッダ行 C 列に「実行種別」を設定
     - 列幅 C = 14 を設定

冪等性: C 列ヘッダが既に「実行種別」なら skip する。

Usage:
    python patch_template_v11.py
"""

import sys
import copy
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"

_HDR_FILL_RGB   = "002E75B6"  # 列ヘッダ（中青）
_THIN_SIDE = Side(border_style="thin", color="00B4C6E7")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE, bottom=_THIN_SIDE,
)
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")


def _save_merges(ws):
    return [(m.min_row, m.max_row, m.min_col, m.max_col)
            for m in list(ws.merged_cells.ranges)]


def _clear_merges(ws):
    for mcr in list(ws.merged_cells.ranges):
        ws.merged_cells.ranges.discard(mcr)


def _find_header_row(ws):
    """■ テストテーブル が入っている行を返す（見つからない場合は None）。"""
    for r in range(1, min(20, ws.max_row + 1)):
        v = ws.cell(r, 1).value
        if v and "テストテーブル" in str(v):
            return r
    return None


def patch_test_sheet(ws):
    """テスト・検証シートに C 列「実行種別」を挿入する（7 列 → 8 列）。"""

    # 冪等チェック: C 列ヘッダが既に「実行種別」かどうか
    hdr_row = _find_header_row(ws)
    if hdr_row is None:
        # フォールバック: r6 をヘッダ行と仮定
        hdr_row = 5
    col_hdr_row = hdr_row + 1  # ■テストテーブル の直下が列ヘッダ

    if str(ws.cell(col_hdr_row, 3).value or "").strip() == "実行種別":
        print("[SKIP] テスト・検証: C 列ヘッダが既に「実行種別」 → スキップ")
        return

    # マージ保存 → 全解除 → 列挿入 → マージ再構築
    merges_snap = _save_merges(ws)
    _clear_merges(ws)

    ws.insert_cols(3, 1)  # 列 C に 1 列挿入（旧 C〜G が 1 つ右へ）

    # row_dimensions は列操作で影響しないのでシフト不要
    # column_dimensions を C 以降に shift
    new_col_dims = {}
    for col_letter, dim in list(ws.column_dimensions.items()):
        from openpyxl.utils import column_index_from_string, get_column_letter
        idx = column_index_from_string(col_letter)
        if idx >= 3:
            new_col_dims[get_column_letter(idx + 1)] = dim
        else:
            new_col_dims[col_letter] = dim
    ws.column_dimensions.clear()
    for k, v in new_col_dims.items():
        ws.column_dimensions[k] = v

    # マージ再構築（min_col/max_col が 3 以上なら +1）
    for (min_r, max_r, min_c, max_c) in merges_snap:
        adj_min_c = min_c + 1 if min_c >= 3 else min_c
        adj_max_c = max_c + 1 if max_c >= 3 else max_c
        if adj_min_c > adj_max_c:
            continue
        try:
            ws.merge_cells(start_row=min_r, end_row=max_r,
                           start_column=adj_min_c, end_column=adj_max_c)
        except Exception:
            pass

    # 列ヘッダ行 C 列に「実行種別」を設定
    # スタイルは隣接する列ヘッダ（B 列）からコピー
    src_cell = ws.cell(col_hdr_row, 2)  # B 列ヘッダ（タイミング/区分）
    dst_cell = ws.cell(col_hdr_row, 3)  # 新 C 列ヘッダ

    dst_cell.value = "実行種別"
    dst_cell.font = copy.copy(src_cell.font) if src_cell.font else _HEADER_FONT
    dst_cell.alignment = copy.copy(src_cell.alignment) if src_cell.alignment else _HEADER_ALIGN
    dst_cell.fill = copy.copy(src_cell.fill) if src_cell.fill else PatternFill("solid", fgColor=_HDR_FILL_RGB)
    dst_cell.border = copy.copy(src_cell.border) if src_cell.border else _THIN_BORDER

    # 列幅: C 列を 14 に設定
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(3)].width = 14

    print(f"[OK  ] テスト・検証: C 列「実行種別」挿入（列ヘッダ行 r{col_hdr_row}）")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)
    print("=== patch_template_v11: Group O テスト列追加 ===")
    print(f"現在のシート: {wb.sheetnames}\n")

    ws_name = "テスト・検証"
    if ws_name in wb.sheetnames:
        patch_test_sheet(wb[ws_name])
    else:
        print(f"[WARN] {ws_name} シートが見つかりません")
    print()

    try:
        wb.save(TEMPLATE)
        print(f"[OK  ] テンプレート保存完了: {TEMPLATE}")
    except Exception as e:
        print(f"[ERROR] 保存失敗: {e}")
        sys.exit(1)

    # ── 自動検証 ──────────────────────────────────────────────────────────────
    print("\n=== 自動検証 ===")
    wb2 = load_workbook(TEMPLATE)
    ok = True

    if ws_name in wb2.sheetnames:
        ws2 = wb2[ws_name]
        hdr_row = None
        for r in range(1, min(20, ws2.max_row + 1)):
            v = ws2.cell(r, 1).value
            if v and "テストテーブル" in str(v):
                hdr_row = r
                break
        col_hdr_row = (hdr_row + 1) if hdr_row else 6

        c3_val = str(ws2.cell(col_hdr_row, 3).value or "").strip()
        ok1 = (c3_val == "実行種別")
        _v(ok1, f"テスト・検証: r{col_hdr_row} C 列 = 「実行種別」（現在: {c3_val!r}）")
        if not ok1:
            ok = False

        # B 列（タイミング/区分）が残っているか
        b3_val = str(ws2.cell(col_hdr_row, 2).value or "").strip()
        ok2 = bool(b3_val)
        _v(ok2, f"テスト・検証: r{col_hdr_row} B 列 = {b3_val!r}（消えていないこと）")
        if not ok2:
            ok = False

        # D 列以降が D・E・F・G・H に広がっているか（D 列に値があること）
        d3_val = str(ws2.cell(col_hdr_row, 4).value or "").strip()
        ok3 = bool(d3_val)
        _v(ok3, f"テスト・検証: r{col_hdr_row} D 列 = {d3_val!r}（シフト済みこと）")
        if not ok3:
            ok = False

    print("\n" + ("✅ 全検証 PASS" if ok else "❌ 一部 FAIL — 上記ログを確認してください"))
    sys.exit(0 if ok else 1)


def _v(cond, msg):
    status = "✅" if cond else "❌"
    print(f"  {status} {msg}")


if __name__ == "__main__":
    main()
