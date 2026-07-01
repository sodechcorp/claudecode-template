# -*- coding: utf-8 -*-
"""backlog-xlsx / generate_evidence_xlsx.py
機械実行用エビデンス.xlsx を生成する（テスト仕様 → 実行結果の自動記入版）。

旧 create_evidence.py（人手貼付前提）の後継。
test-spec.md（9列スキーマ）を読み、コンパクト表＋証跡別シートの 2 シート構成で生成する。

Sheet 1「テスト結果」  : 1 行 = 1 ケース。実際の結果・判定（OK/NG 色分け）を自動記入。
                          再テスト時は回次別判定列（R1/R2…）を動的追加し OK/NG 推移を一覧化。
Sheet 2「証跡」 or     : 各ケースの証跡（スクショ PNG / SOQL テキスト）を縦に配置し、
「証跡_R1」「証跡_R2」…  「テスト結果」シートからハイパーリンクで紐付け。
                          再テスト時は回次別シートを生成（証跡_R1 / 証跡_R2 …）。
「要手動」ケースのみ旧来の「貼付枠＋指示文」を証跡シートに残す（ハイブリッド）。

Usage（初回・単一回次）:
    python generate_evidence_xlsx.py \\
      --folder /path/to/xlsx_folder \\
      --issue-id GF-350 \\
      --spec /path/to/docs/logs/GF-350/test-spec.md \\
      --evidence-dir /path/to/docs/logs/GF-350/evidence/after \\
      --judgment /path/to/docs/logs/GF-350/judgment-result.json

回次履歴は judgment-result.R1.json / evidence/after_R1/ として自動退避済みの場合、
--judgment / --evidence-dir に現在のパスを指定するだけで過去回次を自動発見・表示する。
退避コマンド例（test.md Phase A で実行）:
    cp judgment-result.json judgment-result.R1.json
    cp -r evidence/after    evidence/after_R1
"""

import argparse
import json
import os
import re
import sys
from collections import Counter
from io import BytesIO
from pathlib import Path

from _common import (
    validate_folder, _stripe_fill,
    _font, _align, _thin_border, _thin_side,
    _set_row_height, _set_col_width, _freeze,
    _auto_col_width, _row_height_by_lines,
    patch_preserve_space,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl がインストールされていません。`pip install openpyxl` を実行してください。")
    sys.exit(1)

# Pillow は add_image に必要（PNG 読み込み）
try:
    from openpyxl.drawing.image import Image as XLImage
    _PIL_OK = True
    try:
        from PIL import Image as PILImage
    except ImportError:
        _PIL_OK = False
except ImportError:
    _PIL_OK = False

# openpyxl CellRichText（部分赤字に必要、バージョン依存）
# 旧コメント「lxml 経由なら安全」は誤りと判明: 空白のみ／edge-空白の run は
# lxml ライターでも xml:space="preserve" 無しで出力され Excel の
# 「修復が必要」ダイアログを引き起こす（openpyxl 3.1.5 で確認）。
# リッチテキスト自体は import 可否のみで有効化し、xml:space 欠落は
# 保存後に _common.patch_preserve_space() でパッチして是正する。
try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    _RICHTEXT_OK = True
except ImportError:
    _RICHTEXT_OK = False

# ── スタイル定数 ─────────────────────────────────────────────────────────────
HDR_FILL    = PatternFill("solid", fgColor="1F3864")   # 濃紺ヘッダー
HDR_FONT    = _font(fg="FFFFFF", bold=True, size=10)
OK_FILL     = PatternFill("solid", fgColor="C6EFCE")   # 緑
NG_FILL     = PatternFill("solid", fgColor="FFC7CE")   # 赤
MANUAL_FILL = PatternFill("solid", fgColor="FFEB9C")   # 黄（要手動）
EVIDENCE_BG = PatternFill("solid", fgColor="FFF3CD")   # エビデンス枠
LIGHT_BLUE  = PatternFill("solid", fgColor="D6E4F7")   # サブヘッダー
WHITE       = PatternFill("solid", fgColor="FFFFFF")

THIN_BORDER = _thin_border("AAAAAA")
WRAP        = _align("left", "top", wrap=True)
CENTER_MID  = _align("center", "center", wrap=False)

# 列幅初期値（auto-fit の下限として機能）
_MIN_COL_WIDTHS_RESULT   = [8, 8, 18, 12, 20, 18, 18, 10, 24, 14]   # No/対応要求/観点/種別/テスト手順/期待/実際/判定/NG原因/証跡リンク
_MIN_COL_WIDTHS_EVIDENCE = [8, 58]

# 印刷設定
_PAPER_SIZE   = 9   # A4
_ORIENTATION  = "landscape"
_MARGIN_CM    = 0.8  # 上下左右余白 cm（openpyxl は inch 単位）
_CM_PER_INCH  = 2.54


# ── パーサ ───────────────────────────────────────────────────────────────────

def parse_test_spec(spec_path: str) -> list:
    text = Path(spec_path).read_text(encoding="utf-8")
    headers = []
    rows = []
    in_table = False
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            if in_table:
                break
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if all(re.match(r"^[-: ]+$", c) for c in cells):
            in_table = True
            continue
        if not headers:
            headers = cells
            in_table = True
        else:
            rows.append(dict(zip(headers, cells)))
    return rows


def load_judgment(judgment_path: str) -> dict:
    """judge_results.py の出力 JSON を読む。なければ空辞書。"""
    if not judgment_path or not os.path.exists(judgment_path):
        return {}
    data = json.loads(Path(judgment_path).read_text(encoding="utf-8"))
    return {r["no"]: r for r in data.get("results", [])}



# ── 証跡ファイル探索 ─────────────────────────────────────────────────────────

def _tc_normalize(tc_str: str) -> int:
    """TC番号文字列から数値部を抽出（桁差を吸収: TC-004 / TC-04 / TC-4 → 4）。"""
    m = re.match(r'(?:[Tt][Cc])-0*(\d+)', tc_str.strip())
    return int(m.group(1)) if m else -1


def find_evidence_file(evidence_dir: str, tc_no: str, shubetsu: str, judgment: dict = None) -> str:
    """後方互換: 最初の1ファイルのみ返す。"""
    files = find_evidence_files(evidence_dir, tc_no, shubetsu, judgment)
    return files[0] if files else ""


def build_evidence_index(evidence_dir: str) -> dict:
    """evidence_dir を1回だけ os.walk し、TC番号 → ファイルパス一覧の索引を返す。

    find_evidence_files を TC 件数ぶん繰り返し呼ぶと同じディレクトリを毎回
    再帰走査してしまう（O(TC × ファイル数)）ため、TC ループの外で1回だけ
    構築し find_evidence_files(..., index=...) に渡す。走査順・除外条件は
    find_evidence_files の os.walk 版と同一（証跡順序・正確性は不変）。
    """
    index: dict = {}
    if not os.path.isdir(evidence_dir):
        return index
    for dirpath, _dirs, filenames in os.walk(evidence_dir):
        for fname in sorted(filenames):
            # before / リサイズ済みサムネイルは対象外
            if "_before." in fname or "_resized." in fname:
                continue
            # ファイル名先頭の "TC-NNN" 部分を抽出して TC番号を索引キーにする
            fname_prefix = fname.split("_")[0]  # "TC-001" / "TC-04" / "tc-4"
            tc_num = _tc_normalize(fname_prefix)
            if tc_num < 0:
                continue
            index.setdefault(tc_num, []).append(os.path.join(dirpath, fname))
    return index


def find_evidence_files(evidence_dir: str, tc_no: str, shubetsu: str, judgment: dict = None,
                        index: dict = None) -> list:
    """証跡ディレクトリから TC 番号に対応する全ファイルを返す。
    優先順:
      ① judgment-result.json の evidence パス（検証済み正本・絶対パスで直参照）
      ② index 索引引き（build_evidence_index 済みの場合）または os.walk 再帰探索
         （複合種別・before/after サブディレクトリ・桁差 TC-04 vs TC-004 対応）
    before ファイル（"_before." を含む）は除外。

    index: build_evidence_index(evidence_dir) の返り値。TC ループの外で1回構築して
           渡すと os.walk の重複走査を避けられる。None の場合は従来どおり毎回 os.walk する。
    """
    found = []
    seen = set()

    # ① judgment の evidence パスを最優先（正本が既にある場合は即採用）
    if judgment is not None:
        ev_path = judgment.get(tc_no, {}).get("evidence", "")
        if ev_path and os.path.isfile(ev_path) and ev_path not in seen:
            seen.add(ev_path)
            found.append(ev_path)

    tc_num = _tc_normalize(tc_no)
    if tc_num < 0:
        return found

    if index is not None:
        # ② index 索引引き（build_evidence_index で1回だけ os.walk 済み）
        for fpath in index.get(tc_num, []):
            if fpath in seen:
                continue
            seen.add(fpath)
            found.append(fpath)
    elif os.path.isdir(evidence_dir):
        # ② os.walk 再帰探索（TC番号正規化で桁差・複合種別・サブディレクトリを吸収）
        for dirpath, _dirs, filenames in os.walk(evidence_dir):
            for fname in sorted(filenames):
                # before / リサイズ済みサムネイルは対象外
                if "_before." in fname or "_resized." in fname:
                    continue
                fpath = os.path.join(dirpath, fname)
                if fpath in seen:
                    continue
                # ファイル名先頭の "TC-NNN" 部分を抽出して TC番号と比較
                fname_prefix = fname.split("_")[0]  # "TC-001" / "TC-04" / "tc-4"
                if _tc_normalize(fname_prefix) == tc_num:
                    seen.add(fpath)
                    found.append(fpath)

    return found


# ── テキスト証跡ヘルパー ─────────────────────────────────────────────────────

def _read_text_safe(path: str) -> str:
    """UTF-16 LE 自動検出＋制御文字除去してテキストを返す（打ち切りなし）。"""
    try:
        raw = Path(path).read_bytes()
        if raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
            text = raw.decode("utf-16", errors="replace")
        elif len(raw) > 1 and raw[1] == 0x00:
            text = raw.decode("utf-16-le", errors="replace")
        else:
            text = raw.decode("utf-8", errors="replace")
    except Exception:
        text = ""
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)


def _highlight_terms_from_tc(tc: dict) -> list:
    """テストケースの期待結果から赤字化するトークン一覧を返す。
    否定観点（含まない/非表示/なし確認）はトークンを返さない（誤誘導防止）。
    """
    hanteihoo = tc.get("判定方法", "")
    neg_patterns = ["含まない", "非表示", "なし確認", "存在しない", "NG確認"]
    if any(p in hanteihoo for p in neg_patterns):
        return []
    kitai = tc.get("期待結果", "").strip()
    if not kitai:
        return []
    # before:X / after:Y 形式の状態遷移は after 部分を抽出
    m = re.search(r'after[:：]\s*(.+?)(?:\s*/|$)', kitai)
    if m:
        kitai = m.group(1)
    # 「...」 形式の引用句を優先抽出（句内にセパレータが含まれるケースに対応）
    quoted = re.findall(r'[「『"]([^「『」』"]{1,40})[」』"]', kitai)
    # 残テキストを一般セパレータで分割
    remainder = re.sub(r'[「『"][^「『」』"]*[」』"]', '', kitai)
    raw_tokens = re.split(r'[/,、\s]+', remainder)
    tokens = list(quoted)
    for t in raw_tokens:
        t = t.strip().strip('「」『』"\'')
        if not t or t in ('-', '—', '…') or len(t) <= 1:
            continue
        tokens.append(t)
    # 「3件」→「3」も追加（SOQL件数ヘッダーの数値に一致させる）
    extra = []
    for t in tokens:
        nm = re.match(r'^(\d+)\s*件', t)
        if nm:
            extra.append(nm.group(1))
    tokens.extend(extra)
    return list(dict.fromkeys(tokens))  # 重複除去・順序保持


def _build_richtext_line(line: str, terms: list):
    """1行を CellRichText に変換し、terms に一致する部分を赤太字にする。
    未対応環境（_RICHTEXT_OK=False）または terms 空のときは文字列をそのまま返す。
    """
    if not _RICHTEXT_OK or not terms:
        return line
    if not any(t.lower() in line.lower() for t in terms):
        return line
    # 最長一致優先でトークンをソート
    sorted_terms = sorted(terms, key=len, reverse=True)
    pattern = re.compile(
        "(" + "|".join(re.escape(t) for t in sorted_terms) + ")",
        re.IGNORECASE,
    )
    parts = pattern.split(line)
    if len(parts) <= 1:
        return line
    base_font  = InlineFont(rFont="Courier New", sz=9)
    red_font   = InlineFont(rFont="Courier New", sz=9, color="FFCC0000", b=True)
    blocks = []
    for i, part in enumerate(parts):
        if not part:
            continue
        font = red_font if (i % 2 == 1) else base_font
        blocks.append(TextBlock(font, part))
    return CellRichText(blocks)


def _append_text_block(ws, path: str, start_row: int, highlight_terms: list = None) -> int:
    """テキストファイルを ws の start_row から全行展開し、書いた行数を返す（打ち切りなし）。
    highlight_terms が指定されると期待結果に一致する部分を赤太字にする（S2）。
    """
    text = _read_text_safe(path)
    lines = text.splitlines()
    for i, line in enumerate(lines):
        # "=" 始まりの行は openpyxl が数式セルとして書き出し Excel 修復ダイアログが出る
        safe = (" " + line) if line.startswith("=") else line
        cell_value = _build_richtext_line(safe, highlight_terms)
        c = ws.cell(row=start_row + i, column=2, value=cell_value)
        c.font = Font(name="Courier New", size=9)
        c.border = THIN_BORDER
        c.alignment = _align("left", "top", wrap=True)
    return len(lines)


def _append_dom_excerpt(ws, path: str, start_row: int, highlight_terms: list = None) -> int:
    """DOM テキストから期待値一致行だけ抜粋して ws の start_row から展開し、書いた行数を返す。

    highlight_terms が空（否定観点）の場合は確認ノート1行のみ書く。
    一致行がある場合は一致行のみ赤字抜粋（証跡の主役はスクショ・DOMは補助）。
    SOQL/Apex の .txt 全文展開（_append_text_block）とは別物。
    """
    if not highlight_terms:
        # 否定観点: 期待文字列なしを確認（スクショ参照）
        note = "DOM照合: 期待文字列なしを確認（スクショ参照）"
        c = ws.cell(row=start_row, column=2, value=note)
        c.font = _font(fg="888888", italic=True, size=9)
        c.alignment = _align("left", "top", wrap=True)
        c.border = THIN_BORDER
        ws.cell(start_row, 1).border = THIN_BORDER
        _set_row_height(ws, start_row, 15)
        return 1

    text = _read_text_safe(path)
    lines = text.splitlines()
    matched = [ln for ln in lines
               if any(t.lower() in ln.lower() for t in highlight_terms)]

    if not matched:
        note = "DOM照合: 期待値一致行なし（スクショ参照）"
        c = ws.cell(row=start_row, column=2, value=note)
        c.font = _font(fg="888888", italic=True, size=9)
        c.alignment = _align("left", "top", wrap=True)
        c.border = THIN_BORDER
        ws.cell(start_row, 1).border = THIN_BORDER
        _set_row_height(ws, start_row, 15)
        return 1

    for i, line in enumerate(matched):
        safe = (" " + line) if line.startswith("=") else line
        cell_value = _build_richtext_line(safe, highlight_terms)
        c = ws.cell(row=start_row + i, column=2, value=cell_value)
        c.font = Font(name="Courier New", size=9)
        c.border = THIN_BORDER
        c.alignment = _align("left", "top", wrap=True)
        ws.cell(start_row + i, 1).border = THIN_BORDER
        _set_row_height(ws, start_row + i, 15)
    return len(matched)


def _count_lines(path: str) -> int:
    return len(_read_text_safe(path).splitlines())


# ── エビデンス可読性ヘルパー ─────────────────────────────────────────────────

def _derive_focus(tc: dict) -> str:
    """「着眼点」列が無い場合に期待結果＋判定方法から確認ポイント導出文を生成する。"""
    kitai   = tc.get("期待結果",  "").strip()
    hantei  = tc.get("判定方法",  "").strip()
    neg_patterns = ["含まない", "非表示", "なし確認", "存在しない"]
    if any(p in hantei for p in neg_patterns):
        return f"「{kitai}」が表示・存在しないことを確認（{hantei}）"
    if kitai and hantei:
        return f"「{kitai}」が {hantei} で確認できること"
    if kitai:
        return f"「{kitai}」であることを確認"
    return ""


def _write_reading_header(ws, tc: dict, judgment_entry: dict, row_ptr: int,
                          result_ws_name: str = "テスト結果") -> int:
    """証跡シートの TC セクション先頭に読み方ガイドブロックを書き、次の row_ptr を返す。

    出力イメージ（簡素化版・全行動的行高）:
        ■ TC-003: ②I-797「いいえ」で専用相談メッセージのみ表示（UI）
          確認観点: {確認ポイント（着眼点）or 導出文}
          判定   : {判定方法}  ✅OK / ❌NG / ⬜要手動

    重複排除: 期待結果・判定方法の独立行は廃止（テスト結果シートと test-spec.md に既出のため）。
    """
    no       = tc.get("No", "")
    kanpoin  = tc.get("観点", "")
    shubetsu = tc.get("種別", "")
    hantei   = tc.get("判定方法", "")
    focus    = tc.get("確認ポイント（着眼点）") or tc.get("着眼点") or _derive_focus(tc)

    status = judgment_entry.get("status", "") if judgment_entry else ""
    if status == "OK":
        judge_icon = "✅ OK"
        judge_fill = OK_FILL
    elif status == "NG":
        judge_icon = "❌ NG"
        judge_fill = NG_FILL
    else:
        judge_icon = "⬜ 要手動"
        judge_fill = MANUAL_FILL

    GUIDE_FILL = PatternFill("solid", fgColor="EFF3FB")
    # 証跡シートの B 列幅（≒58文字）から動的行高を計算するための近似値
    _EV_COL_W = 55

    # 行1: ■ No: 観点（種別）— 濃紺ヘッダー・折り返し対応・動的行高
    ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=2)
    hdr_text = f"■ {no}: {kanpoin} （{shubetsu}）"
    hdr = ws.cell(row=row_ptr, column=1, value=hdr_text)
    hdr.fill = PatternFill("solid", fgColor="1F3864")
    hdr.font = _font(fg="FFFFFF", bold=True, size=10)
    hdr.alignment = WRAP
    hdr.border = THIN_BORDER
    _set_row_height(ws, row_ptr, max(20, _row_height_by_lines(hdr_text, col_width=60, base_pt=14)))
    anchor_cell_addr = f"A{row_ptr}"
    row_ptr += 1

    def _guide_row(label: str, value: str, fill=None, bold_val: bool = False):
        nonlocal row_ptr
        lc = ws.cell(row=row_ptr, column=1, value=label)
        lc.fill = fill or GUIDE_FILL
        lc.font = _font(bold=True, size=9, fg="444444")
        lc.alignment = _align("right", "top", wrap=False)
        lc.border = THIN_BORDER
        vc = ws.cell(row=row_ptr, column=2, value=value)
        vc.fill = fill or GUIDE_FILL
        vc.font = _font(bold=bold_val, size=9)
        vc.alignment = WRAP
        vc.border = THIN_BORDER
        # 動的行高: 折り返した日本語が切れないよう内容量に応じて伸ばす
        h = max(15, _row_height_by_lines(str(value), col_width=_EV_COL_W, base_pt=13))
        _set_row_height(ws, row_ptr, h)
        row_ptr += 1

    # 行2: 確認観点（着眼点または導出文）— 水色背景
    _guide_row("確認観点", focus, fill=LIGHT_BLUE, bold_val=True)
    # 行3: 判定（判定方法 + 判定アイコン）— 判定色背景
    _guide_row("判定", f"{hantei}  {judge_icon}", fill=judge_fill)

    return row_ptr, anchor_cell_addr


# ── 結果シートヘルパー ────────────────────────────────────────────────────────

def _extract_req_label(kanpoin: str) -> str:
    """観点テキストから要求ラベル（①②③ / 回帰）を抽出する。"""
    m = re.match(r'^([①-⑳]+)', kanpoin.strip())
    if m:
        return m.group(1)
    if "回帰" in kanpoin:
        return "回帰"
    return ""


def _build_ng_action(j_result: dict) -> str:
    """NG判定結果からユーザー向けのアクション文を生成する（NG原因/次アクション列）。"""
    reason = j_result.get("reason", "")
    ng_type = j_result.get("ng_type", "")
    if ng_type == "未実行":
        return "再テスト要" + (f": {reason}" if reason else "（証跡なし）")
    if ng_type == "要確認":
        return "判定方法修正要" + (f": {reason}" if reason else "")
    return reason or "要確認"


# ── 回次探索 ─────────────────────────────────────────────────────────────────

def discover_rounds(judgment_path: str, evidence_dir: str) -> list:
    """archived rounds を検出して [(label, j_path, ev_dir), ...] を返す。

    judgment-result.R1.json / evidence/after_R1 が存在する場合:
        [("R1", ".../judgment-result.R1.json", ".../after_R1"),
         ("R2", ".../judgment-result.json",    ".../after")]   # 現回次
    単一回次（アーカイブなし）の場合:
        [("R1", ".../judgment-result.json", ".../after")]
    """
    folder = os.path.dirname(os.path.abspath(judgment_path))
    ev_parent = os.path.dirname(os.path.abspath(evidence_dir))
    base = os.path.splitext(os.path.basename(judgment_path))[0]  # "judgment-result"

    archived = []
    try:
        for f in sorted(os.listdir(folder)):
            m = re.match(rf'^{re.escape(base)}\.R(\d+)\.json$', f)
            if m:
                archived.append((int(m.group(1)), os.path.join(folder, f)))
    except OSError:
        pass
    archived.sort(key=lambda x: x[0])

    rounds = []
    for n, j_path in archived:
        label = f"R{n}"
        ev_d = os.path.join(ev_parent, f"after_R{n}")
        rounds.append((label, j_path, ev_d))

    # 現回次番号: アーカイブ数 + 1
    current_n = len(archived) + 1
    rounds.append((f"R{current_n}", judgment_path, evidence_dir))
    return rounds


# ── Sheet 1: テスト結果 ───────────────────────────────────────────────────────

def build_result_sheet(ws, test_cases: list, judgment: dict, evidence_dir: str,
                       all_rounds: list = None) -> tuple:
    """テスト結果シートを構築し (tc_to_row, link_col) を返す。

    all_rounds: [(round_label, judgment_dict), ...] を渡すと回次別判定列を追加する。
                None または 1 要素の場合は従来どおり（回次列なし）。
    link_col  : 証跡ハイパーリンクの列番号（回次列追加で変わるため返却）。
    """
    is_multi = all_rounds is not None and len(all_rounds) > 1

    # ── 動的列定義 ────────────────────────────────────────────────────────────
    base_left  = ["No", "対応要求", "確認観点", "種別", "テスト手順", "期待結果"]
    round_cols = [lbl for lbl, _ in all_rounds] if is_multi else []
    base_right = ["実際の結果", "判定", "NG原因/次アクション", "証跡"]
    headers    = base_left + round_cols + base_right

    n_left  = len(base_left)                        # 6
    n_round = len(round_cols)                       # 0 or ≥2
    col_actual = n_left + n_round + 1               # 実際の結果
    col_judge  = col_actual + 1                     # 判定
    col_ng     = col_judge  + 1                     # NG原因
    col_link   = col_ng     + 1                     # 証跡リンク（動的）

    # 回次別 judgment を dict に整理（all_rounds は (label, j_dict) リスト）
    round_judgments = all_rounds if is_multi else []

    ws.title = "テスト結果"

    # ヘッダー行
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER_MID
        c.border = THIN_BORDER
        # 回次列はやや幅を狭める（6pt）
        if n_left < j <= n_left + n_round:
            _set_col_width(ws, j, 6)
    _set_row_height(ws, 1, 22)

    # 固定列の列幅初期値（auto-fit 下限）
    for j, w in enumerate(_MIN_COL_WIDTHS_RESULT, start=1):
        # 回次列が挿入されるため右側の列はインデックスをずらす
        actual_j = j if j <= n_left else j + n_round
        _set_col_width(ws, actual_j, w)

    tc_to_row = {}
    for i, tc in enumerate(test_cases):
        row = i + 2
        no = tc.get("No", "")
        shubetsu = tc.get("種別", "")
        kanpoin = tc.get("観点", "")
        tesuji = tc.get("テスト手順", "")
        kiki = tc.get("期待結果", "")
        auto = tc.get("自動化可否", "自動").strip()
        is_manual = "要手動" in auto

        # 最新回次の判定結果から実際の結果・ステータスを取得
        j_result = judgment.get(no, {})
        actual = j_result.get("actual", "")
        status = j_result.get("status", "SKIP" if is_manual else "")

        # 判定セルの表示と色
        if status == "OK":
            judge_text = "✅ OK"
            row_fill = OK_FILL
        elif status == "NG":
            judge_text = "❌ NG"
            row_fill = NG_FILL
        else:
            judge_text = "⬜ 要手動"
            row_fill = MANUAL_FILL

        fill = _stripe_fill(i) if status == "OK" else row_fill

        req_label = _extract_req_label(kanpoin)
        ng_action = _build_ng_action(j_result) if status == "NG" else ""

        # 回次別ステータスを事前計算（is_multi 時のみ）
        round_statuses = []
        if is_multi:
            for _, r_j in round_judgments:
                r_res = r_j.get(no, {})
                round_statuses.append(r_res.get("status", ""))

        # vals: base_left + round_cols（判定アイコン）+ base_right
        left_vals  = [no, req_label, kanpoin, shubetsu, tesuji, kiki]
        round_vals = []
        for rs in round_statuses:
            if rs == "OK":
                round_vals.append("✅")
            elif rs == "NG":
                round_vals.append("❌")
            else:
                round_vals.append("⬜")
        right_vals = [actual, judge_text, ng_action, "→証跡"]
        vals = left_vals + round_vals + right_vals

        for j_col, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=j_col, value=val)
            c.border = THIN_BORDER
            if j_col == 1:
                # No 列: 中央揃え・游ゴシック
                c.alignment = CENTER_MID
                c.font = _font(bold=True)
                c.fill = fill
            elif j_col == 2:
                # 対応要求列: 中央揃え・小さめフォント
                c.alignment = CENTER_MID
                c.font = _font(bold=True, size=9)
                c.fill = fill
            elif n_left < j_col <= n_left + n_round:
                # 回次別判定列: アイコン中央揃え・回次ステータス色
                r_idx = j_col - n_left - 1
                rs = round_statuses[r_idx]
                c.alignment = CENTER_MID
                c.font = _font(size=9)
                c.fill = (OK_FILL if rs == "OK" else
                          NG_FILL if rs == "NG" else MANUAL_FILL)
            elif j_col == col_judge:
                # 判定列: 中央揃え・太字
                c.alignment = CENTER_MID
                c.fill = row_fill
                c.font = _font(bold=True)
            elif j_col == col_ng:
                # NG原因/次アクション列: NG のみ赤背景
                c.alignment = WRAP
                c.font = _font(size=9)
                c.fill = row_fill if status == "NG" else fill
            else:
                c.alignment = WRAP
                c.font = _font()
                c.fill = fill

        # 行高: テスト手順・実際の結果のどちらか長い方に合わせて確保（下限 30pt）
        h = max(30.0,
                _row_height_by_lines(actual, col_width=18),
                _row_height_by_lines(tesuji, col_width=20))
        _set_row_height(ws, row, h)
        tc_to_row[no] = row

    # 観点・期待結果・実際の結果列は内容に応じて幅を自動調整
    # col_actual は回次列挿入でシフト済みの絶対列番号
    for col_idx in [3, 5, 6, col_actual]:  # 確認観点, テスト手順, 期待結果, 実際の結果
        # _MIN_COL_WIDTHS_RESULT の元のインデックスに戻してから参照
        orig_idx = col_idx if col_idx <= n_left else col_idx - n_round
        _auto_col_width(ws, col_idx, min_w=_MIN_COL_WIDTHS_RESULT[orig_idx - 1], max_w=40)

    # ── 要件カバレッジ・サマリー（末尾に追記）──────────────────────────────
    req_count: Counter = Counter()
    req_ok: Counter = Counter()
    req_ng: Counter = Counter()
    for tc in test_cases:
        lbl = _extract_req_label(tc.get("観点", "")) or "その他"
        j_r = judgment.get(tc.get("No", ""), {})
        st = j_r.get("status", "")
        req_count[lbl] += 1
        if st == "OK":
            req_ok[lbl] += 1
        elif st == "NG":
            req_ng[lbl] += 1
    summary_parts = []
    for lbl in sorted(req_count.keys()):
        n = req_count[lbl]
        ok_ = req_ok.get(lbl, 0)
        ng_ = req_ng.get(lbl, 0)
        summary_parts.append(f"{lbl}: {n}TC (OK={ok_}/NG={ng_})")
    summary_row = len(test_cases) + 3  # ヘッダー行 + データ行数 + 空白行1
    ncols = len(headers)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=ncols)
    sc = ws.cell(row=summary_row, column=1,
                 value="■ 要件カバレッジ: " + " | ".join(summary_parts))
    sc.fill = LIGHT_BLUE
    sc.font = _font(bold=True, size=9)
    sc.alignment = _align("left", "center")
    sc.border = THIN_BORDER
    _set_row_height(ws, summary_row, 20)

    _freeze(ws, 2)

    # 印刷設定
    ws.page_setup.paperSize = _PAPER_SIZE
    ws.page_setup.orientation = _ORIENTATION
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    margin_inch = _MARGIN_CM / _CM_PER_INCH
    ws.page_margins.top    = margin_inch
    ws.page_margins.bottom = margin_inch
    ws.page_margins.left   = margin_inch
    ws.page_margins.right  = margin_inch
    ws.print_title_rows = "1:1"  # ヘッダー行を全ページに繰り返し

    return tc_to_row, col_link  # col_link は回次列追加でシフトするため返却


# ── Sheet 2: 証跡 ────────────────────────────────────────────────────────────

def build_evidence_sheet(ws, test_cases: list, judgment: dict, evidence_dir: str, tc_to_row: dict) -> None:
    """証跡シートを構築する。シート名は呼び出し側で設定済みのため上書きしない。"""
    # ws.title はシート作成時に呼び出し側が設定する（単一回次="証跡" / 複数回次="証跡_R{N}"）

    headers = ["No", "証跡内容"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER_MID
        c.border = THIN_BORDER
    for j, w in enumerate(_MIN_COL_WIDTHS_EVIDENCE, start=1):
        _set_col_width(ws, j, w)
    _set_row_height(ws, 1, 22)

    row_ptr = 2
    result_ws_name = "テスト結果"

    # evidence_dir の os.walk は TC 件数ぶん繰り返さず1回だけ実行し索引化する
    evidence_index = build_evidence_index(evidence_dir)

    for tc in test_cases:
        no = tc.get("No", "")
        shubetsu = tc.get("種別", "")
        auto = tc.get("自動化可否", "自動").strip()
        is_manual = "要手動" in auto

        # 複数証跡ファイルを全件取得（judgment 正本→索引引きの二段構え）
        all_evidence = find_evidence_files(evidence_dir, no, shubetsu, judgment, index=evidence_index)

        # 読み方ガイドヘッダー（S1-a: 何を確認/期待結果/判定/確認ポイント）
        judgment_entry = judgment.get(no, {})
        hl_terms = _highlight_terms_from_tc(tc)  # S2: 期待結果トークン（否定観点は空）
        row_ptr, anchor_cell_addr = _write_reading_header(ws, tc, judgment_entry, row_ptr)

        if is_manual:
            # 要手動: 貼付枠を残す
            ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=2)
            c = ws.cell(row=row_ptr, column=1,
                        value="⬜ 要手動確認 — ユーザーがここにスクリーンショットを貼り付けてください")
            c.fill = EVIDENCE_BG
            c.font = _font()
            c.alignment = WRAP
            c.border = THIN_BORDER
            row_ptr += 1
            paste_top = row_ptr
            paste_bottom = paste_top + 9
            for r in range(paste_top, paste_bottom + 1):
                for col in range(1, 3):
                    cell = ws.cell(r, col)
                    cell.fill = WHITE
                    cell.border = THIN_BORDER
            ws.merge_cells(start_row=paste_top, start_column=1, end_row=paste_bottom, end_column=2)
            inst = ws.cell(paste_top, 1, "（スクリーンショット貼付エリア）")
            inst.alignment = _align("center", "center", wrap=False)
            inst.font = _font(fg="888888", italic=True)
            row_ptr = paste_bottom + 2

        elif not all_evidence:
            c = ws.cell(row_ptr, 2, "（証跡ファイルなし）")
            c.alignment = WRAP
            c.font = _font(fg="888888", italic=True)
            c.border = THIN_BORDER
            c1 = ws.cell(row_ptr, 1)
            c1.border = THIN_BORDER
            row_ptr += 2

        else:
            # 証跡ファイルを1ファイル1ブロックで展開（PNG は before 除く / txt は DOM 含む）
            # PNG + 同名 txt のペアは PNG -> txt の順で同一ブロックに入れる
            processed = set()
            for ep in all_evidence:
                if ep in processed:
                    continue
                processed.add(ep)
                fname = os.path.basename(ep)

                # ブロック小見出し（複数ある場合のみ）
                if len(all_evidence) > 1:
                    ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=2)
                    sub_hdr = ws.cell(row=row_ptr, column=1, value=f"  └ {fname}")
                    sub_hdr.font = _font(italic=True, size=9, fg="444444")
                    sub_hdr.alignment = WRAP
                    sub_hdr.border = THIN_BORDER
                    _set_row_height(ws, row_ptr, 16)
                    row_ptr += 1

                if ep.lower().endswith(".png") and _PIL_OK:
                    # PNG: 自動貼付
                    try:
                        pil_img = PILImage.open(ep)
                        max_w = 800
                        w, h = pil_img.size
                        if w > max_w:
                            ratio = max_w / w
                            pil_img = pil_img.resize((int(w * ratio), int(h * ratio)), PILImage.LANCZOS)
                            buf = BytesIO()
                            pil_img.save(buf, format="PNG")
                            buf.seek(0)
                            ep_use = buf   # BytesIO で渡し、_resized.png を証跡フォルダに残さない
                        else:
                            ep_use = ep
                        disp_w, disp_h = pil_img.size  # 実際に貼り付ける表示サイズ（px）
                        xl_img = XLImage(ep_use)
                        xl_img.width  = disp_w          # DPIメタデータを無視し表示サイズを固定
                        xl_img.height = disp_h
                        xl_img.anchor = f"B{row_ptr}"
                        ws.add_image(xl_img)
                        # 行高を 15pt (≒20px@96dpi) に固定し、画像が占める行数を決定論的に算出
                        ROW_PX = 20
                        img_rows = max(10, (disp_h + ROW_PX - 1) // ROW_PX + 2)  # ceil + 余白2行
                        for r in range(row_ptr, row_ptr + img_rows):
                            _set_row_height(ws, r, 15)  # 15pt = 20px 固定
                            ws.cell(r, 1).border = THIN_BORDER
                            ws.cell(r, 2).border = THIN_BORDER
                        row_ptr += img_rows + 1
                    except Exception as e:
                        c = ws.cell(row_ptr, 2, f"[WARN] 画像読込失敗: {e}")
                        c.alignment = WRAP
                        c.font = _font(fg="CC0000")
                        c.border = THIN_BORDER
                        ws.cell(row_ptr, 1).border = THIN_BORDER
                        row_ptr += 2

                    # 同名 DOM スナップショット (.txt) を期待値一致行だけ抜粋して展開（S2）
                    # .txt ファイル自体は judge_results.py の DOM 照合判定に使うためディスクに残す
                    snap_path = re.sub(r'\.png$', '.txt', ep, flags=re.IGNORECASE)
                    if os.path.exists(snap_path) and snap_path not in processed:
                        processed.add(snap_path)
                        n_dom = _append_dom_excerpt(ws, snap_path, row_ptr, highlight_terms=hl_terms)
                        row_ptr += n_dom + 1

                elif ep.lower().endswith(".png") and not _PIL_OK:
                    c = ws.cell(row_ptr, 2, f"[PNG] {fname} — Pillow 未インストールのため未貼付")
                    c.alignment = WRAP
                    c.font = _font(fg="888888")
                    c.border = THIN_BORDER
                    ws.cell(row_ptr, 1).border = THIN_BORDER
                    row_ptr += 2

                else:
                    # テキスト証跡: 全行展開（S2: 期待値を赤字）
                    n = _append_text_block(ws, ep, row_ptr, highlight_terms=hl_terms)
                    ws.cell(row_ptr, 1).border = THIN_BORDER
                    row_ptr += n + 1

        # 証跡シート先頭アドレスをメモ（後でリンク設定）
        tc_to_row[no + "_ev"] = anchor_cell_addr

    _freeze(ws, 2)

    # 印刷設定
    ws.page_setup.paperSize = _PAPER_SIZE
    ws.page_setup.orientation = _ORIENTATION
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    margin_inch = _MARGIN_CM / _CM_PER_INCH
    ws.page_margins.top    = margin_inch
    ws.page_margins.bottom = margin_inch
    ws.page_margins.left   = margin_inch
    ws.page_margins.right  = margin_inch
    ws.print_title_rows = "1:1"  # ヘッダー行を全ページに繰り返し


# ── ハイパーリンク設定 ────────────────────────────────────────────────────────

def set_hyperlinks(result_ws, evidence_ws_name: str, tc_to_row: dict,
                   test_cases: list, link_col: int = 9) -> None:
    """テスト結果シートの「証跡」列から証跡シートへの内部ハイパーリンクを設定。

    link_col: 証跡列の絶対列番号（回次列追加でシフトするため動的に渡す）。
    """
    for tc in test_cases:
        no = tc.get("No", "")
        result_row = tc_to_row.get(no)
        ev_addr = tc_to_row.get(no + "_ev", "A2")
        if not result_row:
            continue
        cell = result_ws.cell(row=result_row, column=link_col)
        cell.hyperlink = f"#{evidence_ws_name}!{ev_addr}"
        cell.font = _font(color="0563C1", underline="single")
        cell.value = "→ 証跡を見る"
        cell.alignment = CENTER_MID


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="機械記入用エビデンス.xlsx を生成する")
    parser.add_argument("--folder",       required=True, help="xlsx 出力フォルダ")
    parser.add_argument("--issue-id",     required=True, dest="issue_id")
    parser.add_argument("--spec",         required=True, help="test-spec.md のパス")
    parser.add_argument("--evidence-dir", required=True, dest="evidence_dir",
                        help="証跡ファイルの after/ ディレクトリ（最新回次）")
    parser.add_argument("--judgment",     default="",    help="judge_results.py 出力 JSON のパス（最新回次）")
    args = parser.parse_args()

    args.folder = validate_folder(args.folder)
    if not _PIL_OK:
        print("[WARN] Pillow がインストールされていません。PNG の自動貼付をスキップします。")
        print("       `pip install Pillow` でインストールするとスクショが自動貼付されます。")

    test_cases = parse_test_spec(args.spec)
    if not test_cases:
        print("[WARN] test-spec.md にテストケースが見つかりませんでした。空のエビデンスファイルを生成します。")

    # 回次を自動探索（archived R1/R2... + 現在の最新）
    # --judgment 未指定時は単一回次として扱う
    if args.judgment:
        round_paths = discover_rounds(args.judgment, args.evidence_dir)
    else:
        round_paths = [("R1", args.judgment, args.evidence_dir)]

    # 全回次の judgment を読み込む → [(label, j_dict, ev_dir), ...]
    all_rounds_data = []
    for label, j_path, ev_dir in round_paths:
        j_dict = load_judgment(j_path) if j_path else {}
        all_rounds_data.append((label, j_dict, ev_dir))

    # 最新回次
    latest_label, latest_judgment, latest_ev_dir = all_rounds_data[-1]

    # build_result_sheet に渡す all_rounds は (label, j_dict) のみ
    all_rounds_for_sheet = [(lbl, j) for lbl, j, _ in all_rounds_data]

    wb = Workbook()
    result_ws = wb.active

    tc_to_row, link_col = build_result_sheet(
        result_ws, test_cases, latest_judgment, latest_ev_dir,
        all_rounds=all_rounds_for_sheet,
    )

    # 証跡シート: 複数回次ならシート分割、単一なら従来どおり「証跡」
    is_multi_ev = len(all_rounds_data) > 1
    if is_multi_ev:
        for label, j_dict, ev_dir in all_rounds_data:
            ev_ws = wb.create_sheet(f"証跡_{label}")
            if label == latest_label:
                # 最新回次: anchor アドレスを main tc_to_row に書き込む
                build_evidence_sheet(ev_ws, test_cases, j_dict, ev_dir, tc_to_row)
            else:
                build_evidence_sheet(ev_ws, test_cases, j_dict, ev_dir, {})
        latest_ev_ws_name = f"証跡_{latest_label}"
    else:
        evidence_ws = wb.create_sheet("証跡")
        build_evidence_sheet(evidence_ws, test_cases, latest_judgment, latest_ev_dir, tc_to_row)
        latest_ev_ws_name = "証跡"

    set_hyperlinks(result_ws, latest_ev_ws_name, tc_to_row, test_cases, link_col=link_col)

    out_path = os.path.join(args.folder, f"{args.issue_id}_エビデンス.xlsx")
    os.makedirs(args.folder, exist_ok=True)
    try:
        wb.save(out_path)
    except PermissionError as e:
        print(f"[ERROR] xlsx の保存失敗（ファイルが開かれている可能性）: {out_path}\n{e}")
        sys.exit(1)

    # CellRichText の xml:space 欠落バグ対策（Excel 修復ダイアログ回避）
    n_patched = patch_preserve_space(out_path)
    if n_patched:
        print(f"[INFO] xml:space=\"preserve\" 補正: {n_patched} 箇所")

    ok_count   = sum(1 for r in latest_judgment.values() if r.get("status") == "OK")
    ng_count   = sum(1 for r in latest_judgment.values() if r.get("status") == "NG")
    skip_count = sum(1 for r in latest_judgment.values() if r.get("status") == "SKIP")
    round_labels = [lbl for lbl, _, _ in all_rounds_data]

    print(f"生成完了: {out_path}")
    print(f"  テストケース: {len(test_cases)} 件 (OK={ok_count} / NG={ng_count} / 要手動={skip_count})")
    print(f"  回次: {len(all_rounds_data)} 回（{'・'.join(round_labels)}）")
    if not _PIL_OK:
        print("  ※ Pillow 未インストールのため PNG 自動貼付はスキップ。手動で貼り付けてください。")


if __name__ == "__main__":
    main()
