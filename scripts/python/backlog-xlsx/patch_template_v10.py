# -*- coding: utf-8 -*-
"""patch_template_v10.py
対応記録テンプレート.xlsx を Group N 仕様に移行する。

変更内容:
  1. 対応方針シート: r11 空ヘッダ行（A列のみ 002E75B6 fill）を削除
  2. 調査・影響範囲シート: A1:C1 マージ追加

冪等性: 各ステップは既に適用済みなら skip する。

Usage:
    python patch_template_v10.py
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"

_HEADER_FILL_RGB = "002E75B6"  # 中青（列ヘッダ）


def _save_merges(ws):
    return [(m.min_row, m.max_row, m.min_col, m.max_col)
            for m in list(ws.merged_cells.ranges)]


def _clear_merges(ws):
    for mcr in list(ws.merged_cells.ranges):
        ws.merged_cells.ranges.discard(mcr)


# ── Step 1: 対応方針シート ────────────────────────────────────────────────────

def patch_approach_sheet(ws):
    """r11 空ヘッダ行（A列のみ 002E75B6 fill）を削除する。"""

    r11_cell = ws.cell(11, 1)
    r11_val = r11_cell.value
    r11_fill_rgb = ""
    if r11_cell.fill and r11_cell.fill.patternType == "solid" and r11_cell.fill.fgColor:
        r11_fill_rgb = r11_cell.fill.fgColor.rgb

    if r11_val is not None or r11_fill_rgb != _HEADER_FILL_RGB:
        print(f"[SKIP] 対応方針: r11 は空ヘッダ行でない（value={r11_val!r}, fill={r11_fill_rgb}）→ スキップ")
        return

    merges_snap = _save_merges(ws)
    _clear_merges(ws)
    ws.delete_rows(11, 1)

    # row_dimensions シフト（openpyxl は delete_rows で row_dimensions を移動しないため手動対応）
    max_row = ws.max_row + 1  # 削除後なのでデータ行数は元より1少ない
    for r in range(11, max_row + 1):
        if r + 1 in ws.row_dimensions:
            ws.row_dimensions[r].height = ws.row_dimensions[r + 1].height

    # マージ再構築（r11 を含むマージはスキップ・r12 以降を -1 ずらす）
    for (min_r, max_r, min_c, max_c) in merges_snap:
        if min_r <= 11 <= max_r:
            continue  # 削除行を含むマージはスキップ
        adj_min_r = min_r - 1 if min_r > 11 else min_r
        adj_max_r = max_r - 1 if max_r > 11 else max_r
        if adj_min_r > adj_max_r:
            continue
        try:
            ws.merge_cells(start_row=adj_min_r, end_row=adj_max_r,
                           start_column=min_c, end_column=max_c)
        except Exception:
            pass

    print("[OK  ] 対応方針: r11 空ヘッダ行を削除（r12以降が1行繰り上がり）")


# ── Step 2: 調査・影響範囲シート ──────────────────────────────────────────────

def patch_investigation_sheet(ws):
    """A1:C1 マージを追加する（値・fill はそのまま保持）。"""

    a1c1_exists = any(
        m.min_row == 1 and m.max_row == 1 and m.min_col == 1 and m.max_col >= 3
        for m in ws.merged_cells.ranges
    )
    if a1c1_exists:
        print("[SKIP] 調査・影響範囲: A1:C1 マージが既に存在 → スキップ")
        return

    ws.merge_cells("A1:C1")
    print("[OK  ] 調査・影響範囲: A1:C1 マージ追加")


# ── main ────────────────────────────────────────────────────────────────────

def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)
    print("=== patch_template_v10: Group N 残バグ修正 ===")
    print(f"現在のシート: {wb.sheetnames}\n")

    # Step 1: 対応方針シート
    if "対応方針" in wb.sheetnames:
        patch_approach_sheet(wb["対応方針"])
    else:
        print("[WARN] 対応方針シートが見つかりません")
    print()

    # Step 2: 調査・影響範囲シート
    if "調査・影響範囲" in wb.sheetnames:
        patch_investigation_sheet(wb["調査・影響範囲"])
    else:
        print("[WARN] 調査・影響範囲シートが見つかりません")
    print()

    # 保存
    try:
        wb.save(TEMPLATE)
        print(f"[OK  ] テンプレート保存完了: {TEMPLATE}")
    except Exception as e:
        print(f"[ERROR] 保存失敗: {e}")
        sys.exit(1)

    # ── 自動検証 ──────────────────────────────────────────────────────────────
    print("\n=== 自動検証 ===")
    wb2 = load_workbook(TEMPLATE)
    ok = True

    # 対応方針: r11 が空ヘッダ行でない
    if "対応方針" in wb2.sheetnames:
        ws_a = wb2["対応方針"]
        r11_cell_after = ws_a.cell(11, 1)
        r11_val_after = r11_cell_after.value
        r11_fill_after = ""
        if r11_cell_after.fill and r11_cell_after.fill.patternType == "solid" and r11_cell_after.fill.fgColor:
            r11_fill_after = r11_cell_after.fill.fgColor.rgb
        has_empty_hdr = (r11_val_after is None and r11_fill_after == _HEADER_FILL_RGB)
        _v(not has_empty_hdr,
           f"対応方針: r11 空ヘッダ行が存在しない（val={r11_val_after!r}, fill={r11_fill_after}）", ok)
        if has_empty_hdr:
            ok = False

    # 調査・影響範囲: A1:C1 マージ存在
    if "調査・影響範囲" in wb2.sheetnames:
        ws_i = wb2["調査・影響範囲"]
        a1c1 = any(
            m.min_row == 1 and m.max_row == 1 and m.min_col == 1 and m.max_col >= 3
            for m in ws_i.merged_cells.ranges
        )
        _v(a1c1, "調査・影響範囲: A1:C1 マージ存在", ok)
        if not a1c1:
            ok = False

    print("\n" + ("✅ 全検証 PASS" if ok else "❌ 一部 FAIL — 上記ログを確認してください"))
    sys.exit(0 if ok else 1)


def _v(cond, msg, _ok):
    status = "✅" if cond else "❌"
    print(f"  {status} {msg}")


if __name__ == "__main__":
    main()
