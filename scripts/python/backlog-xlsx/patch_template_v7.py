# -*- coding: utf-8 -*-
"""patch_template_v7.py
対応記録テンプレート.xlsx の対応方針シートを 6列→5列に変更する（方針名列削除）。

変更内容:
  - B列「方針名」を delete_cols で削除（C→B / D→C / E→D / F→E にシフト）
  - A:F の全幅マージを A:E に再構築

冪等性: r3 c2 が既に「概要」であればスキップする。

Usage:
    python patch_template_v7.py
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"


def patch_approach_sheet(ws):
    """対応方針シートの 方針名 列（B列）を削除して 5列化する。"""
    current_c2 = ws.cell(3, 2).value

    # 冪等チェック: 既に 5列化済みなら skip
    if str(current_c2 or "").strip() == "概要":
        print("[SKIP] 対応方針シート: 既に 5 列化済み（r3 c2 = '概要'）")
        return

    if str(current_c2 or "").strip() != "方針名":
        print(f"[WARN] r3 c2 の値が想定外: {current_c2!r} — スキップします")
        return

    # A:F の全幅マージ範囲を収集・解除
    full_width_merges = []
    for mcr in list(ws.merged_cells.ranges):
        if mcr.min_col == 1 and mcr.max_col == 6:
            full_width_merges.append((mcr.min_row, mcr.max_row))
            ws.merged_cells.ranges.discard(mcr)

    # B:F の部分マージ（実施前確認事項 B11:F11 相当）も収集・解除
    partial_merges = []
    for mcr in list(ws.merged_cells.ranges):
        if mcr.min_col == 2 and mcr.max_col == 6:
            partial_merges.append((mcr.min_row, mcr.max_row))
            ws.merged_cells.ranges.discard(mcr)

    # B列（方針名）を削除 → C..F が左に 1 シフトして B..E になる
    ws.delete_cols(2, 1)

    # 全幅マージを A:E（max_col=5）で再構築
    for (min_r, max_r) in full_width_merges:
        ws.merge_cells(start_row=min_r, end_row=max_r,
                       start_column=1, end_column=5)

    # 部分マージを B:E（旧 B:F から 1 列縮小）で再構築
    for (min_r, max_r) in partial_merges:
        ws.merge_cells(start_row=min_r, end_row=max_r,
                       start_column=2, end_column=5)

    print(
        f"[DEL ] 対応方針シート: B列「方針名」を削除\n"
        f"       全幅マージ A:F → A:E: {len(full_width_merges)} 箇所\n"
        f"       部分マージ B:F → B:E: {len(partial_merges)} 箇所"
    )


def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)

    if "対応方針" not in wb.sheetnames:
        print("[ERROR] 「対応方針」シートが見つかりません")
        sys.exit(1)

    print("=== patch_template_v7: 対応方針シート 5 列化 ===")
    patch_approach_sheet(wb["対応方針"])

    try:
        wb.save(TEMPLATE)
        print(f"\n[完了] テンプレート更新: {TEMPLATE}")
    except PermissionError as e:
        print(f"[ERROR] 保存失敗（Excel でファイルが開かれている可能性）: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
