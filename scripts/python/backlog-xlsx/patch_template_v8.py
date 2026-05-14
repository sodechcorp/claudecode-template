# -*- coding: utf-8 -*-
"""patch_template_v8.py
対応記録テンプレート.xlsx を Group L 仕様に移行する。

変更内容:
  1. 対応方針シート: 列幅調整（B=60, C=22, D=22, E=18）
  2. 調査・影響範囲シート: ■コード根拠テーブル・■関連コンポーネント一覧を削除
                          影響範囲テーブルの列ヘッダを刷新（No/種別/対象/問題ない根拠・対応内容）
  3. 対応内容シート: ■対応内容（言語記述）セクションをバックアップ情報の前に追加
  4. リリース・ロールバックシートを削除
  5. 残対応・懸念・保留シートを新設

冪等性: 各ステップは既に適用済みなら skip する。

Usage:
    python patch_template_v8.py
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"

_THIN = Side(border_style="thin", color="00B4C6E7")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="00C6EFCE")  # 薄緑（ヘッダ用）
_WRAP = Alignment(wrap_text=True, vertical="top")
_HEADER_FONT = Font(name="游ゴシック", size=10, bold=True)
_BODY_FONT   = Font(name="游ゴシック", size=10)


def _write_header_cell(ws, row, col, value, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = _WRAP
    c.font = _HEADER_FONT
    c.border = _THIN_BORDER
    if fill:
        c.fill = fill
    return c


def _write_body_cell(ws, row, col, value=""):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = _WRAP
    c.font = _BODY_FONT
    c.border = _THIN_BORDER
    return c


def _find_header_row(ws, keyword):
    """A列を走査してキーワードを含む最初の行番号を返す。見つからなければ None。"""
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and keyword in str(cell.value):
            return cell.row
    return None


# ── Step 1: 対応方針シート 列幅調整 ────────────────────────────────────────────

def patch_approach_col_widths(ws):
    current_b = ws.column_dimensions["B"].width or 0
    if current_b >= 55:
        print("[SKIP] 対応方針: B 列幅が既に 55 以上 → スキップ")
        return
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    print("[OK  ] 対応方針: 列幅調整 A=8 / B=60 / C=22 / D=22 / E=18")


# ── Step 2: 調査・影響範囲シート 再設計 ───────────────────────────────────────

def patch_investigation_sheet(ws):
    """コード根拠テーブルと関連コンポーネント一覧を削除し、影響範囲テーブルのヘッダを刷新。"""
    # 冪等チェック: 影響範囲テーブルの col ヘッダ行がすでに新形式かどうか
    impact_header = _find_header_row(ws, "■ 影響範囲テーブル") or _find_header_row(ws, "■ 影響範囲")
    if impact_header:
        col_hdr_row = impact_header + 1
        a_val = str(ws.cell(col_hdr_row, 1).value or "").strip()
        if a_val == "No":
            print("[SKIP] 調査・影響範囲: 既に新形式ヘッダ（A1='No'）→ スキップ")
            return

    # ── 2-1. ■コード根拠テーブル 行を削除 ──────────────────────────────────
    code_header = _find_header_row(ws, "■ コード根拠")
    if code_header:
        # ■影響範囲テーブルの直前まで削除（スナップショット方式）
        impact_header = _find_header_row(ws, "■ 影響範囲テーブル") or _find_header_row(ws, "■ 影響範囲")
        if impact_header and impact_header > code_header:
            delete_count = impact_header - code_header
            # merged cells を全クリアしてから delete_rows
            for mcr in list(ws.merged_cells.ranges):
                ws.merged_cells.ranges.discard(mcr)
            ws.delete_rows(code_header, delete_count)
            print(f"[OK  ] 調査・影響範囲: ■コード根拠テーブル行を削除 (r{code_header}〜r{code_header + delete_count - 1})")
        else:
            print("[WARN] ■コード根拠テーブルの削除範囲を特定できません")

    # ── 2-2. ■関連コンポーネント一覧 行を削除 ──────────────────────────────
    comp_header = _find_header_row(ws, "■ 関連コンポーネント")
    if comp_header:
        # comp_header 以降を全削除
        total_rows = ws.max_row
        delete_count = total_rows - comp_header + 1
        for mcr in list(ws.merged_cells.ranges):
            ws.merged_cells.ranges.discard(mcr)
        ws.delete_rows(comp_header, delete_count)
        print(f"[OK  ] 調査・影響範囲: ■関連コンポーネント一覧を削除 (r{comp_header}〜r{total_rows})")

    # ── 2-3. 影響範囲テーブルの列ヘッダを刷新 ────────────────────────────────
    impact_header = _find_header_row(ws, "■ 影響範囲テーブル") or _find_header_row(ws, "■ 影響範囲")
    if not impact_header:
        print("[WARN] ■影響範囲テーブルが見つかりません")
        return
    col_hdr_row = impact_header + 1

    # 既存 C:D マージを解除
    for mcr in list(ws.merged_cells.ranges):
        if mcr.min_row == col_hdr_row and mcr.max_row == col_hdr_row:
            ws.merged_cells.ranges.discard(mcr)

    # 新ヘッダ書き込み: No / 種別 / 対象 / 問題ない根拠・対応内容
    _write_header_cell(ws, col_hdr_row, 1, "No")
    _write_header_cell(ws, col_hdr_row, 2, "種別")
    _write_header_cell(ws, col_hdr_row, 3, "対象")
    _write_header_cell(ws, col_hdr_row, 4, "問題ない根拠・対応内容")

    # データ行のヘッダも清書（A 列の旧「種別」値をクリア）
    data_start = col_hdr_row + 1
    for r in range(data_start, ws.max_row + 1):
        a_val = ws.cell(r, 1).value
        if a_val and str(a_val).strip() in ("種別", "ファイル"):
            ws.cell(r, 1).value = None

    print(f"[OK  ] 調査・影響範囲: 影響範囲テーブルのヘッダを刷新 (r{col_hdr_row})")

    # 列幅調整
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 50


# ── Step 3: 対応内容シート 言語記述セクション追加 ──────────────────────────────

def patch_content_sheet(ws):
    """■対応内容（言語記述）セクションをバックアップ情報の前に挿入する。"""
    # 冪等チェック
    lang_header = _find_header_row(ws, "■ 対応内容（言語記述）")
    if lang_header:
        print("[SKIP] 対応内容: ■対応内容（言語記述）セクション既存 → スキップ")
        return

    backup_header = _find_header_row(ws, "■ バックアップ情報")
    if not backup_header:
        print("[WARN] 対応内容: ■バックアップ情報が見つかりません → スキップ")
        return

    # バックアップ情報の直前に 5 行挿入
    # (■ 対応内容（言語記述）ヘッダ 1行 + コンテンツ 3行 + spacer 1行)
    INSERT_AT = backup_header
    INSERT_COUNT = 5

    # merged cells スナップショット → クリア → 挿入 → 再構築
    all_merges = [(m.min_row, m.max_row, m.min_col, m.max_col)
                  for m in list(ws.merged_cells.ranges)]
    for mcr in list(ws.merged_cells.ranges):
        ws.merged_cells.ranges.discard(mcr)

    ws.insert_rows(INSERT_AT, INSERT_COUNT)

    # マージ再構築
    for (min_r, max_r, min_c, max_c) in all_merges:
        if min_r >= INSERT_AT:
            ws.merge_cells(start_row=min_r + INSERT_COUNT, end_row=max_r + INSERT_COUNT,
                           start_column=min_c, end_column=max_c)
        elif max_r >= INSERT_AT:
            ws.merge_cells(start_row=min_r, end_row=max_r + INSERT_COUNT,
                           start_column=min_c, end_column=max_c)
        else:
            ws.merge_cells(start_row=min_r, end_row=max_r,
                           start_column=min_c, end_column=max_c)

    # ■ 対応内容（言語記述）ヘッダ行
    h_row = INSERT_AT
    ws.row_dimensions[h_row].height = 24
    hcell = ws.cell(h_row, 1, value="■ 対応内容（言語記述）")
    hcell.font = Font(name="游ゴシック", size=10, bold=True)
    hcell.alignment = _WRAP
    hcell.border = _THIN_BORDER
    ws.merge_cells(start_row=h_row, end_row=h_row, start_column=1, end_column=4)

    # コンテンツ行（3行）: A:D マージ・空のデータ行
    for r in range(h_row + 1, h_row + 4):
        ws.row_dimensions[r].height = 40
        c = ws.cell(r, 1, value="")
        c.alignment = _WRAP
        c.font = _BODY_FONT
        c.border = _THIN_BORDER
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=4)

    # spacer
    ws.row_dimensions[h_row + 4].height = 6

    print(f"[OK  ] 対応内容: ■対応内容（言語記述）セクションを r{INSERT_AT} に挿入 (5行)")


# ── Step 4: リリース・ロールバックシートを削除 ──────────────────────────────────

def patch_delete_release_sheet(wb):
    for name in list(wb.sheetnames):
        if "リリース" in name or "ロールバック" in name:
            del wb[name]
            print(f"[OK  ] シート削除: '{name}'")
            return
    print("[SKIP] リリース・ロールバックシートが既に存在しない → スキップ")


# ── Step 5: 残対応・懸念・保留シートを新設 ──────────────────────────────────────

def patch_add_pending_sheet(wb):
    SHEET_NAME = "残対応・懸念・保留"
    if SHEET_NAME in wb.sheetnames:
        print(f"[SKIP] '{SHEET_NAME}' シートが既に存在 → スキップ")
        return

    ws = wb.create_sheet(SHEET_NAME)

    # タイトル行
    ws.row_dimensions[1].height = 28
    t = ws.cell(1, 1, value=SHEET_NAME)
    t.font = Font(name="游ゴシック", size=12, bold=True)
    t.alignment = _WRAP
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)

    # ■ 残対応・懸念事項一覧 ヘッダ
    ws.row_dimensions[2].height = 22
    hcell = ws.cell(2, 1, value="■ 残対応・懸念事項一覧")
    hcell.font = Font(name="游ゴシック", size=10, bold=True)
    hcell.alignment = _WRAP
    hcell.border = _THIN_BORDER
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=6)

    # 列ヘッダ行
    headers = ["No", "種別", "内容", "関連", "ステータス", "次アクション"]
    ws.row_dimensions[3].height = 22
    for col, h in enumerate(headers, start=1):
        _write_header_cell(ws, 3, col, h)

    # データ行テンプレート（3行）
    kinds = ["懸念", "許容した影響", "後回しの残対応"]
    statuses = ["未対応", "許容済", "保留"]
    for i in range(3):
        r = 4 + i
        ws.row_dimensions[r].height = 40
        _write_body_cell(ws, r, 1, str(i + 1))
        _write_body_cell(ws, r, 2, kinds[i])
        _write_body_cell(ws, r, 3, "")
        _write_body_cell(ws, r, 4, "")
        _write_body_cell(ws, r, 5, statuses[i])
        _write_body_cell(ws, r, 6, "")

    # 列幅
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 30

    print(f"[OK  ] '{SHEET_NAME}' シートを新設")


# ── テスト・検証記録 → テスト・検証 リネーム ──────────────────────────────────

def patch_rename_test_sheet(wb):
    OLD = "テスト・検証記録"
    NEW = "テスト・検証"
    if NEW in wb.sheetnames:
        print(f"[SKIP] '{NEW}' シートが既に存在 → スキップ")
        return
    if OLD in wb.sheetnames:
        wb[OLD].title = NEW
        print(f"[OK  ] シートリネーム: '{OLD}' → '{NEW}'")
    else:
        print(f"[SKIP] '{OLD}' シートが見つからない → スキップ")


# ── main ────────────────────────────────────────────────────────────────────

def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)
    print(f"=== patch_template_v8: Group L テンプレート全面刷新 ===")
    print(f"現在のシート: {wb.sheetnames}\n")

    # Step 1: 対応方針 列幅
    if "対応方針" in wb.sheetnames:
        patch_approach_col_widths(wb["対応方針"])
    else:
        print("[WARN] 対応方針シートが見つかりません")

    # Step 2: 調査・影響範囲 再設計
    if "調査・影響範囲" in wb.sheetnames:
        patch_investigation_sheet(wb["調査・影響範囲"])
    else:
        print("[WARN] 調査・影響範囲シートが見つかりません")

    # Step 3: 対応内容 言語記述追加
    if "対応内容" in wb.sheetnames:
        patch_content_sheet(wb["対応内容"])
    else:
        print("[WARN] 対応内容シートが見つかりません")

    # Step 4: リリース・ロールバック削除
    patch_delete_release_sheet(wb)

    # Step 5: テスト・検証記録 リネーム
    patch_rename_test_sheet(wb)

    # Step 6: 残対応・懸念・保留 新設
    patch_add_pending_sheet(wb)

    try:
        wb.save(TEMPLATE)
        print(f"\n[完了] テンプレート更新: {TEMPLATE}")
        print(f"更新後のシート: {wb.sheetnames}")
    except PermissionError as e:
        print(f"[ERROR] 保存失敗（Excel でファイルが開かれている可能性）: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
