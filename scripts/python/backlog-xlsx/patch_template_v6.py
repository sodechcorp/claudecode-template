# -*- coding: utf-8 -*-
"""patch_template_v6.py
対応記録テンプレート.xlsx に以下の修正を一度だけ適用する:
  1. サマリー・経緯シート: 「修正前（現状）」「修正後（期待挙動）」行を削除
  2. リリース・ロールバックシート: 「■ リリース前確認事項」セクション削除
  3. リリース・ロールバックシート: 「■ デプロイ後確認事項」セクション削除
  4. リリース・ロールバックシート: 「■ リリース実施記録」セクションを末尾に追加（スタイル統一）

Usage:
    python patch_template_v6.py
"""

import copy
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"

WRAP = Alignment(wrap_text=True, vertical="top")

# テンプレートの共通ヘッダー背景色（既存セクションヘッダーから採用: 淡い青灰 = 0xD0DCE4 相当）
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
COL_HEADER_FILL = PatternFill(start_color="E8EDF4", end_color="E8EDF4", fill_type="solid")


def find_header_row(ws, keywords, start_row=1):
    """キーワードをいずれか含む最初の行番号を返す。見つからなければ None。"""
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.value and any(kw in str(cell.value) for kw in keywords):
                return cell.row
    return None


def find_label_row(ws, label, col=1, start_row=1):
    """col 列で label に完全一致する行を返す。"""
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v and str(v).strip() == label.strip():
            return r
    return None


def safe_delete_rows(ws, row_start, count):
    """マージ保全付き行削除。"""
    if count <= 0:
        return
    row_end = row_start + count - 1
    all_merges = [(m.min_row, m.max_row, m.min_col, m.max_col)
                  for m in list(ws.merged_cells.ranges)]
    for mcr in list(ws.merged_cells.ranges):
        ws.merged_cells.ranges.discard(mcr)

    ws.delete_rows(row_start, count)

    for (min_r, max_r, min_c, max_c) in all_merges:
        if min_r >= row_start and max_r <= row_end:
            continue  # 削除範囲内: 破棄
        elif max_r < row_start:
            ws.merge_cells(start_row=min_r, end_row=max_r,
                           start_column=min_c, end_column=max_c)
        elif min_r > row_end:
            ws.merge_cells(start_row=min_r - count, end_row=max_r - count,
                           start_column=min_c, end_column=max_c)
        else:
            # 部分重複は削除範囲外にクリップして維持
            new_min = min_r if min_r < row_start else row_start
            new_max = max_r - count if max_r > row_end else row_start - 1
            if new_max >= new_min:
                ws.merge_cells(start_row=new_min, end_row=new_max,
                               start_column=min_c, end_column=max_c)


def _get_section_font(ws):
    """既存セクションヘッダー（■ で始まるセル）からフォントを取得する。"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith("■") and cell.font and cell.font.name:
                return cell.font
    return Font(name="游ゴシック", size=10, bold=True)


def _delete_section(ws, section_keywords, next_keywords, label=""):
    """セクションヘッダーから次のセクションの直前まで行を削除する。"""
    start = find_header_row(ws, section_keywords)
    if start is None:
        print(f"[SKIP] {label}: ヘッダーが見つかりません（既に削除済みの可能性）")
        return
    end = find_header_row(ws, next_keywords, start_row=start + 1)
    if end is None:
        # 最後のセクションの場合はシートの末尾まで
        end = ws.max_row + 1
    count = end - start
    print(f"[DEL ] {label}: 行 {start}〜{end - 1} ({count} 行) を削除")
    safe_delete_rows(ws, start, count)


def patch_summary_sheet(ws):
    """サマリー・経緯シート: 修正前/修正後 行を削除。"""
    before_row = find_label_row(ws, "修正前（現状）")
    after_row = find_label_row(ws, "修正後（期待挙動）")

    if before_row is None and after_row is None:
        print("[SKIP] サマリー・経緯: 修正前/修正後 行が見つかりません（既に削除済み）")
        return

    # 連続している前提で上から削除（before → after の順）
    # before_row を先に削除すると after_row が 1 シフトするため調整
    if before_row and after_row:
        if before_row < after_row:
            # before を先に 1 行削除、after は -1 シフト
            print(f"[DEL ] サマリー・経緯: 行 {before_row}「修正前（現状）」を削除")
            safe_delete_rows(ws, before_row, 1)
            after_row_adjusted = after_row - 1
            print(f"[DEL ] サマリー・経緯: 行 {after_row_adjusted}「修正後（期待挙動）」を削除")
            safe_delete_rows(ws, after_row_adjusted, 1)
        else:
            print(f"[DEL ] サマリー・経緯: 行 {after_row}「修正後（期待挙動）」を削除")
            safe_delete_rows(ws, after_row, 1)
            before_row_adjusted = before_row - 1
            print(f"[DEL ] サマリー・経緯: 行 {before_row_adjusted}「修正前（現状）」を削除")
            safe_delete_rows(ws, before_row_adjusted, 1)
    elif before_row:
        print(f"[DEL ] サマリー・経緯: 行 {before_row}「修正前（現状）」を削除")
        safe_delete_rows(ws, before_row, 1)
    elif after_row:
        print(f"[DEL ] サマリー・経緯: 行 {after_row}「修正後（期待挙動）」を削除")
        safe_delete_rows(ws, after_row, 1)


def patch_release_sheet(ws):
    """リリース・ロールバックシート: 不要セクション削除 + リリース実施記録追加。"""
    # 削除は下から上の順（行番号シフトを防ぐ）
    # 1. デプロイ後確認事項 削除
    _delete_section(
        ws,
        section_keywords=("■ デプロイ後確認事項",),
        next_keywords=("■ 注意事項", "■ ロールバック手順", "■ リリース実施記録"),
        label="デプロイ後確認事項",
    )
    # 2. リリース前確認事項 削除
    _delete_section(
        ws,
        section_keywords=("■ リリース前確認事項",),
        next_keywords=("■ デプロイ手順", "■ デプロイ後確認", "■ 注意事項", "■ ロールバック手順"),
        label="リリース前確認事項",
    )

    # 3. リリース実施記録セクションが既に存在するか確認
    existing = find_header_row(ws, ("■ リリース実施記録",))
    if existing:
        print(f"[SKIP] リリース実施記録: 既に行 {existing} に存在します → スタイルのみ修正")
        _fix_release_record_style(ws, existing)
        return

    # 4. リリース実施記録セクションを末尾に追加
    _add_release_record_section(ws)


def _fix_release_record_style(ws, header_row):
    """既存のリリース実施記録セクションのスタイルを修正する。"""
    base_font = _get_section_font(ws)
    max_col = 4

    # ヘッダー行のフォント・マージ修正
    h_cell = ws.cell(header_row, 1)
    h_cell.font = Font(name=base_font.name, size=base_font.size or 10, bold=True)
    h_cell.alignment = WRAP
    if not any(m.min_row == header_row and m.min_col == 1 and m.max_col == max_col
               for m in ws.merged_cells.ranges):
        ws.merge_cells(start_row=header_row, end_row=header_row, start_column=1, end_column=max_col)

    # 列ヘッダー行のフォント修正
    col_header_row = header_row + 1
    for c, label in enumerate(["No", "実施日時", "対象環境", "デプロイ結果"], start=1):
        cell = ws.cell(col_header_row, c)
        if cell.value is None:
            cell.value = label
        cell.font = Font(name=base_font.name, size=base_font.size or 10, bold=True)
        cell.alignment = WRAP

    # データ行のフォント修正（Calibri になっている場合に游ゴシックへ統一）
    data_font = Font(name=base_font.name, size=base_font.size or 10, bold=False)
    data_row = col_header_row + 1
    while data_row <= ws.max_row:
        v = ws.cell(data_row, 1).value
        if v is not None and str(v).startswith("■"):
            break  # 次のセクションヘッダーに到達
        for c in range(1, max_col + 1):
            cell = ws.cell(data_row, c)
            if cell.font.name != base_font.name:
                cell.font = data_font
            cell.alignment = WRAP
        data_row += 1

    print(f"[FIX ] リリース実施記録: 行 {header_row} のスタイルを修正しました")


def _add_release_record_section(ws):
    """リリース実施記録セクションをロールバック手順の後に追加する。"""
    rb_header = find_header_row(ws, ("■ ロールバック手順",))
    if rb_header is None:
        print("[WARN] ロールバック手順ヘッダーが見つかりません → max_row の後に追加します")
        insert_at = ws.max_row + 2
    else:
        # ロールバック手順の後ろを探す（次のセクションヘッダーまたはシート末尾）
        next_sec = find_header_row(ws, ("■ リリース実施記録",), start_row=rb_header + 1)
        if next_sec:
            insert_at = ws.max_row + 2  # 既にあればその後ろ
        else:
            insert_at = ws.max_row + 2  # シート末尾の次の行

    base_font = _get_section_font(ws)
    max_col = 4

    # スペーサー行
    spacer_row = insert_at - 1
    if spacer_row >= 1:
        ws.row_dimensions[spacer_row].height = 6

    # ヘッダー行: ■ リリース実施記録
    h_row = insert_at
    h_cell = ws.cell(h_row, 1, value="■ リリース実施記録")
    h_cell.font = Font(name=base_font.name, size=base_font.size or 10, bold=True)
    h_cell.fill = HEADER_FILL
    h_cell.alignment = WRAP
    ws.merge_cells(start_row=h_row, end_row=h_row, start_column=1, end_column=max_col)
    ws.row_dimensions[h_row].height = 20

    # 列ヘッダー行
    ch_row = h_row + 1
    for c, label in enumerate(["No", "実施日時", "対象環境", "デプロイ結果"], start=1):
        cell = ws.cell(ch_row, c, value=label)
        cell.font = Font(name=base_font.name, size=base_font.size or 10, bold=True)
        cell.fill = COL_HEADER_FILL
        cell.alignment = WRAP
    ws.row_dimensions[ch_row].height = 18

    # データ行 2 件（ステージング / 本番）
    data_font = Font(name=base_font.name, size=base_font.size or 10, bold=False)
    envs = ["ステージング", "本番"]
    for i, env in enumerate(envs):
        d_row = ch_row + 1 + i
        for c, val in enumerate(["", "", env, ""], start=1):
            cell = ws.cell(d_row, c, value=val if val else None)
            cell.alignment = WRAP
            cell.font = data_font
        ws.row_dimensions[d_row].height = 28

    print(f"[ADD ] リリース実施記録: 行 {h_row}〜{ch_row + len(envs)} に追加しました")


def patch_kaito_sheet(ws):
    """対応内容シート: Before/After 見出しから「（実装後に記入）」を削除。"""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and "■ Before / After" in str(v) and "（実装後に記入）" in str(v):
            ws.cell(r, 1).value = "■ Before / After"
            print(f"[FIX ] 対応内容: row {r} を「■ Before / After」に変更")
            return
    print("[SKIP] 対応内容: 「（実装後に記入）」が見つかりません（既に修正済みの可能性）")


def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)

    print("=== patch_template_v6: サマリー・経緯 ===")
    patch_summary_sheet(wb["サマリー・経緯"])

    print("\n=== patch_template_v6: リリース・ロールバック ===")
    patch_release_sheet(wb["リリース・ロールバック"])

    print("\n=== patch_template_v6: 対応内容 ===")
    patch_kaito_sheet(wb["対応内容"])

    try:
        wb.save(TEMPLATE)
        print(f"\n[完了] テンプレート更新: {TEMPLATE}")
    except PermissionError as e:
        print(f"[ERROR] 保存失敗（Excel でファイルが開かれている可能性）: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
