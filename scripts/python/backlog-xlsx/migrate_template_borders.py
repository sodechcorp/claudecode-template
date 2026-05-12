# -*- coding: utf-8 -*-
"""migrate_template_borders.py
対応記録テンプレート.xlsx に以下の修正を一度だけ適用する:
  1. サマリー・経緯: r11-r15（対応サマリー行）に thin border 追加
  2. サマリー・経緯: r18-r21（判断保留事項 列ヘッダ+データ行）に thin border 追加
  3. 調査・影響範囲: r2-r9（仮説検証ブロック）を delete_rows で削除
  4. 全 6 シートの tabColor を None（無色）にクリア

Usage:
    python migrate_template_borders.py
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles.borders import Border, Side
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"

_thin = Side(style="thin")
_FULL  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_TB    = Border(top=_thin, bottom=_thin)            # merged interior
_RTB   = Border(right=_thin, top=_thin, bottom=_thin)  # merged right-end


def _apply_row_border_summary(ws, row):
    """サマリー・経緯のデータ行（A 単独 + B:F マージ）パターンで thin border を適用。
    r3-r9 の border 設定を模倣。
    """
    ws.cell(row, 1).border = _FULL
    ws.cell(row, 2).border = _FULL   # マージ先頭: left+top+bottom は確実に
    ws.cell(row, 3).border = _TB
    ws.cell(row, 4).border = _TB
    ws.cell(row, 5).border = _TB
    ws.cell(row, 6).border = _RTB


def _apply_row_border_plain(ws, row, max_col):
    """独立セル（マージなし）の行に thin border を全列適用。"""
    for col in range(1, max_col + 1):
        ws.cell(row, col).border = _FULL


def _delete_rows_with_merge_fix(ws, first_row, count):
    """delete_rows + マージ再構築（snapshot → clear → delete → rebuild）。
    _shrink_table (create_records.py:308-342) と同じパターン。
    """
    all_merges = [
        (m.min_row, m.max_row, m.min_col, m.max_col)
        for m in list(ws.merged_cells.ranges)
    ]
    for mcr in list(ws.merged_cells.ranges):
        ws.merged_cells.ranges.discard(mcr)

    ws.delete_rows(first_row, count)

    last_deleted = first_row + count - 1
    for (min_r, max_r, min_c, max_c) in all_merges:
        if min_r >= first_row and max_r <= last_deleted:
            # 削除範囲に完全収容 → 破棄
            continue
        elif min_r >= first_row:
            # 削除位置以降 → シフト
            ws.merge_cells(start_row=min_r - count, end_row=max_r - count,
                           start_column=min_c, end_column=max_c)
        elif max_r >= first_row:
            # 削除範囲またがり → 上端だけ残してクリップ
            new_max_r = max(min_r, first_row - 1)
            ws.merge_cells(start_row=min_r, end_row=new_max_r,
                           start_column=min_c, end_column=max_c)
        else:
            # 削除位置より前 → そのまま
            ws.merge_cells(start_row=min_r, end_row=max_r,
                           start_column=min_c, end_column=max_c)


def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)

    # ── 1. サマリー・経緯: r11-r15 に thin border（対応サマリー行）──────────
    ws_sum = wb["サマリー・経緯"]
    for r in range(11, 16):  # r11, r12, r13, r14, r15
        _apply_row_border_summary(ws_sum, r)
    print("[OK] サマリー・経緯 r11-r15 border 適用")

    # ── 2. サマリー・経緯: r18-r21 に thin border（判断保留事項）────────────
    # r18 = 列ヘッダ（No/内容/影響範囲/期待する判断者/関連ファイル 5列）
    # r19-r21 = データ行 (5列)
    for r in range(18, 22):  # r18, r19, r20, r21
        _apply_row_border_plain(ws_sum, r, max_col=5)
    print("[OK] サマリー・経緯 r18-r21 border 適用")

    # ── 3. 調査・影響範囲: r2-r9 削除（仮説検証ブロック）──────────────────
    ws_inv = wb["調査・影響範囲"]
    _delete_rows_with_merge_fix(ws_inv, first_row=2, count=8)
    print("[OK] 調査・影響範囲 r2-r9 削除（仮説検証ブロック）")

    # ── 4. 全シートの tabColor をクリア（無色化）──────────────────────────
    for sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_properties.tabColor = None
    print(f"[OK] 全 {len(wb.sheetnames)} シートの tabColor クリア")

    # ── 保存 ─────────────────────────────────────────────────────────────────
    try:
        wb.save(TEMPLATE)
        print(f"\n[完了] テンプレート更新: {TEMPLATE}")
    except PermissionError as e:
        print(f"[ERROR] 保存失敗（Excel でファイルが開かれている可能性）: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
