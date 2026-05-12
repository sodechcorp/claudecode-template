# -*- coding: utf-8 -*-
"""patch_template_v5.py
対応記録テンプレート.xlsx に以下の修正を一度だけ適用する:
  1. 調査・影響範囲シート コード根拠テーブル: D 列「説明」ヘッダーをクリア
  2. リリース・ロールバックシート リリース対象テーブル: D 列「デプロイ方法」ヘッダーをクリア

Usage:
    python patch_template_v5.py
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"


def find_header_row(ws, keywords):
    """キーワードをいずれか含む最初の行番号を返す。見つからなければ None。"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and any(kw in str(cell.value) for kw in keywords):
                return cell.row
    return None


def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)

    # ── 1. 調査・影響範囲シート コード根拠テーブル D 列ヘッダーをクリア ──────
    ws_inv = wb["調査・影響範囲"]
    code_header = find_header_row(ws_inv, ("■ コード根拠テーブル", "■ コード根拠"))
    if code_header:
        col_header_row = code_header + 1  # 列見出し行
        cell = ws_inv.cell(col_header_row, 4)
        old_val = cell.value
        cell.value = None
        print(f"[OK] 調査・影響範囲 r{col_header_row} D 列「{old_val}」→ クリア")
        # データ行（標準 5 行分）も念のちクリア
        data_start = code_header + 2
        for r in range(data_start, data_start + 5):
            ws_inv.cell(r, 4).value = None
    else:
        print("[WARN] 調査・影響範囲シートのコード根拠テーブルヘッダーが見つかりませんでした")

    # ── 2. リリース・ロールバックシート リリース対象テーブル D 列ヘッダーをクリア ──
    ws_rel = wb["リリース・ロールバック"]
    # リリース対象テーブルの列見出しは RELEASE_START(=4) の 1 つ前 = r3
    RELEASE_COL_HEADER_ROW = 3
    cell = ws_rel.cell(RELEASE_COL_HEADER_ROW, 4)
    old_val = cell.value
    cell.value = None
    print(f"[OK] リリース・ロールバック r{RELEASE_COL_HEADER_ROW} D 列「{old_val}」→ クリア")
    # データ行（標準 2 行分）も念のちクリア
    for r in range(4, 6):
        ws_rel.cell(r, 4).value = None

    # ── 保存 ──────────────────────────────────────────────────────────────────
    try:
        wb.save(TEMPLATE)
        print(f"\n[完了] テンプレート更新: {TEMPLATE}")
    except PermissionError as e:
        print(f"[ERROR] 保存失敗（Excel でファイルが開かれている可能性）: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
