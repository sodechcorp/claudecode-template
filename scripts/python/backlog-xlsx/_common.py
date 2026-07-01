# -*- coding: utf-8 -*-
"""backlog-xlsx 共通ユーティリティ"""
import os
import re
import zipfile
from pathlib import Path

try:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# リリース・ロールバックシート「リリース実施記録」の開始行（xlsx テンプレート固定値）
# テンプレートを更新した場合はここを変更するだけで全スクリプトに反映される
RELEASE_HISTORY_START_ROW = 38

_STRIPE_A_RGB = "FFFFFF"  # 偶数インデックス行（0, 2, 4, ...）
_STRIPE_B_RGB = "F2F7FB"  # 奇数インデックス行（1, 3, 5, ...）（薄青）


def _stripe_fill(i):
    """0-indexed の行番号 i に対応する縞模様 PatternFill を毎回 fresh に生成して返す。

    openpyxl の style index aliasing バグ（singleton を使うと白代入が青セルで
    silent no-op になる）を回避するため、呼び出し毎に新規インスタンスを生成する。
    """
    rgb = _STRIPE_A_RGB if i % 2 == 0 else _STRIPE_B_RGB
    return PatternFill("solid", fgColor=rgb)


# ── 書式ヘルパー（游ゴシック統一） ───────────────────────────────────────────
# sf-doc-mcp/writer.py の _font/_align/_thin/_medium 相当を移植。
# generate_evidence_xlsx.py 等が import して Calibri 混在を解消する。

def _font(fg: str = "1A1A1A", bold: bool = False, size: int = 10,
          name: str = "游ゴシック", italic: bool = False,
          underline: str = None, color: str = None) -> "Font":
    """游ゴシック既定のフォント生成。color は fg の別名（互換）。"""
    actual_fg = color if color else fg
    kwargs = dict(color=actual_fg, bold=bold, size=size, name=name, italic=italic)
    if underline:
        kwargs["underline"] = underline
    return Font(**kwargs)


def _align(h: str = "left", v: str = "top", wrap: bool = True) -> "Alignment":
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _thin_side(color: str = "AAAAAA") -> "Side":
    return Side(style="thin", color=color)


def _medium_side(color: str = "888888") -> "Side":
    return Side(style="medium", color=color)


def _thin_border(color: str = "AAAAAA") -> "Border":
    s = _thin_side(color)
    return Border(left=s, right=s, top=s, bottom=s)


def _set_row_height(ws, row: int, h: float) -> None:
    ws.row_dimensions[row].height = h


def _set_col_width(ws, col: int, w: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = w


def _freeze(ws, row: int) -> None:
    ws.freeze_panes = f"A{row}"


def _auto_col_width(ws, col: int, min_w: float = 8, max_w: float = 40) -> None:
    """列内の最大文字数から列幅を推算して設定する（上限・下限付き）。
    openpyxl は実フォント幅を計算できないため、文字数×係数の近似値を使う。
    """
    max_len = 0
    col_letter = get_column_letter(col)
    for cell in ws[col_letter]:
        if cell.value:
            # 改行を含む場合は最大行の長さを使う
            cell_len = max(len(str(line)) for line in str(cell.value).splitlines())
            max_len = max(max_len, cell_len)
    # 全角文字（CJK）は幅 2 なので係数 1.8 で近似
    estimated = max_len * 1.8
    ws.column_dimensions[col_letter].width = max(min_w, min(estimated, max_w))


def _row_height_by_lines(text: str, col_width: float = 30, base_pt: float = 15) -> float:
    """折り返し行数から行高を推算して返す（折り返し列幅を col_width 文字と仮定）。

    1行あたり base_pt pt として、改行＋折り返し行数分の高さを返す。
    """
    if not text:
        return base_pt * 1.5
    lines = str(text).splitlines()
    total = 0
    for line in lines:
        # 全角文字は幅 2 として折り返し行数を推算
        line_width = sum(2 if ord(c) > 0x7F else 1 for c in line)
        wrapped = max(1, (line_width + int(col_width) - 1) // int(col_width))
        total += wrapped
    return max(base_pt * 1.5, base_pt * total)


# ── xlsx 保存後パッチ（openpyxl CellRichText の xml:space 欠落バグ対策） ──────
# openpyxl 3.1.5 は CellRichText の run を書き出す際、run のテキストが
# 空白のみ／先頭or末尾が空白の場合に xml:space="preserve" を付け忘れる
# （lxml ライター経由でも純Python etree ライターでも発生）。
# Excel は XML の仕様どおり両端の空白を除去して解釈するため「文字列が変更
# された」と判定し、開いた瞬間に "修復が必要" ダイアログを出す。
# 保存後の xlsx を ZIP として開き、該当する <t> 要素へ属性を注入して是正する。
_T_TAG_RE = re.compile(r'<t((?:\s+[^<>]*)?)>([^<]*)</t>')


def _t_needs_preserve(attrs: str, inner: str) -> bool:
    return bool(inner) and (inner[0] in " \t\r\n" or inner[-1] in " \t\r\n") and "xml:space" not in attrs


def patch_preserve_space(xlsx_path: str) -> int:
    """保存済み xlsx 内の worksheet / sharedStrings パーツを走査し、
    先頭・末尾が空白なのに xml:space="preserve" が無い <t> 要素を修正して
    ファイルを書き戻す。戻り値は修正した <t> 要素数（0 なら無修正）。
    """
    fixed_total = 0

    def _fix(m: "re.Match") -> str:
        nonlocal fixed_total
        attrs, inner = m.group(1), m.group(2)
        if _t_needs_preserve(attrs, inner):
            fixed_total += 1
            return f'<t{attrs} xml:space="preserve">{inner}</t>'
        return m.group(0)

    tmp_path = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path) as zin, \
         zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            is_target = (
                (item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"))
                or item.filename == "xl/sharedStrings.xml"
            )
            if is_target:
                text = _T_TAG_RE.sub(_fix, data.decode("utf-8"))
                data = text.encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp_path, xlsx_path)
    return fixed_total


def validate_folder(value: str) -> str:
    """--folder 引数のサニティチェック。プレースホルダー残留・相対パスを early-exit。"""
    if "{" in value or "}" in value:
        raise SystemExit(
            f"[FATAL] placeholder not resolved: {value!r}\n"
            "        /backlog Phase 1.5 に戻って {xlsx_folder} を実値で置換してください。"
        )
    p = Path(value)
    if not p.is_absolute():
        raise SystemExit(
            f"[FATAL] --folder must be absolute path: {value!r}"
        )
    return str(p)
