# -*- coding: utf-8 -*-
"""patch_template_v9.py
対応記録テンプレート.xlsx を Group M 仕様に移行する。

変更内容:
  1. 対応内容シート: ■影響確認チェックリスト削除・言語記述エリアを A3:D5 一括マージ
  2. 調査・影響範囲シート: 種別列（B列）削除→3列化・A2:C2 マージ・列幅調整
  3. 残対応・懸念・保留シート: 関連列（D列）削除→5列化・A1 タイトル青背景・A2:E2 マージ
  4. テスト・検証シート: A1 セル値「テスト・検証記録」→「テスト・検証」・注記行挿入

冪等性: 各ステップは既に適用済みなら skip する。

Usage:
    python patch_template_v9.py
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"

_THIN = Side(border_style="thin", color="00B4C6E7")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TITLE_FILL    = PatternFill("solid", fgColor="001F4E79")  # 濃紺（タイトル行）
_SECTION_FILL  = PatternFill("solid", fgColor="00D6E4F0")  # 薄青（■セクション見出し）
_HEADER_FILL   = PatternFill("solid", fgColor="002E75B6")  # 中青（列ヘッダ）
_WRAP = Alignment(wrap_text=True, vertical="top")
_TITLE_FONT  = Font(name="游ゴシック", size=14, bold=True, color="FFFFFFFF")
_HEADER_FONT = Font(name="游ゴシック", size=10, bold=True)
_BODY_FONT   = Font(name="游ゴシック", size=10)


def _find_header_row(ws, keyword):
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and keyword in str(cell.value):
            return cell.row
    return None


def _write_header_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = _WRAP
    c.font = _HEADER_FONT
    c.fill = _HEADER_FILL
    c.border = _THIN_BORDER
    return c


def _write_body_cell(ws, row, col, value=""):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = _WRAP
    c.font = _BODY_FONT
    c.border = _THIN_BORDER
    return c


def _clear_merges(ws):
    for mcr in list(ws.merged_cells.ranges):
        ws.merged_cells.ranges.discard(mcr)


def _save_merges(ws):
    return [(m.min_row, m.max_row, m.min_col, m.max_col)
            for m in list(ws.merged_cells.ranges)]


def _restore_merges(ws, merges):
    for (min_r, max_r, min_c, max_c) in merges:
        try:
            ws.merge_cells(start_row=min_r, end_row=max_r,
                           start_column=min_c, end_column=max_c)
        except Exception:
            pass


def _safe_merge(ws, min_row, max_row, min_col, max_col):
    for mcr in list(ws.merged_cells.ranges):
        if (mcr.min_row <= max_row and mcr.max_row >= min_row and
                mcr.min_col <= max_col and mcr.max_col >= min_col):
            ws.merged_cells.ranges.discard(mcr)
    ws.merge_cells(start_row=min_row, end_row=max_row,
                   start_column=min_col, end_column=max_col)


# ── Step 1: 対応内容シート ───────────────────────────────────────────────────

def patch_content_sheet(ws):
    """■影響確認チェックリストを完全削除・言語記述エリアを A3:D5 一括マージ。"""

    # 冪等チェック: チェックリスト見出しが既に存在しないなら skip
    check_header = _find_header_row(ws, "■ 影響確認チェックリスト")
    if check_header is None:
        print("[SKIP] 対応内容: ■影響確認チェックリストが既に存在しない → スキップ")
    else:
        merges_snap = _save_merges(ws)
        _clear_merges(ws)
        delete_from = check_header
        delete_count = ws.max_row - check_header + 1
        ws.delete_rows(delete_from, delete_count)
        # マージ再構築（削除範囲外のみ）
        for (min_r, max_r, min_c, max_c) in merges_snap:
            if max_r < delete_from:
                try:
                    ws.merge_cells(start_row=min_r, end_row=max_r,
                                   start_column=min_c, end_column=max_c)
                except Exception:
                    pass
        print(f"[OK  ] 対応内容: ■影響確認チェックリスト削除 (r{delete_from}〜末尾 {delete_count}行)")

    # 言語記述エリア: A3:D5 を一括マージに統合
    # 現状は A3:D3 / A4:D4 / A5:D5 の 3 個別マージ → 1 つにまとめる
    lang_header = _find_header_row(ws, "■ 対応内容（言語記述）")
    if lang_header:
        lang_data_start = lang_header + 1
        lang_data_end   = lang_header + 3  # 3 行分

        # 個別マージを解除して一括マージ
        for mcr in list(ws.merged_cells.ranges):
            if (mcr.min_row >= lang_data_start and mcr.max_row <= lang_data_end
                    and mcr.min_col == 1):
                ws.merged_cells.ranges.discard(mcr)

        already_unified = False
        for mcr in ws.merged_cells.ranges:
            if (mcr.min_row == lang_data_start and mcr.max_row == lang_data_end
                    and mcr.min_col == 1):
                already_unified = True
                break

        if not already_unified:
            ws.merge_cells(start_row=lang_data_start, end_row=lang_data_end,
                           start_column=1, end_column=4)
            # 統合セルのスタイル設定
            c = ws.cell(lang_data_start, 1)
            c.alignment = _WRAP
            c.font = _BODY_FONT
            c.border = _THIN_BORDER
            ws.row_dimensions[lang_data_start].height = 80
            print(f"[OK  ] 対応内容: 言語記述エリアを A{lang_data_start}:D{lang_data_end} 一括マージ")
        else:
            print("[SKIP] 対応内容: 言語記述エリアが既に一括マージ済み → スキップ")
    else:
        print("[WARN] 対応内容: ■対応内容（言語記述）が見つかりません")


# ── Step 2: 調査・影響範囲シート ──────────────────────────────────────────────

def patch_investigation_sheet(ws):
    """種別列（B列）を削除して3列化・A2:C2 マージ・列幅調整。"""

    # 冪等チェック: ヘッダ行が既に3列（No/対象/問題ない根拠）なら skip
    impact_header = _find_header_row(ws, "■ 影響範囲テーブル") or _find_header_row(ws, "■ 影響範囲")
    if impact_header:
        col_hdr_row = impact_header + 1
        b_val = str(ws.cell(col_hdr_row, 2).value or "").strip()
        # 新形式は B3="対象"。旧形式は B3="種別"
        if b_val == "対象":
            print("[SKIP] 調査・影響範囲: 既に3列構成（B='対象'）→ スキップ")
            # A2:C2 マージだけ追加確認
            a2_merge_exists = any(
                m.min_row == impact_header and m.max_row == impact_header
                and m.min_col == 1 and m.max_col >= 3
                for m in ws.merged_cells.ranges
            )
            if not a2_merge_exists:
                _safe_merge(ws, impact_header, impact_header, 1, 3)
                print(f"[OK  ] 調査・影響範囲: A{impact_header}:C{impact_header} マージ追加")
            return

    # B列（種別）を delete_cols で削除
    merges_snap = _save_merges(ws)
    _clear_merges(ws)
    ws.delete_cols(2, 1)  # B列を1列削除
    print("[OK  ] 調査・影響範囲: 種別列（旧B列）を削除")

    # マージ再構築（列番号を調整）
    for (min_r, max_r, min_c, max_c) in merges_snap:
        # 削除した列（2）の右側は1つずらす
        adj_min_c = min_c if min_c < 2 else max(1, min_c - 1)
        adj_max_c = max_c if max_c < 2 else max(1, max_c - 1)
        if adj_min_c <= adj_max_c:
            try:
                ws.merge_cells(start_row=min_r, end_row=max_r,
                               start_column=adj_min_c, end_column=adj_max_c)
            except Exception:
                pass

    # 影響範囲テーブルのヘッダを3列に書き直し
    impact_header = _find_header_row(ws, "■ 影響範囲テーブル") or _find_header_row(ws, "■ 影響範囲")
    if not impact_header:
        print("[WARN] 調査・影響範囲: ■影響範囲テーブルが見つかりません")
        return

    col_hdr_row = impact_header + 1
    # 既存マージを解除して3列ヘッダを書く
    for mcr in list(ws.merged_cells.ranges):
        if mcr.min_row == col_hdr_row and mcr.max_row == col_hdr_row:
            ws.merged_cells.ranges.discard(mcr)

    _write_header_cell(ws, col_hdr_row, 1, "No")
    _write_header_cell(ws, col_hdr_row, 2, "対象")
    _write_header_cell(ws, col_hdr_row, 3, "問題ない根拠・対応内容")
    print(f"[OK  ] 調査・影響範囲: 3列ヘッダ（No/対象/問題ない根拠・対応内容）書き直し")

    # A2（■ 影響範囲テーブル）を A2:C2 にマージ
    title_row = impact_header
    _safe_merge(ws, title_row, title_row, 1, 3)
    # 既存スタイルを適用
    t = ws.cell(title_row, 1)
    t.font = Font(name="游ゴシック", size=10, bold=True)
    t.fill = _SECTION_FILL
    t.alignment = _WRAP
    t.border = _THIN_BORDER
    print(f"[OK  ] 調査・影響範囲: A{title_row}:C{title_row} マージ＋スタイル適用")

    # 列幅調整
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 70
    print("[OK  ] 調査・影響範囲: 列幅調整 A=8 / B=35 / C=70")


# ── Step 3: 残対応・懸念・保留シート ──────────────────────────────────────────

def patch_pending_sheet(ws):
    """関連列（D列）を削除して5列化・A1 タイトル青背景・A2:E2 マージ・列幅調整。"""

    # 冪等チェック: ヘッダが既に5列（D列=ステータス）なら skip
    header_row = _find_header_row(ws, "■ 残対応・懸念事項一覧")
    if header_row:
        col_hdr_row = header_row + 1
        d_val = str(ws.cell(col_hdr_row, 4).value or "").strip()
        if d_val in ("ステータス", "状態"):
            print("[SKIP] 残対応・懸念・保留: 既に5列構成（D='ステータス'）→ スキップ")
            # A1 スタイルだけ確認
            _fix_pending_title(ws)
            return

    # D列（関連）を delete_cols で削除
    merges_snap = _save_merges(ws)
    _clear_merges(ws)
    ws.delete_cols(4, 1)  # D列を1列削除
    print("[OK  ] 残対応・懸念・保留: 関連列（旧D列）を削除")

    # マージ再構築（列番号を調整）
    for (min_r, max_r, min_c, max_c) in merges_snap:
        adj_min_c = min_c if min_c < 4 else max(1, min_c - 1)
        adj_max_c = max_c if max_c < 4 else max(1, max_c - 1)
        if adj_min_c <= adj_max_c:
            try:
                ws.merge_cells(start_row=min_r, end_row=max_r,
                               start_column=adj_min_c, end_column=adj_max_c)
            except Exception:
                pass

    # ヘッダ行を5列に書き直し
    header_row = _find_header_row(ws, "■ 残対応・懸念事項一覧")
    if header_row:
        col_hdr_row = header_row + 1
        for mcr in list(ws.merged_cells.ranges):
            if mcr.min_row == col_hdr_row and mcr.max_row == col_hdr_row:
                ws.merged_cells.ranges.discard(mcr)
        headers = ["No", "種別", "内容", "ステータス", "次アクション"]
        for i, h in enumerate(headers, start=1):
            _write_header_cell(ws, col_hdr_row, i, h)
        print("[OK  ] 残対応・懸念・保留: 5列ヘッダ書き直し")

        # A2（■ 残対応・懸念事項一覧）を A2:E2 にマージ
        _safe_merge(ws, header_row, header_row, 1, 5)
        s = ws.cell(header_row, 1)
        s.font = Font(name="游ゴシック", size=10, bold=True)
        s.fill = _SECTION_FILL
        s.alignment = _WRAP
        s.border = _THIN_BORDER
        print(f"[OK  ] 残対応・懸念・保留: A{header_row}:E{header_row} マージ＋スタイル適用")

        # データ行のサンプル値を削除（旧6列目が残っていると混乱するため）
        data_start = col_hdr_row + 1
        for r in range(data_start, data_start + 5):
            cell_f = ws.cell(r, 6)
            if cell_f.value:
                cell_f.value = None

    # A1 タイトルスタイルを他シートと統一
    _fix_pending_title(ws)

    # 列幅調整
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 30
    print("[OK  ] 残対応・懸念・保留: 列幅調整 A=5 / B=16 / C=50 / D=14 / E=30")


def _fix_pending_title(ws):
    """A1 の背景色・フォントを他シート（001F4E79 濃紺）と統一する。"""
    c = ws.cell(1, 1)
    current_fill = c.fill.fgColor.rgb if c.fill and c.fill.fgColor else ""
    if current_fill == "001F4E79":
        return  # 既に正しい
    c.fill = _TITLE_FILL
    c.font = Font(name="游ゴシック", size=14, bold=True, color="FFFFFFFF")
    c.alignment = _WRAP
    _safe_merge(ws, 1, 1, 1, 5)
    ws.row_dimensions[1].height = 28
    print("[OK  ] 残対応・懸念・保留: A1 タイトル背景色を濃紺（001F4E79）に統一")


# ── Step 4: テスト・検証シート ────────────────────────────────────────────────

def patch_test_sheet(ws):
    """A1 セル値を修正・ヘッダ行直下に注記行を挿入。"""

    # A1 セル値修正
    a1_val = str(ws.cell(1, 1).value or "").strip()
    if "記録" in a1_val:
        ws.cell(1, 1).value = "テスト・検証"
        print("[OK  ] テスト・検証: A1 セル値「テスト・検証記録」→「テスト・検証」")
    else:
        print("[SKIP] テスト・検証: A1 セル値が既に修正済み → スキップ")

    # 注記行: A7 に既に注記があれば skip
    note_keyword = "※ 区分=実装前"
    test_table_header = _find_header_row(ws, "■ テストテーブル")
    if not test_table_header:
        print("[WARN] テスト・検証: ■テストテーブルが見つかりません → 注記行スキップ")
        return

    col_hdr_row = test_table_header + 1  # 列ヘッダ（No/区分/...）行
    note_row = col_hdr_row + 1           # 注記行の挿入位置（元のデータ先頭）

    # 注記が既にあるか確認
    existing = str(ws.cell(note_row, 1).value or "").strip()
    if existing.startswith("※"):
        print("[SKIP] テスト・検証: 注記行が既に存在 → スキップ")
        return

    # マージスナップショット
    merges_snap = _save_merges(ws)
    _clear_merges(ws)
    ws.insert_rows(note_row, 1)

    # マージ再構築
    for (min_r, max_r, min_c, max_c) in merges_snap:
        adj_min_r = min_r + 1 if min_r >= note_row else min_r
        adj_max_r = max_r + 1 if max_r >= note_row else max_r
        try:
            ws.merge_cells(start_row=adj_min_r, end_row=adj_max_r,
                           start_column=min_c, end_column=max_c)
        except Exception:
            pass

    # 注記行に内容を書き込む
    note_text = (
        "※ 区分=実装前 の F/G 列は Phase 3.5（validator）で事前記入済み。"
        "区分=実装後 の F/G 列（実際の結果・判定）のみ Phase 5（tester）がテスト実施後に記入する。"
    )
    c = ws.cell(note_row, 1, value=note_text)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(name="游ゴシック", size=9, italic=True, color="FF808080")
    ws.merge_cells(start_row=note_row, end_row=note_row, start_column=1, end_column=7)
    ws.row_dimensions[note_row].height = 30
    print(f"[OK  ] テスト・検証: 注記行を r{note_row} に挿入")


# ── main ────────────────────────────────────────────────────────────────────

def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)
    print("=== patch_template_v9: Group M 実害修正 ===")
    print(f"現在のシート: {wb.sheetnames}\n")

    # Step 1: 対応内容シート
    if "対応内容" in wb.sheetnames:
        patch_content_sheet(wb["対応内容"])
    else:
        print("[WARN] 対応内容シートが見つかりません")
    print()

    # Step 2: 調査・影響範囲シート
    if "調査・影響範囲" in wb.sheetnames:
        patch_investigation_sheet(wb["調査・影響範囲"])
    else:
        print("[WARN] 調査・影響範囲シートが見つかりません")
    print()

    # Step 3: 残対応・懸念・保留シート
    if "残対応・懸念・保留" in wb.sheetnames:
        patch_pending_sheet(wb["残対応・懸念・保留"])
    else:
        print("[WARN] 残対応・懸念・保留シートが見つかりません")
    print()

    # Step 4: テスト・検証シート
    if "テスト・検証" in wb.sheetnames:
        patch_test_sheet(wb["テスト・検証"])
    else:
        print("[WARN] テスト・検証シートが見つかりません")
    print()

    # 保存
    try:
        wb.save(TEMPLATE)
        print(f"[OK  ] テンプレート保存完了: {TEMPLATE}")
    except Exception as e:
        print(f"[ERROR] 保存失敗: {e}")
        sys.exit(1)

    # ── 自動検証 ────────────────────────────────────────────────────────────
    print("\n=== 自動検証 ===")
    wb2 = load_workbook(TEMPLATE)
    ok = True

    # 対応内容: ■影響確認チェックリストが存在しない
    ws_c = wb2["対応内容"]
    has_check = any(
        "影響確認チェックリスト" in str(ws_c.cell(r, 1).value or "")
        for r in range(1, ws_c.max_row + 1)
    )
    _v(not has_check, "対応内容: ■影響確認チェックリスト削除済み", ok)
    if has_check: ok = False

    # 対応内容: A2:D2 マージ
    a2d2 = any(m.min_row == 2 and m.max_row == 2 and m.min_col == 1 and m.max_col == 4
               for m in ws_c.merged_cells.ranges)
    _v(a2d2, "対応内容: A2:D2 マージ済み", ok)
    if not a2d2: ok = False

    # 調査・影響範囲: ヘッダが3列
    ws_i = wb2["調査・影響範囲"]
    inv_hdr = next((r for r in range(1, ws_i.max_row + 1)
                    if "■ 影響範囲テーブル" in str(ws_i.cell(r, 1).value or "") or
                       "■ 影響範囲" in str(ws_i.cell(r, 1).value or "")), None)
    if inv_hdr:
        b_val = str(ws_i.cell(inv_hdr + 1, 2).value or "").strip()
        _v(b_val == "対象", f"調査・影響範囲: ヘッダB列='対象' (got '{b_val}')", ok)
        if b_val != "対象": ok = False
        c_val = str(ws_i.cell(inv_hdr + 1, 3).value or "").strip()
        _v("問題ない根拠" in c_val, f"調査・影響範囲: ヘッダC列に「問題ない根拠」(got '{c_val}')", ok)
        if "問題ない根拠" not in c_val: ok = False

    # 残対応: D列=ステータス
    ws_p = wb2["残対応・懸念・保留"]
    pend_hdr = next((r for r in range(1, ws_p.max_row + 1)
                     if "■ 残対応" in str(ws_p.cell(r, 1).value or "")), None)
    if pend_hdr:
        d_val = str(ws_p.cell(pend_hdr + 1, 4).value or "").strip()
        _v(d_val == "ステータス", f"残対応: D列='ステータス' (got '{d_val}')", ok)
        if d_val != "ステータス": ok = False
    # A1 fill
    a1_fill = ws_p.cell(1, 1).fill.fgColor.rgb if ws_p.cell(1, 1).fill else ""
    _v(a1_fill == "001F4E79", f"残対応: A1 fill=001F4E79 (got '{a1_fill}')", ok)
    if a1_fill != "001F4E79": ok = False

    # テスト・検証: A1 値
    ws_t = wb2["テスト・検証"]
    a1_test = str(ws_t.cell(1, 1).value or "").strip()
    _v(a1_test == "テスト・検証", f"テスト・検証: A1='{a1_test}'", ok)
    if a1_test != "テスト・検証": ok = False

    # テスト・検証: 注記行存在
    has_note = any("※ 区分=実装前" in str(ws_t.cell(r, 1).value or "")
                   for r in range(1, ws_t.max_row + 1))
    _v(has_note, "テスト・検証: 注記行存在", ok)
    if not has_note: ok = False

    print("\n" + ("✅ 全検証 PASS" if ok else "❌ 一部 FAIL — 上記ログを確認してください"))
    sys.exit(0 if ok else 1)


def _v(cond, msg, _ok):
    status = "✅" if cond else "❌"
    print(f"  {status} {msg}")


if __name__ == "__main__":
    main()
