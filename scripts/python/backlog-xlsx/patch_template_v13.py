# -*- coding: utf-8 -*-
"""patch_template_v13.py
対応記録テンプレート.xlsx を Group Q 仕様に移行する。

変更内容:
  1. テスト・検証シート: H 列「判定」を廃止し「実際の結果」に統合
     - F(期待結果)→G、G(実際の結果)→H へ値をシフト
     - E-F をセル結合して「確認方法」を 2 列幅に拡張
     - H 列ヘッダを「実際の結果」に変更
     - 最終列構成: A=No|B=区分|C=実行種別|D=テスト項目|E-F=確認方法|G=期待結果|H=実際の結果
  2. テスト・検証シート: C 列「実行種別」データ行フォントを游ゴシックに統一
     (v11 でデータ行の設定が漏れ、v12 の冪等チェックがヘッダのみ確認のため修正)
  3. 注記行のテキストを新列構成に合わせて更新

冪等性:
  - H 列シフトは「H 列ヘッダが実際の結果」なら skip
  - E-F 結合・C 列フォント・列幅は行単位/セル単位で冪等

Usage:
    python patch_template_v13.py               # テンプレートに適用
    python patch_template_v13.py --target <path>  # 既存ファイルに適用
"""

import sys
import copy
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.cell.cell import MergedCell as _MergedCell
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

_DEFAULT_TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"
_YAGOTHIC = "游ゴシック"
_NOTE_ROW_TEXT = (
    "※ 区分=実装前 の H 列（実際の結果）は Phase 3.5（validator）が記入。"
    "実装後行は Phase 5（tester）が記入。値は「OK」または「NG: <理由>」で始めること（空欄禁止）。"
)


def _find_col_hdr_row(ws):
    """■ テストテーブル 行の直下を列ヘッダ行として返す。見つからなければ 6。"""
    for r in range(1, min(20, ws.max_row + 1)):
        v = ws.cell(r, 1).value
        if v and "テストテーブル" in str(v):
            return r + 1
    return 6


def _is_merged(ws, row, col1, col2):
    """row 行の col1〜col2 が既にマージ範囲に含まれているか確認。"""
    for rng in ws.merged_cells.ranges:
        if (rng.min_row <= row <= rng.max_row
                and rng.min_col <= col1 and col2 <= rng.max_col):
            return True
    return False


def patch_test_sheet(ws):
    """テスト・検証シートをGroup Q仕様に更新する。"""
    col_hdr_row = _find_col_hdr_row(ws)
    data_start = col_hdr_row + 1
    data_end = ws.max_row

    # ─ Step 1: H 列シフト（冪等チェック付き）────────────────────────────────
    h_val = str(ws.cell(col_hdr_row, 8).value or "").strip()
    if h_val == "実際の結果":
        print("[SKIP] テスト・検証: H 列ヘッダが既に「実際の結果」 → シフト処理をスキップ")
    else:
        h_hdr = str(ws.cell(col_hdr_row, 8).value or "").strip()
        g_hdr = str(ws.cell(col_hdr_row, 7).value or "").strip()
        if h_hdr != "判定":
            print(f"[WARN] テスト・検証: H 列ヘッダが「判定」でない ({h_hdr!r}) → patch_v11 未適用の可能性")
            return

        # 列ヘッダ行: G ← F (期待結果), H ← G (実際の結果)
        f6 = ws.cell(col_hdr_row, 6)
        g6 = ws.cell(col_hdr_row, 7)
        h6 = ws.cell(col_hdr_row, 8)
        g6.value = f6.value
        for attr in ("font", "fill", "alignment", "border"):
            src = getattr(f6, attr)
            if src: setattr(g6, attr, copy.copy(src))
        h6.value = g_hdr  # "実際の結果"
        for attr in ("font", "fill", "alignment", "border"):
            src = getattr(g6, attr)
            if src: setattr(h6, attr, copy.copy(src))
        f6.value = None
        print(f"[OK  ] テスト・検証: 列ヘッダ行 r{col_hdr_row} の F→G→H シフト完了")

        # データ行: 各行 G ← F, H ← G
        shifted = 0
        for r in range(data_start, data_end + 1):
            if ws.cell(r, 1).value is None and ws.cell(r, 2).value is None:
                continue
            if isinstance(ws.cell(r, 6), _MergedCell):
                continue  # 全幅マージ行 (注記行等) はスキップ
            f_val = ws.cell(r, 6).value
            g_val = ws.cell(r, 7).value
            ws.cell(r, 7).value = f_val
            ws.cell(r, 8).value = g_val
            ws.cell(r, 6).value = None
            shifted += 1
        if shifted:
            print(f"[OK  ] テスト・検証: データ行 {shifted} 行の F→G→H 値シフト完了")

    # ─ Step 2: E-F セル結合（全データ行枠・行単位冪等）─────────────────────
    # テンプレート空白行を含む全行を対象にする（create_records のコピー元になるため）
    all_rows = [col_hdr_row] + list(range(data_start, data_end + 1))
    merged_count = 0
    for r in all_rows:
        if _is_merged(ws, r, 5, 6):
            continue  # 全幅マージ行またはE-F結合済み → スキップ
        ws.merge_cells(start_row=r, end_row=r, start_column=5, end_column=6)
        merged_count += 1
    if merged_count:
        print(f"[OK  ] テスト・検証: E-F セル結合 {merged_count} 行に適用")
    else:
        print("[SKIP] テスト・検証: E-F セル結合は全行適用済み")

    # ─ Step 3: C 列データ行フォントを游ゴシックに統一───────────────────────
    data_font = Font(name=_YAGOTHIC, size=10)
    hdr_font  = Font(name=_YAGOTHIC, bold=True, color="FFFFFFFF", size=10)
    ws.cell(col_hdr_row, 3).font = hdr_font
    c_fixed = 0
    for r in range(data_start, data_end + 1):
        cell = ws.cell(r, 3)
        if isinstance(cell, _MergedCell):
            continue
        fname = cell.font.name if cell.font else ""
        fsize = cell.font.size if cell.font else None
        if fname != _YAGOTHIC or fsize != 10:
            cell.font = data_font
            c_fixed += 1
    if c_fixed:
        print(f"[OK  ] テスト・検証: C 列データ行フォントを游ゴシックに修正 ({c_fixed} セル)")
    else:
        print("[SKIP] テスト・検証: C 列データ行フォントは既に全て游ゴシック")

    # ─ Step 4: 列幅調整────────────────────────────────────────────────────
    ws.column_dimensions["E"].width = 60   # 旧E(30)+旧F(30) 相当
    ws.column_dimensions["G"].width = 30   # 期待結果 (旧F幅=30)
    ws.column_dimensions["H"].width = 30   # 実際の結果 (旧G=10 → 30)
    print("[OK  ] テスト・検証: 列幅 E=60, G=30, H=30 に設定")

    # ─ Step 5: 注記行テキスト更新────────────────────────────────────────────
    note_row = col_hdr_row + 1  # 列ヘッダ直下の注記行
    note_cell = ws.cell(note_row, 1)
    if note_cell.value and "Phase 3" in str(note_cell.value):
        note_cell.value = _NOTE_ROW_TEXT
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        print(f"[OK  ] テスト・検証: 注記行 r{note_row} を更新")


def _v(ok, msg):
    prefix = "[OK  ]" if ok else "[FAIL]"
    print(f"  {prefix} {msg}")


def main():
    parser = argparse.ArgumentParser(
        description="対応記録テンプレートをGroup Q仕様に移行する"
    )
    parser.add_argument(
        "--target", default=None,
        help="テンプレート以外のファイルパス（省略時は対応記録テンプレート.xlsx）"
    )
    args = parser.parse_args()

    target = Path(args.target) if args.target else _DEFAULT_TEMPLATE
    if not target.exists():
        print(f"[ERROR] ファイルが見つかりません: {target}")
        sys.exit(1)

    wb = load_workbook(str(target))
    print(f"=== patch_template_v13: Group Q 判定列廃止・確認方法2列化 ===")
    print(f"対象ファイル: {target}")
    print(f"現在のシート: {wb.sheetnames}\n")

    ws_name = "テスト・検証"
    if ws_name in wb.sheetnames:
        patch_test_sheet(wb[ws_name])
    else:
        print(f"[WARN] {ws_name} シートが見つかりません")
    print()

    try:
        wb.save(str(target))
        print(f"[OK  ] 保存完了: {target}")
    except Exception as e:
        print(f"[ERROR] 保存失敗: {e}")
        sys.exit(1)

    # ── 自動検証 ──────────────────────────────────────────────────────────────
    print("\n=== 自動検証 ===")
    wb2 = load_workbook(str(target))
    ok = True

    if ws_name in wb2.sheetnames:
        ws2 = wb2[ws_name]
        col_hdr_row2 = _find_col_hdr_row(ws2)
        data_start2 = col_hdr_row2 + 1

        h_val = str(ws2.cell(col_hdr_row2, 8).value or "").strip()
        ok1 = h_val == "実際の結果"
        _v(ok1, f"H{col_hdr_row2} ヘッダ = 「{h_val}」 (期待: 実際の結果)")
        if not ok1: ok = False

        g_val = str(ws2.cell(col_hdr_row2, 7).value or "").strip()
        ok2 = g_val == "期待結果"
        _v(ok2, f"G{col_hdr_row2} ヘッダ = 「{g_val}」 (期待: 期待結果)")
        if not ok2: ok = False

        ok3 = _is_merged(ws2, col_hdr_row2, 5, 6)
        _v(ok3, f"E{col_hdr_row2}:F{col_hdr_row2} がセル結合されている")
        if not ok3: ok = False

        # データ行枠の E-F 結合確認 (全幅マージ行以外)
        ef_ok_rows = []
        ef_ng_rows = []
        for r in range(data_start2, ws2.max_row + 1):
            if _is_merged(ws2, r, 1, 8):
                continue  # 全幅マージ行（注記行等）はスキップ
            if _is_merged(ws2, r, 5, 6):
                ef_ok_rows.append(r)
            else:
                ef_ng_rows.append(r)
        ok4 = len(ef_ng_rows) == 0
        msg4 = f"データ行枠の E-F 結合 OK ({len(ef_ok_rows)} 行)" if ok4 else f"未結合データ行: {ef_ng_rows[:5]}"
        _v(ok4, msg4)
        if not ok4: ok = False

        bad_fonts = []
        for r in range(data_start2, ws2.max_row + 1):
            c = ws2.cell(r, 3)
            if not isinstance(c, _MergedCell) and c.value is not None:
                fname = c.font.name if c.font else ""
                fsize = c.font.size if c.font else None
                if fname != _YAGOTHIC or fsize != 10:
                    bad_fonts.append(f"C{r}={fname!r}(size={fsize})")
        ok5 = len(bad_fonts) == 0
        msg5 = "C 列データ行フォントが全て游ゴシック(10pt)" if ok5 else f"NG セル: {bad_fonts[:5]}"
        _v(ok5, msg5)
        if not ok5: ok = False

    print()
    print("[OK  ] 全検証 PASS" if ok else "[FAIL] 検証に失敗した項目があります")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
