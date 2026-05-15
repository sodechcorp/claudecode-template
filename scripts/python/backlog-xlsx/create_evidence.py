# -*- coding: utf-8 -*-
"""backlog-xlsx / create_evidence.py
エビデンス.xlsx を生成する（implementation-plan.md のテスト仕様からテストケース埋め）

Usage:
    python create_evidence.py \\
      --folder FOLDER --issue-id ID \\
      --implementation-plan PATH
"""

import argparse
import os
import re
import sys
from pathlib import Path

from _common import validate_folder

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl がインストールされていません。`pip install openpyxl` を実行してください。")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "エビデンステンプレート.xlsx"
WRAP = Alignment(wrap_text=True, vertical="top")
STRIPE_A = PatternFill("solid", fgColor="FFFFFF")
STRIPE_B = PatternFill("solid", fgColor="F2F7FB")
BLUE_HDR  = PatternFill("solid", fgColor="1F3864")   # ヘッダー行の濃紺（テンプレ準拠）
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F7")  # サブヘッダー（テンプレ準拠）
CHECKLIST_BG = PatternFill("solid", fgColor="E8F0FE")  # チェックリスト行
EVIDENCE_BG  = PatternFill("solid", fgColor="FFF3CD")  # エビデンス貼付枠
WHITE = PatternFill("solid", fgColor="FFFFFF")


# ── パースユーティリティ（create_records.py と同じ） ────────────────────────

def read_md(path):
    if path and Path(path).exists():
        try:
            return Path(path).read_text(encoding="utf-8")
        except UnicodeDecodeError as e:
            print(f"[ERROR] ファイルのエンコーディングが UTF-8 ではありません: {path}\n{e}")
            sys.exit(1)
    return ""


def extract_section(md, *headings):
    for h in headings:
        pat = r"^#{1,3}\s+" + re.escape(h) + r"\s*$"
        m = re.search(pat, md, re.MULTILINE)
        if m:
            start = m.end()
            rest = md[start:]
            end_m = re.search(r"^#{1,3}\s", rest, re.MULTILINE)
            body = rest[: end_m.start()] if end_m else rest
            return body.strip()
    return ""


def parse_md_table(section_text):
    rows = []
    headers = []
    for line in section_text.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if all(re.match(r"^[-: ]+$", c) for c in cells):
            continue
        if not headers:
            headers = cells
        else:
            rows.append(dict(zip(headers, cells)))
    return rows


# ── テスト仕様シート ────────────────────────────────────────────────────────

def fill_test_spec(ws, test_cases):
    """テスト仕様シートにテストケースを書き込む（r3 以降）。UI手動テスト（対象）のみ。"""
    for i, tc in enumerate(test_cases):
        r = 3 + i
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        vals = [
            tc.get("No", str(i + 1)),
            tc.get("確認観点", ""),
            tc.get("タイミング", ""),
            tc.get("実行種別", "UI手動"),
            tc.get("確認手順", ""),
            tc.get("期待結果", ""),
            tc.get("貼付先シート", "実装前エビデンス" if tc.get("タイミング") == "実装前" else "実装後エビデンス"),
        ]
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.alignment = WRAP
            cell.fill = fill


# ── エビデンスシート（実装前 / 実装後 共通ロジック） ────────────────────────

def fill_evidence_sheet(ws, cases, sheet_label):
    """
    cases: タイミングでフィルタ済みのテストケースリスト
    sheet_label: 「実装前」or「実装後」
    """
    # r2 説明文はテンプレ時点で設定済み（「両方」削除済み）なのでそのまま

    # テンプレートは3件固定想定のマージ構造を持つため、データエリア (r5以降) のマージを
    # 事前に全解除してから動的書き込みする（r1-r4のタイトル/ヘッダー行マージは保持）
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 5:
            ws.unmerge_cells(str(mr))

    # チェックリスト（r6 以降）
    checklist_start = 6
    for i, tc in enumerate(cases):
        r = checklist_start + i
        cell_cb = ws.cell(row=r, column=1, value="□")
        cell_cb.alignment = WRAP
        cell_cb.fill = CHECKLIST_BG

        label = f"No.{tc.get('No', i + 1)}: {tc.get('確認観点', '')}"
        # 確認手順が複数行あれば短縮して括弧で補足
        steps = tc.get("確認手順", "")
        if steps and "\n" in steps:
            first_step = steps.split("\n")[0].strip().lstrip("0123456789. ")
            label += f"（{first_step[:30]}）"
        cell_obs = ws.cell(row=r, column=2, value=label)
        cell_obs.alignment = WRAP
        cell_obs.fill = CHECKLIST_BG

        ws.cell(row=r, column=3).fill = WHITE  # 結果欄
        ws.cell(row=r, column=4).fill = WHITE  # メモ欄

    # エビデンス貼付欄（チェックリスト末尾の次の行から）
    paste_start = checklist_start + len(cases) + 1

    # セクション見出し行: A:D 横幅マージ
    ws.merge_cells(start_row=paste_start, start_column=1, end_row=paste_start, end_column=4)
    hdr_cell = ws.cell(row=paste_start, column=1, value="■ エビデンス貼付欄")
    hdr_cell.fill = LIGHT_BLUE
    hdr_cell.alignment = WRAP

    # 各テストケースにエビデンス枠を作る
    row_ptr = paste_start + 1
    for i, tc in enumerate(cases):
        # ラベル行: A:D 横幅マージ
        label = f"エビデンス{chr(0x2460 + i)}: No.{tc.get('No', i + 1)} {tc.get('確認観点', '')}"
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=4)
        cell = ws.cell(row=row_ptr, column=1, value=label)
        cell.fill = EVIDENCE_BG
        cell.alignment = WRAP
        row_ptr += 1

        # 貼付エリア: A:D × 10行をまとめてマージ。先頭行に説明文を書く
        paste_top = row_ptr
        paste_bottom = paste_top + 9
        # マージ前に範囲内の全セルへ白塗りを設定（マージ後は非アンカーへの style 反映が消えるため）
        for r in range(paste_top, paste_bottom + 1):
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = WHITE
        ws.cell(row=paste_top, column=1, value="ここにスクリーンショットを貼り付けてください")
        ws.cell(row=paste_top, column=1).alignment = WRAP
        ws.merge_cells(start_row=paste_top, start_column=1, end_row=paste_bottom, end_column=4)

        row_ptr = paste_bottom + 2  # 1行空けて次のラベルへ


# ── main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="エビデンス.xlsx を生成する（テスト仕様埋め込み版）")
    parser.add_argument("--folder",              required=True)
    parser.add_argument("--issue-id",            required=True, dest="issue_id")
    parser.add_argument("--implementation-plan", required=True, dest="implementation_plan",
                        help="docs/logs/{issueID}/implementation-plan.md のパス")
    parser.add_argument("--allow-blank",         action="store_true", dest="allow_blank",
                        help="タイミング列が空のテストケースを 'after' 扱いにする（従来挙動）。未指定時は警告で abort")
    args = parser.parse_args()
    args.folder = validate_folder(args.folder)

    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    impl_md = read_md(args.implementation_plan)
    if not impl_md:
        print(f"[ERROR] ファイルが見つかりません: {args.implementation_plan}")
        sys.exit(1)

    # テスト仕様テーブル抽出
    spec_text = extract_section(impl_md, "テスト仕様", "テストケース", "テスト仕様テーブル")
    test_cases = parse_md_table(spec_text)

    if not test_cases:
        print("[WARN] テスト仕様テーブルが見つかりませんでした。空のエビデンスファイルを生成します。")

    # タイミング空文字チェック
    blank_timing_cases = [tc for tc in test_cases if not tc.get("タイミング", "").strip()]
    if blank_timing_cases:
        ids = [tc.get("番号", "?") for tc in blank_timing_cases]
        if args.allow_blank:
            print(f"[WARN] タイミング未設定のテストケース: {ids} → 'after' 扱いで続行（--allow-blank 指定）")
            for tc in blank_timing_cases:
                tc["タイミング"] = "after"
        else:
            print(f"[ERROR] タイミング未設定のテストケースがあります: {ids}")
            print("        タイミング列（'実装前' / 'after' 等）を記入するか --allow-blank で続行してください。")
            sys.exit(1)

    # 実行種別=UI手動 の行のみエビデンスファイルに記録（ClaudeCode 自動行は対応記録xlsx側）
    # 実行種別列がない場合（旧フォーマット）は全行を対象にする（後方互換）
    has_shubetsu = any(tc.get("実行種別", "").strip() for tc in test_cases)
    if has_shubetsu:
        test_cases = [tc for tc in test_cases if tc.get("実行種別", "").strip() == "UI手動"]
    before_cases = [tc for tc in test_cases if tc.get("タイミング", "").strip() == "実装前"]
    after_cases  = [tc for tc in test_cases if tc.get("タイミング", "").strip() != "実装前"]

    os.makedirs(args.folder, exist_ok=True)
    try:
        wb = load_workbook(TEMPLATE)
    except Exception as e:
        print(f"[ERROR] テンプレートファイルの読み込みに失敗しました: {TEMPLATE}\n{e}")
        sys.exit(1)

    fill_test_spec(wb["テスト仕様"], test_cases)
    fill_evidence_sheet(wb["実装前エビデンス"], before_cases, "実装前")
    fill_evidence_sheet(wb["実装後エビデンス"], after_cases, "実装後")

    path = os.path.join(args.folder, f"{args.issue_id}_エビデンス.xlsx")
    try:
        wb.save(path)
    except PermissionError as e:
        print(f"[ERROR] xlsx の保存に失敗しました（ファイルが開かれている可能性があります）: {path}\n{e}")
        sys.exit(1)
    print(f"生成完了: {path}")
    print(f"  テスト仕様: {len(test_cases)} 件（実装前: {len(before_cases)} 件 / 実装後: {len(after_cases)} 件）")


if __name__ == "__main__":
    main()
