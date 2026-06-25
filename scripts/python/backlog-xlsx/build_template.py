# -*- coding: utf-8 -*-
"""backlog-xlsx / build_template.py
対応記録テンプレート.xlsx を2シート構成で再生成する。

Usage:
    python build_template.py
    # スクリプトと同じフォルダに「対応記録テンプレート.xlsx」を生成する。

シート構成（2シート）:
  ① 課題と対応方針  — 課題整理 / 経緯・対応方針 / タイムライン
  ② 対応内容        — 実施した対応 / 変更を加えた資材一覧 / Before/After（任意）
                      / NG対応履歴（/test NG 時の回次別修正記録）
"""

import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl がインストールされていません。`pip install openpyxl` を実行してください。")
    sys.exit(1)

OUTPUT = Path(__file__).parent / "対応記録テンプレート.xlsx"

# ── カラーパレット ────────────────────────────────────────────────────────────
_SECTION_BG  = "D6E4F0"   # セクションヘッダー（薄い鋼青）
_COLHDR_BG   = "EBF3FB"   # 列ヘッダー行（さらに薄い青）
_STRIPE_A    = "FFFFFF"   # 偶数データ行
_STRIPE_B    = "F2F7FB"   # 奇数データ行（薄青）
_LABEL_BG    = "FAFAFA"   # ラベルセル

# ── フォント / アライメント ─────────────────────────────────────────────────
_FONT_BASE   = Font(name="游ゴシック", size=10, color="1A1A1A")
_FONT_BOLD   = Font(name="游ゴシック", size=10, color="1A1A1A", bold=True)
_FONT_HEADER = Font(name="游ゴシック", size=10, color="1A1A1A", bold=True)
_WRAP        = Alignment(horizontal="left", vertical="top",  wrap_text=True)
_WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── 枠線 ──────────────────────────────────────────────────────────────────────
def _thin(color="AAAAAA"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _medium(color="888888"):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

_THIN_BORDER   = _thin()
_MEDIUM_BORDER = _medium()


# ── セル書き込みヘルパー ──────────────────────────────────────────────────────

def _fill(rgb):
    return PatternFill("solid", fgColor=rgb)

def _w(ws, row, col, value="", bold=False, bg=None, border=None,
       align=None, font=None):
    """セル書き込みユーティリティ"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font   = font or (_FONT_BOLD if bold else _FONT_BASE)
    cell.alignment = align or _WRAP
    if bg:
        cell.fill = _fill(bg)
    if border:
        cell.border = border
    return cell

def _section_header(ws, row, text, max_col=6):
    """セクションヘッダー行: A:max_col をマージしてタイトルを書く"""
    ws.merge_cells(start_row=row, end_row=row,
                   start_column=1, end_column=max_col)
    _w(ws, row, 1, text, bold=True, bg=_SECTION_BG, border=_MEDIUM_BORDER)
    ws.row_dimensions[row].height = 22

def _col_header_row(ws, row, headers, bg=_COLHDR_BG):
    """列ヘッダー行: 各列を太字で書く"""
    for col, h in enumerate(headers, start=1):
        _w(ws, row, col, h, bold=True, bg=bg, border=_THIN_BORDER)
    ws.row_dimensions[row].height = 20

def _data_row(ws, row, ncols, stripe_idx):
    """空データ行を縞模様で初期化する"""
    bg = _STRIPE_A if stripe_idx % 2 == 0 else _STRIPE_B
    for col in range(1, ncols + 1):
        _w(ws, row, col, "", bg=bg, border=_THIN_BORDER)
    ws.row_dimensions[row].height = 28

def _label_value(ws, row, label, max_col=6, val_merge_start=2):
    """ラベル（A列）+ 値セル（B:max_col マージ）を書く"""
    _w(ws, row, 1, label, bold=True, bg=_LABEL_BG, border=_THIN_BORDER)
    if val_merge_start < max_col:
        ws.merge_cells(start_row=row, end_row=row,
                       start_column=val_merge_start, end_column=max_col)
    _w(ws, row, val_merge_start, "", bg=_STRIPE_A, border=_THIN_BORDER)
    ws.row_dimensions[row].height = 28

def _label_value_large(ws, row, label, height, max_col=6, val_merge_start=2):
    """ラベル（A列）+ 大テキスト値セル（B:max_col マージ・tall）"""
    _w(ws, row, 1, label, bold=True, bg=_LABEL_BG, border=_THIN_BORDER)
    ws.merge_cells(start_row=row, end_row=row,
                   start_column=val_merge_start, end_column=max_col)
    _w(ws, row, val_merge_start, "", bg=_STRIPE_A, border=_THIN_BORDER)
    ws.row_dimensions[row].height = height


# ── Sheet ① 課題と対応方針 ──────────────────────────────────────────────────

def _build_sheet_plan(wb):
    ws = wb.create_sheet("課題と対応方針")

    # 列幅
    widths = {1: 22, 2: 18, 3: 12, 4: 14, 5: 32, 6: 20}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    MAX_COL = 6  # A〜F

    # ── ■ 課題の整理 ──────────────────────────────────────────────
    _section_header(ws, 1, "■ 課題の整理", max_col=MAX_COL)

    # メタ情報（シンプルラベル＋値）
    meta_rows = [
        (2, "課題ID"),
        (3, "件名"),
        (4, "優先度・期限"),
        (5, "種別"),
        (6, "ステータス"),
    ]
    for row, label in meta_rows:
        _label_value(ws, row, label, max_col=MAX_COL)

    # 大テキスト欄
    _label_value_large(ws, 7,  "課題の内容・詳細",  height=80, max_col=MAX_COL)
    _label_value_large(ws, 8,  "原因・現状",         height=80, max_col=MAX_COL)

    # ── ■ 経緯・対応方針 ──────────────────────────────────────────
    _section_header(ws, 9,  "■ 経緯・対応方針", max_col=MAX_COL)
    _label_value_large(ws, 10, "対応方針（結論）",         height=60, max_col=MAX_COL)
    _label_value_large(ws, 11, "方針決定の経緯・根拠",     height=80, max_col=MAX_COL)

    # ── ■ 対応経緯タイムライン ────────────────────────────────────
    _section_header(ws, 12, "■ 対応経緯タイムライン", max_col=MAX_COL)
    _col_header_row(ws, 13, ["No", "日時", "発生元", "フェーズ", "内容", "理由"])

    for i in range(8):
        _data_row(ws, 14 + i, MAX_COL, i)

    # タブ色
    ws.sheet_properties.tabColor = "4472C4"
    return ws


# ── Sheet ② 対応内容 ────────────────────────────────────────────────────────

def _build_sheet_content(wb):
    ws = wb.create_sheet("対応内容")

    # 列幅
    widths = {1: 6, 2: 32, 3: 14, 4: 45}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    MAX_COL = 4  # A〜D

    # ── ■ 実施した対応 ────────────────────────────────────────────
    _section_header(ws, 1, "■ 実施した対応", max_col=MAX_COL)

    # 大テキスト用マージセル（A:D）
    ws.merge_cells(start_row=2, end_row=5,
                   start_column=1, end_column=MAX_COL)
    cell = ws.cell(row=2, column=1, value="")
    cell.alignment = _WRAP
    cell.fill = _fill(_STRIPE_A)
    cell.border = _THIN_BORDER
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 30

    # ── ■ 変更を加えた資材一覧 ───────────────────────────────────
    _section_header(ws, 6, "■ 変更を加えた資材一覧", max_col=MAX_COL)
    _col_header_row(ws, 7, ["No", "資材名（表示名）", "変更種別", "変更内容"])

    for i in range(5):
        _data_row(ws, 8 + i, MAX_COL, i)

    # ── ■ Before / After（任意） ─────────────────────────────────
    _section_header(ws, 13, "■ Before / After（任意・コード変更がある場合のみ）", max_col=MAX_COL)

    # 3行分の空領域
    for r in range(14, 17):
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=MAX_COL)
        cell = ws.cell(row=r, column=1, value="")
        cell.alignment = _WRAP
        cell.fill = _fill(_STRIPE_A)
        cell.border = _THIN_BORDER
        ws.row_dimensions[r].height = 28

    # ── ■ NG対応履歴 ─────────────────────────────────────────────
    # /test NG 時の回次別修正記録（回次 / TC番号 / NG原因 / 修正内容）
    _section_header(ws, 18, "■ NG対応履歴（/test NG 修正ループ記録）", max_col=MAX_COL)
    _col_header_row(ws, 19, ["回次", "TC番号", "NG原因", "修正内容"])

    for i in range(4):
        _data_row(ws, 20 + i, MAX_COL, i)

    # タブ色
    ws.sheet_properties.tabColor = "70AD47"
    return ws


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    # デフォルトシートを削除
    wb.remove(wb.active)

    _build_sheet_plan(wb)
    _build_sheet_content(wb)

    try:
        wb.save(str(OUTPUT))
        print(f"[OK] テンプレート生成完了: {OUTPUT}")
    except PermissionError as e:
        print(f"[ERROR] ファイルが開かれている可能性があります: {OUTPUT}\n{e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
