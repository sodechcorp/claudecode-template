# -*- coding: utf-8 -*-
"""patch_template_v3.py
対応記録テンプレート.xlsx に以下の修正を一度だけ適用する:
  1. サマリー・経緯 r18-r21 F 列（関連ファイル）のフォント・塗り・配置を他列と統一
  2. サマリー・経緯 r18-r21 列構造変更:
     旧: A=No / B:C=内容 / D:E=影響範囲 / F=関連ファイル
     新: A=No / B:D=内容 / E=影響範囲 / F=関連ファイル

Usage:
    python patch_template_v3.py
"""

import sys
from copy import copy
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles.borders import Border, Side
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"
COLOR = "00B4C6E7"

_thin = Side(style="thin", color=COLOR)
_FULL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_TB   = Border(top=_thin, bottom=_thin)
_LTB  = Border(left=_thin, top=_thin, bottom=_thin)
_RTB  = Border(right=_thin, top=_thin, bottom=_thin)


def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)
    ws = wb["サマリー・経緯"]

    # ── 1. r18-r21 の既存マージを全解除（unmerge_cells で MergedCell 状態も解消）──
    to_unmerge = [
        str(mc) for mc in list(ws.merged_cells.ranges)
        if mc.min_row >= 18 and mc.max_row <= 21
    ]
    for rng in to_unmerge:
        ws.unmerge_cells(rng)

    # ── 2. r18 ヘッダー値を新構造に書き換え（マージ前に実施）─────────────────────
    ws.cell(18, 1).value = "No"
    ws.cell(18, 2).value = "内容"
    ws.cell(18, 3).value = None
    ws.cell(18, 4).value = None
    ws.cell(18, 5).value = "影響範囲"
    ws.cell(18, 6).value = "関連ファイル"
    print("[OK] r18 ヘッダー値を新構造に更新")

    # ── 3. F 列のスタイルを B 列からコピー（マージ前に実施）──────────────────────
    # copy.copy で独立化（StyleArray 共有参照バグ対策）
    for r in range(18, 22):
        src = ws.cell(r, 2)
        dst = ws.cell(r, 6)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)

    # ── 4. 新マージを設定: B:D=内容, E=単独 ───────────────────────────────────
    for r in range(18, 22):
        ws.merge_cells(start_row=r, end_row=r, start_column=2, end_column=4)

    # ── 5. r18-r21 border 再適用 ──────────────────────────────────────────────
    for r in range(18, 22):
        ws.cell(r, 1).border = _FULL   # A 単独
        ws.cell(r, 2).border = _LTB    # B (B:D マージ開始)
        ws.cell(r, 3).border = _TB     # C (B:D マージ中央)
        ws.cell(r, 4).border = _RTB    # D (B:D マージ終端)
        ws.cell(r, 5).border = _FULL   # E 単独
        ws.cell(r, 6).border = _FULL   # F 単独
    print("[OK] r18-r21 マージ構造変更（B:D=内容 / E=影響範囲 / F=関連ファイル）+ border 適用")

    # ── 保存 ──────────────────────────────────────────────────────────────────
    try:
        wb.save(TEMPLATE)
        print(f"\n[完了] テンプレート更新: {TEMPLATE}")
    except PermissionError as e:
        print(f"[ERROR] 保存失敗（Excel でファイルが開かれている可能性）: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
