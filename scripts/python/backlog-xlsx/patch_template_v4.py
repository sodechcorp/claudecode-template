# -*- coding: utf-8 -*-
"""patch_template_v4.py
対応記録テンプレート.xlsx に以下の修正を一度だけ適用する:
  1. サマリー・経緯 r18-r21 E 列（影響範囲）のフォント・塗り・配置を他列と統一
     （v3 で F 列のみ修正済み、E 列が未修正だったため同パターンで追加）

Usage:
    python patch_template_v4.py
"""

import sys
from copy import copy
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("[ERROR] openpyxl がインストールされていません: pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"


def main():
    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)
    ws = wb["サマリー・経緯"]

    # r18-r21 の E 列スタイルを B 列からコピー
    # （v3 の F 列処理と同パターン。border は v3 で _FULL 適用済みのため不要）
    for r in range(18, 22):
        src = ws.cell(r, 2)
        dst = ws.cell(r, 5)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)

    print("[OK] r18-r21 E 列フォント・塗り・配置を B 列から統一コピー")

    try:
        wb.save(TEMPLATE)
        print(f"\n[完了] テンプレート更新: {TEMPLATE}")
    except PermissionError as e:
        print(f"[ERROR] 保存失敗（Excel でファイルが開かれている可能性）: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
