"""
機能一覧.xlsx を生成する（テンプレート駆動・種別別シート）。

テンプレート: 機能一覧テンプレート.xlsx
  - 改版履歴 / サマリー / __SHEET_TEMPLATE__（種別別シートの雛形）

出力:
  - 改版履歴 / サマリー / Apex / Batch / Flow / 画面フロー / LWC / ...（入力JSONに含まれる種別のみ）

Usage:
  python generate_feature_list.py \
    --input features.json \
    --output-dir /path/to/output \
    --author 作成者名 \
    --project-name プロジェクト名 \
    [--system-name システム名] \
    [--template /path/to/機能一覧テンプレート.xlsx]

Input JSON:
[
  { "id":"F-001", "type":"Apex", "name":"機能名", "api_name":"ClassName",
    "overview":"処理概要" },
  ...
]
"""
import argparse
import json
import shutil
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from meta_store import read_meta, write_meta
from version_manager import increment_version
from text_cleaning import clean_tech_business

DEFAULT_TEMPLATE = Path(__file__).parent / "機能一覧テンプレート.xlsx"

C_ALT_ROW   = "F5F5F5"
C_FONT_D    = "000000"
C_FONT_LINK = "0563C1"
FONT_NAME   = "游ゴシック"
THIN        = Side(style="thin", color="8B9DC3")

TYPE_ORDER = ["Apex", "Batch", "Trigger",
              "Flow", "画面フロー", "LWC", "Aura",
              "Visualforce", "その他"]

# ── テンプレートと一致させる定数 ────────────────────────────────
# 改版履歴
REV_META_ROW       = 3
REV_META_PROJECT_V = (7, 18)
REV_META_DATE_V    = (23, 31)
REV_DATA_ROW_START = 6
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6, 11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}

# サマリー
SUM_META_ROW_1 = 3
SUM_META_ROW_2 = 4
# 値セル: (cs, ce)
SUM_PROJECT_V = (6, 18)
SUM_DATE_V    = (23, 31)
SUM_AUTHOR_V  = (6, 12)   # 作成者
SUM_VERSION_V = (17, 22)  # バージョン（追加）
SUM_TOTAL_V   = (26, 31)  # 合計件数
SUM_DATA_ROW_START = 8
SUM_COLS = {
    "No":         (2,  3),
    "種別":       (4,  11),
    "件数":       (12, 15),
    "対応シート": (16, 31),
}

# 種別別シート（メタ行なし・行番号を詰める）
ST_SEC_ROW        = 3
ST_HEAD_ROW       = 4
ST_DATA_ROW_START = 5
ST_COLS = {
    "ID":              (2,  4),
    "API名/ファイル名": (5,  11),
    "機能名":          (12, 18),
    "処理概要":        (19, 31),
}

C_FONT_R  = "C00000"  # 赤字（変更・追加行）
C_ADD_BG  = "FFF9C4"  # 薄黄（追加行背景）

ANNOT_COL_START = 33  # 注記列の開始列（col31 + 1gap）


def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)
def _fill(c): return PatternFill("solid", fgColor=c)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def B_all(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def MW(ws, row, cs, ce, value="", border=None, bg=None, bold=False,
       fg=C_FONT_D, h="left", v="center", wrap=True, size=10):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    cell = ws.cell(row=row, column=cs, value=value)
    cell.font = _fnt(bold=bold, color=fg, size=size)
    cell.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg: cell.fill = _fill(bg)
    if border: cell.border = border
    return cell

def set_h(ws, row, h):
    ws.row_dimensions[row].height = h


def clone_sheet(wb, source_name: str, new_name: str) -> Worksheet:
    src = wb[source_name]
    new = wb.copy_worksheet(src)
    new.title = new_name
    return new


# ── 差分計算 ──────────────────────────────────────────────────
def _feature_comparable(f: dict) -> tuple:
    """差分検出用に比較対象キーのみ抽出。
    overview / name はエージェントが生成するため非決定的。毎回微妙に変わりうるので除外する。
    type / api_name の構造的変化のみを差分として扱う。
    """
    return (
        f.get("type", ""),
        f.get("api_name", ""),
    )


def compare_features(old_features: list, new_features: list) -> dict:
    """前回と今回の機能一覧を比較し差分を返す。
    キーは id（feature_ids.yml で安定）を使う。
    """
    old_map = {f.get("id"): f for f in (old_features or []) if f.get("id")}
    new_map = {f.get("id"): f for f in new_features if f.get("id")}

    added    = [new_map[k] for k in new_map if k not in old_map]
    removed  = [old_map[k] for k in old_map if k not in new_map]
    modified = []
    for k in new_map:
        if k in old_map and _feature_comparable(old_map[k]) != _feature_comparable(new_map[k]):
            modified.append({"id": k, "old": old_map[k], "new": new_map[k]})

    return {"added": added, "removed": removed, "modified": modified}


def has_any_diff(diffs: dict) -> bool:
    return bool(diffs["added"] or diffs["removed"] or diffs["modified"])


def build_revision_entries(current_version: str, diffs: dict, author: str,
                           today: str, start_no: int, is_major: bool,
                           is_initial: bool) -> list[dict]:
    """改版履歴エントリを1行に集約して構築する。"""
    if is_initial:
        return [{
            "項番": start_no, "版数": current_version, "変更箇所": "全シート",
            "変更内容": "新規作成", "変更理由": "", "変更日": today,
            "変更者": author, "備考": "",
        }]
    if is_major:
        return [{
            "項番": start_no, "版数": current_version, "変更箇所": "全シート",
            "変更内容": "メジャーバージョンアップ", "変更理由": "",
            "変更日": today, "変更者": author, "備考": "",
        }]

    # 影響種別を集約
    types: set[str] = set()
    for f in diffs["added"] + diffs["removed"]:
        types.add(f.get("type", "その他"))
    for m in diffs["modified"]:
        types.add(m["new"].get("type", "その他"))

    parts = []
    if diffs["added"]:    parts.append(f"追加{len(diffs['added'])}件")
    if diffs["removed"]:  parts.append(f"削除{len(diffs['removed'])}件")
    if diffs["modified"]: parts.append(f"変更{len(diffs['modified'])}件")

    return [{
        "項番":     start_no,
        "版数":     current_version,
        "変更箇所": "、".join(sorted(types)) if types else "—",
        "変更内容": "・".join(parts) if parts else "変更なし",
        "変更理由": "",
        "変更日":   today,
        "変更者":   author,
        "備考":     "",
    }]


# ── 埋め込み ───────────────────────────────────────────────────
def fill_revision(ws, history: list, project_name: str, today: str):
    """history 配列を改版履歴テーブルに動的書き込みする（行数制限なし）。"""
    vs, ve = REV_META_PROJECT_V
    ws.cell(row=REV_META_ROW, column=vs, value=project_name)
    vs, ve = REV_META_DATE_V
    ws.cell(row=REV_META_ROW, column=vs, value=today)

    CENTER_LABELS = {"項番", "版数", "変更日", "変更者"}
    r = REV_DATA_ROW_START
    for h in history:
        ws.row_dimensions[r].height = 22
        for label, (cs, ce) in REV_COLS.items():
            for c in range(cs, ce + 1):
                ws.cell(row=r, column=c).border = B_all()
            ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
            cell = ws.cell(row=r, column=cs, value=h.get(label, ""))
            cell.font = _fnt()
            cell.alignment = _aln(h="center" if label in CENTER_LABELS else "left")
            cell.border = B_all()
        r += 1


def fill_summary(ws, groups: dict, project_name: str, author: str, today: str,
                 sheet_name_map: dict, current_version: str = ""):
    # メタ行1: プロジェクト名 / 作成日
    vs, ve = SUM_PROJECT_V
    ws.cell(row=SUM_META_ROW_1, column=vs, value=project_name)
    vs, ve = SUM_DATE_V
    ws.cell(row=SUM_META_ROW_1, column=vs, value=today)
    # メタ行2: 作成者 / バージョン / 合計件数
    vs, ve = SUM_AUTHOR_V
    ws.cell(row=SUM_META_ROW_2, column=vs, value=author)
    vs, ve = SUM_VERSION_V
    ws.cell(row=SUM_META_ROW_2, column=vs, value=f"v{current_version}" if current_version else "")
    vs, ve = SUM_TOTAL_V
    total = sum(len(v) for v in groups.values())
    ws.cell(row=SUM_META_ROW_2, column=vs, value=f"{total}件")

    sorted_types = ([t for t in TYPE_ORDER if t in groups]
                    + [t for t in groups.keys() if t not in TYPE_ORDER])
    r = SUM_DATA_ROW_START
    for i, type_key in enumerate(sorted_types):
        count = len(groups[type_key])
        sheet_name = sheet_name_map[type_key]
        set_h(ws, r, 24)
        MW(ws, r, *SUM_COLS["No"], value=str(i + 1),
           border=B_all(), h="center")
        MW(ws, r, *SUM_COLS["種別"], value=type_key, border=B_all())
        MW(ws, r, *SUM_COLS["件数"], value=count,
           border=B_all(), h="center")
        cell = MW(ws, r, *SUM_COLS["対応シート"], value=sheet_name,
                  border=B_all(), fg=C_FONT_LINK)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        r += 1


def fill_type_sheet(ws, type_key: str, features: list,
                    added_ids: set = None, modified_ids: set = None,
                    field_changes: dict = None, current_major: str = "1"):
    from openpyxl.utils import get_column_letter
    added_ids    = added_ids    or set()
    modified_ids = modified_ids or set()
    field_changes = field_changes or {}

    # 注記列の準備: このシートに関係するバージョンを収集
    feature_ids = {f.get("id") for f in features if f.get("id")}
    seen: list[str] = []
    for fid in feature_ids:
        for c in field_changes.get(fid, []):
            if c["version"] not in seen:
                seen.append(c["version"])
    all_versions = sorted(seen, key=lambda v: [int(x) for x in v.split(".")])
    version_col_map = {v: ANNOT_COL_START + i for i, v in enumerate(all_versions)}
    for col in version_col_map.values():
        ws.column_dimensions[get_column_letter(col)].width = 14

    # タイトル・セクション帯（件数入り）
    ws.cell(row=1, column=2, value=f"機能一覧 — {type_key}")
    ws.cell(row=ST_SEC_ROW, column=2, value=f"■ 機能一覧（{len(features)}件）")

    r = ST_DATA_ROW_START
    for i, feat in enumerate(features):
        feat_id    = feat.get("id", "")
        is_added   = feat_id in added_ids
        is_modified = feat_id in modified_ids

        bg = C_ADD_BG if is_added else (C_ALT_ROW if i % 2 == 1 else None)
        overview = clean_tech_business(feat.get("overview", "") or "")
        set_h(ws, r, max(26, min(150, (len(overview) // 30) * 16 + 28)))

        vals = {
            "ID":              (feat.get("id", ""), "center"),
            "API名/ファイル名": (feat.get("api_name", ""), "left"),
            "機能名":          (feat.get("name", ""), "left"),
            "処理概要":        (overview, "left"),
        }
        for label, (cs, ce) in ST_COLS.items():
            val, ha = vals[label]
            cell = MW(ws, r, cs, ce, value=val, border=B_all(), bg=bg, h=ha, v="top")
            if is_modified or is_added:
                cell.font = _fnt(color=C_FONT_R)

        # 注記列に ver{version} {author} を書き込む
        for change in field_changes.get(feat_id, []):
            col = version_col_map.get(change["version"])
            if col is None:
                continue
            color = C_FONT_R if change["version"].split(".")[0] == current_major else C_FONT_D
            cc = ws.cell(row=r, column=col,
                         value=f"ver{change['version']} {change['author']}")
            cc.font = Font(name=FONT_NAME, color=color, size=8)
            cc.alignment = Alignment(horizontal="center", vertical="center")

        r += 1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",        required=True)
    parser.add_argument("--output-dir",   required=True)
    parser.add_argument("--author",       required=True)
    parser.add_argument("--project-name", default="")
    parser.add_argument("--system-name",  default="")
    parser.add_argument("--template",     default=str(DEFAULT_TEMPLATE))
    parser.add_argument("--source-file",  default="",
                        help="更新時: 既存の機能一覧.xlsx のパス")
    parser.add_argument("--version-increment", default="minor",
                        choices=["minor", "major"],
                        help="minor: x.1増 / major: 1.0増")
    parser.add_argument("--force", action="store_true",
                        help="構造差分ゼロでも xlsx を強制再生成する（overview/name 変更反映用。"
                             "バージョンは据え置きで history は更新しない）")
    parser.add_argument("--project-dir", default="",
                        help="Salesforce プロジェクトルート。指定するとカスタムオブジェクト/フィールド名を日本語ラベルに置換する")
    args = parser.parse_args()

    if args.project_dir:
        try:
            from build_detail_design_json import load_obj_labels
            import text_cleaning as _tc
            _obj_labels, _fld_labels = load_obj_labels(Path(args.project_dir))
            _tc.set_sf_labels(_obj_labels, _fld_labels)
            print(f"  [INFO] オブジェクトラベル {len(_obj_labels)}件・フィールドラベル {sum(len(v) for v in _fld_labels.values())}件を読み込み")
        except Exception as e:
            print(f"  [WARN] カスタムラベル読み込みに失敗しました（処理は続行）: {e}", file=sys.stderr)

    template = Path(args.template)
    if not template.exists():
        print(f"[ERROR] テンプレートが見つかりません: {template}", file=sys.stderr)
        sys.exit(1)

    features = json.loads(Path(args.input).read_text(encoding="utf-8"))
    today = date.today().strftime("%Y-%m-%d")

    # ── バージョン判定 ──────────────────────────────────────
    is_major    = (args.version_increment == "major")
    source_file = args.source_file.strip()
    prev_meta   = read_meta(source_file) if source_file else None

    if prev_meta:
        prev_history_len = len(prev_meta.get("history", []))
        # 改版履歴 20 行制限: 既存履歴が 20 以上なら minor 指定でも major に強制昇格し履歴リセット
        forced_major = False
        if prev_history_len >= 20 and args.version_increment == "minor":
            print(f"  [WARN] 改版履歴が {prev_history_len} 件に達しているため minor → major に強制昇格し、履歴をリセットします")
            args.version_increment = "major"
            is_major = True
            forced_major = True
        current_version = increment_version(prev_meta.get("version", "1.0"),
                                            args.version_increment)
        # major 時は履歴リセット（手動・強制問わず。メジャーUP 1行だけ残す）
        history         = [] if is_major else prev_meta.get("history", [])
        old_features    = prev_meta.get("features", [])
        is_initial      = False
        if forced_major:
            print(f"メジャー昇格モード（履歴リセット）: {prev_meta.get('version', '?')} → {current_version}")
        else:
            print(f"更新モード: {prev_meta.get('version', '?')} → {current_version}"
                  + (" (メジャー)" if is_major else ""))
    else:
        current_version = "1.0"
        history         = []
        old_features    = []
        is_initial      = True
        print("新規作成モード: v1.0")

    # ── 部分更新: 入力にない機能を _meta から保持してマージ ──────
    # 全件処理時は features = old_features と同じ範囲なので影響なし。
    # 一部機能のみ指定時は、指定外の機能を保持して削除されるのを防ぐ。
    # ただし ledger で deprecated=true になっている機能は保持せず除外する
    # （削除済み機能が xlsx に残り続けるのを防ぐ）。
    if prev_meta and not is_major and old_features:
        input_ids  = {f.get("id") for f in features if f.get("id")}
        # 台帳から deprecated ID を取得（feature_list.json の隣にある feature_ids.yml を参照）
        deprecated_ids: set = set()
        try:
            import yaml
            ledger_path = Path(args.input).parent / "feature_ids.yml"
            if ledger_path.exists():
                ledger = yaml.safe_load(ledger_path.read_text(encoding="utf-8")) or {}
                deprecated_ids = {e.get("id") for e in ledger.get("features", [])
                                  if e.get("deprecated")}
        except ImportError:
            print("[WARN] pyyaml が未インストールのため deprecated_ids を取得できません。"
                  "pip install pyyaml を実行してください。", file=sys.stderr)
        except Exception as e:
            print(f"[WARN] feature_ids.yml の読み込みに失敗しました: {e}", file=sys.stderr)
        preserved  = [f for f in old_features
                      if f.get("id") not in input_ids
                      and f.get("id") not in deprecated_ids]
        old_subset = [f for f in old_features if f.get("id") in input_ids]
        # 差分は「今回処理した機能」のみで計算する
        diffs = compare_features(old_subset, features)
        # 削除対象（deprecated）を diffs.removed に反映
        dropped = [f for f in old_features
                   if f.get("id") not in input_ids and f.get("id") in deprecated_ids]
        if dropped:
            diffs["removed"] = diffs.get("removed", []) + dropped
        # xlsx に書く完全リスト = 今回の更新分 + 保持分
        features = features + preserved
        n_preserved = len(preserved)
        n_dropped = len(dropped)
        if n_preserved:
            print(f"  [INFO] {n_preserved}件の機能を _meta から保持（入力外）")
        if n_dropped:
            print(f"  [INFO] {n_dropped}件の deprecated 機能を _meta から除外（削除扱い）")
    else:
        diffs = compare_features(old_features, features)

    if prev_meta and not is_major and not has_any_diff(diffs):
        if not args.force:
            print("差分なし: 既存ファイルと一致しているため更新をスキップしました")
            sys.exit(0)
        # --force 指定時: xlsx のみ再生成（version は据え置き・history 追記しない）
        print("--force 指定: 構造差分ゼロだが xlsx を再生成します（version 据え置き）")
        current_version = prev_meta.get("version", "1.0")
        new_entries = []
    else:
        last_no = max((h["項番"] for h in history
                       if isinstance(h.get("項番"), int)), default=0)
        new_entries = build_revision_entries(
            current_version, diffs, args.author, today,
            start_no=last_no + 1, is_major=is_major, is_initial=is_initial,
        )
    history = history + new_entries

    # ── 差分IDセット（赤字・背景用）──────────────────────────
    # 初回作成時は全件"追加"扱いになるが、比較対象がないため色付けは不要
    if is_initial or is_major:
        added_ids, modified_ids = set(), set()
    else:
        added_ids    = {f.get("id") for f in diffs["added"]   if f.get("id")}
        modified_ids = {m["id"]     for m in diffs["modified"] if m.get("id")}

    # ── 注記列（ver{version} {author}）用の変更履歴 ─────────────
    prev_field_changes: dict = {} if is_major else (
        (prev_meta or {}).get("field_changes", {})
    )
    current_major = current_version.split(".")[0]

    if is_initial or is_major:
        field_changes = {}
    else:
        field_changes = dict(prev_field_changes)
        entry = {"version": current_version, "author": args.author}
        for fid in added_ids | modified_ids:
            field_changes[fid] = field_changes.get(fid, []) + [entry]

    # ── xlsx 生成 ─────────────────────────────────────────────
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "機能一覧.xlsx"
    shutil.copy(template, out_path)

    # 種別ごとにグループ化
    groups: dict = defaultdict(list)
    for feat in features:
        groups[feat.get("type", "その他")].append(feat)

    wb = load_workbook(out_path)

    fill_revision(wb["改版履歴"], history, args.project_name, today)

    sheet_name_map = {}
    sorted_types = ([t for t in TYPE_ORDER if t in groups]
                    + [t for t in groups.keys() if t not in TYPE_ORDER])
    for type_key in sorted_types:
        sheet_name = type_key
        for ch in r'[]:*?/\\':
            sheet_name = sheet_name.replace(ch, "_")
        new_ws = clone_sheet(wb, "__SHEET_TEMPLATE__", sheet_name)
        sheet_name_map[type_key] = sheet_name
        fill_type_sheet(new_ws, type_key, groups[type_key],
                        added_ids=added_ids, modified_ids=modified_ids,
                        field_changes=field_changes, current_major=current_major)

    fill_summary(wb["サマリー"], groups, args.project_name, args.author, today,
                 sheet_name_map, current_version=current_version)

    wb["__SHEET_TEMPLATE__"].sheet_state = "hidden"

    # ── _meta 保存（次回差分判定用）────────────────────────
    write_meta(wb, {
        "version":      current_version,
        "date":         today,
        "project_name":   args.project_name,
        "system_name":    args.system_name,
        "author":         args.author,
        "features":       features,
        "history":        history,
        "field_changes":  field_changes,
    })

    wb.save(out_path)
    print(f"機能一覧生成完了: v{current_version} → {out_path}")
    print(f"  差分: 追加{len(diffs['added'])} / 削除{len(diffs['removed'])} / 変更{len(diffs['modified'])}")
    print(f"  シート: {wb.sheetnames}")


if __name__ == "__main__":
    main()
