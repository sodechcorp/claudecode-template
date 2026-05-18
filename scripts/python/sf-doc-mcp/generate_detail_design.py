# scripts/python/sf-doc-mcp/generate_detail_design.py
# -*- coding: utf-8 -*-
"""
詳細設計書.xlsx を1機能分生成する（テンプレート読込方式・新JSONスキーマ対応）。

  詳細設計書テンプレート.xlsx（build_detail_design_template.py で生成した「器」）を
  コピーしてセル値 + 図形PNGを流し込む。

6シート構成:
  1. 改版履歴           : メタ + 履歴テーブル
  2. 概要               : 機能名 / 機能概要 / 目的 / 利用者 / 起点画面 / 操作トリガー
  3. 業務フロー         : スイムレーン図PNG + フロー表(No/アクター/処理内容/分岐条件)
  4. 対象オブジェクト   : ER図PNG + 項目表(オブジェクト名/項目API名/項目ラベル/読み書き区分/備考)
  5. 処理概要           : フローチャートPNG + 処理表(No/処理内容/コンポーネント/分岐条件)
  6. 関連コンポーネント : コンポーネント図PNG + 一覧表(コンポーネント名/種別/役割/依存方向)

Usage:
  python generate_detail_design.py \\
    --input  detail_design.json \\
    --template "C:/.../詳細設計書テンプレート.xlsx" \\
    --output-dir "C:/.../出力先" \\
    [--version-increment minor]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
from datetime import date as _date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import design_revision as dr
from build_detail_design_template import (
    section_band, diagram_area, data_rows, setup_grid, set_h,
    GRID_LEFT, GRID_RIGHT,
    C_BAND_BLUE, C_TITLE_DARK,
)
from meta_store import read_meta, write_meta
from tmp_utils import get_project_tmp_dir, set_project_tmp_dir
from version_manager import increment_version

# ── 色定数 ─────────────────────────────────────────────────────────
C_HDR_BLUE  = "2E75B6"
C_BAND_BLUE = "0070C0"
C_LABEL_BG  = "D9E1F2"
C_FONT_D    = "000000"
C_FONT_W    = "FFFFFF"

THIN = Side(style="thin", color="8B9DC3")

# ── テンプレート行番号定数 ─────────────────────────────────────────
# build_detail_design_template.py の構造から算出
# 改版履歴
REV_META_ROW       = 3
REV_META_PROJECT_V = (6, 18)
REV_META_DATE_V    = (23, 31)
REV_DATA_ROW_START = 6
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6,  11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 31),
}

# 概要（row3〜row8: 機能名/機能概要/目的/利用者/起点画面/操作トリガー）
OV_LABEL_VAL_CS = 7
OV_LABEL_VAL_CE = 31
OV_ROWS = {
    "name_ja":        3,
    "summary":        4,
    "purpose":        5,
    "users":          6,
    "trigger_screen": 7,
    "trigger":        8,
}

# 業務フロー: テーブルヘッダは row4、データは row5 から動的追加
BF_DATA_ROW_START  = 5
BF_STEP_CS,  BF_STEP_CE  = 2,  3
BF_ACTOR_CS, BF_ACTOR_CE = 4,  8
BF_ACT_CS,   BF_ACT_CE   = 9,  31
# BF_COL_GROUPS: データ行のマージセル定義
BF_COL_GROUPS = [
    (BF_STEP_CS, BF_STEP_CE),
    (BF_ACTOR_CS, BF_ACTOR_CE),
    (BF_ACT_CS, BF_ACT_CE),
]

# 対象オブジェクト: 表上・図下レイアウト（動的行）
# ※ 項目ラベルが左（8-14）、API名が右（15-20）の順
OBJ_DATA_ROW_START = 5
OBJ_NAME_CS,  OBJ_NAME_CE  = 2,  7
OBJ_FLBL_CS,  OBJ_FLBL_CE  = 8,  14   # 項目ラベル（左）
OBJ_FAPI_CS,  OBJ_FAPI_CE  = 15, 20   # 項目API名（右）
OBJ_ACC_CS,   OBJ_ACC_CE   = 21, 23
OBJ_NOTE_CS,  OBJ_NOTE_CE  = 24, 31
OBJ_COL_GROUPS = [
    (OBJ_NAME_CS, OBJ_NAME_CE),
    (OBJ_FLBL_CS, OBJ_FLBL_CE),
    (OBJ_FAPI_CS, OBJ_FAPI_CE),
    (OBJ_ACC_CS,  OBJ_ACC_CE),
    (OBJ_NOTE_CS, OBJ_NOTE_CE),
]

# 処理概要: テーブルヘッダは row4、データは row5 から動的追加
PROC_DATA_ROW_START = 5
PROC_STEP_CS, PROC_STEP_CE = 2,  3
PROC_DESC_CS, PROC_DESC_CE = 4,  19
PROC_COMP_CS, PROC_COMP_CE = 20, 31
PROC_COL_GROUPS = [
    (PROC_STEP_CS, PROC_STEP_CE),
    (PROC_DESC_CS, PROC_DESC_CE),
    (PROC_COMP_CS, PROC_COMP_CE),
]

# 関連コンポーネント: テーブルヘッダは row4、データは row5 から動的追加
COMP_DATA_ROW_START = 5
COMP_NAME_CS, COMP_NAME_CE = 2,  9
COMP_TYPE_CS, COMP_TYPE_CE = 10, 13
COMP_ROLE_CS, COMP_ROLE_CE = 14, 31  # 依存方向列廃止のため役割を31まで拡張
COMP_COL_GROUPS = [
    (COMP_NAME_CS, COMP_NAME_CE),
    (COMP_TYPE_CS, COMP_TYPE_CE),
    (COMP_ROLE_CS, COMP_ROLE_CE),
]

GRID_RIGHT_C = 31

SCALAR_FIELDS  = ["summary", "purpose", "users", "trigger_screen", "trigger"]
SECTION_SHEETS = {
    "business_flow":    "業務フロー",
    "process_steps":    "処理概要",
    "related_objects":  "対象オブジェクト",
    "components":       "関連コンポーネント",
}

# 動的シートの図エリア高さ（行数）
DIAGRAM_AREA_ROWS = 30
# 動的シートの空行追加数（手動入力用）— 0=データ行のみ
DYNAMIC_EMPTY_ROWS = 0


# ── スタイルヘルパー ────────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg: c.fill = _fill(bg)
    if border: c.border = border
    return c

def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)

# set_h is imported from build_detail_design_template


# ── PNG埋め込み ────────────────────────────────────────────────────
def _pad_png_to_aspect(png_path: str, target_w_h: float, align: str = "center") -> str:
    """PNG を target_w_h（幅/高さ）のアスペクトに白背景で padding して返す。
    FG ごとに異なる graphviz 出力アスペクトを正規化し、全 FG で同一表示サイズを保証する。
    元ファイルは保持し、padding 済みファイルを隣接 .pad.png として返す。
    align="left" にすると画像を左寄せで貼り付け（右端に白帯）。デフォルトは中央。
    """
    try:
        from PIL import Image as _PILImage
        img = _PILImage.open(png_path)
        w, h = img.width, img.height
        cur = w / h if h else 1.0
        if abs(cur - target_w_h) < 0.02:
            return png_path
        if cur < target_w_h:
            new_w = int(h * target_w_h)
            new_h = h
        else:
            new_w = w
            new_h = int(w / target_w_h)
        canvas = _PILImage.new("RGBA", (new_w, new_h), (255, 255, 255, 255))
        paste_x = 0 if align == "left" else (new_w - w) // 2
        paste_y = (new_h - h) // 2
        canvas.paste(img, (paste_x, paste_y))
        out = png_path[:-4] + ".pad.png"
        canvas.convert("RGB").save(out, "PNG")
        return out
    except Exception as e:
        print(f"  [WARN] PNG padding 失敗({png_path}): {e}")
        return png_path


def _resize_png_to_width(png_path: str, target_width: int) -> str:
    """PNG を target_width にアスペクト維持リサイズ。FG 間でノード密度を揃えるため
    graphviz の size 強制スケーリングを撤廃した代わりに、埋め込み前に幅を固定する。"""
    try:
        from PIL import Image as _PILImage
        img = _PILImage.open(png_path)
        if img.width == target_width:
            return png_path
        scale = target_width / img.width
        new_h = max(1, int(img.height * scale))
        resized = img.resize((target_width, new_h), _PILImage.LANCZOS)
        out = png_path[:-4] + ".wfix.png"
        resized.save(out, "PNG")
        return out
    except Exception as e:
        print(f"  [WARN] PNG リサイズ失敗({png_path}): {e}")
        return png_path


def _embed_image(ws, png_path: str, anchor: str,
                 img_w: int = 840, img_h: int | None = None,
                 max_h: int | None = None,
                 max_w: int | None = None,
                 target_aspect_wh: float | None = None,
                 target_width: int | None = None,
                 zoom: float | None = None,
                 align: str = "center",
                 display_width: int | None = None):
    """PNGをExcelシートに埋め込む。

    動作モード（優先順）:
      0. target_aspect_wh が指定されていれば _pad_png_to_aspect で正規化（align で左寄せも可）
         target_width が指定されていれば _resize_png_to_width でリサイズ（後方互換）
      1. img_h が指定されていれば固定サイズ（img_w × img_h・アスペクト比無視）
      2. display_width: PNG を物理リサイズせず Excel 表示幅のみ固定。高解像度 PNG を
         Excel 側でレンダリングさせることで PIL 縮小によるぼやけを回避する
      3. zoom が指定されていれば自然サイズに zoom を掛けた表示サイズ:
         全 FG で同一 zoom をかけることでノード絶対サイズを統一できる
      4. max_w と max_h が両方指定 → fit-in-box:
         アスペクト比を維持したまま max_w × max_h の枠に収まる最大サイズ
      5. それ以外は従来互換: 幅を img_w に合わせて高さをアスペクト比で決める
    """
    try:
        if not Path(png_path).exists():
            return
        if target_aspect_wh is not None:
            png_path = _pad_png_to_aspect(png_path, target_aspect_wh, align=align)
        if target_width is not None:
            png_path = _resize_png_to_width(png_path, target_width)
        img = XLImage(png_path)
        img.anchor = anchor
        if img_h is not None:
            img.width = img_w
            img.height = img_h
        elif display_width is not None:
            orig_w = float(img.width or 1)
            orig_h = float(img.height or 1)
            ratio = orig_h / orig_w if orig_w else 1.0
            img.width = display_width
            img.height = max(1, int(display_width * ratio))
        elif zoom is not None:
            orig_w = float(img.width or 1)
            orig_h = float(img.height or 1)
            img.width = max(1, int(orig_w * zoom))
            img.height = max(1, int(orig_h * zoom))
        elif max_w is not None and max_h is not None:
            orig_w = float(img.width or 1)
            orig_h = float(img.height or 1)
            scale = min(max_w / orig_w, max_h / orig_h)
            img.width = int(orig_w * scale)
            img.height = int(orig_h * scale)
        else:
            # target_width が指定されていれば PNG リサイズ後の幅を表示幅にも使う
            _display_w = target_width if target_width is not None else img_w
            ratio = img.height / img.width if img.width else 1.0
            w, h = _display_w, int(_display_w * ratio)
            if max_h and h > max_h:
                h = max_h
                w = int(h / ratio)
            img.width = w
            img.height = h
        ws.add_image(img)
    except Exception as e:
        print(f"  [WARN] 画像埋め込み失敗({anchor}): {e}")


# ── シート埋め込み ─────────────────────────────────────────────────
def fill_revision(ws, data: dict, history: list[dict]):
    vs, _ = REV_META_PROJECT_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("project_name", ""))
    vs, _ = REV_META_DATE_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("date", ""))
    dr.fill_revision_table(ws, history, REV_COLS, REV_DATA_ROW_START)


def fill_overview(ws, data: dict, changed_fields: set):
    """概要シートに値を書き込む。"""
    for key, row in OV_ROWS.items():
        val = data.get(key, "")
        if val:
            cell = ws.cell(row=row, column=OV_LABEL_VAL_CS, value=val)
            if key in changed_fields:
                dr.apply_red(cell, size=10)


def fill_business_flow(ws, data: dict, changed_step_nos: set,
                       png_path: str | None):
    """業務フローシート: テーブル(動的行) + スイムレーン図。"""
    flows = data.get("business_flow", [])
    n_data = len(flows)
    total_rows = n_data + DYNAMIC_EMPTY_ROWS

    # データ行 + 空行の枠を作成
    r = BF_DATA_ROW_START
    data_rows(ws, r, r + total_rows - 1, BF_COL_GROUPS, row_h=24)

    # データ書き込み
    for i, flow in enumerate(flows):
        step_no = flow.get("step", i + 1)
        is_changed = step_no in changed_step_nos
        set_h(ws, r, 24)

        nexts = flow.get("next", [])
        c1 = MW(ws, r, BF_STEP_CS,  BF_STEP_CE,  step_no,
                border=B_all(), h="center")
        c2 = MW(ws, r, BF_ACTOR_CS, BF_ACTOR_CE, flow.get("actor", ""),
                border=B_all(), h="center")
        c3 = MW(ws, r, BF_ACT_CS,   BF_ACT_CE,   flow.get("action", ""),
                border=B_all(), wrap=True, v="top")
        if is_changed:
            for c in (c1, c2, c3):
                dr.apply_red(c)
        r += 1

    # スペーサー + 図エリア
    spacer_row = BF_DATA_ROW_START + total_rows
    set_h(ws, spacer_row, 8)
    diagram_start = spacer_row + 1
    diagram_area(ws, diagram_start, "業務フロー図（自動生成）", section_no=2)

    # 図埋め込み（fit-in-box: 横長・縦長どちらでも枠内に収まる最大サイズ）
    # Z-4b: target_aspect_wh=0.60 でアスペクト正規化（FG-010 REF1/REF2 の中央値）→ 全 FG で 500×800 枠に統一
    img_anchor = f"B{diagram_start + 1}"
    if png_path:
        # AA-2: zoom 方式。graphviz 自然 PNG に同一倍率をかけてノード絶対サイズを FG 間で統一
        _embed_image(ws, png_path, img_anchor, zoom=0.50)


def _compute_obj_note(obj_api: str, field_access: str, data: dict) -> str:
    """object_access + process_steps から項目ごとの備考テキストを生成する。

    フィールドの読み書き区分に応じて関連するアクセスのみ抽出する。
    例（読み取り専用項目）: 「プリチェック判定時に参照。」
    例（書き込み項目）: 「コンサルテーション依頼時に更新・登録。見積作成時に新規作成。」
    """
    _OP_JA = {"R": "参照", "W": "更新・登録", "RW": "参照・更新", "INSERT": "新規作成"}
    _WRITE_OPS = {"W", "RW", "INSERT"}
    _READ_OPS  = {"R", "RW"}

    object_access = data.get("object_access", [])
    process_steps = data.get("process_steps", [])

    # comp_api_name（API名）でルックアップ。なければ component（タイプ名）でフォールバック
    comp_to_step: dict[str, str] = {}
    for ps in process_steps:
        key = ps.get("comp_api_name") or ps.get("component", "")
        if key and key not in comp_to_step:
            comp_to_step[key] = ps.get("title", "")

    accesses = [a for a in object_access if a.get("object") == obj_api]
    if not accesses:
        return ""

    # フィールドの読み書き区分に応じて関連するアクセスのみ抽出
    relevant = [
        a for a in accesses
        if (field_access == "R" and a.get("operation", "") in _READ_OPS)
        or (field_access in ("W", "INSERT") and a.get("operation", "") in _WRITE_OPS)
        or field_access == "RW"
    ]
    if not relevant:
        relevant = accesses  # フォールバック

    # 備考には「どの処理ステップで使われるか」のみ記載（操作種別は読み書き区分列と重複するため省く）
    step_titles = []
    for acc in relevant:
        step_title = comp_to_step.get(acc.get("component", ""), "")
        # M-5b: 機械付与「を行う」で終わる title は備考欄に混入させない
        if step_title and step_title not in step_titles and not step_title.endswith('を行う'):
            step_titles.append(step_title)
    if step_titles:
        return "・".join(step_titles)

    # X-4e: process_steps に対応タイトルがない場合は object_access の操作種別から備考を構築
    ops_ja_parts: list[str] = []
    ops_seen: set[str] = set()
    for acc in relevant:
        op = acc.get("operation", "")
        op_ja = _OP_JA.get(op, "")
        if op_ja and op_ja not in ops_seen:
            ops_seen.add(op_ja)
            ops_ja_parts.append(op_ja)
    return "・".join(ops_ja_parts) if ops_ja_parts else ""


def fill_target_objects(ws, data: dict, changed_obj_keys: set,
                        png_path: str | None):
    """対象オブジェクトシート: 項目テーブル(動的行・縦結合) + 図。

    列順: オブジェクト名(縦結合) | 項目ラベル | 項目API名 | 読み書き区分(日本語) | 備考
    """
    _ACCESS_JA = {
        "R": "参照", "W": "更新", "RW": "参照・更新", "INSERT": "新規作成",
    }
    objects = data.get("related_objects", [])

    r = OBJ_DATA_ROW_START

    for obj in objects:
        obj_api   = obj.get("api_name", "")
        obj_label = f"{obj.get('label', '')} ({obj_api})"
        is_changed = obj_api in changed_obj_keys
        fields = obj.get("fields", [])
        if not fields:
            continue
        n_fields      = len(fields)
        obj_start_row = r

        # 読み書き区分・備考はオブジェクト単位で統一（フィールドごとに変わらないため縦結合）
        # obj['access_ja'] が設定されていればそれを優先（R+INSERT 等の複合操作を正しく表示）
        access    = fields[0].get("access", "") if fields else ""
        access_ja = obj.get("access_ja") or _ACCESS_JA.get(access, access)
        note      = fields[0].get("note", "") if fields else ""
        if not note:
            note = _compute_obj_note(obj_api, access, data)

        for fi, field in enumerate(fields):
            set_h(ws, r, 22)

            # オブジェクト名列: 全行書くが後で縦結合する
            MW(ws, r, OBJ_NAME_CS, OBJ_NAME_CE,
               obj_label if fi == 0 else "", border=B_all())

            c2 = MW(ws, r, OBJ_FLBL_CS, OBJ_FLBL_CE, field.get("label", ""),
                    border=B_all())
            c3 = MW(ws, r, OBJ_FAPI_CS, OBJ_FAPI_CE, field.get("api_name", ""),
                    border=B_all())
            # 読み書き区分・備考は初行のみ書き込み（後で縦結合）
            c4 = MW(ws, r, OBJ_ACC_CS,  OBJ_ACC_CE,  access_ja if fi == 0 else "",
                    border=B_all(), h="center")
            c5 = MW(ws, r, OBJ_NOTE_CS, OBJ_NOTE_CE, note if fi == 0 else "",
                    border=B_all(), wrap=True, v="top")
            if is_changed:
                for c in (c2, c3, c4, c5):
                    dr.apply_red(c)
            r += 1

        def _merge_col(ws, row_start, row_end, cs, ce, value, is_changed_flag, h="left", v="center"):
            for ri in range(row_start, row_end + 1):
                try:
                    ws.unmerge_cells(start_row=ri, start_column=cs,
                                     end_row=ri, end_column=ce)
                except Exception:
                    pass
            ws.merge_cells(start_row=row_start, start_column=cs,
                           end_row=row_end, end_column=ce)
            mc = ws.cell(row=row_start, column=cs)
            mc.value     = value
            mc.font      = _fnt()
            mc.alignment = _aln(h=h, v=v, wrap=True)
            mc.border    = B_all()
            if is_changed_flag:
                dr.apply_red(mc)

        # オブジェクト名・読み書き区分・備考を縦結合
        if n_fields > 1:
            _merge_col(ws, obj_start_row, r - 1, OBJ_NAME_CS, OBJ_NAME_CE,
                       obj_label, is_changed, v="center")
            _merge_col(ws, obj_start_row, r - 1, OBJ_ACC_CS, OBJ_ACC_CE,
                       access_ja, is_changed, h="center", v="center")
            _merge_col(ws, obj_start_row, r - 1, OBJ_NOTE_CS, OBJ_NOTE_CE,
                       note, is_changed, v="top")

    # スペーサー + 図エリア（動的位置）
    spacer_row = r
    set_h(ws, spacer_row, 8)
    diagram_start = spacer_row + 1
    diagram_area(ws, diagram_start, "オブジェクト関連図（自動生成）", section_no=2)

    img_anchor = f"B{diagram_start + 1}"
    if png_path:
        # AA-3g: zoom=0.4 で Phase 1 より少し大きめ表示（全 FG 統一倍率）
        _embed_image(ws, png_path, img_anchor, zoom=0.4)


def _estimate_row_height(text: str, chars_per_line: int = 34,
                         line_pt: int = 14, min_h: int = 24, max_h: int = 300) -> int:
    """テキストの折り返しを考慮して行の高さ（ポイント）を推定する。"""
    import math as _math
    if not text:
        return min_h
    lines = text.split("\n")
    total = sum(_math.ceil(max(len(ln), 1) / chars_per_line) for ln in lines)
    return min(max_h, max(min_h, total * line_pt + 8))


def fill_process_overview(ws, data: dict, changed_step_nos: set,
                          png_path: str | None):
    """処理概要シート: テーブル(動的行・動的高) + フローチャート。"""
    steps = data.get("process_steps", [])
    n_data = len(steps)
    total_rows = n_data + DYNAMIC_EMPTY_ROWS

    # データ行の枠を作成（デフォルト高は後で上書き）
    r = PROC_DATA_ROW_START
    data_rows(ws, r, r + total_rows - 1, PROC_COL_GROUPS, row_h=30)

    for i, ps in enumerate(steps):
        step_no = ps.get("step", i + 1)
        is_changed = step_no in changed_step_nos

        desc_text = ps.get("description", "").strip()
        row_h = _estimate_row_height(desc_text)
        set_h(ws, r, row_h)

        component = ps.get("component") or ""
        branch = ps.get("branch") or ""

        c1 = MW(ws, r, PROC_STEP_CS, PROC_STEP_CE, step_no,
                border=B_all(), h="center")
        c2 = MW(ws, r, PROC_DESC_CS, PROC_DESC_CE, desc_text,
                border=B_all(), wrap=True, v="top")
        c3 = MW(ws, r, PROC_COMP_CS, PROC_COMP_CE, component,
                border=B_all(), wrap=True, v="top")
        if is_changed:
            for c in (c1, c2, c3):
                dr.apply_red(c)
        r += 1

    # スペーサー + 図エリア
    spacer_row = PROC_DATA_ROW_START + total_rows
    set_h(ws, spacer_row, 8)
    diagram_start = spacer_row + 1
    diagram_area(ws, diagram_start, "処理フロー図（自動生成）", section_no=2)

    img_anchor = f"B{diagram_start + 1}"
    if png_path:
        # AA-2c: Excel 表示幅のみ 1120px 固定（PIL 物理リサイズなし）。
        # graphviz 高解像度 PNG をそのまま埋め込み Excel 側レンダリングに任せる → 文字のぼやけ回避。
        # graphviz 自然レイアウトが既に左寄せのため「図形は左寄り」は自動で成立。
        _embed_image(ws, png_path, img_anchor, display_width=1120)


def fill_related_components(ws, data: dict, changed_comp_keys: set,
                            png_path: str | None):
    """関連コンポーネントシート: テーブル(動的行) + コンポーネント図。"""
    components = data.get("components", [])
    n_data = len(components)
    total_rows = n_data + DYNAMIC_EMPTY_ROWS

    r = COMP_DATA_ROW_START
    data_rows(ws, r, r + total_rows - 1, COMP_COL_GROUPS, row_h=24)

    for comp in components:
        api_name = comp.get("api_name", "")
        is_changed = api_name in changed_comp_keys
        set_h(ws, r, 24)

        c1 = MW(ws, r, COMP_NAME_CS, COMP_NAME_CE, api_name,
                border=B_all())
        c2 = MW(ws, r, COMP_TYPE_CS, COMP_TYPE_CE, comp.get("type", ""),
                border=B_all(), h="center")
        c3 = MW(ws, r, COMP_ROLE_CS, COMP_ROLE_CE, comp.get("role", ""),
                border=B_all(), wrap=True, v="top")
        if is_changed:
            for c in (c1, c2, c3):
                dr.apply_red(c)
        r += 1

    # スペーサー + 図エリア
    spacer_row = COMP_DATA_ROW_START + total_rows
    set_h(ws, spacer_row, 8)
    diagram_start = spacer_row + 1
    diagram_area(ws, diagram_start, "コンポーネント関連図（自動生成）", section_no=2)

    img_anchor = f"B{diagram_start + 1}"
    if png_path:
        # AA: 幅固定・縦伸長。コンポーネント図は幅 1050px に固定し高さは内容量で伸長
        _embed_image(ws, png_path, img_anchor, target_width=1050, max_w=662, max_h=9999)



# ── GFスキーマ正規化 ────────────────────────────────────────────────
import re as _re

# ── SFプロジェクト → メタデータパス マッピング ───────────────────────────
# flows/classes は greenfield、フィールドラベル翻訳は両方から取得（GF_UATが補完）
_SF_PROJECT_PATHS: dict[str, str] = {
    # キーは小文字で統一。data.get('project_name') は .lower() して照合する
    "greenfield": "C:/workspace/16_グリーンフィールド/greenfield",
    "link_prod":  "C:/workspace/21_リンク/link_prod",
    "link":       "C:/workspace/21_リンク/link_prod",
}
_SF_EXTRA_LABEL_PATHS: list[str] = [
    "C:/workspace/16_グリーンフィールド/GF_UAT",
]
# メタデータから構築するフィールドラベルマップ {obj_api: {field_api: ja_label}}
_SF_FIELD_LABELS: dict[str, dict[str, str]] = {}
# オブジェクトラベルマップ {obj_api: ja_label}
_SF_OBJ_LABELS: dict[str, str] = {}
# コンポーネント別フィールドマップ {comp_api_name: {obj_api: {field_api}}}
_SF_COMP_FIELDS: dict[str, dict[str, set]] = {}
# K-C: コンポーネント別 operation マップ {comp_api_name: {obj_api: "R" | "W" | "INSERT"}}
# Apex の DML / VF の inputField / Site.* 検出結果を格納する
_SF_COMP_OPS: dict[str, dict[str, str]] = {}
# 直近にロードしたSFプロジェクトのベースパス（VF→Apex 補強で参照）
_CURRENT_SF_BASE_PATH: str = ""


def _merge_op(dst: dict[str, str], obj: str, op: str) -> None:
    """op を dst[obj] に合成する（INSERT > W > R の優先度）。"""
    rank = {"R": 1, "W": 2, "INSERT": 3}
    cur = dst.get(obj)
    if not cur or rank.get(op, 0) > rank.get(cur, 0):
        dst[obj] = op


def _parse_flow_fields(flow_path: Path) -> dict[str, set]:
    """Flow XMLから {obj_api: {field_api}} を抽出する。"""
    try:
        content = flow_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return {}
    result: dict[str, set] = {}
    tags = ("recordCreates", "recordUpdates", "recordLookups", "recordDeletes")
    for tag in tags:
        for block in _re.findall(rf'<{tag}>(.*?)</{tag}>', content, _re.DOTALL):
            obj_m = _re.search(r'<object>([A-Za-z0-9_]+)</object>', block)
            if not obj_m:
                continue
            obj_api = obj_m.group(1)
            fields = set(_re.findall(r'<field>([A-Za-z0-9_]+)</field>', block))
            if fields:
                result.setdefault(obj_api, set()).update(fields)
    return result


def _parse_apex_ops(content: str, fields_by_obj: dict[str, set]) -> dict[str, str]:
    """Apex クラスから {obj_api: op} を推定する（op は "R" / "W" / "INSERT"）。

    優先度:
      1. `insert xxx;` / `Database.insert(...)` → INSERT
      2. `update xxx;` / `upsert xxx;` / `Database.update(...)` / `Database.upsert(...)` → W
      3. `Site.setPassword` / `System.setPassword` / `Site.changePassword` / `Site.validatePassword` → W (User)
      4. `Site.createPortalUser` → INSERT (User, Contact)
      5. SOQL で SELECT のみ → R
    """
    ops: dict[str, str] = {}
    # 型宣言とローカル変数名を収集: List<User> users; Map<Id, Contact> cmap; User u; Contact c;
    var_types: dict[str, str] = {}
    for m in _re.finditer(
        r'\b(?:List|Set|Iterable|Queue|Map<[^>,]+,)\s*<\s*([A-Za-z][A-Za-z0-9_]*)\s*>\s+(\w+)\s*[=;,)]',
        content):
        obj_t, var = m.group(1), m.group(2)
        if obj_t[0].isupper():
            var_types[var] = obj_t
    # 単体型宣言: User u = ...;  Contact c = new Contact();
    for m in _re.finditer(
        r'\b([A-Z][A-Za-z0-9_]*(?:__c)?)\s+(\w+)\s*=\s*(?:new\s+|\[|[\w.])',
        content):
        obj_t, var = m.group(1), m.group(2)
        if obj_t in ("String", "Integer", "Long", "Decimal", "Double", "Boolean",
                     "Date", "Datetime", "Time", "Id", "Object", "Blob", "PageReference",
                     "ApexPages", "System", "Database", "Schema", "JSON", "Test"):
            continue
        var_types.setdefault(var, obj_t)

    # DML 動詞による op 推定
    for m in _re.finditer(r'\b(insert|update|upsert|delete)\s+(\w+)\s*[;,]', content):
        verb, var = m.group(1).lower(), m.group(2)
        obj_t = var_types.get(var)
        if not obj_t:
            continue
        op = "INSERT" if verb in ("insert", "upsert") else "W"
        _merge_op(ops, obj_t, op)
    # Database.insert / update / upsert
    for m in _re.finditer(r'\bDatabase\.(insert|update|upsert|delete)\s*\(\s*(\w+)', content):
        verb, var = m.group(1).lower(), m.group(2)
        obj_t = var_types.get(var)
        if not obj_t:
            continue
        op = "INSERT" if verb in ("insert", "upsert") else "W"
        _merge_op(ops, obj_t, op)

    # Site.* / System.setPassword による User 更新/新規作成
    if _re.search(r'\bSite\.(?:login|forgotPassword|validatePassword|changePassword|passwordless)\b'
                  r'|\bSystem\.setPassword\b', content):
        _merge_op(ops, "User", "W")
    if _re.search(r'\bSite\.createPortalUser\b', content):
        _merge_op(ops, "User", "INSERT")
        _merge_op(ops, "Contact", "INSERT")

    # SOQL で参照のみの場合、R を埋める（既に W/INSERT があれば上書きしない）
    for obj_api in fields_by_obj.keys():
        if obj_api in ("__any__", "Id"):
            continue
        _merge_op(ops, obj_api, "R")
    return ops


def _parse_apex_fields(cls_path: Path) -> dict[str, set]:
    """Apexクラスから SOQL + DML + トリガーハンドラーのフィールドを抽出する。
    Returns {obj_api: {field_api}}
    """
    try:
        content = cls_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return {}
    result: dict[str, set] = {}

    # SOQL: SELECT f1, f2 FROM ObjectName
    for m in _re.finditer(
        r'SELECT\s+(.*?)\s+FROM\s+([A-Za-z0-9_]+)', content, _re.IGNORECASE | _re.DOTALL
    ):
        fields_str, obj_api = m.group(1), m.group(2)
        fields = {
            f.strip() for f in _re.split(r'[\s,]+', fields_str)
            if f.strip() and f.strip().lower() != "id" and _re.match(r'^[A-Za-z]\w*$', f.strip())
        }
        if fields:
            result.setdefault(obj_api, set()).update(fields)

    # DML: variable.FieldName__c = ...（オブジェクト特定不可のため __any__ に保持）
    for fapi in _re.findall(r'\.\s*([A-Za-z][A-Za-z0-9_]*__c)\s*=', content):
        result.setdefault("__any__", set()).add(fapi)

    # トリガーハンドラー: メソッド引数・ローカル変数の型宣言からプライマリオブジェクトを特定
    # 例: Map<Id, ViewAblePerson__c> や List<ViewAblePerson__c> → ViewAblePerson__c がプライマリ
    trigger_obj = None
    for pat in [r'Map<Id,\s*([A-Za-z][A-Za-z0-9]*__c)>',
                r'List<([A-Za-z][A-Za-z0-9]*__c)>',
                r'([A-Za-z][A-Za-z0-9]*__c)\s+\w+\s*[=;,)]']:
        m = _re.search(pat, content)
        if m:
            trigger_obj = m.group(1)
            break
    if trigger_obj:
        # そのオブジェクト型変数へのプロパティアクセス (.Field__c) を収集
        prop_fields = set(_re.findall(
            r'\b\w+\.\s*([A-Za-z][A-Za-z0-9]*__[cr])\b', content
        ))
        # Id系・リレーション(__r)は除外
        prop_fields = {f for f in prop_fields if not f.endswith('__r')}
        if prop_fields:
            result.setdefault(trigger_obj, set()).update(prop_fields)

    # G-6: Site.* / System.setPassword 呼び出しがあれば User オブジェクトを操作とみなす
    if _re.search(
        r'\bSite\.(?:login|forgotPassword|validatePassword|createPortalUser|changePassword|passwordless)\b'
        r'|\bSystem\.setPassword\b', content):
        result.setdefault("User", set()).update({"Username", "Email", "IsActive"})
        # Contact もポータル連携で暗黙に参照される（login → Contact を辿る）
        if _re.search(r'\bSite\.(?:login|createPortalUser)\b', content):
            result.setdefault("Contact", set()).update({"Email", "Name"})

    return result


def _parse_vf_fields(vf_path: Path) -> dict[str, set]:
    """Visualforce ページから使用オブジェクト+フィールドを抽出する。
    Returns {obj_api: {field_api}}
    """
    try:
        content = vf_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return {}
    result: dict[str, set] = {}
    # standardController="Contact" → オブジェクト特定
    ctrl_m = _re.search(r'standardController="([A-Za-z][A-Za-z0-9_]*)"', content)
    if ctrl_m:
        obj_api = ctrl_m.group(1)
        for m in _re.finditer(
            r'(?:inputField|outputField)[^>]+value="\{!(?:[A-Za-z_]\w*\.)?([A-Za-z][A-Za-z0-9_]*(?:__c)?)\}"',
            content,
        ):
            field = m.group(1)
            if _re.match(r'^[A-Za-z]', field):
                result.setdefault(obj_api, set()).add(field)
    # カスタムオブジェクト直参照: {!MyObj__c.Field__c}
    for m in _re.finditer(
        r'\{!([A-Za-z][A-Za-z0-9_]*__c)\.([A-Za-z][A-Za-z0-9_]*(?:__c)?)\}', content
    ):
        obj_api2, field = m.group(1), m.group(2)
        result.setdefault(obj_api2, set()).add(field)
    return result


def _parse_vf_ops(content: str, fields_by_obj: dict[str, set]) -> dict[str, str]:
    """VF ページから {obj_api: op} を推定する。
    - <apex:inputField> を含む → W（ユーザー入力で更新される）
    - <apex:outputField> / <apex:inputText> などのみ → R
    - <apex:commandButton action="{!save}"> / "{!create}" 等が並ぶ → INSERT/W
    """
    ops: dict[str, str] = {}
    # standardController の対象オブジェクト
    ctrl_m = _re.search(r'standardController="([A-Za-z][A-Za-z0-9_]*)"', content)
    ctrl_obj = ctrl_m.group(1) if ctrl_m else None
    has_input = bool(_re.search(r'<apex:inputField\b', content, _re.IGNORECASE))
    has_save = bool(_re.search(r'action="\{!\s*(?:save|update|upsert|edit)\b', content, _re.IGNORECASE))
    has_create = bool(_re.search(r'action="\{!\s*(?:create|insert|register|signUp|selfReg)\b', content, _re.IGNORECASE))
    if ctrl_obj:
        if has_create:
            _merge_op(ops, ctrl_obj, "INSERT")
        elif has_input or has_save:
            _merge_op(ops, ctrl_obj, "W")
        else:
            _merge_op(ops, ctrl_obj, "R")
    # fields_by_obj 由来は R（outputField など）
    for obj_api in fields_by_obj.keys():
        if obj_api == "__any__":
            continue
        _merge_op(ops, obj_api, "R")
    return ops


def _load_sf_metadata(sf_project_path: str) -> None:
    """SFプロジェクトの objectTranslations/flows/classes からメタデータを構築してグローバルに格納する。"""
    global _SF_FIELD_LABELS, _SF_OBJ_LABELS, _SF_COMP_FIELDS, _SF_COMP_OPS, _CURRENT_SF_BASE_PATH
    _CURRENT_SF_BASE_PATH = sf_project_path
    base = Path(sf_project_path) / "force-app/main/default"

    # objectTranslations: フィールド・オブジェクトの日本語ラベル
    trans_dir = base / "objectTranslations"
    if trans_dir.exists():
        for obj_dir in trans_dir.iterdir():
            if not obj_dir.name.endswith("-ja"):
                continue
            obj_api = obj_dir.name[:-3]
            obj_trans = obj_dir / f"{obj_dir.name}.objectTranslation-meta.xml"
            if obj_trans.exists():
                content = obj_trans.read_text(encoding="utf-8")
                m = _re.search(r'<value>([^<]+)</value>', content)
                if m:
                    _SF_OBJ_LABELS[obj_api] = m.group(1).strip()
            for fxml in obj_dir.glob("*.fieldTranslation-meta.xml"):
                field_api = fxml.name.replace(".fieldTranslation-meta.xml", "")
                content = fxml.read_text(encoding="utf-8")
                m = _re.search(r'<label><!--\s*(.*?)\s*--></label>', content)
                if m and m.group(1):
                    _SF_FIELD_LABELS.setdefault(obj_api, {})[field_api] = m.group(1).strip()

    # flows: コンポーネント別のオブジェクト+フィールド
    flows_dir = base / "flows"
    if flows_dir.exists():
        for flow_file in flows_dir.glob("*.flow-meta.xml"):
            comp_api = flow_file.name.replace(".flow-meta.xml", "")
            obj_fields = _parse_flow_fields(flow_file)
            if obj_fields:
                _SF_COMP_FIELDS[comp_api] = {k: v for k, v in obj_fields.items()}

    # classes: コンポーネント別のオブジェクト+フィールド（テストクラスは除外）
    classes_dir = base / "classes"
    if classes_dir.exists():
        for cls_file in classes_dir.glob("*.cls"):
            if cls_file.name.endswith("Test.cls"):
                continue
            comp_api = cls_file.name.replace(".cls", "")
            obj_fields = _parse_apex_fields(cls_file)
            if obj_fields:
                _SF_COMP_FIELDS[comp_api] = {k: v for k, v in obj_fields.items()}
                # K-C: Apex DML / Site.* から op を推定
                try:
                    cls_content = cls_file.read_text(encoding="utf-8", errors="ignore")
                    ops = _parse_apex_ops(cls_content, obj_fields)
                    if ops:
                        _SF_COMP_OPS[comp_api] = ops
                except Exception:
                    pass

    # pages: Visualforce コンポーネント別フィールド
    # G-6: controller="ClassName" を辿って Apex クラス側のフィールド（Site.* 検出を含む）も引き込む
    pages_dir = base / "pages"
    if pages_dir.exists():
        for vf_file in pages_dir.glob("*.page"):
            comp_api = vf_file.name.replace(".page", "")
            obj_fields = _parse_vf_fields(vf_file)
            try:
                vf_content = vf_file.read_text(encoding="utf-8", errors="ignore")
            except Exception:
                vf_content = ""
            ctrl_name = None
            ctrl_m = _re.search(r'(?<!standard)[Cc]ontroller="([A-Za-z][A-Za-z0-9_]*)"', vf_content)
            if ctrl_m:
                ctrl_name = ctrl_m.group(1)
                ctrl_fields = _SF_COMP_FIELDS.get(ctrl_name, {})
                for k, v in ctrl_fields.items():
                    obj_fields.setdefault(k, set()).update(v)
            if obj_fields:
                _SF_COMP_FIELDS[comp_api] = {k: v for k, v in obj_fields.items()}
            # K-C: VF 自身の op + controller Apex の op をマージ
            vf_ops = _parse_vf_ops(vf_content, obj_fields) if vf_content else {}
            if ctrl_name:
                for k, v in _SF_COMP_OPS.get(ctrl_name, {}).items():
                    _merge_op(vf_ops, k, v)
            if vf_ops:
                _SF_COMP_OPS[comp_api] = vf_ops

    # objects: field-meta.xml からラベルを補完（objectTranslations にないカスタムフィールド対応）
    objects_dir = base / "objects"
    if objects_dir.exists():
        for obj_dir in objects_dir.iterdir():
            if not obj_dir.is_dir():
                continue
            obj_api = obj_dir.name
            fields_dir = obj_dir / "fields"
            if not fields_dir.exists():
                continue
            for fxml in fields_dir.glob("*.field-meta.xml"):
                field_api = fxml.name.replace(".field-meta.xml", "")
                if field_api in _SF_FIELD_LABELS.get(obj_api, {}):
                    continue  # 翻訳ファイルで既に取得済み
                try:
                    content = fxml.read_text(encoding="utf-8")
                except Exception:
                    continue
                m = _re.search(r'<label>([^<]+)</label>', content)
                if m:
                    _SF_FIELD_LABELS.setdefault(obj_api, {})[field_api] = m.group(1).strip()

    # 追加翻訳パス（GF_UAT等）からobjectTranslationsのみ補完ロード
    for extra_path in _SF_EXTRA_LABEL_PATHS:
        extra_trans = Path(extra_path) / "force-app/main/default/objectTranslations"
        if not extra_trans.exists():
            continue
        for obj_dir in extra_trans.iterdir():
            if not obj_dir.name.endswith("-ja"):
                continue
            obj_api = obj_dir.name[:-3]
            for fxml in obj_dir.glob("*.fieldTranslation-meta.xml"):
                field_api = fxml.name.replace(".fieldTranslation-meta.xml", "")
                if field_api in _SF_FIELD_LABELS.get(obj_api, {}):
                    continue  # 既に取得済み
                try:
                    content = fxml.read_text(encoding="utf-8")
                except Exception:
                    continue
                m = _re.search(r'<label><!--\s*(.*?)\s*--></label>', content)
                if m and m.group(1):
                    _SF_FIELD_LABELS.setdefault(obj_api, {})[field_api] = m.group(1).strip()

    # 標準オブジェクトの主要フィールド日本語ラベル（翻訳ファイルが存在しない場合の補完）
    _STD_FIELD_LABELS_BUILTIN: dict[str, dict[str, str]] = {
        "User": {
            "Username": "ユーザー名", "Email": "メールアドレス", "IsActive": "有効",
            "FirstName": "名", "LastName": "姓", "Name": "フルネーム",
            "ContactId": "取引先責任者", "AccountId": "取引先", "ProfileId": "プロファイル",
            "UserRoleId": "ロール", "Title": "役職", "Department": "部門",
            "Phone": "電話", "MobilePhone": "携帯", "LanguageLocaleKey": "言語",
            "LocaleSidKey": "ロケール", "TimeZoneSidKey": "タイムゾーン",
        },
        "Account": {
            "Name": "取引先名", "Phone": "電話", "Website": "Webサイト",
            "BillingAddress": "請求先住所", "ShippingAddress": "配送先住所",
            "Industry": "業種", "AnnualRevenue": "年間売上", "NumberOfEmployees": "従業員数",
            "OwnerId": "所有者", "Type": "取引先種別", "ParentId": "親取引先",
        },
        "Contact": {
            "FirstName": "名", "LastName": "姓", "Name": "氏名",
            "Email": "メールアドレス", "Phone": "電話", "MobilePhone": "携帯",
            "AccountId": "取引先", "Title": "役職", "Department": "部門",
            "OwnerId": "所有者", "Birthdate": "生年月日",
        },
        "Lead": {
            "FirstName": "名", "LastName": "姓", "Name": "氏名",
            "Company": "会社名", "Email": "メールアドレス", "Phone": "電話",
            "MobilePhone": "携帯", "Title": "役職", "Industry": "業種",
            "LeadSource": "リードソース", "Status": "ステータス",
            "Street": "住所（番地）", "City": "市区町村", "State": "都道府県",
            "PostalCode": "郵便番号", "Country": "国", "OwnerId": "所有者",
            "HasOptedOutOfEmail": "メール配信停止", "IsConverted": "取引先に変換済み",
        },
        "Opportunity": {
            "Name": "商談名", "AccountId": "取引先", "CloseDate": "完了予定日",
            "StageName": "フェーズ", "Amount": "金額", "Probability": "確度",
            "OwnerId": "所有者", "LeadSource": "リードソース", "Type": "種別",
        },
    }
    for obj_api, fld_map in _STD_FIELD_LABELS_BUILTIN.items():
        for fapi, flabel in fld_map.items():
            _SF_FIELD_LABELS.setdefault(obj_api, {}).setdefault(fapi, flabel)


def _sf_field_label(obj_api: str, field_api: str) -> str:
    """フィールドAPI名を日本語ラベルに変換する（メタデータ優先、なければ標準予備辞書→加工済みAPI名）。"""
    label = _SF_FIELD_LABELS.get(obj_api, {}).get(field_api)
    if label:
        return label
    # G-5: 標準オブジェクトの主要項目は予備辞書から日本語化
    for f in _STD_OBJ_FIELDS_FALLBACK.get(obj_api, []):
        if f["api_name"] == field_api:
            return f["label"]
    return field_api.replace("__c", "").replace("__", "_")


def _sf_obj_label(obj_api: str) -> str:
    """オブジェクトAPI名を日本語ラベルに変換する。"""
    if obj_api in _STD_OBJ_LABELS:
        return _STD_OBJ_LABELS[obj_api]
    return _SF_OBJ_LABELS.get(obj_api) or _obj_label_from_api(obj_api)


# ── Salesforce 標準オブジェクト API名 → 日本語ラベルマップ ───────────────
_STD_OBJ_LABELS = {
    "ContentDocumentLink": "コンテンツ紐付けレコード",
    "ContentDocument":     "コンテンツドキュメント",
    "ContentVersion":      "コンテンツファイル",
    "EmailMessage":        "メールメッセージ",
    "Attachment":          "添付ファイル",
    "Opportunity":         "商談",
    "Contact":             "取引先責任者",
    "Account":             "取引先",
    "Lead":                "リード",
    "Case":                "ケース",
    "Task":                "ToDo",
    "Event":               "行動",
    "User":                "ユーザー",
    "Quote":               "見積",
}

# 標準オブジェクトの主要項目フォールバック（_SF_FIELD_LABELS に翻訳が無い場合の予備）
_STD_OBJ_FIELDS_FALLBACK: dict[str, list[dict[str, str]]] = {
    "User": [
        {"api_name": "Username",  "label": "ユーザー名",     "note": "ログイン ID"},
        {"api_name": "Email",     "label": "メールアドレス", "note": ""},
        {"api_name": "Name",      "label": "氏名",           "note": ""},
        {"api_name": "IsActive",  "label": "有効",           "note": ""},
        {"api_name": "ContactId", "label": "取引先責任者",   "note": "ポータル連携"},
    ],
    "Contact": [
        {"api_name": "Name",      "label": "氏名",           "note": ""},
        {"api_name": "Email",     "label": "メールアドレス", "note": ""},
        {"api_name": "Phone",     "label": "電話番号",       "note": ""},
        {"api_name": "AccountId", "label": "取引先",         "note": ""},
    ],
    "Account": [
        {"api_name": "Name",  "label": "取引先名", "note": ""},
        {"api_name": "Phone", "label": "電話番号", "note": ""},
    ],
    "Lead": [
        {"api_name": "Name",    "label": "氏名",           "note": ""},
        {"api_name": "Email",   "label": "メールアドレス", "note": ""},
        {"api_name": "Company", "label": "会社名",         "note": ""},
    ],
    "Case": [
        {"api_name": "CaseNumber", "label": "ケース番号",   "note": ""},
        {"api_name": "Subject",    "label": "件名",         "note": ""},
        {"api_name": "Status",     "label": "ステータス",   "note": ""},
    ],
}

# 標準オブジェクトの canonical 名（lowercase variant → PascalCase API 名）
# VF の {!user.firstname} のような小文字参照を正規 API 名に寄せるための辞書
_STD_OBJ_CANONICAL_NAMES: dict[str, dict[str, str]] = {
    "User": {
        "firstname": "FirstName", "lastname": "LastName", "email": "Email",
        "username": "Username", "name": "Name", "isactive": "IsActive",
        "contactid": "ContactId", "accountid": "AccountId", "profileid": "ProfileId",
        "userroleid": "UserRoleId", "phone": "Phone", "mobilephone": "MobilePhone",
        "title": "Title", "department": "Department", "usertype": "UserType",
        "languagelocalekey": "LanguageLocaleKey", "localesidkey": "LocaleSidKey",
        "timezonesidkey": "TimeZoneSidKey", "emailencodingkey": "EmailEncodingKey",
        "city": "City", "country": "Country", "state": "State", "street": "Street",
        "postalcode": "PostalCode", "fax": "Fax", "extension": "Extension",
        "communitynickname": "CommunityNickname", "alias": "Alias",
        "lastlogindate": "LastLoginDate", "id": "Id",
    },
    "Contact": {
        "firstname": "FirstName", "lastname": "LastName", "email": "Email",
        "name": "Name", "phone": "Phone", "mobilephone": "MobilePhone",
        "accountid": "AccountId", "title": "Title", "department": "Department",
        "id": "Id",
    },
    "Account": {
        "name": "Name", "phone": "Phone", "id": "Id",
    },
}

# 技術英語ジャーゴン → 日本語変換（SF オブジェクト名以外の技術用語）
# NOTE: re.ASCII により \b が ASCII 文字のみを word char として扱う（日本語隣接でも正しくマッチ）
# NOTE: "before insert" 等の複合パターンは単体 "insert" より前に置く（置換順依存を防ぐ）
_A = _re.ASCII  # shorthand
_AI = _re.ASCII | _re.IGNORECASE
_JARGON_JA: list[tuple] = [
    # Apex トリガー複合表現（単体 insert/delete/update より先に処理）
    (_re.compile(r'\bbefore\s+insert\b', _AI), '処理前（新規）'),
    (_re.compile(r'\bafter\s+insert\b',  _AI), '処理後（新規）'),
    (_re.compile(r'\bbefore\s+update\b', _AI), '処理前（更新）'),
    (_re.compile(r'\bafter\s+update\b',  _AI), '処理後（更新）'),
    (_re.compile(r'\bbefore\s+delete\b', _AI), '処理前（削除）'),
    (_re.compile(r'\bafter\s+delete\b',  _AI), '処理後（削除）'),
    # SF API / 外部サービス
    (_re.compile(r'\bOPROARTS\s+API\b', _A), '外部帳票サービス'),
    (_re.compile(r'\bOPROARTS\b', _A), '外部帳票サービス'),
    # コレクション型
    (_re.compile(r'\bList<Id>\b', _A), 'IDリスト'),
    (_re.compile(r'\bList<[A-Za-z]+>\b', _A), 'リスト'),
    # データ型
    (_re.compile(r'\bBlob\b', _A), 'バイナリデータ'),
    (_re.compile(r'\bvoid\b', _A), 'なし'),
    # DML 操作（単体）
    (_re.compile(r'\binsert\b', _AI), '新規作成'),
    (_re.compile(r'\bupsert\b', _AI), '登録・更新'),
    (_re.compile(r'\bdelete\b', _AI), '削除'),
    (_re.compile(r'\bupdate\b', _AI), '更新'),
    (_re.compile(r'\bquery\b',  _AI), '参照'),
    # Apex 固有クラス名・アノテーション
    (_re.compile(r'\bCustomerUser\b', _A), 'カスタマーユーザー'),
    (_re.compile(r'\b[A-Z][A-Za-z]*Tmp\b', _A), '一時データ'),
    (_re.compile(r'\bInvocableMethod\b', _A), 'フローアクション'),
    (_re.compile(r'\bAuraEnabled\b', _A), 'LWC公開メソッド'),
    # Apex 非同期アノテーション（後続の「で非同期実行する」も一緒に置換して重複を防ぐ）
    (_re.compile(r'@future(?:で非同期実行する|で実行される?)?'), '非同期で実行する'),
    (_re.compile(r'@\w+', _A), ''),
]

# 技術用語→日本語変換ルール（役割・説明文用）
_TECH_REPL = [
    # アノテーション → 日本語フレーズ
    # NOTE: _translate_jargon で InvocableMethod → フローアクション に変換済みのため
    # 変換後パターン "@フローアクション〜" も除去対象に含める
    (_re.compile(r'@InvocableMethod[としてで\s]*'), 'フローから呼び出され、'),
    (_re.compile(r'@フローアクション(?:として)?(?:Flowから呼ばれ)?[、\s]*'), 'フローから呼び出され、'),
    (_re.compile(r'@AuraEnabled[としてで\s]*'), 'LWCから呼び出され、'),
    (_re.compile(r'@LWC公開メソッド(?:として)?[、\s]*'), 'LWCから呼び出され、'),
    (_re.compile(r'@RemoteAction[としてで\s]*'), '非同期処理として呼び出され、'),
    (_re.compile(r'@\w+[としてで\s]*'), ''),  # 残った@アノテーションを除去
    # SOQL文を丸ごと除去（SELECT〜FROM〜を含む記述）
    (_re.compile(r'SELECT\s+.+?\s+FROM\s+\w+(?:\s+WHERE\s+[^。\n]+)?', _re.DOTALL | _re.IGNORECASE), ''),
    (_re.compile(r':\w+'), ''),  # SOQL bind変数
    # HTTPプロトコル技術記述（POST→302→GET等）を除去
    (_re.compile(r'[A-Z]{2,6}→\d{3}→[A-Z]{2,6}'), ''),
    # boolean変数が括弧で補足されているケース: （isXxx）（hasXxx）
    (_re.compile(r'[（(](?:is|has)[A-Z]\w*[）)]'), ''),
    # boolean条件文: "isXxxがtrueの場合は" → 変数名を除去し条件の結果のみ残す
    (_re.compile(r'(?:is|has)[A-Z]\w*が(?:false|true|null)(?:の場合[はに]?)?'), ''),
    # 単独のboolean変数（isXxx / hasXxx）を除去
    (_re.compile(r'(?<![A-Za-z])(?:is|has)[A-Z][A-Za-z]+(?![A-Za-z])'), ''),
    # 小文字始まりのcamelCase技術名（コンポーネント名等）が日本語助詞の前にある場合
    (_re.compile(r'(?<![A-Za-z])[a-z][a-zA-Z]{3,}(?=[をがはにでへのもと])'), ''),
    # 大文字始まりの長いCamelCaseクラス名が日本語に隣接する場合（VisaApplicationTypeMaster等）
    (_re.compile(r'[A-Z][a-zA-Z0-9]{6,}(?=[ぁ-ん一-龥ァ-ヶーをがはにでへのもと]|により|によって)'), ''),
    # クラス名.メソッド名 を除去
    (_re.compile(r'[A-Z][A-Za-z0-9]+\.[A-Za-z]\w+\([^)]*\)'), ''),
    (_re.compile(r'[A-Z][A-Za-z0-9]+\.[A-Za-z]\w+'), ''),
    # Apex トリガーイベント文脈（処理前/後）を除去
    (_re.compile(r'[^\s。]{2,20}(?:作成|更新|削除)時（処理前（(?:新規|更新|削除)）(?:/処理後（(?:新規|更新|削除)）)?）[にので]?'), ''),
    # 技術的なサーバーサイド表現を簡潔な日本語に
    (_re.compile(r'のサーバーサイドロジック'), ''),
    (_re.compile(r'のメインコンポーネント'), ''),
    (_re.compile(r'単一責務クラス'), 'クラス'),
    # "呼び出され、Flow〜から呼ばれ、" の重複を除去（ルックビハインド不可のため捕捉グループで実装）
    (_re.compile(r'(呼び出され)[、\s]{0,2}Flow[^\s、。]{0,20}から呼ばれ[、\s]*'), r'\1、'),
    # 同一表現の繰り返し（サービス名など）が括弧内外で重複: "外部帳票サービス（外部帳票サービス）"
    (_re.compile(r'([^\s（]{4,30})（\1）'), r'\1'),
    # メソッドが重複: "公開メソッドメソッド" → "公開メソッド"
    (_re.compile(r'(メソッド){2,}'), 'メソッド'),
    # 空括弧を除去: （）
    (_re.compile(r'（\s*）'), ''),
    (_re.compile(r'\(\s*\)'), ''),
    # 主語なし「・による〜」「/による〜」を除去（CamelCase除去後に残る）
    (_re.compile(r'[・、/]\s*による[^\s、。]{1,20}'), ''),
    # "/作成" を "・作成" に正規化し、重複した "・作成・作成" をまとめる
    (_re.compile(r'/作成'), '・作成'),
    (_re.compile(r'(?:・作成){2,}'), '・作成'),
    # 助詞・接続詞が文節先頭に孤立するケース（クラス名除去後）を修正
    (_re.compile(r'[・、]\s*を[ぁ-ん一-龥A-Za-z]{1,10}て(?=の)'), ''),  # を〜ての → 除去
    (_re.compile(r'[はがをにでへのも][。．]'), '。'),
    (_re.compile(r'^[、。\s]+'), ''),  # 文頭の余分な記号
    # 連続記号・空白の整理
    (_re.compile(r'[ \t]{2,}'), ' '),
    (_re.compile(r'[・、]{2,}'), '・'),
    (_re.compile(r'(、){2,}'), '、'),
    (_re.compile(r'(。){2,}'), '。'),
]

# 業務フロー・タイトル・概要用: SF標準オブジェクトは翻訳済み前提でAPIを除去
_TECH_REPL_BIZ = [
    (_re.compile(r'@InvocableMethod[としてで\s]*'), ''),
    (_re.compile(r'@AuraEnabled[としてで\s]*'), ''),
    (_re.compile(r'@RemoteAction[としてで\s]*'), ''),
    (_re.compile(r'@\w+'), ''),
    # SOQL・boolean変数・HTTPプロトコル除去（タイトル用）
    (_re.compile(r'SELECT\s+.+?\s+FROM\s+\w+(?:\s+WHERE\s+[^。\n]+)?', _re.DOTALL | _re.IGNORECASE), ''),
    (_re.compile(r':\w+'), ''),
    (_re.compile(r'[A-Z]{2,6}→\d{3}→[A-Z]{2,6}'), ''),
    (_re.compile(r'[（(](?:is|has)[A-Z]\w*[）)]'), ''),
    (_re.compile(r'(?:is|has)[A-Z]\w*が(?:false|true|null)(?:の場合[はに]?)?'), ''),
    (_re.compile(r'(?<![A-Za-z])(?:is|has)[A-Z][A-Za-z]+(?![A-Za-z])'), ''),
    (_re.compile(r'(?<![A-Za-z])[a-z][a-zA-Z]{3,}(?=[をがはにでへのもと])'), ''),
    (_re.compile(r'[A-Z][A-Za-z0-9]+\.[A-Za-z]\w+\([^)]*\)'), ''),
    (_re.compile(r'[A-Z][A-Za-z0-9]+\.[A-Za-z]\w+'), ''),
    (_re.compile(r'のサーバーサイドロジック'), ''),
    (_re.compile(r'のメインコンポーネント'), ''),
    (_re.compile(r'単一責務クラス'), 'クラス'),
    # List<CustomObj__c> → __c削除の前に除去（先に処理しないと List<> が残る）
    (_re.compile(r'List<[A-Z][A-Za-z0-9]*__[cepr]>'), 'レコードリスト'),
    (_re.compile(r'List<[A-Za-z]+>'), 'リスト'),
    # （trigger xxx） 等のApexトリガー技術的記述を除去
    (_re.compile(r'（trigger\s+\w+）'), ''),
    (_re.compile(r'\(trigger\s+\w+\)'), ''),
    # __c/__e/__r カスタムオブジェクト名（上記 List<> 除去後）
    (_re.compile(r'\b[A-Z][A-Za-z0-9]*__[cepr]\b'), ''),
    # Apex クラス名（Controller/Service/Handler/Manager/Batch/Trigger で終わるもの）
    (_re.compile(r'\b[A-Z][A-Za-z0-9]{2,}(?:Controller|Service|Handler|Manager|Batch|Trigger)\b'), ''),
    # 残ったCamelCase英語（日本語助詞に挟まれていない単独の英単語）
    (_re.compile(r'(?<![ぁ-ん一-龥ァ-ヶーa-z_])([A-Z][a-zA-Z]{3,})(?![ぁ-ん一-龥ァ-ヶーa-z_])'), ''),
    # 括弧内が英語・記号で始まる（技術情報）は除去
    (_re.compile(r'（[A-Z@#][^）]{0,60}）'), ''),
    (_re.compile(r'\([A-Z@#][^)]{0,60}\)'), ''),
    # 同一表現の繰り返し（括弧内外重複）: "外部帳票サービス（外部帳票サービス）"
    (_re.compile(r'([^\s（]{4,30})（\1）'), r'\1'),
    # メソッドが重複: "公開メソッドメソッド" → "公開メソッド"
    (_re.compile(r'(メソッド){2,}'), 'メソッド'),
    # 空括弧を除去
    (_re.compile(r'（\s*）'), ''),
    (_re.compile(r'\(\s*\)'), ''),
    # 主語なし「・による〜」を除去（CamelCase除去後に残る場合）
    (_re.compile(r'[・、/]\s*による[^\s、。]{1,20}'), ''),
    # "/作成" 正規化・重複まとめ
    (_re.compile(r'/作成'), '・作成'),
    (_re.compile(r'(?:・作成){2,}'), '・作成'),
    # Salesforce フィールド API名（__c/__r 等）を除去（\bはUnicode文字前で効かないため (?![A-Za-z0-9_]) を使用）
    (_re.compile(r'[A-Za-z][A-Za-z0-9_]*(?:__c|__r|__C|__R)(?![A-Za-z0-9_])'), ''),
    # フィールド名除去後の「の等」「の、」 → 整理
    (_re.compile(r'の(?=[等や・、。\s])'), ''),
    # 整理
    (_re.compile(r'[ \t]{2,}'), ' '),
    (_re.compile(r'[・、]{2,}'), '・'),
    (_re.compile(r'(、){2,}'), '、'),
    (_re.compile(r'^[、。・/\s]+|[、。・/\s]+$'), ''),
]

# 業務フロー step の先頭 preamble（呼び出し起点の技術的説明）を除去
_PREAMBLE_RE = _re.compile(
    r'^(?:[^、。]{0,25}として(?:呼ばれ|呼び出され)、|'
    r'Flowから呼ばれ、|'
    r'[^、。]{0,15}[がは]Flowから[^、。]{0,15}、)'
)


def _translate_sf_objects(text: str) -> str:
    """Salesforce 標準オブジェクト + ロード済みカスタムオブジェクト API名を日本語ラベルに置換する。

    NOTE: Python3 の re は Unicode モードで日本語文字も \\w 扱いするため、
    日本語に隣接する英語 API 名が word boundary にマッチしない。
    ASCII 専用の lookahead/lookbehind に変更。
    """
    for api, ja in _STD_OBJ_LABELS.items():
        text = _re.sub(rf'(?<![A-Za-z0-9_]){_re.escape(api)}(?![A-Za-z0-9_])', ja, text)
    # カスタムオブジェクト（メタデータ読み込み済みの場合のみ）
    for api, ja in _SF_OBJ_LABELS.items():
        text = _re.sub(rf'(?<![A-Za-z0-9_]){_re.escape(api)}(?![A-Za-z0-9_])', ja, text)
    return text


def _translate_jargon(text: str) -> str:
    """技術英語ジャーゴンを日本語に変換する。"""
    for pat, repl in _JARGON_JA:
        text = pat.sub(repl, text)
    return text


def _clean_tech(text: str) -> str:
    """役割・説明文用: アノテーション・クラス名.メソッド名を除去して日本語説明にする。"""
    for pattern, repl in _TECH_REPL:
        text = pattern.sub(repl, text)
    return text.strip()


_EC_PLACEHOLDER = "\x01EC\x01"

def _clean_tech_business(text: str) -> str:
    """業務フロー・タイトル・概要用: SF標準オブジェクトを日本語化→技術用語を全除去する。

    "Experience Cloud" はブランド名のため、CamelCase 除去パターンに巻き込まれないよう
    プレースホルダーで保護してから処理し、最後に復元する。
    """
    # Protect "Experience Cloud" before CamelCase removal
    text = text.replace("Experience Cloud", _EC_PLACEHOLDER)
    text = _translate_sf_objects(text)
    text = _translate_jargon(text)
    for pattern, repl in _TECH_REPL_BIZ:
        text = pattern.sub(repl, text)
    text = text.replace(_EC_PLACEHOLDER, "Experience Cloud")
    return text.strip()


def _translate_sf_fields(text: str) -> str:
    """ロード済みメタデータの全フィールドAPI名（__c等）を日本語ラベルに置換する。
    メタデータ未ロードの場合はそのまま返す（__c除去は _TECH_REPL_BIZ で後処理）。
    """
    for _obj_api, fields in _SF_FIELD_LABELS.items():
        for field_api, ja_label in fields.items():
            text = _re.sub(
                rf'(?<![A-Za-z0-9_]){_re.escape(field_api)}(?![A-Za-z0-9_])',
                ja_label, text,
            )
    return text


def _clean_io_text(text: str) -> str:
    """inputs/outputs テキストの技術用語を日本語化する（処理概要の説明文用）。"""
    text = text.replace("Experience Cloud", _EC_PLACEHOLDER)
    text = _translate_sf_objects(text)
    text = _translate_sf_fields(text)
    text = _translate_jargon(text)
    for pattern, repl in _TECH_REPL_BIZ:
        text = pattern.sub(repl, text)
    text = text.replace(_EC_PLACEHOLDER, "Experience Cloud")
    return text.strip()


def _short_title(responsibility: str, max_len: int = 18) -> str:
    """責務テキストから短い日本語アクションタイトルを生成する（フローチャートノード用）。

    方針:
      - 「〜を起点に、」「〜で起動される」「〜を受けて、」等のプリアンブルを除去
      - 語の途中で切らない（粒度: 日本語連続ラン / 助詞境界）
      - パターン1: `〜する画面フロー` → 「画面フローを〜する」
      - パターン2: `〜を[キーワード]` → 「〜を[キーワード]する」（最後の出現を優先）
      - パターン3: `[キーワード]` 単独 → 直前の日本語名詞句を補完
      - 18 文字以内に収める
    """
    clean = _clean_tech_business(responsibility)
    # 「Flowのアクションとして〜から」「〜クラスとして〜から」等のプリアンブルを除去
    clean = _re.sub(r'^[^。・\n]*?(?:アクション|クラス|ハンドラ|ハンドラー)として[^。・\n]*?から[、]?', '', clean).strip()
    clean = _PREAMBLE_RE.sub('', clean).strip()
    # 追加プリアンブル: トリガー記述に限定（読点付きのみ／Phase 0で拾える"起動する画面フロー"は残す）
    extra_preambles = [
        r'^[^。、]{0,40}?を起点に[、,]\s*',
        r'^[^。、]{0,40}?(?:で|から)起動される[、,]\s*',
        r'^[^。、]{0,40}?により[、,]\s*',
        r'^[^。、]{0,40}?を契機に[、,]\s*',
    ]
    for pat in extra_preambles:
        clean = _re.sub(pat, '', clean)
    # 末尾の「〜の責務を持つ」「〜責務を担う」を除去
    clean = _re.sub(r'(?:の)?責務を(?:持つ|担う|持ち)?。?$', '', clean).strip()

    _KW_PAIRS = [
        ("判定",   "を判定する"),
        ("検証",   "を検証する"),
        ("確認",   "を確認する"),
        ("チェック", "をチェックする"),
        ("作成",   "を作成する"),
        ("登録",   "を登録する"),
        ("更新",   "を更新する"),
        ("削除",   "を削除する"),
        ("同期",   "を同期する"),
        ("集約",   "を集約する"),
        ("送信",   "を送信する"),
        ("通知",   "を通知する"),
        ("制御",   "を制御する"),
        ("管理",   "を管理する"),
        ("実行",   "を実行する"),
        ("検知",   "を検知する"),
        ("委譲",   "を委譲する"),
        ("表示",   "を表示する"),
    ]

    # --- ヘルパー ---
    _BOUNDARIES = set("、。・「」『』（）() \t\n→")

    def _script_class(ch: str) -> str:
        if _re.match(r'[ぁ-んァ-ヶ一-龯々ー]', ch):
            return 'jp'
        if _re.match(r'[a-zA-Z0-9_]', ch):
            return 'en'
        return 'other'

    def _strip_leading_connectives(text: str) -> str:
        """先頭の接続語・助詞・動詞テ形・数字カウンタ等を除去する。"""
        prev = None
        # ループで除去 pass を繰り返す（複合プレフィックスに対応）
        while text != prev:
            prev = text
            text = _re.sub(r'^ー+', '', text)  # orphan 長音
            text = _re.sub(r'^[0-9０-９]+(?:[つ個件本回件枚人]の?)?', '', text)
            text = _re.sub(r'^(?:して|ために|ため|そして|また|により|ので|かつ|または|および)', '', text)
            # 動詞テ形（連用形＋て）: 返して/受けて/分岐して/連動して etc.
            text = _re.sub(r'^[ぁ-ん一-龯]{1,5}?て(?=[ぁ-んァ-ヶ一-龯])', '', text)
            text = _re.sub(r'^(?:て|に|で|と|の|は|が|を|も|や|ず)+', '', text)
            text = text.strip()
        return text

    def _trim_trailing_noise(text: str) -> str:
        """末尾の非日本語ノイズ（"/", スペース等）を除去する。"""
        return _re.sub(r'[^ぁ-んァ-ヶ一-龯々ー]+$', '', text)

    def _refine_prefix(text: str, target_len: int = 10) -> str:
        """プリフィクスを粒度よく整形する。助詞境界で分割し、最右の意味のある名詞句を返す。"""
        text = _strip_leading_connectives(text)
        text = _trim_trailing_noise(text)
        # 助詞境界で分割し、最右のセグメントを優先
        segs = _re.split(r'(?:を|は|が|に|で|と|から|まで|より)', text)
        segs = [_trim_trailing_noise(_strip_leading_connectives(s)) for s in segs if s.strip()]
        for seg in reversed(segs):
            if 2 <= len(seg) <= target_len + 4 and _re.search(r'[一-龯ぁ-んァ-ヶ]', seg):
                return seg
        # 末尾の漢字クラスター
        kanji_clusters = _re.findall(r'[一-龯]{2,}', text)
        if kanji_clusters:
            return kanji_clusters[-1]
        # カタカナクラスター
        kana_clusters = _re.findall(r'[ァ-ヶー]{2,}', text)
        if kana_clusters:
            return kana_clusters[-1]
        return text[:target_len]

    def _extract_jp_noun_backward(text_before: str) -> str:
        """キーワード直前から遡り、スクリプト遷移／区切り文字で停止して名詞句を抽出する。"""
        pos = len(text_before)
        prev_class = None
        stop_pos = 0
        while pos > 0:
            ch = text_before[pos - 1]
            if ch in _BOUNDARIES:
                stop_pos = pos
                break
            cls = _script_class(ch)
            if prev_class == 'jp' and cls != 'jp':
                stop_pos = pos
                break
            prev_class = cls
            pos -= 1
        text = text_before[stop_pos:].strip()
        return _refine_prefix(text)

    def _build(prefix: str, suffix: str, kw: str, clean: str = "") -> str:
        # 日本語を含まないプリフィクスは棄却（"/" のようなノイズ混入を防ぐ）
        if prefix and not _re.search(r'[ぁ-んァ-ヶ一-龯]', prefix):
            prefix = ""
        # prefix が空または貧弱なら、より広い範囲から拾う
        if (not prefix or len(prefix) < 2) and clean:
            m = _re.search(r'([ぁ-んァ-ヶ一-龯々ー]{2,10})を[^。\n]{0,40}?' + kw, clean)
            if m:
                candidate = _strip_leading_connectives(m.group(1))
                if candidate and len(candidate) >= 2:
                    prefix = candidate
        if prefix and prefix.endswith('を') and suffix.startswith('を'):
            prefix = prefix[:-1]
        if prefix:
            t = prefix + suffix
        else:
            t = kw + 'する'
        return _re.sub(r'(を|の){2,}', r'\1', t)

    # --- Phase 0: `(起動|実行|表示|遷移)する{名詞}` 複合パターン ---
    m = _re.search(
        r'(起動|実行|表示|遷移)する([ぁ-んァ-ヶ一-龯々ー]{2,10})(?=[。、・「」（）\s]|$)',
        clean
    )
    if m:
        verb = m.group(1)
        noun = _strip_leading_connectives(m.group(2))
        if noun:
            return (noun + 'を' + verb + 'する')[:max_len]

    # --- Phase 1: 厳格な `を[キーワード]` パターン（最後の出現を優先） ---
    strict_results: list[str] = []
    for kw, suffix in _KW_PAIRS:
        marker = 'を' + kw
        idx = clean.rfind(marker)
        if idx == -1:
            continue
        prefix = _extract_jp_noun_backward(clean[:idx])
        title = _build(prefix, suffix, kw, clean)
        strict_results.append(title)

    if strict_results:
        valid = [t for t in strict_results if _re.match(r'^[ぁ-んァ-ヶ一-龯]', t)]
        prefixed = [t for t in valid if len(t) >= 6]  # 名詞プリフィクス付きを優先
        pool = prefixed or valid or strict_results
        return min(pool, key=len)[:max_len]

    # --- Phase 2: 緩い `[キーワード]` 単独パターン ---
    lax_results: list[str] = []
    for kw, suffix in _KW_PAIRS:
        idx = clean.rfind(kw)
        if idx == -1:
            continue
        # 列挙文脈（"/", "・", "、"）直後のキーワードはスキップ（意味的に動詞でない可能性が高い）
        if idx > 0 and clean[idx - 1] in '/／・、,':
            continue
        prefix = _extract_jp_noun_backward(clean[:idx])
        title = _build(prefix, suffix, kw, clean)
        lax_results.append(title)

    if lax_results:
        valid = [t for t in lax_results if _re.match(r'^[ぁ-んァ-ヶ一-龯]', t)]
        prefixed = [t for t in valid if len(t) >= 6]
        pool = prefixed or valid or lax_results
        return min(pool, key=len)[:max_len]

    # --- Phase 3: フォールバック ---
    # 「〜を担当する」形式 → 先頭項目のみ
    m_tantou = _re.match(r'^(.+?)を担当する', clean)
    if m_tantou:
        items = [x.strip() for x in m_tantou.group(1).split('・') if x.strip()]
        if items:
            return (items[0] + 'を行う')[:max_len]

    # 先頭文の日本語名詞句を抽出
    first_sent = _re.split(r'[。\n]', clean)[0].strip()
    jp_tokens = [t for t in _re.findall(r'[ぁ-んァ-ヶ一-龯々ー]+', first_sent) if len(t) >= 2]
    if jp_tokens:
        base = jp_tokens[0]
        if len(base) > max_len - 3:
            base = base[:max_len - 3]
        return (base + 'を行う')[:max_len]

    # 最後の保険: 18 文字で丸める
    return clean[:max_len]


def _extract_actor(token: str) -> str:
    """フロートークンから日本語アクター名を推定する。"""
    t = _re.sub(r'（[^）]*）|\([^)]*\)', '', token).strip()
    if _re.search(r'お客様|顧客|申請者|依頼者', t):
        return "お客様"
    if _re.search(r'管理者|事務|担当者|スタッフ|GF社', t):
        return "GF社担当者"
    if _re.search(r'Flow|フロー|承認', t):
        return "自動フロー"
    if _re.search(r'画面|フォーム|入力|ページ', t):
        return "お客様"
    if _re.search(r'[A-Z][a-zA-Z]', t):
        return "システム"
    return t[:20] if t else "システム"


def _infer_trigger_screen(data: dict) -> str:
    """起点画面を推定する: screens[] → LWC → テキストキーワード の順に判定。"""
    # 1. screens[] の screen_name を使用（最優先）
    for s in data.get("screens", []):
        name = s.get("screen_name", "") or s.get("component", "")
        if name:
            return name

    # 2. LWC コンポーネントが存在する場合
    lwcs = [c.get("api_name", "") for c in data.get("components", []) if c.get("type") == "LWC"]
    if lwcs:
        return " / ".join(lwcs) + "（Lightningコンポーネント画面）"

    # 3. テキストキーワードから推定
    combined = " ".join([data.get("processing_purpose", ""), data.get("data_flow_overview", "")])
    first_token = data.get("data_flow_overview", "").split("→")[0].strip()

    if _re.search(r'Visualforce|VFページ|\bVF\b', combined):
        return "Visualforceページ（フォーム画面）"
    if _re.search(r'Experience Cloud|Experienceポータル|ポータル画面', combined):
        return "Experience Cloudポータル画面"
    if _re.search(r'Flow|フロー', first_token) and _re.search(r'管理者', first_token):
        return "Salesforce管理画面（またはFlowアクション）"
    if _re.search(r'管理者', first_token):
        return "Salesforce管理画面"
    if _re.search(r'Flow|フロー', first_token):
        return "Salesforce Flow（ボタンアクション）"
    if _re.search(r'お客様|顧客', first_token):
        return "Experience Cloudポータル画面"

    return "Salesforce管理画面"


def _gentle_clean_role(text: str) -> str:
    """role 欄専用の軽量クリーニング: SF オブジェクト/フィールド/ジャーゴンを日本語化するが
    英語 API 識別子の除去はしない（_strip_tech_identifiers を通さない）。
    旧キャッシュ由来の典型的な断片アーティファクトのみピンポイント修復する。
    """
    if not text:
        return text
    text = _translate_sf_fields(text)
    text = _translate_sf_objects(text)
    text = _translate_jargon(text)
    # 旧 _deep_clean_ja のキャッシュ残骸 — ピンポイント修復のみ
    text = _re.sub(r'(?i)\s*(?:を|が|は|に|で|と)\s+(?:apex|vf|lwc|aura)\s*$', '', text)  # 「を apex」
    text = _re.sub(r'ため動作不全[^。]*$', '', text)                                        # 「ため動作不全」
    text = _re.sub(r'\s*=\s*で(?=[ぁ-んァ-ヶ一-龯]|$)', '', text)                             # 「=で〜」
    text = _re.sub(r'[a-z]+(?:/[a-z]*)+\.?\s*', '', text)                                  # 「answers/.」「header/body/」等
    text = _re.sub(r'(?:から|より)\s+(?:apex|vf)\b', '', text, flags=_re.IGNORECASE)        # 「から apex」
    text = _re.sub(r'(?<=[。、])\s*[使呼]\w{0,3}(?:\s+\w+)*\s*$', '', text)               # 末尾の「使用」孤立動詞
    # 末尾の単独助詞を除去
    text = _re.sub(r'(?<=[ぁ-んァ-ヶ一-龯])\s+(?:を|が|は|に)\s*$', '', text)
    text = _re.sub(r'\s{2,}', ' ', text)
    return text.strip()


def _is_desc_fragment(text: str) -> bool:
    """テキストが断片（主語欠落・助詞始まり・極端に短い）かどうか判定する。"""
    if not text or len(text) < 6:
        return True
    # 助詞・接続詞の残骸で始まる
    if _re.match(r'^(?:を|が|は|に|で|と|も|へ|の|から|より|ため|ので|により|によって|経由で|を通じて|ため動作)', text):
        return True
    # 明らかな記号混じりフラグメント
    if _re.search(r'(?:→に|→を|／を|\/を|または遷移|および遷移|\s+遷移$)', text):
        return True
    return False


# O-3②: role 欄の stale cache 断片マーカー
_ROLE_FRAGMENT_MARKERS: tuple[str, ...] = (
    "紐づくが動作不全",
    "紐づくが",
    "動作不全",
    "対で、",
    "へのリダイレクタ",
    "使い ",        # 半角スペース混入（「使い 〜と紐付く」パターン）
    "使っ(た)",
    "使う標準テンプレート",
    "を新規作成で挿入",
    "と紐付く",
    "紐付く",
    "ランタイム解決",
    "$でランタイム",
    "VF。を",
    "VF。紐",
)


def _is_role_fragment(text: str) -> bool:
    """role 欄の stale cache 断片を検出。True なら破棄して辞書 / 型別 fallback で書き直す。"""
    if not text or len(text.strip()) < 4:
        return True
    t = text.strip()
    for marker in _ROLE_FRAGMENT_MARKERS:
        if marker in t:
            return True
    # 文頭が裸助詞のみ（「を」「が」「は」「に」「で」で始まる）
    if _re.match(r'^[をがはにでとへの]', t):
        return True
    # 裸ドット（半角スペース＋ドット）— 「使い 契約申込.と」等
    if _re.search(r'\s+\.\S', t):
        return True
    return False


_APEX_ROLE_SUFFIXES: tuple[tuple[str, str], ...] = (
    ("Controller",  "画面のコントローラー処理を担当する Apex クラス。"),
    ("Extension",   "画面の拡張コントローラー処理を担当する Apex クラス。"),
    ("Handler",     "トリガーまたはイベントのハンドラー処理を担当する Apex クラス。"),
    ("Service",     "業務ロジックを提供するサービス層の Apex クラス。"),
    ("Helper",      "共通処理を担うヘルパー Apex クラス。"),
    ("Trigger",     "レコード変更時に起動する Apex トリガー。"),
    ("Batch",       "大量レコードを分割処理する Apex バッチクラス。"),
    ("Schedulable", "スケジュール起動される Apex クラス。"),
    ("Queueable",   "非同期キューで処理される Apex クラス。"),
)


def _apex_role_from_api_name(api_name: str) -> str:
    """api_name の末尾サフィックスから Apex の役割文を推論する。"""
    if not api_name:
        return ""
    for suffix, role in _APEX_ROLE_SUFFIXES:
        if api_name.endswith(suffix):
            return role
    return ""


# AA-2i: 標準コミュニティ／サイトコントローラの処理フロー図用一言ラベル
_API_TO_FLOW_LABEL_HEURISTIC: dict[str, str] = {
    "CommunitiesLandingController":        "コミュニティ表示",
    "CommunitiesLoginController":          "ログイン認証",
    "CommunitiesSelfRegController":        "セルフ登録受付",
    "CommunitiesSelfRegConfirmController": "登録完了確認",
    "ForgotPasswordController":            "パスワード忘れ",
    "MicrobatchSelfRegController":         "非同期登録補助",
    "SiteRegisterController":              "サイト登録処理",
    "SiteLoginController":                 "サイトログイン",
    "ChangePasswordController":            "パスワード変更",
}


def _flow_label_from_api_name(api_name: str) -> str:
    """API 名から処理フロー図用の短ラベル（6〜10字）を生成する（augmented Apex 用）。"""
    if api_name in _API_TO_FLOW_LABEL_HEURISTIC:
        return _API_TO_FLOW_LABEL_HEURISTIC[api_name]
    base = _re.sub(r'(?:Controller|Extension|Handler|Service|Helper|Manager|Batch|Trigger)$', '', api_name)
    return base[:10] if base else api_name[:10]


def _title_from_desc(desc: str, max_len: int = 18) -> str:
    """description の先頭節から short title を生成する。「を行う」の機械付与をしない。"""
    if not desc:
        return ""
    # 先頭文（。か改行まで）を取る
    first = _re.split(r'[。\n]', desc)[0].strip()
    # 末尾の接続形・助詞を整える
    first = _re.sub(r'(?:して|ており|ていて|たうえで)[^。\n]*$', '', first).strip()
    first = _re.sub(r'(?:は|が|を|に|で|と|の|も|へ)$', '', first).strip()
    if not first or len(first) < 3:
        tokens = _re.findall(r'[ぁ-んァ-ヶ一-龯々ー]{3,}', desc)
        if tokens:
            return tokens[0][:max_len]
        return desc[:max_len]
    return first[:max_len]


# 標準 VF/コンポーネント API 名 → 自然な日本語説明（O-3③: 「〜画面で〜を行う」形式に統一）
_STD_VF_DESCRIPTIONS: dict[str, str] = {
    "SiteLogin":                        "ポータルログイン画面でユーザー認証を行う。",
    "ForgotPassword":                   "パスワードリセット申請画面でメール宛先の入力を受け付け、リセット用メールを送信する。",
    "ForgotPasswordConfirm":            "パスワードリセット申請完了画面でメール送信済みのメッセージを表示する。",
    "ChangePassword":                   "パスワード変更画面で新しいパスワードの入力を受け付け、パスワードを更新する。",
    "FileNotFound":                     "404 エラー画面でリソースが見つからない旨を表示する。",
    "Exception":                        "システム例外エラー画面で想定外のエラーをキャッチしメッセージを表示する。",
    "StdExceptionTemplate":             "例外テンプレート画面で共通のエラー表示レイアウトを提供する。",
    "Unauthorized":                     "アクセス権限エラー画面で権限不足のメッセージを表示する。",
    "InMaintenance":                    "メンテナンス中画面でサービス停止期間のメッセージを表示する。",
    "BandwidthExceeded":                "帯域超過エラー画面で利用上限に達した旨を表示する。",
    "AnswersHome":                      "旧 Answers 機能のトップ画面を表示する（廃止予定）。",
    "IdeasHome":                        "旧 Ideas 機能のトップ画面を表示する（廃止予定）。",
    "SiteRegister":                     "ポータルユーザ登録画面でユーザー情報の入力を受け付け、アカウントを登録する。",
    "SiteRegisterConfirm":              "ポータルユーザ登録完了画面で登録完了のメッセージを表示する。",
    "CommunitiesLogin":                 "コミュニティログイン画面でユーザー認証を行う。",
    "CommunitiesSelfReg":               "コミュニティ自己登録画面でユーザー情報の入力を受け付け、アカウントを登録する。",
    "CommunitiesSelfRegConfirm":        "コミュニティ自己登録完了画面で登録完了のメッセージを表示する。",
    "CommunitiesForgotPassword":        "コミュニティのパスワードリセット申請画面でメール宛先の入力を受け付ける。",
    "CommunitiesForgotPasswordConfirm": "コミュニティのパスワードリセット申請完了画面でメール送信済みのメッセージを表示する。",
    "CommunitiesChangePassword":        "コミュニティのパスワード変更画面で新しいパスワードの入力を受け付け、パスワードを更新する。",
}

# 標準 Apex クラス除外（Apex→Apex callees 解析で誤検知しないためのスキップセット）
_STD_APEX_SKIP: frozenset[str] = frozenset({
    "System", "Database", "Schema", "Test", "Math", "UserInfo",
    "Json", "JSON", "DateTime", "Date", "Http", "HttpRequest", "HttpResponse",
    "List", "Map", "Set", "Iterator", "PageReference", "ApexPages",
    "Pattern", "Matcher", "Integer", "Long", "Double", "Decimal",
    "Boolean", "Blob", "Id", "Time", "Type", "Exception",
    "SObject", "Trigger", "Limits", "EncodingUtil", "URL",
    "Crypto", "Auth", "Site", "Network", "UserManagement",
    "ApexClass", "String", "Object", "Comparable", "Iterable",
})


def _augment_components_with_vf_controllers(data: dict) -> None:
    """VF コンポーネントの controller 属性から Apex クラスを特定し、
    components に未登録のものを自動追加して callees 関係を構築する（防御層）。
    エージェントが Apex を components に含め忘れたケースをカバーする。
    M-5c: さらに既登録 Apex コンポーネント同士の呼び出し関係（new / static call）も補強する。
    """
    if not _CURRENT_SF_BASE_PATH:
        return
    pages_dir = Path(_CURRENT_SF_BASE_PATH) / "force-app/main/default/pages"
    if not pages_dir.exists():
        return
    classes_dir = Path(_CURRENT_SF_BASE_PATH) / "force-app/main/default/classes"

    # 標準ボイラープレートは追加対象外（カスタムコントローラを持たない標準 VF）
    _STD_SKIP = {
        "SiteLogin", "ForgotPassword", "ForgotPasswordConfirm", "ChangePassword",
        "FileNotFound", "Exception", "Unauthorized", "InMaintenance",
        "BandwidthExceeded", "AnswersHome", "SiteRegister", "SiteRegisterConfirm",
        "CommunitiesLogin", "CommunitiesSelfReg", "CommunitiesSelfRegConfirm",
        "CommunitiesForgotPassword", "CommunitiesForgotPasswordConfirm",
        "CommunitiesChangePassword",
    }

    existing_apis = {c.get("api_name", "") for c in data.get("components", [])}
    new_comps: list[dict] = []

    for comp in data.get("components", []):
        api = comp.get("api_name", "")
        comp_type = comp.get("type", "")
        if comp_type not in ("Visualforce", "VF", "VisualForce", "Visualforce Page"):
            continue
        vf_file = pages_dir / f"{api}.page"
        if not vf_file.exists():
            continue
        try:
            vf_content = vf_file.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        ctrl_m = _re.search(r'(?<!standard)[Cc]ontroller="([A-Za-z][A-Za-z0-9_]*)"', vf_content)
        if not ctrl_m:
            continue
        ctrl_name = ctrl_m.group(1)
        if ctrl_name in _STD_SKIP:
            continue
        # VF → Apex の callees 関係を設定（既存 Apex でも callees は必ず更新）
        callees = comp.setdefault("callees", [])
        if ctrl_name not in callees:
            callees.append(ctrl_name)
        if ctrl_name in existing_apis:
            continue
        # Apex クラスからオブジェクト名を取得して role を生成
        role = ""
        cls_file = classes_dir / f"{ctrl_name}.cls"
        if cls_file.exists():
            comp_fields = _SF_COMP_FIELDS.get(ctrl_name, {})
            obj_names = [o for o in comp_fields if o not in ("__any__", "Id")]
            if obj_names:
                role = f"{' / '.join(obj_names[:2])} を操作するコントローラ"
        # AA-2i: responsibility を空にしない。VF の STD 説明を借用するか suffix 推論を使う
        std_desc = _STD_VF_DESCRIPTIONS.get(api, "")
        if std_desc:
            aug_resp = std_desc
        elif role:
            aug_resp = role
        else:
            aug_resp = _apex_role_from_api_name(ctrl_name) or ""
        aug_flow = _flow_label_from_api_name(ctrl_name)
        new_comps.append({
            "api_name": ctrl_name,
            "type": "Apex",
            "role": role or aug_resp[:40] or "ポータル画面コントローラ",
            "callees": [],
            "responsibility": aug_resp,
            "flow_label": aug_flow,
        })
        existing_apis.add(ctrl_name)

    data.setdefault("components", []).extend(new_comps)

    # M-5c: Apex → Apex の callees 補強（new ClassName() / ClassName.staticMethod() を検出）
    all_apis = {c.get("api_name", "") for c in data.get("components", [])}
    for comp in data.get("components", []):
        if comp.get("type") not in ("Apex", "ApexClass"):
            continue
        api = comp.get("api_name", "")
        cls_file = classes_dir / f"{api}.cls"
        if not cls_file.exists():
            continue
        try:
            cls_content = cls_file.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        # new ClassName(/ static ClassName.method( パターンを拾う
        new_calls = _re.findall(r'\bnew\s+([A-Z][A-Za-z0-9_]+)\s*[(<]', cls_content)
        static_calls = _re.findall(r'\b([A-Z][A-Za-z0-9_]+)\.[a-z][A-Za-z0-9_]*\s*\(', cls_content)
        called = (set(new_calls) | set(static_calls)) - _STD_APEX_SKIP
        existing_callees = comp.setdefault("callees", [])
        for cls_name in called:
            if cls_name in all_apis and cls_name != api and cls_name not in existing_callees:
                existing_callees.append(cls_name)

    # X-4b: Flow → Apex/Flow の callees 補強（Flow XML の actionCalls / subflows を解析）
    flows_dir = Path(_CURRENT_SF_BASE_PATH) / "force-app/main/default/flows"
    all_apis_now = {c.get("api_name", "") for c in data.get("components", [])}
    for comp in data.get("components", []):
        comp_type = comp.get("type", "")
        if comp_type not in ("Flow", "AutoLaunchedFlow", "ScheduledFlow", "RecordTriggeredFlow"):
            continue
        api = comp.get("api_name", "")
        # .flow-meta.xml 検索（複数の命名パターンに対応）
        flow_file = None
        for candidate in [f"{api}.flow-meta.xml", f"{api}.flow"]:
            p = flows_dir / candidate
            if p.exists():
                flow_file = p
                break
        if flow_file is None:
            continue
        try:
            flow_content = flow_file.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        existing_callees = comp.setdefault("callees", [])
        # <actionCalls> の Apex action
        for m in _re.finditer(
            r'<actionType>\s*apex\s*</actionType>.*?<actionName>\s*([A-Za-z][A-Za-z0-9_]*)\s*</actionName>',
            flow_content, _re.DOTALL
        ):
            cls_name = m.group(1)
            if cls_name in all_apis_now and cls_name != api and cls_name not in existing_callees:
                existing_callees.append(cls_name)
        # <subflows> の呼び出し先フロー
        for m in _re.finditer(r'<flowName>\s*([A-Za-z][A-Za-z0-9_]*)\s*</flowName>', flow_content):
            sub_name = m.group(1)
            if sub_name in all_apis_now and sub_name != api and sub_name not in existing_callees:
                existing_callees.append(sub_name)

    # X-4b: Trigger → Apex の callees 補強（Trigger ファイルを読んで Apex 呼び出しを検出）
    triggers_dir = Path(_CURRENT_SF_BASE_PATH) / "force-app/main/default/triggers"
    if triggers_dir.exists():
        all_apis_now = {c.get("api_name", "") for c in data.get("components", [])}
        for comp in data.get("components", []):
            comp_type = comp.get("type", "")
            if comp_type not in ("Trigger", "ApexTrigger"):
                continue
            api = comp.get("api_name", "")
            trig_file = triggers_dir / f"{api}.trigger"
            if not trig_file.exists():
                # type が Flow だが .trigger ファイルが実在するケース（type 誤判定の補正）
                trig_file = triggers_dir / f"{api}.trigger"
                if not trig_file.exists():
                    continue
            try:
                trig_content = trig_file.read_text(encoding="utf-8", errors="ignore")
            except Exception:
                continue
            new_calls = _re.findall(r'\bnew\s+([A-Z][A-Za-z0-9_]+)\s*[(<]', trig_content)
            static_calls = _re.findall(r'\b([A-Z][A-Za-z0-9_]+)\.[a-z][A-Za-z0-9_]*\s*\(', trig_content)
            called = (set(new_calls) | set(static_calls)) - _STD_APEX_SKIP
            existing_callees = comp.setdefault("callees", [])
            for cls_name in called:
                if cls_name in all_apis_now and cls_name != api and cls_name not in existing_callees:
                    existing_callees.append(cls_name)


def _infer_callees(data: dict) -> None:
    """data_flow_overview の「→」連鎖からコンポーネント間呼び出し関係を推論して callees に設定する。"""
    flow_text = data.get("data_flow_overview", "")
    if not flow_text:
        return
    comp_names = {c.get("api_name", "") for c in data.get("components", [])}
    if not comp_names:
        return

    # 全文を処理（。で区切らず全体から → を抽出）
    tokens = _re.split(r'→', flow_text.replace("\n", " "))

    def find_comp(token: str) -> str | None:
        for name in comp_names:
            if name and name in token:
                return name
        return None

    callees_map: dict[str, list[str]] = {n: [] for n in comp_names}
    prev = None
    for token in tokens:
        curr = find_comp(token)
        if curr and prev and prev != curr and curr not in callees_map[prev]:
            callees_map[prev].append(curr)
        if curr:
            prev = curr

    for comp in data.get("components", []):
        name = comp.get("api_name", "")
        if callees_map.get(name):
            # M-5c: 上書きではなくマージ（_augment 済みの callees を保持）
            existing = comp.setdefault("callees", [])
            for callee in callees_map[name]:
                if callee not in existing:
                    existing.append(callee)


def _build_business_flow(data: dict) -> list[dict]:
    """processing_purpose/screens/data_flow_overview から業務レベルフローを生成する。

    ルール:
    - 「誰が・何を起点に → 何が起きる → 誰が何を受け取る」を3〜4ステップで表現
    - クラス名・API名・アノテーションは一切使わない
    - 処理概要シートとの重複を避けるため、システム内部処理は書かない
    """
    purpose  = data.get("processing_purpose", "")
    flow_ov  = data.get("data_flow_overview", "")
    notes    = data.get("notes", "")
    screens  = data.get("screens", [])
    combined = purpose + " " + flow_ov + " " + notes

    steps    = []
    step_no  = 1

    # ─── Step 1: 起点アクター＋起動アクション ─────────────────────────────
    first_tok = (flow_ov.split("→")[0] if flow_ov else "").strip()

    # 技術ワードのみのトークン（例「画面フロー（Create_CustomerUser）」「Apexクラス」等）は
    # 業務アクションとして読めないため、適切なデフォルトに差し替える。
    first_tok_core = _re.sub(r'[（(].*?[）)]', '', first_tok).strip()
    _TECH_ONLY = ("画面フロー", "フロー", "Apex", "Apexクラス", "Batch",
                  "Flow", "Trigger", "LWC", "Aura", "Visualforce",
                  "VF", "Integration", "Controller", "Service", "Handler")
    is_tech_only = first_tok_core in _TECH_ONLY
    is_screen_flow_start = "画面フロー" in first_tok

    if screens:
        # 画面あり → お客様が画面から入力・送信
        scr = screens[0].get("screen_name") or screens[0].get("component", "画面")
        scr_clean = _clean_tech_business(scr)
        action1 = f"{scr_clean}から必要情報を入力し、送信する"
        steps.append({"step": step_no, "actor": "お客様", "action": action1, "next": []})
        step_no += 1
    elif is_screen_flow_start:
        # data_flow_overview が「画面フロー（...）→ ...」で始まる場合は画面起点。
        # screens が空でも Flow タイプの画面フローが存在する想定。
        # 管理系業務（ユーザー発行・管理者操作）は GF社担当者、それ以外はお客様。
        admin_hint = _re.search(r'管理者|発行|管理.*ユーザー|GF社担当|事務', purpose + notes)
        actor = "GF社担当者" if admin_hint else "お客様"
        steps.append({"step": step_no, "actor": actor,
                      "action": "画面から必要情報を入力し、処理を起動する", "next": []})
        step_no += 1
    elif is_tech_only:
        # 技術ワードのみ（画面フロー以外）→ 自動起点とみなす
        steps.append({"step": step_no, "actor": "自動フロー",
                      "action": "処理を起動する", "next": []})
        step_no += 1
    elif _re.search(r'管理者|GF社|担当者|事務', first_tok):
        # GF社担当者が直接操作して起動
        clean = _clean_tech_business(first_tok)
        steps.append({"step": step_no, "actor": "GF社担当者",
                      "action": clean or "処理を起動する", "next": []})
        step_no += 1
    elif _re.search(r'承認.*[Ff]low|承認.*フロー|承認フロー', first_tok):
        # 承認フロー起点 → 「誰かが何かを承認した後に動く」→ GF社担当者が承認
        m_ctx = _re.search(r'(.{2,25}(?:確定|承認|依頼|完了))後', purpose)
        ctx = _clean_tech_business(m_ctx.group(1)) if m_ctx else "業務処理"
        steps.append({"step": step_no, "actor": "GF社担当者",
                      "action": f"{ctx}後、処理を承認・起動する", "next": []})
        step_no += 1
    elif _re.search(r'お客様|顧客[^向]|Experience Cloud.*ポータル|ポータル.*画面', first_tok):
        clean = _clean_tech_business(first_tok)
        steps.append({"step": step_no, "actor": "お客様",
                      "action": clean or "操作を行う", "next": []})
        step_no += 1
    else:
        clean = _clean_tech_business(first_tok)
        steps.append({"step": step_no, "actor": "GF社担当者",
                      "action": clean or "処理を起動する", "next": []})
        step_no += 1

    # ─── Step 2: メイン業務処理（processing_purpose の主要部分） ──────────────
    sents = [s.strip() for s in purpose.split("。") if s.strip()]
    if sents:
        first = sents[0]
        # 「〜後の」前置き除去（「コンサルティング本確定後の」等）
        core = _re.sub(r'^.{0,35}後の', '', first).strip()
        main = _clean_tech_business(core or first)
        main = _PREAMBLE_RE.sub('', main).strip()
        # "〜を担う" 等の文末表現を除去（行動として読めるように）
        main = _re.sub(r'を担う$|を行う$|を実施する$', '', main).strip()
        main = _re.sub(r'^[、。・/\s]+|[、。・/\s]+$', '', main).strip()
        if main and len(main) > 5:
            steps.append({"step": step_no, "actor": "自動フロー",
                          "action": main, "next": []})
            step_no += 1

    # ─── Step 3: 顧客向け結果（キーワードで判定） ─────────────────────────────
    outcome = ""
    if _re.search(r'初期パスワード|パスワード.{0,5}メール', combined):
        outcome = "初期パスワードメールを受信し、Experience Cloudポータルにアクセスする"
    elif _re.search(r'thankPage|thanksPage|thank.*page', combined, _re.IGNORECASE):
        outcome = "フォーム送信後、受付完了ページで確認する"
    elif _re.search(r'顧客向け通知|通知.*メール|ダウンロード.*通知|完了通知', combined):
        outcome = "完了通知メールを受信し、書類を確認する"
    elif _re.search(r'ダウンロード', combined) and _re.search(r'Experience Cloud|ポータル', combined):
        outcome = "Experience Cloudポータルで書類をダウンロードする"
    elif _re.search(r'メール.*送信|自動メール|顧客宛', combined) and not screens:
        outcome = "自動送信されたメールを受信する"

    if outcome:
        steps.append({"step": step_no, "actor": "お客様",
                      "action": outcome, "next": []})
        step_no += 1

    # ─── next リンク設定 ───────────────────────────────────────────────
    for i in range(len(steps) - 1):
        steps[i]["next"] = [{"to": steps[i + 1]["step"]}]

    return steps


def _comp_type_label(comp: dict) -> str:
    """コンポーネントの種別ラベルを日本語で返す（クラス名は出さない）。"""
    api   = comp.get("api_name", "")
    ctype = comp.get("type", "Apex")
    if ctype == "Flow":
        return "フロー"
    if ctype == "LWC":
        return "LWCコンポーネント"
    if ctype == "Integration":
        return "外部サービス連携"
    if ctype == "Trigger" or ("Trigger" in api and "Handler" in api):
        return "Apexトリガーハンドラー"
    if "Trigger" in api:
        return "Apexトリガー"
    return "Apexクラス"


def _make_box_label(desc: str, comp_type: str = "") -> str:
    """処理フロー図ボックス用の 2 行ラベルを生成する。
    Q2 確定: 「画面名（1 行目）＋ アクション（2 行目）」の 2 行体裁。
    - description が「〜画面で〜を行う」形式なら「で」を軸に分割。
    - VF/LWC/Aura 以外（Apex/Flow/Trigger 等）は 1 行要約のみ。
    - 「（Experience Cloud〜）」等の括弧注記は除去。
    """
    if not desc:
        return ""
    # 括弧注記を除去（「（Experience Cloud 標準テンプレート）」等）
    clean = _re.sub(r'[（(][^）)]{2,30}[）)]', '', desc).strip()
    clean = _re.sub(r'[。\s]+$', '', clean).strip()

    # 「〜画面で〜する」パターンを分割
    m = _re.search(r'^(.{3,20}画面)で(.{3,})', clean)
    if m:
        screen = m.group(1).strip()
        action = m.group(2).strip()
        # アクションは 22 字以内に圧縮（Q-3c: 18→22 に緩和して情報量を維持）
        action = action[:22] + ('…' if len(action) > 22 else '')
        return f"{screen}\n{action}"

    # 画面系だが「で」なしの場合: 先頭 15 字でラベル化
    if comp_type.lower() in ("visualforce", "vf", "lwc", "aura"):
        return clean[:15] + ('…' if len(clean) > 15 else '')

    # Apex/Flow 等: 先頭節を 1 行で
    return clean[:18] + ('…' if len(clean) > 18 else '')


def _build_process_steps(data: dict) -> list[dict]:
    """components の responsibility から日本語の処理概要 steps を生成する。

    M-5a: _deep_clean_ja（強い除去）を適用 → 断片なら _STD_VF_DESCRIPTIONS / 型別デフォルト にフォールバック。
    O-3④: box_label（処理フロー図ボックス用 2 行ラベル）を追加。business_flow 順に並び替え。
    title は description の先頭節から生成（「を行う」機械付与廃止）。
    """
    # 型別デフォルト説明（責務が空または断片の場合のフォールバック）
    _TYPE_DEFAULTS: dict[str, str] = {
        "Visualforce": "ポータル画面で標準的な処理を行う。",
        "VF": "ポータル画面で標準的な処理を行う。",
        "VisualForce": "ポータル画面で標準的な処理を行う。",
        "Visualforce Page": "ポータル画面で標準的な処理を行う。",
        "Apex": "Apex クラスとして処理を担当する。",
        "Flow": "自動起動フローが処理を担当する。",
        "LWC": "ユーザー操作画面でインタラクションを担当する。",
        "Aura": "ユーザー操作画面でインタラクションを担当する。",
    }

    # business_flow の actor/component 順でソートするための順序マップを作成
    _bf_order: dict[str, int] = {}
    for _step in (data.get("business_flow") or []):
        _actor = _step.get("actor") or _step.get("component") or ""
        if _actor and _actor not in _bf_order:
            _bf_order[_actor] = len(_bf_order)

    comps = list(data.get("components", []))
    if _bf_order:
        def _bf_sort_key(c: dict) -> int:
            _name = c.get("api_name", "") or c.get("name", "")
            # actor との部分一致でも順序を取る
            for _actor, _idx in _bf_order.items():
                if _name in _actor or _actor in _name:
                    return _idx
            return len(_bf_order) + 1  # 未マッチは末尾
        comps = sorted(comps, key=_bf_sort_key)

    # R-2 (コンポーネント列に拡張子付与): type → 拡張子マップ
    _EXT_BY_TYPE: dict[str, str] = {
        "Apex":                   ".cls",
        "Apex Class":             ".cls",
        "Apex Trigger":           ".trigger",
        "Trigger":                ".trigger",
        "Visualforce":            ".page",
        "VF":                     ".page",
        "VisualForce":            ".page",
        "Visualforce Page":       ".page",
        "LWC":                    ".js",
        "Lightning Web Component":".js",
        "Aura":                   ".cmp",
        "Aura Component":         ".cmp",
        "Flow":                   ".flow",
    }

    steps = []
    n_comps = len(comps)
    for i, comp in enumerate(comps, 1):
        responsibility = comp.get("responsibility", "")
        comp_type = comp.get("type", "Apex")
        _raw_api = comp.get("api_name", "")
        comp_name = _re.sub(r'（[^）]+）$', '', _raw_api).strip() or _raw_api
        # R-2: 拡張子を付けて Apex/VF/LWC の区別を明示
        _ext = _EXT_BY_TYPE.get(comp_type, "")
        comp_display = f"{comp_name}{_ext}" if _ext and not comp_name.endswith(_ext) else comp_name

        # 1) 標準 VF 辞書からの固定説明（最優先）
        desc_main = _STD_VF_DESCRIPTIONS.get(_raw_api, "")

        # 2) LLM responsibility を _deep_clean_ja で処理
        if not desc_main and responsibility:
            cleaned = _deep_clean_ja(responsibility)
            # トリガータイミング注釈（処理前（新規）/処理後（新規））等を除去
            cleaned = _re.sub(
                r'[（(](?:処理前|処理後)[（(][^）)]+[）)](?:/(?:処理前|処理後)[（(][^）)]+[）)])?[）)]', '', cleaned
            ).strip()
            cleaned = _re.sub(r'\s{2,}', ' ', cleaned)
            if not _is_desc_fragment(cleaned):
                desc_main = cleaned

        # 3) フォールバック: Apex は API 名の suffix 推論を優先（AA-2i）、それ以外は型別デフォルト
        if not desc_main:
            if comp_type in ("Apex", "ApexClass", "ApexTrigger", "Apex Class", "Apex Trigger"):
                suffix_role = _apex_role_from_api_name(_raw_api)
                desc_main = suffix_role or _TYPE_DEFAULTS.get(comp_type, "処理を担当する。")
            else:
                desc_main = _TYPE_DEFAULTS.get(comp_type, "処理を担当する。")

        # M-5a: title は description の先頭節から生成（「を行う」機械付与廃止）
        title = _title_from_desc(desc_main)

        # O-3④: box_label（処理フロー図ボックス用 2 行ラベル）
        box_label = _make_box_label(desc_main, comp_type)

        steps.append({
            "step": i,
            "title": title,
            "description": desc_main,
            "box_label": box_label,
            "flow_label": (comp.get("flow_label") or "").strip(),
            "component": comp_display,
            "comp_api_name": _raw_api,
            "branch": None,
            "next": [{"to": i + 1}] if i < n_comps else [],
        })

    return steps


# API 名・メソッド呼び出し・拡張子付きファイル名の検出パターン
_RE_METHOD_CALL   = _re.compile(r'\b[A-Za-z_][A-Za-z0-9_]*(?:\.[A-Za-z_][A-Za-z0-9_]*)*\s*\([^)]*\)')
_RE_FILE_EXT      = _re.compile(r'\b[A-Za-z_][A-Za-z0-9_]*\.(?:page|cls|trigger|cmp|app|evt|js|html|css|xml)\b')
_RE_FLOW_META_EXT = _re.compile(r'\b[A-Za-z_][A-Za-z0-9_]*\.flow-meta(?:\.xml)?\b')
_RE_CUSTOM_OBJ    = _re.compile(r'\b[A-Z][A-Za-z0-9_]*__c\b')  # カスタムAPI名（__c 付き）
_RE_CAMEL_IDENT   = _re.compile(r'\b([A-Z][a-z]+){2,}[A-Za-z0-9_]*\b')  # CamelCase 識別子（連続2語以上）
_RE_CMP_PAREN     = _re.compile(r'[（(]CMP[‐\-][0-9]+[）)]')  # 「（CMP-015）」
_RE_MULTI_SPACE   = _re.compile(r'[ \t]{2,}')
_RE_ORPHAN_PARTICLE = _re.compile(r'[ 　][をはがにでとのやもへ][ 　]')
# 追加: 英語残留パターン（_strip_tech_identifiers で適用）
_RE_REL_FIELD   = _re.compile(r'[A-Za-z][A-Za-z0-9_]*__r\.[A-Za-z_]\w*')  # User_portal__r.Field
_RE_DOT_CHAIN   = _re.compile(r'\b[A-Z][A-Za-z0-9]+(?:\.[A-Za-z_]\w*)+\b')  # ClassName.method（括弧なし）
_RE_URL_QUERY   = _re.compile(r'\?[a-zA-Z_]\w*(?:=[^\s。、）)]*)?')         # ?appkbn=0/1
_RE_LOWER_CAMEL = _re.compile(r'(?<![A-Za-z_])[a-z]+[A-Z][a-zA-Z]+(?![A-Za-z_])')  # lowerCamelCase


def _strip_tech_identifiers(text: str) -> str:
    """テキストから API 名・メソッド呼び出し・拡張子付きファイル名・CamelCase 識別子を除去する。
    _clean_tech_business / _translate_sf_fields を通した後の残渣をさらに除去する。"""
    if not text:
        return text
    # K-B: XML/JSX 風タグの裸出し（<c:xxx>, <apex:xxx>, <c:> 等）
    #       → タグだけ除去し、中身の意味語（カスタムコンポーネント等）は LLM 側で保持
    text = _re.sub(r'<\s*[cC]\s*:\s*[A-Za-z0-9_]*\s*/?>', '', text)
    text = _re.sub(r'<\s*/?\s*[aA][pP][eE][xX]\s*:\s*[A-Za-z0-9_]*\s*/?>', '', text)
    text = _re.sub(r'<\s*/\s*[cC]\s*:\s*[A-Za-z0-9_]*\s*>', '', text)
    # K-B: 裸パス残渣（/apex/, /answers/, /apex/answers 等の先頭/末尾スラッシュ付きパス）
    # 日本語句読点・矢印・行頭行末に隣接する裸パスのみ対象（URL として意味ある形は保護しない前提）
    text = _re.sub(r'(?<![A-Za-z0-9])\/[a-z]+(?:\/[a-z]+)*\/?(?![A-Za-z0-9])', '', text)
    # __r リレーション参照（User_portal__r.UserName 等）
    text = _RE_REL_FIELD.sub('', text)
    # ClassName.method 形式（括弧なし: Site.validatePassword / System.setPassword 等）
    text = _RE_DOT_CHAIN.sub('', text)
    # URL クエリパラメータ（?appkbn=0/1 等）
    text = _RE_URL_QUERY.sub('', text)
    text = _RE_METHOD_CALL.sub('', text)
    text = _RE_FILE_EXT.sub('', text)
    text = _RE_FLOW_META_EXT.sub('', text)
    text = _RE_CMP_PAREN.sub('', text)
    text = _RE_CUSTOM_OBJ.sub('', text)
    # CamelCase 識別子（Controller, Handler などのクラス名）も除去。
    # ただし "Experience Cloud" のような既知ブランドは保護（事前にプレースホルダ化してから呼ぶ）
    text = _RE_CAMEL_IDENT.sub('', text)
    # lowerCamelCase 識別子（objId, userName 等）
    text = _RE_LOWER_CAMEL.sub('', text)
    # 除去で生じた連続空白・孤立助詞・読点連続を整える
    text = _RE_MULTI_SPACE.sub(' ', text)
    # 句点直後の孤立助詞（"。の結果" "。を呼び出し" 等）→ 句点のみに
    text = _re.sub(r'。[ \t]*([をはがにでとのやもへ])(?=[ぁ-んァ-ヶ一-龯])', '。', text)
    # 読点直後の孤立助詞（"、の結果" 等）→ 読点のみに
    text = _re.sub(r'[、，][ \t]*([をはがにでとのやもへ])(?=[ぁ-んァ-ヶ一-龯])', '、', text)
    # 「て+助詞」の不自然接続（"受けてに" 等）→ 「て」のみに
    text = _re.sub(r'(?<=[ぁ-んァ-ヶ])て[にでをがはと](?=[ぁ-んァ-ヶ一-龯])', 'て', text)
    # 前後スペースで挟まれた孤立助詞（" の結果"）→ 前スペースだけ除去
    text = _re.sub(r'\s+([をはがにでとのやもへ])\s*(?=[ぁ-んァ-ヶ一-龯])', r'\1', text)
    # 連続矢印（要素除去後に → → → と残るケース）→ 1つに
    text = _re.sub(r'→(?:[ \t]*→)+', '→', text)
    # 接続詞直後の孤立助詞（「またはに」「およびで」等）
    text = _re.sub(r'(または|および|かつ|ただし)[ \t]*([をがはにでへと])(?=[ぁ-んァ-ヶ一-龯])', r'\1', text)
    # 末尾の孤立矢印
    text = _re.sub(r'[ \t]*→[ \t]*$', '', text)
    # 読点連続を1つに
    text = _re.sub(r'[、，]\s*[、，]+', '、', text)
    # スペース + 句読点 → 句読点のみ
    text = _re.sub(r'\s+([、，。])', r'\1', text)
    # 句読点直後の不要スペース
    text = _re.sub(r'([、，。])\s+', r'\1', text)
    # 連続スペースを再圧縮
    text = _RE_MULTI_SPACE.sub(' ', text)
    # G-1: 記号残留対策
    text = _re.sub(r'\.{2,}', '', text)                                  # 連続ドット（__r 除去後の .. 等）
    text = _re.sub(r'／(?:\s*／)+', '／', text)                           # 連続全角スラッシュ → 単一
    text = _re.sub(r'→\s*[+・\-\/／＋=]\s*→', '→', text)               # 矢印間の裸記号（→ + → 等）
    text = _re.sub(r'(?<=[ぁ-んァ-ヶ一-龯て])で([がをにはと])(?=[ぁ-んァ-ヶ一-龯])', 'で', text)  # でが 系二連助詞
    text = _re.sub(r'→\s*(?:=|は)\s*(?=\s|$)', '', text)                 # 矢印末尾の裸記号
    text = _re.sub(r'([。、])\s*[+・\-\/／＋=]\s*', r'\1', text)         # 句読点直後の裸記号
    # G-3: 行頭の孤児助詞（英語除去で前詞が消えたケース "から 契約申込を特定し" 等）
    text = _re.sub(r'^\s*(?:から|まで|より|へ|を|に|で|と|の|は|が)\s+', '', text)
    text = _re.sub(r'^\s*(URL)\s+(?:から|まで|より|へ|を|に|で|と|の|は|が)\s+', r'\1 ', text)
    # G-3b: 行頭の裸スラッシュ（コンポーネント名連結除去後の "／ は..." "／は..." "（／）は..."）
    text = _re.sub(r'^[\s　]*[（(][\s　]*／[\s　]*[）)][\s　]*(?:は|が|を|に|で|と|の)?[\s　]*', '', text)
    text = _re.sub(r'^[\s　]*／[\s　]*(?:は|が|を|に|で|と|の|へ|まで|から)?[\s　]*', '', text)
    # G-3c: 連続助詞の補正（Class名を空置換後の "がに" "にで" など、右側を残す）
    text = _re.sub(r'([がはをにと])([がはをにと])(?=[ぁ-んァ-ヶ一-龯])', r'\2', text)
    # G-3d: 末尾・中間の孤立「=」（"=のため" "=true" 等の残渣）
    text = _re.sub(r'\s*=\s*(?=のため|の結果|のとき|でない|である)', '', text)
    text = _re.sub(r'\s*=\s*$', '', text)
    text = _re.sub(r'[、，]\s*系は\s*', '、', text)  # 「は、系は」の中間「系は」は元 "IsXxx 系は" の残骸
    # G-3e: タイトル行頭の単一助詞（VF名/Class名除去の副作用、仮名/漢字/長音が続く場合のみ）
    text = _re.sub(r'^(?:で|は|を|に|と|が)(?=[一-龯ァ-ヶー][ぁ-んァ-ヶ一-龯ー])', '', text)
    # G-3f: 中間位置の裸「（／）」「（／／）」等（コンポーネント名連結除去後）
    text = _re.sub(r'[（(][\s　]*(?:／[\s　]*)+[）)]', '', text)
    # H-3a: 矢印間の裸記号（G-1 より広い文字セット: 中黒・& も対象）
    text = _re.sub(r'→[\s　]*[・+\-\/／＋=&][\s　]*→', '→', text)
    # H-3b: 接続詞（または/および/かつ）直後の孤立助詞
    text = _re.sub(r'(または|および|かつ)\s*(?:を|が|は|に|で|と|へ|の)(?=[ぁ-んァ-ヶ一-龯])', r'\1', text)
    # H-3c: 行頭「は、ため〜」「が、ため〜」（技術識別子除去で主語が消えたケース）
    text = _re.sub(r'^[、。\s]*(?:(?:は|が)[、，]?\s*)?ため(?=[ぁ-んァ-ヶ一-龯])', '', text)
    # K-B: 文中「。は、ため〜」「、が、ため〜」等のフラグメント除去（句点前までを保持）
    # 例: "VF。は、ため動作不全" → "VF。" / "処理実行。が、ため失敗" → "処理実行。"
    text = _re.sub(r'([。、])\s*(?:は|が|を)[、，]?\s*ため[ぁ-んァ-ヶ一-龯ーA-Za-z0-9]*(?=[。、]|$)', r'\1', text)
    # K-B: 矢印直後の裸助詞（を/が/は/で/と/へ）— 「→を使用」「→が発生」等。
    # 注意: 「に」は「→に遷移」等で意味語として機能するので除外
    text = _re.sub(r'→\s*(?:を|が|は|で|と|へ)(?=[ぁ-んァ-ヶ一-龯])', '→', text)
    text = _re.sub(r'^[\s、，。]+', '', text)
    text = _re.sub(r'[\s、，]+$', '', text)
    return text.strip()


def _deep_clean_ja(text: str) -> str:
    """業務テキスト用の総合クリーニング: 英語のうちに先行除去 → 日本語化 → 残渣除去。"""
    if not text:
        return text
    # Experience Cloud を保護
    text = text.replace("Experience Cloud", _EC_PLACEHOLDER)
    # 英語のうちに除去（日本語訳が混入すると正規表現が日本語始まりになり効かなくなるため）
    text = _RE_REL_FIELD.sub('', text)   # Contact__r.UserName 等
    text = _RE_DOT_CHAIN.sub('', text)   # Site.validatePassword 等
    text = _RE_URL_QUERY.sub('', text)   # ?appkbn=0/1 等
    # 日本語化
    text = _translate_sf_fields(text)
    text = _translate_sf_objects(text)
    text = _translate_jargon(text)
    # 残渣除去
    text = _strip_tech_identifiers(text)
    for pat, repl in _TECH_REPL_BIZ:
        text = pat.sub(repl, text)
    text = _strip_tech_identifiers(text)  # TECH_REPL_BIZ 適用後の残渣を再除去
    text = text.replace(_EC_PLACEHOLDER, "Experience Cloud")
    return text.strip()


def _normalize_process_steps(steps: list[dict]) -> None:
    """既存の process_steps をクリーニングする（in-place）。

    - description: API 名・メソッド呼び出し・拡張子付きファイル名を除去し日本語化
    - title: 空なら description から _short_title で短文生成。ある場合もクリーニング
    """
    for ps in steps:
        desc = ps.get("description", "") or ""
        title = ps.get("title", "") or ""
        if desc:
            desc = _deep_clean_ja(desc)
            ps["description"] = desc
        if title:
            ps["title"] = _deep_clean_ja(title)
        # title が空 or 貧弱（2文字未満）なら description から生成
        if desc and (not ps.get("title") or len(ps.get("title", "")) < 2):
            ps["title"] = _short_title(desc)


def _summarize_action(s: str) -> str:
    """action を 15 文字以内の一言ラベルに要約する。
    - 矢印（→）・区切り記号（／・（・、・。）の先頭節を採用
    - G-2: "Experience Cloud" を単一トークン扱い
    - K-A: キャッシュされた短い action（≤25字）でも、action と label を分離するため
      矢印・区切りで必ず先頭節のみ抽出して label を短くする"""
    if not s:
        return s
    # EC を一時的に短いプレースホルダに差し替えてから節選択・切断
    s_w = s.replace("Experience Cloud", _EC_PLACEHOLDER)
    # 矢印で切り落とし（「A → B」の先頭節）
    head = _re.split(r'\s*→\s*', s_w, 1)[0]
    # 区切り記号で切り落とし（「A／B」「A（B）」「A、B」「A。B」の先頭節）
    head = _re.split(r'[／（、。]', head, 1)[0].strip()
    # 長さ調整: 15 字超は 14+… で切る
    if len(head) > 15:
        head = head[:14] + '…'
    # 元の s が十分短く、かつ結果が同じ（区切りが無かった）場合は最後まで保持
    if not head:
        head = s_w[:15]
    return head.replace(_EC_PLACEHOLDER, "Experience Cloud")


def _normalize_business_flow(flows: list[dict]) -> None:
    """business_flow[] の actor / action / label をクリーニングする（in-place）。

    - actor: コンポーネント名（CamelCase + __c 等）は「システム」に置換
    - action: **丁寧な日本語の処理内容**。API 名・メソッド呼び出し・拡張子は除去するが 長さは保持
    - label: **図形ラベル用の一言まとめ（≤25字）**。action からの自動要約
    """
    _BF_TECH_ONLY = ("画面フロー", "フロー", "Apex", "Apexクラス", "Batch",
                     "Flow", "Trigger", "LWC", "Aura", "Visualforce",
                     "VF", "Integration", "Controller", "Service", "Handler")
    for s in flows:
        # description-only 形式の救済（LLM が process_steps[] と混同して description キーを使った場合）
        if not s.get("action") and s.get("description"):
            s["action"] = s["description"]
        if not s.get("actor"):
            s["actor"] = "業務担当者"
        actor = str(s.get("actor", "") or "").strip()
        # アクターが CamelCase クラス名風、__c 付き、（CMP-XXX）を含む等はコンポーネント名扱い → 「システム」に
        if (_RE_CAMEL_IDENT.search(actor) or _RE_LOWER_CAMEL.search(actor)
                or "__c" in actor or "（CMP" in actor or "(CMP" in actor):
            s["actor"] = "システム"

        act = str(s.get("action", "") or "").strip()
        if act in _BF_TECH_ONLY:
            actor2 = s.get("actor", "")
            if actor2 in ("自動フロー", "システム"):
                s["action"] = "処理を起動する"
            else:
                s["action"] = "画面から必要情報を入力し、処理を起動する"
        elif act:
            cleaned = _deep_clean_ja(act)
            # クリーニングで空・極短になった場合は actor に応じて定型に差し替え
            if not cleaned or len(cleaned) < 4:
                actor2 = s.get("actor", "")
                cleaned = ("処理を起動する" if actor2 in ("自動フロー", "システム")
                           else "画面から必要情報を入力し、処理を起動する")
            s["action"] = cleaned  # 丁寧な処理内容として保持（truncate しない）

        # label: 業務フロー図のノードラベル用に要約（すでに LLM が短く書いている場合はそのまま）
        lbl_src = s.get("label") or s.get("action", "")
        lbl_cleaned = _deep_clean_ja(str(lbl_src))
        s["label"] = _summarize_action(lbl_cleaned) if lbl_cleaned else ""


def _obj_label_from_api(api: str) -> str:
    """オブジェクトAPIから日本語ラベルを推定する（メタデータ優先）。"""
    if api in _STD_OBJ_LABELS:
        return _STD_OBJ_LABELS[api]
    if api in _SF_OBJ_LABELS:
        return _SF_OBJ_LABELS[api]
    raw = api.replace("__c", "").replace("__", "")
    base = _re.sub(r'([A-Z])', r' \1', raw).strip()
    return base


def _build_related_objects_and_access(data: dict) -> tuple[list[dict], list[dict]]:
    """components の inputs/outputs から related_objects と object_access を構築する。

    Returns: (related_objects, object_access)
    - related_objects: [{api_name, label, fields, relations}]
    - object_access:   [{component, object, operation}]
    """
    obj_comp_ops: dict[str, dict[str, str]] = {}   # obj_api → {comp_name → operation}
    obj_fields: dict[str, list[dict]] = {}          # obj_api → [field_dicts]

    def _register(obj_api: str, comp_name: str, op: str):
        obj_comp_ops.setdefault(obj_api, {})
        existing = obj_comp_ops[obj_api].get(comp_name, "")
        if not existing:
            obj_comp_ops[obj_api][comp_name] = op
        elif existing == "R" and op in ("W", "INSERT"):
            obj_comp_ops[obj_api][comp_name] = "RW"

    for comp in data.get("components", []):
        name = comp.get("api_name", "")

        # G-6: _SF_COMP_FIELDS（Apex/VF メタデータ由来）で User/Contact 等が紐付いていれば登録
        comp_key = _re.sub(r'（[^）]+）$', '', name).strip()
        meta_fields = _SF_COMP_FIELDS.get(comp_key, {}) or _SF_COMP_FIELDS.get(name, {})
        # K-C: Apex/VF パース時に推定した op を優先使用
        meta_ops = _SF_COMP_OPS.get(comp_key, {}) or _SF_COMP_OPS.get(name, {})
        for meta_obj in meta_fields.keys():
            if meta_obj in ("__any__", "Id"):
                continue
            # 優先1: メタデータ由来の op (DML / Site.* / VF save)
            est_op = meta_ops.get(meta_obj)
            if est_op:
                _register(meta_obj, name, est_op)
                continue
            # 優先2: outputs 文脈で「更新／登録／作成」言及があれば W、なければ R
            out_text = comp.get("outputs", "") or ""
            if _re.search(r'新規作成|新規登録|insert', out_text, _re.IGNORECASE):
                _register(meta_obj, name, "INSERT")
            elif _re.search(r'更新|登録|作成|保存|設定', out_text):
                _register(meta_obj, name, "W")
            else:
                _register(meta_obj, name, "R")

        # inputs → R（参照）。ただし trigger newList / trigger.new 文脈では INSERT
        for text in [comp.get("inputs", "")]:
            if not text:
                continue
            is_trigger_new = bool(_re.search(r'trigger\s+new|trigger\.new|newList', text, _re.IGNORECASE))
            for m in _re.finditer(r'(?<![A-Za-z0-9_])([A-Z][A-Za-z0-9]*__c)(?![A-Za-z0-9_])', text):
                _register(m.group(1), name, "INSERT" if is_trigger_new else "R")
            for std_api in _STD_OBJ_LABELS:
                if _re.search(rf'(?<![A-Za-z0-9_]){_re.escape(std_api)}(?![A-Za-z0-9_])', text):
                    _register(std_api, name, "R")

        # outputs → W または INSERT
        for text in [comp.get("outputs", "")]:
            if not text:
                continue

            # "OBJ更新（FIELD1・FIELD2）" → W + フィールド名抽出
            for m in _re.finditer(r'([A-Z][A-Za-z0-9]*__c)(?:[^\n（]*?)更新[^\n（]*?(?:（([^）]+)）)?', text):
                obj_api = m.group(1)
                _register(obj_api, name, "W")
                if m.group(2):
                    fnames = _re.findall(r'([A-Za-z][A-Za-z0-9]*__c)', m.group(2))
                    obj_fields.setdefault(obj_api, [])
                    for fn in fnames:
                        if not any(f["api_name"] == fn for f in obj_fields[obj_api]):
                            obj_fields[obj_api].append({
                                "api_name": fn,
                                "label": _sf_field_label(obj_api, fn),
                                "access": "W", "note": "",
                            })

            # "OBJ（insert）" や "insert" を含む → INSERT
            if _re.search(r'(?i)\binsert\b|新規作成|新規登録', text):
                for m in _re.finditer(r'(?<![A-Za-z0-9_])([A-Z][A-Za-z0-9]*__c)(?![A-Za-z0-9_])', text):
                    _register(m.group(1), name, "INSERT")
                for std_api in _STD_OBJ_LABELS:
                    if _re.search(rf'(?<![A-Za-z0-9_]){_re.escape(std_api)}(?![A-Za-z0-9_])', text):
                        _register(std_api, name, "INSERT")
            else:
                for m in _re.finditer(r'(?<![A-Za-z0-9_])([A-Z][A-Za-z0-9]*__c)(?![A-Za-z0-9_])', text):
                    _register(m.group(1), name, "W")
                for std_api in _STD_OBJ_LABELS:
                    if _re.search(rf'(?<![A-Za-z0-9_]){_re.escape(std_api)}(?![A-Za-z0-9_])', text):
                        _register(std_api, name, "W")

    # data_flow_overview から「ObjName（Field__c等フラグ更新）」形式のフィールド情報を抽出
    dfo = data.get("data_flow_overview", "")
    if dfo:
        # 標準オブジェクト: "Contact（IsConsignee__c等フラグ更新）" → Contact.IsConsignee__c = W
        for std_api in _STD_OBJ_LABELS:
            for m in _re.finditer(rf'{std_api}（([^）]{{1,120}})）', dfo):
                parens = m.group(1)
                fnames = _re.findall(r'([A-Za-z][A-Za-z0-9]*__c)', parens)
                is_w = bool(_re.search(r'更新|フラグ|設定|保存', parens))
                op = "W" if is_w else "R"
                for fn in fnames:
                    label = _sf_field_label(std_api, fn)
                    obj_fields.setdefault(std_api, [])
                    if not any(f["api_name"] == fn for f in obj_fields[std_api]):
                        obj_fields[std_api].append(
                            {"api_name": fn, "label": label, "access": op, "note": ""})
        # カスタムオブジェクト: "Quote__c更新（QuoteLinkId__c・...）" → 既存パターンと同様
        for m in _re.finditer(r'([A-Z][A-Za-z0-9]*__c)（([^）]+)）', dfo):
            obj_api = m.group(1)
            parens = m.group(2)
            fnames = _re.findall(r'([A-Za-z][A-Za-z0-9]*__c)', parens)
            is_w = bool(_re.search(r'更新|フラグ|設定|保存', parens))
            op = "W" if is_w else "R"
            for fn in fnames:
                label = _sf_field_label(obj_api, fn)
                obj_fields.setdefault(obj_api, [])
                if not any(f["api_name"] == fn for f in obj_fields[obj_api]):
                    obj_fields[obj_api].append(
                        {"api_name": fn, "label": label, "access": op, "note": ""})

    # 他テキストから追加オブジェクトを収集（標準オブジェクトのみ）
    # ※ カスタムオブジェクトは inputs/outputs から取得済み。__c のスキャンは
    #   フィールド名（IsConsignee__c等）を誤ってオブジェクトとして追加する恐れがあるため除外する
    purpose_text = data.get("processing_purpose", "")
    combined_extra = " ".join([dfo, purpose_text])
    comp_names_real = [c.get("api_name", "") for c in data.get("components", []) if c.get("api_name")]

    for std_api in _STD_OBJ_LABELS:
        pat = rf'(?<![A-Za-z0-9_]){_re.escape(std_api)}(?![A-Za-z0-9_])'
        if not _re.search(pat, combined_extra):
            continue
        if std_api in obj_comp_ops:
            continue

        # 操作種別を推定（INSERT > W > R の優先順位）
        _insert_ctx = rf'(?:{_re.escape(std_api)}.{{0,20}}(?:作成|登録|新規)|(?:作成|登録|新規).{{0,20}}{_re.escape(std_api)})'
        _update_ctx = rf'(?:{_re.escape(std_api)}.{{0,20}}(?:更新|変更|保存)|(?:更新|変更|保存).{{0,20}}{_re.escape(std_api)})'
        if _re.search(_insert_ctx, combined_extra):
            inferred_op = "INSERT"
        elif _re.search(_update_ctx, combined_extra):
            inferred_op = "W"
        else:
            inferred_op = "R"

        # 前後200文字以内に実コンポーネント名があれば紐付ける
        linked_comp = None
        for m in _re.finditer(pat, combined_extra):
            window = combined_extra[max(0, m.start() - 200): m.end() + 200]
            for cname in comp_names_real:
                if cname in window:
                    linked_comp = cname
                    break
            if linked_comp:
                break

        # 紐付け先が見つかればそのコンポーネントへ。なければテキスト検出として仮登録
        assigned_comp = linked_comp if linked_comp else "（テキスト検出）"
        obj_comp_ops[std_api] = {assigned_comp: inferred_op}

    # related_objects 構築
    related_objects = []
    for obj_api, comp_ops in obj_comp_ops.items():
        # オブジェクトレベルの合算アクセス種別（マトリクスと一致させる）
        # R/W/INSERT を全て独立に保持し、表示はそれらを「・」連結する
        all_ops = list(comp_ops.values())
        has_read   = any(o in ("R", "RW") for o in all_ops)
        has_write  = any(o in ("W", "RW") for o in all_ops)
        has_insert = any(o == "INSERT" for o in all_ops)
        parts: list[str] = []
        if has_read:   parts.append("参照")
        if has_write:  parts.append("更新")
        if has_insert: parts.append("新規作成")
        obj_combined_ja = "・".join(parts) if parts else ""
        # 後方互換用の単一キー（最も強い操作を代表値として残す）
        if has_insert:
            obj_combined_op = "INSERT"
        elif has_write and has_read:
            obj_combined_op = "RW"
        elif has_write:
            obj_combined_op = "W"
        elif has_read:
            obj_combined_op = "R"
        else:
            obj_combined_op = all_ops[0] if all_ops else ""

        fields = obj_fields.get(obj_api, [])

        if not fields:
            # Flow/Apexメタデータからフィールドを補完
            meta = _SF_FIELD_LABELS.get(obj_api, {})
            # G-5: 大小混在の重複（Username/username）を避けるため小文字キーで dedup。
            # 標準オブジェクトは正規 API 名（_STD_OBJ_FIELDS_FALLBACK/_STD_OBJ_CANONICAL_NAMES）を優先採用。
            canonical_names: dict[str, str] = {}  # lowercase → canonical api_name
            for std in _STD_OBJ_FIELDS_FALLBACK.get(obj_api, []):
                canonical_names[std["api_name"].lower()] = std["api_name"]
            for fapi in meta.keys():
                canonical_names.setdefault(fapi.lower(), fapi)
            for _low, _pascal in _STD_OBJ_CANONICAL_NAMES.get(obj_api, {}).items():
                canonical_names[_low] = _pascal

            seen_lower: set[str] = set()
            raw_fields: list[tuple[str, str]] = []
            for comp in data.get("components", []):
                raw_api = comp.get("api_name", "")
                comp_key = _re.sub(r'（[^）]+）$', '', raw_api).strip()
                comp_field_map = _SF_COMP_FIELDS.get(comp_key, {})
                for fapi in comp_field_map.get(obj_api, set()):
                    lower = fapi.lower()
                    if lower in seen_lower:
                        continue
                    seen_lower.add(lower)
                    canonical = canonical_names.get(lower, fapi)
                    raw_fields.append((canonical, comp_key))

            # 標準オブジェクトの場合はノイズフィルタ（主要フィールド + __c カスタム項目のみ）
            if obj_api in _STD_OBJ_FIELDS_FALLBACK:
                allowed = {f["api_name"].lower() for f in _STD_OBJ_FIELDS_FALLBACK[obj_api]}
                # objectTranslations にラベルがあるものも許可
                allowed.update(k.lower() for k in meta.keys())
                raw_fields = [(n, c) for (n, c) in raw_fields
                              if n.lower() in allowed or n.endswith("__c")]

            for fapi, _comp_key in raw_fields:
                label = meta.get(fapi) or _sf_field_label(obj_api, fapi)
                fields.append({"api_name": fapi, "label": label, "access": obj_combined_op, "note": ""})

        if not fields:
            # G-4/G-5: プレースホルダ行は挿入せず、_normalize_schema の最終段で
            # _STD_OBJ_FIELDS_FALLBACK / _SF_FIELD_LABELS から補完する
            fields = []
        else:
            # 既存フィールドの access もオブジェクト合算値に統一
            for f in fields:
                f["access"] = obj_combined_op

        related_objects.append({
            "api_name": obj_api,
            "label": _obj_label_from_api(obj_api),
            "fields": fields,
            "relations": [],
            "access_ja": obj_combined_ja,
        })

    # object_access 構築（テキスト検出の仮コンポーネントは除外）
    object_access = []
    for obj_api, comp_ops in obj_comp_ops.items():
        for comp_name, op in comp_ops.items():
            if comp_name and op and comp_name != "（テキスト検出）":
                object_access.append({"component": comp_name, "object": obj_api, "operation": op})

    return related_objects, object_access


def _hydrate_from_feature_groups(data: dict, project_dir: str) -> None:
    """feature_groups.yml から related_objects を補完する（JSON が空の場合）。

    FG 単位の設計書で、対象 FG の related_objects がプロジェクト定義に
    登録されているが JSON に含まれていない場合、最低限のレコード（api_name のみ）
    として追加する。fields は _build_related_objects_and_access が別途埋める。
    """
    if data.get("related_objects"):
        return
    feature_id = data.get("feature_id", "")
    if not feature_id or not feature_id.startswith("FG-"):
        return
    try:
        import yaml as _yaml
    except ImportError:
        return
    fg_path = Path(project_dir) / "docs" / ".sf" / "feature_groups.yml"
    if not fg_path.exists():
        return
    try:
        raw = _yaml.safe_load(fg_path.read_text(encoding="utf-8")) or []
        fg_data = raw.get("groups", []) if isinstance(raw, dict) else raw
    except Exception:
        return
    group = next((g for g in fg_data if isinstance(g, dict) and g.get("group_id") == feature_id), None)
    if not group:
        return
    obj_apis = group.get("related_objects") or []
    if not obj_apis:
        return
    seed = []
    for obj_api in obj_apis:
        seed.append({
            "api_name": obj_api,
            "label": _obj_label_from_api(obj_api),
            "fields": [],      # _build_related_objects_and_access が埋める or フォールバックで "—" 行
            "relations": [],
        })
    data["related_objects"] = seed
    # fields 補完のために access_ja は後段（_enrich_related_object_fields or _build_~）に委ねる


def _enrich_related_object_fields(data: dict) -> None:
    """既存の related_objects + object_access を保持したまま、fields のみ補完する。

    キャッシュされた related_objects がオブジェクト構成は正しいが fields が空のとき、
    object_access と SF メタデータ（_SF_COMP_FIELDS / _SF_FIELD_LABELS）から
    コンポーネント単位で関連フィールドを抽出して埋める。
    """
    oa = data.get("object_access") or []
    ro = data.get("related_objects") or []
    if not oa or not ro:
        return

    # obj_api → 合算アクセス種別（R/W/INSERT/RW）を再計算
    obj_ops: dict[str, dict[str, bool]] = {}
    obj_comps: dict[str, set[str]] = {}
    for entry in oa:
        obj_api = entry.get("object", "")
        comp = entry.get("component", "")
        op = (entry.get("operation", "") or "").upper()
        if not obj_api or not op:
            continue
        flags = obj_ops.setdefault(obj_api, {"R": False, "W": False, "INSERT": False})
        if op in ("R", "RW"):
            flags["R"] = True
        if op in ("W", "RW"):
            flags["W"] = True
        if op == "INSERT":
            flags["INSERT"] = True
        if comp:
            obj_comps.setdefault(obj_api, set()).add(comp)

    def _combined(flags: dict) -> tuple[str, str]:
        parts: list[str] = []
        if flags.get("R"):      parts.append("参照")
        if flags.get("W"):      parts.append("更新")
        if flags.get("INSERT"): parts.append("新規作成")
        ja = "・".join(parts) if parts else ""
        if flags.get("INSERT"):
            key = "INSERT"
        elif flags.get("W") and flags.get("R"):
            key = "RW"
        elif flags.get("W"):
            key = "W"
        elif flags.get("R"):
            key = "R"
        else:
            key = ""
        return key, ja

    for obj in ro:
        if obj.get("fields"):
            continue
        obj_api = obj.get("api_name", "")
        if not obj_api:
            continue
        flags = obj_ops.get(obj_api, {})
        op_key, op_ja = _combined(flags)

        # SF メタデータからフィールドを収集
        fields: list[dict] = []
        seen: set[str] = set()
        for comp_name in obj_comps.get(obj_api, set()):
            comp_field_map = _SF_COMP_FIELDS.get(comp_name, {})
            for fapi in comp_field_map.get(obj_api, set()):
                if fapi in seen:
                    continue
                seen.add(fapi)
                label = _sf_field_label(obj_api, fapi)
                fields.append({"api_name": fapi, "label": label, "access": op_key, "note": ""})

        if not fields:
            # X-4d: プレースホルダ注入を廃止。std fallback → label dict → 空配列の順に試す
            std_fields = _STD_OBJ_FIELDS_FALLBACK.get(obj_api)
            if std_fields:
                fields = [dict(f, access=op_key) for f in std_fields]
            else:
                labels = _SF_FIELD_LABELS.get(obj_api, {})
                if labels:
                    picks = [(k, v) for k, v in labels.items()
                             if k == "Name" or k.endswith("__c")][:8]
                    fields = [{"api_name": k, "label": v, "access": op_key, "note": ""}
                              for k, v in picks]
                # labels も空なら fields=[] のまま（xlsx 側が行スキップ）

        obj["fields"] = fields
        if op_ja and not obj.get("access_ja"):
            obj["access_ja"] = op_ja


def _infer_users(data: dict) -> str:
    """processing_purpose / data_flow_overview / components からユーザー/利用部門を推定する。"""
    comp_resp = " ".join(
        c.get("responsibility", "") for c in data.get("components", [])
    )
    combined = " ".join([
        data.get("processing_purpose", ""),
        data.get("data_flow_overview", ""),
        comp_resp,
    ])
    parts = []
    if _re.search(r'お客様|顧客|申請者|Experience Cloud|見込み客|問い合わせ者|HP問い合わせ|WebTo', combined):
        if _re.search(r'Experience Cloud|申請者', combined):
            parts.append("お客様（Experience Cloudユーザー）")
        else:
            parts.append("見込み客・問い合わせ者")
    if _re.search(r'管理者|GF社|担当者|事務|社内|内部', combined):
        parts.append("GF社担当者")
    if _re.search(r'コンサル', combined):
        parts.append("GF社コンサル部")
    if _re.search(r'営業', combined):
        parts.append("GF社営業部")
    if not parts:
        parts.append("GF社担当者")
    return "・".join(parts)


def _normalize_schema(data: dict) -> dict:
    """GFプロジェクト固有スキーマを標準スキーマに変換する。全テキストを日本語化する。"""
    # feature_id
    if not data.get("feature_id") and data.get("group_id"):
        data["feature_id"] = data["group_id"]

    # SFプロジェクトメタデータの自動ロード（project_name から判定、大小無視）
    proj = (data.get("project_name", "") or "").strip()
    if proj and not _SF_FIELD_LABELS:
        sf_path = _SF_PROJECT_PATHS.get(proj.lower())
        if sf_path:
            _load_sf_metadata(sf_path)

    # ── 概要フィールド（すべてAPI名・クラス名なしの日本語で） ────────────
    purpose_raw = data.get("processing_purpose", "")

    if not data.get("summary") and purpose_raw:
        # 最初の文のみ使用（2文目以降は技術実装詳細が多い）
        first_sent = purpose_raw.split("。")[0]
        data["summary"] = _clean_tech_business(first_sent)
    elif data.get("summary"):
        data["summary"] = _deep_clean_ja(data["summary"])  # G-extra: キャッシュ値も再クリーン

    if not data.get("purpose") and purpose_raw:
        # 全文をクリーンして目的とする（data_flow_overview は技術的すぎるので使わない）
        cleaned_purpose = _clean_tech_business(purpose_raw)
        if data.get("notes"):
            cleaned_notes = _clean_tech_business(data["notes"])
            if cleaned_notes:
                cleaned_purpose += f"\n【前提・補足】{cleaned_notes}"
        data["purpose"] = cleaned_purpose
    elif data.get("purpose"):
        data["purpose"] = _deep_clean_ja(data["purpose"])  # G-extra: キャッシュ値も再クリーン

    # 利用者推定
    if not data.get("users"):
        data["users"] = _infer_users(data)

    # 起点画面（screens[] → LWC → キーワード 推定）
    if not data.get("trigger_screen"):
        data["trigger_screen"] = _infer_trigger_screen(data)

    # 操作トリガー: prerequisites を日本語化してセット
    if not data.get("trigger") and data.get("prerequisites"):
        data["trigger"] = _clean_tech_business(data["prerequisites"])
    elif data.get("trigger"):
        data["trigger"] = _deep_clean_ja(data["trigger"])  # G-extra: キャッシュ値も再クリーン

    # H-2b: VF controller 属性から Apex コンポーネントを自動補強（防御層）
    _augment_components_with_vf_controllers(data)

    # components: responsibility → role（日本語化）& callees 初期化
    # M-5d: _gentle_clean_role（translate 系のみ、識別子削除しない）で role の破損を防ぐ
    # O-3②: _is_role_fragment で stale cache 断片を検出→ _STD_VF_DESCRIPTIONS / 型別 fallback でリセット
    for comp in data.get("components", []):
        api_name = comp.get("api_name", "")
        if not comp.get("role"):
            role_src = comp.get("responsibility", "")
            comp["role"] = _gentle_clean_role(role_src)
        else:
            comp["role"] = _gentle_clean_role(comp["role"])
        # 断片検出 → STD 辞書または型別 fallback で上書き
        if _is_role_fragment(comp["role"]):
            std = _STD_VF_DESCRIPTIONS.get(api_name)
            if std:
                comp["role"] = std
            else:
                # responsibility を再取得して gentle clean を試みる
                resp = comp.get("responsibility", "")
                if resp and not _is_role_fragment(_gentle_clean_role(resp)):
                    comp["role"] = _gentle_clean_role(resp)
                else:
                    comp_type = (comp.get("type") or "").lower()
                    apex_role = _apex_role_from_api_name(api_name) if "apex" in comp_type else ""
                    comp["role"] = apex_role or ""
        if "callees" not in comp:
            comp["callees"] = []

    # M-5c: callees を data_flow_overview から補強（ガード撤去: 常に実行してマージ）
    _infer_callees(data)

    # M-5a: process_steps は常に _build_process_steps で作り直す（キャッシュ破棄）
    # キャッシュに Phase H 以前の断片が残留しているため、責務から再構築する
    data["process_steps"] = _build_process_steps(data)

    # business_flow: 業務レベルフロー生成 or 既存をクリーニング
    _bf = data.get("business_flow") or []
    if not _bf:
        data["business_flow"] = _build_business_flow(data)
    else:
        _normalize_business_flow(_bf)

    # related_objects + object_access: components の inputs/outputs から構築
    # さらに既存 related_objects（hydrate 由来など）とマージしてフィールド補完する。
    rel_objs_auto, obj_access_auto = _build_related_objects_and_access(data)
    existing_ro = data.get("related_objects") or []

    # G-4: キャッシュ済のプレースホルダ行（api_name="—" / label に「別途設計書」）を除去
    # さらに大小混在の重複（Username / username 等）を canonical 名でまとめる
    _ro_canonical: dict[str, dict[str, str]] = {}  # obj_api → {lower → canonical}
    for _obj_api_c, _std_fs in _STD_OBJ_FIELDS_FALLBACK.items():
        _ro_canonical[_obj_api_c] = {f["api_name"].lower(): f["api_name"] for f in _std_fs}
    for _obj_api_c, _meta_labels in _SF_FIELD_LABELS.items():
        _ro_canonical.setdefault(_obj_api_c, {})
        for _k in _meta_labels.keys():
            _ro_canonical[_obj_api_c].setdefault(_k.lower(), _k)
    # 標準オブジェクトの canonical 名辞書を上書きマージ（VF の小文字参照対応）
    for _obj_api_c, _canon_map in _STD_OBJ_CANONICAL_NAMES.items():
        _ro_canonical.setdefault(_obj_api_c, {})
        for _low, _pascal in _canon_map.items():
            _ro_canonical[_obj_api_c][_low] = _pascal
    for o in existing_ro:
        api = o.get("api_name", "")
        cleaned: list[dict] = []
        seen_lower: set[str] = set()
        for f in (o.get("fields") or []):
            fapi = f.get("api_name") or ""
            # プレースホルダ除去
            if fapi in ("—", "", None):
                continue
            if "別途設計書" in (f.get("label") or "") or "別途設計書" in (f.get("note") or ""):
                continue
            if "（レコード新規登録）" in (f.get("label") or ""):
                continue
            # 大小混在の重複を除去し、canonical 名に正規化
            low = fapi.lower()
            if low in seen_lower:
                continue
            seen_lower.add(low)
            canonical = _ro_canonical.get(api, {}).get(low, fapi)
            if canonical != fapi:
                f = dict(f)
                f["api_name"] = canonical
                # label が元 API 名と同一なら再ラベルを試みる
                if (f.get("label") or "") in (fapi, canonical):
                    f["label"] = _sf_field_label(api, canonical)
            cleaned.append(f)
        o["fields"] = cleaned

    if existing_ro or rel_objs_auto:
        merged: dict[str, dict] = {}
        for o in existing_ro:
            api = o.get("api_name")
            if api:
                merged[api] = o
        for ao in rel_objs_auto:
            api = ao.get("api_name")
            if not api:
                continue
            if api in merged:
                cur = merged[api]
                # 既存エントリの fields が空なら自動生成のものを採用
                if not (cur.get("fields") or []) and ao.get("fields"):
                    cur["fields"] = ao["fields"]
                # access_ja / label は既存優先、空なら auto
                if not cur.get("access_ja") and ao.get("access_ja"):
                    cur["access_ja"] = ao["access_ja"]
                if not cur.get("label") and ao.get("label"):
                    cur["label"] = ao["label"]
            else:
                merged[api] = ao
        data["related_objects"] = list(merged.values())

    # K-C: キャッシュされた object_access があっても、ソースコード由来の W/INSERT を優先して反映する。
    # 既存エントリと obj_access_auto を (component, object) キーでマージし、
    # operation 優先度 INSERT > W > R で上書き。
    if not data.get("object_access"):
        data["object_access"] = obj_access_auto
    elif obj_access_auto:
        _op_priority = {"INSERT": 3, "W": 2, "R": 1, "": 0, None: 0}
        existing_map: dict[tuple[str, str], dict] = {}
        for entry in data["object_access"]:
            key = (entry.get("component", ""), entry.get("object", ""))
            existing_map[key] = entry
        for auto in obj_access_auto:
            key = (auto.get("component", ""), auto.get("object", ""))
            auto_op = auto.get("operation", "")
            if key in existing_map:
                cur = existing_map[key]
                cur_op = cur.get("operation", "")
                if _op_priority.get(auto_op, 0) > _op_priority.get(cur_op, 0):
                    cur["operation"] = auto_op
            else:
                data["object_access"].append(auto)
                existing_map[key] = auto
        # related_objects の access_ja / fields[].access も同様に再計算。
        # obj_access_auto 側で算出した obj['access_ja']（R+INSERT → 参照・新規作成 等）を優先採用
        auto_access_ja = {o.get("api_name"): o.get("access_ja")
                          for o in rel_objs_auto if o.get("access_ja")}
        auto_fields_access = {}  # {obj_api: {field_api: access}}
        for o in rel_objs_auto:
            auto_fields_access[o.get("api_name", "")] = {
                f.get("api_name", ""): f.get("access", "")
                for f in (o.get("fields") or [])
            }
        for o in data.get("related_objects", []):
            api = o.get("api_name", "")
            if api in auto_access_ja:
                o["access_ja"] = auto_access_ja[api]
            auto_fa = auto_fields_access.get(api, {})
            for f in (o.get("fields") or []):
                fapi = f.get("api_name", "")
                new_access = auto_fa.get(fapi)
                if new_access and _op_priority.get(new_access, 0) > _op_priority.get(f.get("access", ""), 0):
                    f["access"] = new_access

    # fields が空のオブジェクトを SF メタデータから補完
    if any(not (o.get("fields") or []) for o in data.get("related_objects", [])):
        _enrich_related_object_fields(data)

    # フォールバック: fields が空のオブジェクトは以下の順で補完
    # 1) SF ラベル辞書（objectTranslations／field-meta 由来）
    # 2) _STD_OBJ_FIELDS_FALLBACK（User/Contact 等の標準オブジェクト予備辞書）
    for obj in data.get("related_objects", []):
        if not (obj.get("fields") or []):
            obj_api = obj.get("api_name", "")
            access_default = "W" if "更新" in (obj.get("access_ja") or "") \
                             else ("R" if "参照" in (obj.get("access_ja") or "") else "R")
            labels = _SF_FIELD_LABELS.get(obj_api, {})
            picks_by_label: list[tuple[str, str]] = []
            if labels:
                picks_by_label = [(k, v) for k, v in labels.items()
                                  if k == "Name" or k.endswith("__c")][:8]
            if picks_by_label:
                obj["fields"] = [{"api_name": k, "label": v, "access": access_default, "note": ""}
                                 for k, v in picks_by_label]
            else:
                # G-5: 標準オブジェクト予備辞書
                std_fields = _STD_OBJ_FIELDS_FALLBACK.get(obj_api, [])
                if std_fields:
                    obj["fields"] = [{"api_name": f["api_name"],
                                      "label": f["label"],
                                      "access": access_default,
                                      "note": f.get("note", "")}
                                     for f in std_fields]
            if not obj.get("access_ja"):
                obj["access_ja"] = "参照"

    return data


# ── 差分計算 ────────────────────────────────────────────────────────
def _compute_diffs(prev_data: dict | None, new_data: dict) -> dict:
    if prev_data is None:
        return {"scalars": [], "lists": {}}
    return {
        "scalars": dr.diff_scalars(prev_data, new_data, SCALAR_FIELDS),
        "lists": {
            "business_flow": dr.diff_list(
                prev_data.get("business_flow", []),
                new_data.get("business_flow", []), "step"),
            "related_objects": dr.diff_list(
                prev_data.get("related_objects", []),
                new_data.get("related_objects", []), "api_name"),
            "process_steps": dr.diff_list(
                prev_data.get("process_steps", []),
                new_data.get("process_steps", []), "step"),
            "components": dr.diff_list(
                prev_data.get("components", []),
                new_data.get("components", []), "api_name"),
        },
    }


# ── PNG生成 ────────────────────────────────────────────────────────
def _generate_diagrams(data: dict, tmp_dir: str) -> dict[str, str | None]:
    """4種の図形PNGを生成し、パスを返す。失敗したらNone。"""
    import diagram_gen as dg

    paths: dict[str, str | None] = {
        "swimlane":  None,
        "er":        None,
        "flowchart": None,
        "component": None,
    }

    # 1. スイムレーン図（業務フロー）
    flows = data.get("business_flow", [])
    if flows:
        try:
            sl_path = str(Path(tmp_dir) / "swimlane.png")
            dg.render_swimlane(_business_flow_to_swimlane(flows), sl_path)
            paths["swimlane"] = sl_path
            print("  [OK] スイムレーン図")
        except Exception as e:
            print(f"  [WARN] スイムレーン図: {e}")

    # 2. オブジェクト参照マトリクス（対象オブジェクト）
    objects = data.get("related_objects", [])
    object_access = data.get("object_access", [])
    components = data.get("components", [])
    if objects:
        try:
            er_path = str(Path(tmp_dir) / "er.png")
            if object_access:
                from diagram_utils import generate_object_component_matrix
                # テキスト検出のみのオブジェクトはマトリクスから除外
                oa_obj_apis = {oa["object"] for oa in object_access}
                matrix_objects = [o for o in objects if o["api_name"] in oa_obj_apis]
                generate_object_component_matrix(
                    object_access, components, matrix_objects, er_path)
            else:
                # object_access がない場合は従来のER図にフォールバック
                from er_utils import generate_er_image
                boxes, arrows = _related_objects_to_er_boxes(objects)
                n_obj = len(boxes)
                er_w = min(14, max(8, n_obj * 3.0))
                er_h = min(10, max(6, n_obj * 2.0))
                generate_er_image(boxes, arrows, er_path,
                                  title="オブジェクト関連図",
                                  slide_w=er_w, slide_h=er_h)
            paths["er"] = er_path
            print("  [OK] オブジェクト参照マトリクス")
        except Exception as e:
            print(f"  [WARN] オブジェクト参照マトリクス: {e}")

    # 3. フローチャート（処理概要）
    steps = data.get("process_steps", [])
    if steps:
        try:
            fc_path = str(Path(tmp_dir) / "flowchart.png")
            dg.render_flowchart(steps, fc_path)
            paths["flowchart"] = fc_path
            print("  [OK] フローチャート")
        except Exception as e:
            print(f"  [WARN] フローチャート: {e}")

    # 4. コンポーネント図（関連コンポーネント）
    components = data.get("components", [])
    if components:
        try:
            cm_path = str(Path(tmp_dir) / "component.png")
            dg.render_component_diagram(components, cm_path, steps=steps)
            paths["component"] = cm_path
            print("  [OK] コンポーネント図")
        except Exception as e:
            print(f"  [WARN] コンポーネント図: {e}")

    return paths


def _business_flow_to_swimlane(flows: list[dict]) -> dict:
    """business_flow リストを render_swimlane 用のフロー dict に変換する。

    図形ノードは label を優先（丁寧な action を図に突っ込むと横長・読み難いため）。
    cross-lane 遷移（アクターが変わる遷移）は constraint=false で描画させ、
    LR モードでのランク強制によるレーン横並び（横長画像）を防ぐ。
    """
    lane_names = list(dict.fromkeys(f.get("actor", "不明") for f in flows))
    # actor→step の逆引き辞書（step ID → actor）
    step_to_actor: dict[str, str] = {str(f.get("step", i + 1)): f.get("actor", "不明")
                                      for i, f in enumerate(flows)}
    steps = [
        {"id": str(f.get("step", i + 1)), "lane": f.get("actor", "不明"),
         "label": f.get("label") or f.get("action", "")}
        for i, f in enumerate(flows)
    ]
    transitions = []
    for i, f in enumerate(flows):
        nexts = f.get("next", [])
        src = str(f.get("step", i + 1))
        src_actor = f.get("actor", "不明")
        if nexts:
            for n in nexts:
                dst_step = n.get("to") or (flows[i + 1].get("step", i + 2) if i + 1 < len(flows) else None)
                if dst_step:
                    dst_actor = step_to_actor.get(str(dst_step), src_actor)
                    cross = dst_actor != src_actor
                    transitions.append({"from": src, "to": str(dst_step),
                                        "condition": n.get("condition", ""), "cross": cross})
        elif i + 1 < len(flows):
            dst_step = str(flows[i + 1].get("step", i + 2))
            dst_actor = flows[i + 1].get("actor", "不明")
            cross = dst_actor != src_actor
            transitions.append({"from": src, "to": dst_step, "cross": cross})
    return {
        "title": "業務フロー",
        "lanes": [{"name": n} for n in lane_names],
        "steps": steps,
        "transitions": transitions,
    }


def _related_objects_to_er(objects: list[dict]) -> tuple[list, list]:
    """related_objects を render_er_diagram 用の (objects, relations) に変換する。"""
    obj_list = [
        {"api": o.get("api_name", ""), "label": o.get("label", ""),
         "type": "カスタム" if "__c" in o.get("api_name", "") else "標準"}
        for o in objects
    ]
    rels = []
    for o in objects:
        for r in o.get("relations", []):
            rels.append({
                "parent": o.get("api_name", ""),
                "rel":    r.get("type", "lookup"),
                "child":  r.get("to", ""),
                "field":  r.get("field", ""),
            })
    return obj_list, rels


def _related_objects_to_er_boxes(objects: list[dict]) -> tuple[list, list]:
    """related_objects を er_utils.generate_er_image 用の (boxes, arrows) に変換する。

    自動グリッドレイアウト: オブジェクトを横に並べ、収まらなければ折り返す。
    """
    import math

    n = len(objects)
    cols = min(n, 3)
    rows = math.ceil(n / cols)

    # レイアウト定数
    box_w = 3.0
    box_h_base = 1.2
    field_h = 0.32
    x_gap = 1.0
    y_gap = 1.0
    x_start = 1.0
    y_start = 1.5  # タイトルバー分

    boxes = []
    obj_id_map: dict[str, str] = {}  # api_name -> box id

    for i, obj in enumerate(objects):
        api = obj.get("api_name", "")
        label = obj.get("label", "")
        fk_fields = [f for f in obj.get("fields", []) if f.get("access") in ("RW", "W")]
        n_fields = min(len(fk_fields), 4)  # 表示上限4フィールド
        h = box_h_base + n_fields * field_h

        col_idx = i % cols
        row_idx = i // cols
        x = x_start + col_idx * (box_w + x_gap)
        y = y_start + row_idx * (box_h_base + 4 * field_h + y_gap)

        style = "primary" if "__c" in api else "secondary"
        box_id = api.replace("__c", "_c").replace("__", "_")
        obj_id_map[api] = box_id

        box_fields = []
        for fi, f in enumerate(fk_fields[:4]):
            box_fields.append({
                "name": f.get("api_name", ""),
                "label": f.get("label", ""),
                "is_fk": any(
                    r.get("to") == f.get("api_name", "")
                    or r.get("field", "") == f.get("api_name", "")
                    for r in obj.get("relations", [])
                ),
            })

        boxes.append({
            "id": box_id,
            "api": api,
            "label": label,
            "x": x, "y": y, "w": box_w, "h": h,
            "style": style,
            "fields": box_fields,
        })

    arrows = []
    for obj in objects:
        src_api = obj.get("api_name", "")
        src_id = obj_id_map.get(src_api, "")
        for rel in obj.get("relations", []):
            to_api = rel.get("to", "")
            dst_id = obj_id_map.get(to_api, "")
            if not dst_id:
                continue
            rel_type = rel.get("type", "lookup").lower()
            arrows.append({
                "from": src_id,
                "to": dst_id,
                "rel": rel_type,
                "field": rel.get("field", rel.get("label", "")),
            })

    return boxes, arrows


# ── メイン ──────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="詳細設計書 Excel 生成（新スキーマ対応）")
    parser.add_argument("--input",      required=True, help="詳細設計 JSON ファイルパス")
    parser.add_argument("--template",   required=True, help="詳細設計書テンプレート.xlsx パス")
    parser.add_argument("--output-dir", required=True, help="出力先ディレクトリ")
    parser.add_argument("--version-increment", default="minor",
                        choices=["minor", "major"])
    parser.add_argument("--source-hash", default="",
                        help="グループ内ソースのSHA256。_meta に保存して次回差分判定に使う")
    parser.add_argument("--author", default="",
                        help="作成者名。JSON の author が空の場合にフォールバックで使用")
    parser.add_argument("--force", action="store_true",
                        help="差分ゼロ・ソースハッシュ一致でも再生成する（ロジック改修検証用）")
    parser.add_argument("--project-dir", default="",
                        help="プロジェクトルート。指定時は docs/.sf/feature_groups.yml を参照して "
                             "related_objects 等を補完する")
    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text(encoding="utf-8"))

    # --project-dir が指定されていれば SF メタデータを先にロードし、
    # feature_groups.yml から related_objects を補完する
    if args.project_dir:
        try:
            _load_sf_metadata(args.project_dir)
        except Exception as _e:
            print(f"  [WARN] project_dir からのメタデータロード失敗: {_e}")
        _hydrate_from_feature_groups(data, args.project_dir)

    data = _normalize_schema(data)   # GFスキーマ → 標準スキーマ変換
    today  = _date.today().strftime("%Y-%m-%d")
    # --author が指定された場合は JSON 側に反映（改版履歴の変更者と表紙の両方に効く）
    if args.author:
        data["author"] = args.author
    author = data.get("author", "")

    feature_id = data.get("feature_id", "")
    name_ja    = data.get("name_ja", "機能")
    safe_name  = re.sub(r'[\\/:*?"<>|]', "_", name_ja)

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    set_project_tmp_dir(out_dir)
    # 他の設計書（プログラム設計・オブジェクト定義書）と統一: 【{feature_id}】{safe_name}_詳細設計.xlsx
    if feature_id:
        out_path = out_dir / f"【{feature_id}】{safe_name}_詳細設計.xlsx"
    else:
        out_path = out_dir / f"{safe_name}_詳細設計.xlsx"

    # ── バージョン判定 ────────────────────────────────────────────
    # feature_id ベースで既存ファイルを検出（機能名変更に追随できる）
    source_file = ""
    if feature_id:
        existing = sorted(out_dir.glob(f"【{feature_id}】*.xlsx"),
                          key=lambda f: f.stat().st_mtime, reverse=True)
        if existing:
            source_file = str(existing[0])
            print(f"  [AUTO] 既存ファイルを検出: {existing[0].name}")
    if not source_file:
        # 後方互換: 旧命名（FG-prefixなし）も探す
        existing = sorted(out_dir.glob(f"*{safe_name}*詳細設計*.xlsx"),
                          key=lambda f: f.stat().st_mtime, reverse=True)
        if existing:
            source_file = str(existing[0])
            print(f"  [AUTO] 既存ファイル（旧命名）を検出: {existing[0].name}")

    prev_meta   = read_meta(source_file) if source_file else None
    prev_data   = prev_meta.get("data") if prev_meta else None

    # 第1ゲート: ソースハッシュ一致ならLLM呼び出しもExcel再生成もスキップ（プログラム設計と同じ方式）
    if (prev_meta and args.source_hash
            and args.version_increment == "minor"
            and prev_meta.get("source_hash") == args.source_hash
            and not args.force):
        print("差分なし: ソースハッシュが既存ファイルと一致しているため更新をスキップしました")
        sys.exit(0)

    if prev_meta:
        prev_history_len = len(prev_meta.get("history", []))
        # 改版履歴 20 行制限 (Phase N+1): 既存履歴が 20 以上なら minor 指定でも major に強制昇格し履歴リセット
        forced_major = False
        if prev_history_len >= 20 and args.version_increment == "minor":
            print(f"  [WARN] 改版履歴が {prev_history_len} 件に達しているため minor → major に強制昇格し、履歴をリセットします")
            args.version_increment = "major"
            forced_major = True
        is_major = (args.version_increment == "major")
        current_version = increment_version(
            prev_meta.get("version", "1.0"), args.version_increment)
        # major 時は履歴リセット（手動・強制問わず。メジャーUP 1行だけ残す）
        history    = [] if is_major else prev_meta.get("history", [])
        is_initial = False
        if forced_major:
            print(f"メジャー昇格モード（履歴リセット）: {prev_meta.get('version', '?')} -> {current_version}")
        else:
            print(f"更新モード: {prev_meta.get('version', '?')} -> {current_version}")
    else:
        current_version = data.get("version") or "1.0"
        history    = []
        is_initial = True
        is_major   = (args.version_increment == "major")
        print(f"新規作成モード: v{current_version}")

    data["version"] = current_version
    if not data.get("date"):
        data["date"] = today

    # 第2ゲート: JSONフィールド差分ゼロならExcel再生成スキップ
    diffs = _compute_diffs(prev_data, data)
    if (prev_meta and args.version_increment == "minor"
            and not dr.has_any_diff(diffs) and not args.force):
        print("差分なし: 既存ファイルと一致しているため更新をスキップしました")
        sys.exit(0)

    last_no = max((h["項番"] for h in history
                   if isinstance(h.get("項番"), int)), default=0)
    new_entries = dr.build_entries(
        current_version, diffs, author, today,
        start_no=last_no + 1,
        is_major=is_major,
        is_initial=is_initial,
        section_sheet_map=SECTION_SHEETS,
        scalar_sheet="概要",
    )
    history = history + new_entries

    changed_scalars    = dr.changed_scalar_fields(diffs)
    changed_flows      = dr.changed_ids(diffs, "business_flow")
    changed_objs       = dr.changed_ids(diffs, "related_objects")
    changed_proc_steps = dr.changed_ids(diffs, "process_steps")
    changed_comps      = dr.changed_ids(diffs, "components")

    # ── 図形PNG生成 ──────────────────────────────────────────────
    with tempfile.TemporaryDirectory(dir=get_project_tmp_dir()) as tmp_dir:
        png_paths = _generate_diagrams(data, tmp_dir)

        # ── テンプレ読込 -> セル流し込み ──────────────────────────
        wb = load_workbook(args.template)

        fill_revision(wb["改版履歴"], data, history)

        fill_overview(
            wb["概要"], data,
            changed_fields=set() if is_major else changed_scalars)

        fill_business_flow(
            wb["業務フロー"], data,
            changed_step_nos=set() if is_major else changed_flows,
            png_path=png_paths.get("swimlane"))

        fill_process_overview(
            wb["処理概要"], data,
            changed_step_nos=set() if is_major else changed_proc_steps,
            png_path=png_paths.get("flowchart"))

        fill_target_objects(
            wb["対象オブジェクト"], data,
            changed_obj_keys=set() if is_major else changed_objs,
            png_path=png_paths.get("er"))

        fill_related_components(
            wb["関連コンポーネント"], data,
            changed_comp_keys=set() if is_major else changed_comps,
            png_path=png_paths.get("component"))

        meta_payload = {
            "version": current_version,
            "date":    today,
            "author":  author,
            "data":    data,
            "history": history,
        }
        if args.source_hash:
            meta_payload["source_hash"] = args.source_hash
        write_meta(wb, meta_payload)

        wb.save(str(out_path))

    sys.stdout.buffer.write(
        f"[OK] 詳細設計書を生成しました: v{current_version} -> {out_path}\n".encode("utf-8"))

    # 同一 feature_id の旧ファイルのみクリーンアップ（別名・旧命名ファイルは触らない）
    if feature_id:
        for old_f in out_dir.glob(f"【{feature_id}】*.xlsx"):
            if old_f.resolve() != out_path.resolve():
                old_f.unlink()
                sys.stdout.buffer.write(
                    f"  [CLEANUP] 旧ファイルを削除: {old_f.name}\n".encode("utf-8"))


if __name__ == "__main__":
    main()
