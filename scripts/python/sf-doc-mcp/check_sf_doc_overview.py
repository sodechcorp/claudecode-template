#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
プロジェクト概要書.xlsx の生成後 内容検証チェッカー（sf-doc-overview-writer Phase 2 post-check）。

generate_basic_doc.py は docs/ から決定論的に xlsx を組み立てる。ソースファイルが存在するのに
パース結果が空だと、対応シートは無警告のままフォールバック文言（「〜が見つかりません」等）で
生成される。本スクリプトは「ソースファイルは存在するのに対応シートがフォールバック文言のまま」
というギャップのみを検出する（ソース自体が無いケースはフォールバックが意図した挙動のため対象外）。

ソースが決定論的（LLM が内容を創作しない）であることを前提にした軽量チェックであり、
sf-design 系の check_design_json.py のような「記述品質」の判定は行わない。

Usage:
  python check_sf_doc_overview.py --docs-dir <path> --xlsx <path/to/プロジェクト概要書.xlsx>

終了コード:
  常に 0（非ブロッキング）。WARNING は標準出力に列挙するのみで生成フローを止めない。
  xlsx 自体が存在しない場合のみ 1（この場合は Phase 2 の exists() チェックで既に検出されるはず）。
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

# (シート名, フォールバック文言, ソースファイル群（いずれか1つでも存在すればチェック対象）, 表示ラベル)
_CHECKS = [
    ("システム概要", "（system.json が見つかりません）",
     ["architecture/system.json"], "システム全体構成図"),
    ("業務フロー図", "（フローデータなし）",
     ["flow/swimlanes.json"], "業務フロー図（As-Is/To-Be）"),
    ("ER図", "（カタログデータなし）",
     ["catalog/_index.md", "catalog/_data-model.md"], "ER図（オブジェクト関連図）"),
]


def _sheet_has_marker(ws, marker: str) -> bool:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and marker in cell.value:
                return True
    return False


def check(docs_dir: Path, xlsx_path: Path) -> list[str]:
    """WARNING メッセージのリストを返す（空リスト = 異常なし）。"""
    wb = load_workbook(str(xlsx_path), data_only=True)
    warnings: list[str] = []

    for sheet_name, marker, src_rel_paths, label in _CHECKS:
        existing_sources = [p for p in src_rel_paths if (docs_dir / p).exists()]
        if not existing_sources:
            continue  # ソース自体が無い → フォールバックは想定挙動

        if sheet_name not in wb.sheetnames:
            warnings.append(f"{label}: シート「{sheet_name}」が存在しません")
            continue

        ws = wb[sheet_name]
        if _sheet_has_marker(ws, marker):
            src_list = "、".join(existing_sources)
            warnings.append(
                f"{label}: ソース（{src_list}）は存在するが「{marker}」のまま生成されています"
                f"（パース結果が空の可能性）"
            )

    return warnings


def main():
    ap = argparse.ArgumentParser(description="プロジェクト概要書.xlsx 内容検証（非ブロッキング）")
    ap.add_argument("--docs-dir", required=True)
    ap.add_argument("--xlsx", required=True)
    args = ap.parse_args()

    docs_dir = Path(args.docs_dir)
    xlsx_path = Path(args.xlsx)

    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} が存在しません", file=sys.stderr)
        sys.exit(1)

    warnings = check(docs_dir, xlsx_path)

    if warnings:
        print(f"[WARN] 内容検証WARNING {len(warnings)}件（生成は継続されています。内容を確認してください）:")
        for w in warnings:
            print(f"  - {w}")
    else:
        print("OK: 内容検証で異常なし（ソース存在シートにフォールバック文言なし）")

    sys.exit(0)  # 非ブロッキング: 常に exit 0


if __name__ == "__main__":
    main()
