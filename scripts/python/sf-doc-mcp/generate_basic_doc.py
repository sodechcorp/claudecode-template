"""
プロジェクト概要書.xlsx を生成する（直接生成方式 + 高品質図埋め込み）。

入力 (docs/ 配下):
  docs/overview/org-profile.md         — 組織・プロジェクト基本情報・用語集
  docs/requirements/requirements.md    — 導入背景・目的
  docs/architecture/system.json        — システム構成図データ
  docs/flow/swimlanes.json             — 業務フロー（As-Is/To-Be）
  docs/catalog/_index.md               — オブジェクト一覧
  docs/catalog/_data-model.md          — オブジェクト関連定義（ER図）

出力:
  プロジェクト概要書.xlsx（5シート: 表紙/システム概要/業務フロー図/ER図/用語集）
  ※ 図は diagram_gen.py (graphviz/drawsvg) で高品質PNG生成して埋め込み

Usage:
  python generate_basic_doc.py \\
    --docs-dir <path/to/project/docs> \\
    --output <output/プロジェクト概要書.xlsx> \\
    --author "作成者名" \\
    [--project-name "プロジェクト名"]
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
from datetime import date
from pathlib import Path

# Graphviz on Windows is often installed but not in PATH — add the default location
_GV_BIN = r"C:\Program Files\Graphviz\bin"
if os.path.isdir(_GV_BIN) and _GV_BIN not in os.environ.get("PATH", ""):
    os.environ["PATH"] = _GV_BIN + os.pathsep + os.environ.get("PATH", "")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import diagram_gen as dg
from tmp_utils import get_project_tmp_dir, set_project_tmp_dir
from meta_store import read_meta, write_meta, compute_source_hash
from version_manager import increment_version

# ── デザインシステム（build_basic_doc_template.py と統一）────────
C_TITLE_DARK = "1F3864"
C_HDR_BLUE   = "2E75B6"
C_BAND_BLUE  = "0070C0"
C_LABEL_BG   = "D9E1F2"
C_FONT_W     = "FFFFFF"
C_FONT_D     = "000000"
FONT_NAME    = "游ゴシック"
THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")
GRID_LEFT  = 2
GRID_RIGHT = 31


def _fill(c):   return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def B_all(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def B_med(): return Border(left=MED,  right=MED,  top=MED,  bottom=MED)

def _set_h(ws, row, h): ws.row_dimensions[row].height = h
def _setup_grid(ws):
    ws.column_dimensions["A"].width = 2.0
    for i in range(GRID_LEFT, GRID_RIGHT + 1):
        ws.column_dimensions[get_column_letter(i)].width = 4.2
    ws.sheet_view.showGridLines = False

def _MW(ws, row, cs, ce, value="", bold=False, fg=C_FONT_D, bg=None,
        h="left", v="center", wrap=True, border=None, size=10):
    if border:
        for c in range(cs, ce + 1): ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1): ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    cell = ws.cell(row=row, column=cs, value=value)
    cell.font = _fnt(bold=bold, color=fg, size=size)
    cell.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg:     cell.fill = _fill(bg)
    if border: cell.border = border
    return cell

def _title_row(ws, row, text):
    _MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
        bold=True, fg=C_FONT_W, bg=C_TITLE_DARK,
        h="center", size=14, border=B_med())
    _set_h(ws, row, 28)
    return row + 1

def _section_row(ws, row, text):
    _MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
        bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, border=B_all())
    _set_h(ws, row, 18)
    return row + 1

def _sub_section_row(ws, row, text):
    _MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
        bold=True, fg=C_FONT_D, bg=C_LABEL_BG, border=B_all())
    _set_h(ws, row, 16)
    return row + 1

def _meta_row(ws, row, label, value="", col_label_end=8):
    _MW(ws, row, GRID_LEFT, col_label_end, label,
        bold=True, bg=C_LABEL_BG, border=B_all())
    _MW(ws, row, col_label_end + 1, GRID_RIGHT, value, border=B_all())
    _set_h(ws, row, 16)
    return row + 1

def _hdr_row(ws, row, cols: list[tuple[int, int, str]]):
    for cs, ce, label in cols:
        _MW(ws, row, cs, ce, label,
            bold=True, fg=C_FONT_W, bg=C_HDR_BLUE, h="center", border=B_all())
    _set_h(ws, row, 18)
    return row + 1

def _data_row(ws, row, cols_vals: list[tuple[int, int, str]], row_h=16):
    for cs, ce, val in cols_vals:
        _MW(ws, row, cs, ce, val or "", border=B_all())
    _set_h(ws, row, row_h)
    return row + 1

def _empty_rows(ws, row, count, col_groups: list[tuple[int, int]], row_h=16):
    for r in range(row, row + count):
        for cs, ce in col_groups:
            for c in range(cs, ce + 1):
                ws.cell(row=r, column=c).border = B_all()
            ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
        _set_h(ws, r, row_h)
    return row + count

def _text_area(ws, row, n_rows, value="", row_h=18):
    for r in range(row, row + n_rows):
        for c in range(GRID_LEFT, GRID_RIGHT + 1):
            ws.cell(row=r, column=c).border = B_all()
        _set_h(ws, r, row_h)
    ws.merge_cells(start_row=row, start_column=GRID_LEFT,
                   end_row=row + n_rows - 1, end_column=GRID_RIGHT)
    cell = ws.cell(row=row, column=GRID_LEFT, value=value)
    cell.font = _fnt()
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = B_all()
    return row + n_rows

def _margin(ws, row, h=6):
    _set_h(ws, row, h)
    return row + 1


# ── docs パーサー ─────────────────────────────────────────────────

def _table_val(text: str, keys) -> str:
    """表行 `| key | value |` から value を取り出す。
    keys は str または候補リスト。最初に見つかった値を返す。
    """
    if isinstance(keys, str):
        keys = [keys]
    for key in keys:
        m = re.search(rf'\|\s*{re.escape(key)}\s*\|\s*(.+?)\s*\|', text)
        if m:
            return m.group(1).strip()
    return ""


def _section_text(text: str, headings) -> str:
    """H2〜H4 見出しにキーワードを含むセクション本文を返す。
    headings は str または候補リスト。大文字小文字を無視し、最初にヒットしたものを返す。
    終端は**同じ階層以上の次見出し**（ヒットした見出しが H2 なら次の H1/H2 まで、
    H3 なら次の H1/H2/H3 まで）。つまり下位見出し（H3/H4）は本文に含める。
    """
    if isinstance(headings, str):
        headings = [headings]
    for h in headings:
        for level in (2, 3, 4):
            # 開始: 行頭から `#` が level 個、後ろは見出し以外の文字（空白・文字）
            # 終端: 改行＋`#` が 1〜level 個、その直後は `#` ではない見出しテキスト
            # f-string では {n,m} が tuple 展開されるため string 連結で組む
            start_pat = r'(?m)^#{' + str(level) + r'}\s[^\n]*' + re.escape(h) + r'[^\n]*\n'
            end_pat   = r'(?=\n#{1,' + str(level) + r'}[^#\n]|\Z)'
            pat = start_pat + r'(.*?)' + end_pat
            m = re.search(pat, text, re.DOTALL | re.IGNORECASE)
            if m:
                return m.group(1).strip()
    return ""


def _parse_stakeholder_table(sec: str) -> list[dict]:
    """ステークホルダー表を柔軟にパースする。
    - ヘッダ行検出: `役割`/`氏名`/`担当`/`会社`/`組織`/`メンバー`/`ベンダー` のいずれかを含む行
    - 区切り行（`---` のみ）はスキップ
    - データ行は 2〜5 列対応、最初の3列を role/name/area、4列目以降を note に寄せる
    """
    rows: list[dict] = []
    header_found = False
    header_kw = ("役割", "氏名", "担当", "会社", "組織", "メンバー", "ベンダー", "名前", "所属")
    for line in sec.splitlines():
        stripped = line.strip()
        if not stripped.startswith("|"):
            continue
        cols = [c.strip() for c in stripped.strip("|").split("|")]
        if len(cols) < 2:
            continue
        # 区切り行（`|---|---|`）はスキップ
        if all(re.fullmatch(r'[-:\s]+', c) for c in cols if c):
            continue
        if not header_found:
            if any(kw in c for c in cols for kw in header_kw):
                header_found = True
            continue
        role = cols[0]
        # 空行や連続区切り行は無視
        if not role or re.fullmatch(r'[-:\s]+', role):
            continue
        rows.append({
            "role": role,
            "name": cols[1] if len(cols) > 1 else "",
            "area": cols[2] if len(cols) > 2 else "",
            "note": " / ".join(c for c in cols[3:] if c) if len(cols) > 3 else "",
        })
    return rows


def parse_org(path: Path) -> dict:
    if not path.exists(): return {}
    t = path.read_text(encoding="utf-8")

    # 用語集
    # ヘッダー行を先にスキャンして列構成を検出する（2列: biz|desc, 3列: biz|sf|desc）
    glossary = []
    sec = _section_text(t, ["用語集", "Glossary"])
    _SF_HDR_WORDS = {"sf対応", "sf名称", "api名", "salesforce", "オブジェクト", "項目名", "対応項目"}
    _DESC_HDR_WORDS = {"説明", "補足", "備考", "description", "内容", "詳細"}
    biz_col, sf_col, desc_col = 0, 1, 2  # デフォルト: 3列想定
    _header_detected = False
    for line in sec.splitlines():
        if not line.strip().startswith("|"):
            continue
        cols = [c.strip() for c in line.strip().strip("|").split("|")]
        if not cols or not cols[0]:
            continue
        cols_lower = [c.lower() for c in cols]
        # 区切り行（`| --- |` 等）はスキップ
        if all(re.fullmatch(r'[-:\s]*', c) for c in cols):
            continue
        # ヘッダー行を検出（"業務用語"/"用語" 等が最初の列にある行）
        if not _header_detected and cols[0] in ("業務用語", "用語", "Term", "term"):
            _header_detected = True
            if len(cols) == 2:
                # 2列構成: biz | desc（SF対応列なし）
                sf_col, desc_col = -1, 1
            elif len(cols) >= 3:
                # 3列構成: 2列目が SF対応かどうかをヘッダーラベルで判定
                is_sf_col1 = any(w in cols_lower[1] for w in _SF_HDR_WORDS)
                is_desc_col1 = any(w in cols_lower[1] for w in _DESC_HDR_WORDS)
                if is_desc_col1 and not is_sf_col1:
                    sf_col, desc_col = -1, 1
                else:
                    sf_col, desc_col = 1, 2
            continue
        # データ行
        if len(cols) >= 2 and cols[0] and not re.fullmatch(r'[-:\s]+', cols[0]):
            sf_val = cols[sf_col] if sf_col >= 0 and sf_col < len(cols) else ""
            desc_val = cols[desc_col] if desc_col < len(cols) else ""
            glossary.append({"biz": cols[0], "sf": sf_val, "desc": desc_val})

    # 体制（セクション名の候補を横断探索）
    stakeholders: list[dict] = []
    for sec_name in ("ステークホルダーマップ", "担当ベンダー", "ステークホルダー",
                     "プロジェクト体制", "体制", "関係者", "プロジェクトメンバー", "チーム体制"):
        sec = _section_text(t, sec_name)
        if not sec:
            continue
        parsed = _parse_stakeholder_table(sec)
        if parsed:
            stakeholders = parsed
            break

    # 背景・目的（AS-IS課題 + TO-BE目的 を結合）
    bg_asis = _section_text(t, ["導入背景", "AS-IS課題", "AS-IS", "背景"])
    bg_tobe = _section_text(t, ["導入目的", "TO-BE", "目指す姿", "目的"])
    background = "\n\n".join(s for s in [bg_asis, bg_tobe] if s)

    system_name  = _table_val(t, ["システム名", "組織名", "会社名"])
    project_name = _table_val(t, ["プロジェクト名", "案件名"]) \
                   or (f"{system_name} Salesforce導入プロジェクト" if system_name else "")

    return {
        "project_name": project_name,
        "system_name":  system_name,
        "sf_edition":   _table_val(t, ["Salesforce Edition", "Edition", "エディション"]),
        "start_date":   _table_val(t, ["開始日", "プロジェクト開始日", "Salesforce利用開始", "初期導入時期"]),
        "end_date":     _table_val(t, ["終了予定日", "リリース予定日"]),
        "go_live_date": _table_val(t, ["本番公開日", "稼働日", "カットオーバー"]),
        "target_biz":   _table_val(t, ["対象業務", "業務領域"]),
        "stakeholders": stakeholders[:6],
        "glossary":     glossary[:30],
        "background":   background,
    }


def parse_requirements(path: Path) -> dict:
    if not path.exists(): return {}
    t = path.read_text(encoding="utf-8")

    # 背景（候補を横断探索）
    # 注: 「プロジェクト概要」は要件定義書でテーブル形式になりがちで本文が取れないため候補から除外
    bg = _section_text(t, ["背景・目的", "導入背景", "背景", "目的",
                            "AS-IS（現状）", "AS-IS", "導入前"])

    # スコープ（対象）: 見出し候補から探して表行 or 箇条書きから項目を抽出
    # 部分一致で「スコープ外」にもヒットしないよう、対象系を先に試す
    scope_in_sec = _section_text(t, ["対応スコープ", "対象スコープ", "1stステップ",
                                      "対象範囲", "スコープ定義", "対象業務", "対象"])
    # 見出し名に「対象外」系が紛れた場合のガード（最初の H2/H3 内だけを採用）
    if scope_in_sec:
        scope_in_sec = re.split(r'\n#{2,4}[^\n]*(?:対象外|スコープ外)', scope_in_sec)[0]
    scope_in_items = _extract_list_items(scope_in_sec)
    scope_in = "\n".join(f"・{i}" for i in scope_in_items) if scope_in_items else scope_in_sec[:300]

    # スコープ（対象外）: 候補から探す
    scope_out_sec = _section_text(t, ["対象外スコープ", "スコープ外", "2ndステップ",
                                       "非対象", "対象外"])
    scope_out_items = _extract_list_items(scope_out_sec)
    scope_out = "\n".join(f"・{i}" for i in scope_out_items) if scope_out_items else scope_out_sec[:300]

    return {"background": bg[:600], "scope_in": scope_in[:400], "scope_out": scope_out[:300]}


def _extract_list_items(sec: str) -> list[str]:
    """表行（最初の列）または箇条書き（-*・）から項目を抽出する。"""
    items: list[str] = []
    _SKIP = {"スコープ項目", "スコープ", "項目", "---", ""}
    for line in sec.splitlines():
        stripped = line.strip()
        if stripped.startswith("|"):
            cols = [c.strip() for c in stripped.strip("|").split("|")]
            if len(cols) >= 2 and cols[0] and cols[0] not in _SKIP \
                    and not re.fullmatch(r'[-:\s]+', cols[0]):
                items.append(cols[0])
        elif re.match(r'^[-*・]\s+', stripped):
            items.append(re.sub(r'^[-*・]\s+', '', stripped))
    return items


def parse_system_json(path: Path) -> dict:
    if not path.exists(): return {}
    try: return json.loads(path.read_text(encoding="utf-8"))
    except Exception: return {}


def parse_swimlanes(path: Path) -> dict:
    if not path.exists(): return {}
    try: return json.loads(path.read_text(encoding="utf-8"))
    except Exception: return {}


def _normalize_system_json(sys: dict) -> dict:
    """sf-analyst-cat1 生成の system.json（label/systems[] 形式）を
    diagram_gen.py が期待するスキーマ（name/core/external_systems[] 形式）に変換する。
    すでに core キーがある場合はそのまま返す。
    external_systems[] が直接形式で存在する場合（sf-org-analyst 生成）はそのまま使用する。"""
    if not sys or "core" in sys:
        return sys

    # external_systems[] 直接形式（sf-org-analyst 生成）の場合
    # systems[]/integrations[] からの再生成は行わず、既存値をそのまま使用する
    if sys.get("external_systems"):
        core = {"name": "Salesforce", "role": sys.get("description", "")}
        actors = [
            {"name": a.get("label", a.get("name", "")), "count": a.get("count", 0)}
            for a in sys.get("actors", [])
        ]
        data_stores = [
            {"name": ds.get("label", ds.get("name", "")), "purpose": ds.get("description", "")}
            for ds in sys.get("data_stores", [])
        ]
        return {**sys, "core": core, "actors": actors, "data_stores": data_stores}

    # systems[] から Salesforce を core として抽出
    core_sys = next(
        (s for s in sys.get("systems", []) if s.get("id") == "salesforce"), {}
    )
    core = {
        "name": core_sys.get("label", "Salesforce"),
        "role": core_sys.get("description", ""),
    }

    # actors[].label → name
    actors = [
        {"name": a.get("label", a.get("name", "")), "count": a.get("count", 0)}
        for a in sys.get("actors", [])
    ]

    # systems[] から外部システムマップを生成（salesforce 以外）
    ext_sys_map = {
        s["id"]: s
        for s in sys.get("systems", [])
        if s.get("id") and s.get("id") != "salesforce"
    }

    # integrations[] から方向・方式・頻度・用途を収集
    direction_map: dict[str, str] = {}
    proto_map: dict[str, str] = {}
    freq_map: dict[str, str] = {}
    purpose_map: dict[str, str] = {}
    for intg in sys.get("integrations", []):
        frm = intg.get("from", "")
        to = intg.get("to", "")
        method = intg.get("method", "")
        freq = intg.get("frequency", "")
        data = intg.get("data", "")[:60]
        if frm != "salesforce" and to == "salesforce":
            ext_id = frm
            direction_map[ext_id] = "both" if direction_map.get(ext_id) == "out" else "in"
            proto_map.setdefault(ext_id, method)
            freq_map.setdefault(ext_id, freq)
            purpose_map.setdefault(ext_id, data)
        elif frm == "salesforce" and to not in ("salesforce", ""):
            ext_id = to
            direction_map[ext_id] = "both" if direction_map.get(ext_id) == "in" else "out"
            proto_map.setdefault(ext_id, method)
            freq_map.setdefault(ext_id, freq)
            purpose_map.setdefault(ext_id, data)

    external_systems = [
        {
            "name": ext.get("label", ext.get("id", eid)),
            "direction": direction_map.get(eid, "out"),
            "protocol": proto_map.get(eid, ext.get("integration_method", "")),
            "frequency": freq_map.get(eid, ""),
            "purpose": purpose_map.get(eid, ext.get("description", "")[:60]),
        }
        for eid, ext in ext_sys_map.items()
    ]

    # data_stores[].label → name
    data_stores = [
        {"name": ds.get("label", ds.get("name", "")), "purpose": ds.get("description", "")}
        for ds in sys.get("data_stores", [])
    ]

    return {
        **sys,
        "core": core,
        "actors": actors,
        "external_systems": external_systems,
        "data_stores": data_stores,
    }


def _normalize_flow(flow: dict) -> dict:
    """sf-analyst-cat1 生成の swimlanes.json フロー（step/action/next:int 形式）を
    diagram_gen.py が期待するスキーマ（id/label/transitions[] 形式）に変換する。"""
    if not flow:
        return flow
    lanes_raw = flow.get("lanes", [])
    steps_in = flow.get("steps", [])

    # lane id → name 解決マップを構築（id 有無どちらにも対応）
    # lanes に id がある場合: {id: name} と {name: name} を両方登録
    # lanes に id がない場合: {name: name} のみ（name 直接参照で動作する）
    lane_resolve: dict[str, str] = {}
    for i, ll in enumerate(lanes_raw):
        name = ll.get("name", f"Lane{i+1}")
        if ll.get("id"):
            lane_resolve[ll["id"]] = name
        lane_resolve[name] = name

    # steps: step/action → id/label、lane を name に事前解決（diagram_gen に渡す前に確定）
    normalized_steps = []
    for s in steps_in:
        step_id = str(s.get("id", s.get("step", "")))
        label = (
            s.get("label")
            or s.get("title")
            or s.get("name")
            or s.get("action", "")
        )
        lane_raw = str(s.get("lane", ""))
        lane_resolved = lane_resolve.get(lane_raw, lane_raw)
        normalized_steps.append({**s, "id": step_id, "label": label, "lane": lane_resolved})

    # transitions: step.next（int または list）から自動生成
    transitions = list(flow.get("transitions", []))
    if not transitions:
        for s in steps_in:
            step_id = str(s.get("id", s.get("step", "")))
            next_val = s.get("next")
            if next_val is None:
                continue
            if isinstance(next_val, (int, str)):
                transitions.append({"from": step_id, "to": str(next_val)})
            elif isinstance(next_val, list):
                for dst in next_val:
                    if isinstance(dst, dict):
                        transitions.append({
                            "from": step_id,
                            "to": str(dst.get("to", "")),
                            "condition": dst.get("condition", ""),
                        })
                    else:
                        transitions.append({"from": step_id, "to": str(dst)})

    return {
        **flow,
        "title": flow.get("title") or flow.get("name", ""),
        "steps": normalized_steps,
        "transitions": transitions,
    }


def _clean_cell(s: str) -> str:
    """テーブルセルの値を正規化する。
    - `[text](url)` マークダウンリンクは `text` に剥がす
    - 前後のバッククォート・空白を除去
    """
    s = s.strip()
    m = re.match(r'^\[(.+?)\]\([^)]*\)$', s)
    if m:
        s = m.group(1)
    return s.strip().strip("`").strip()


def parse_catalog_index(path: Path) -> list[dict]:
    """
    _index.md をパースする。
    表の列順はプロジェクトによって異なる（`| API名 | ラベル | ...` または
    `| ラベル | API名 | ...`）ため、ヘッダ行のキーワードから
    api / label / type 列位置を**動的に検出**する。
    """
    if not path.exists(): return []
    t = path.read_text(encoding="utf-8")
    objs: list[dict] = []

    # 検出できるまでの既定列順（見つからなかった場合のフォールバック）
    api_col, label_col, type_col = 0, 1, 2
    header_found = False
    in_table = False
    _API_HDR   = {"api名", "api", "apiname", "オブジェクトapi"}
    _LABEL_HDR = {"ラベル", "表示名", "オブジェクト", "名称", "label"}
    _TYPE_HDR  = {"種別", "タイプ", "type", "オブジェクト種別"}

    for line in t.splitlines():
        stripped = line.strip()
        if not stripped.startswith("|"):
            in_table = False
            continue
        cols = [_clean_cell(c) for c in stripped.strip("|").split("|")]
        if len(cols) < 2:
            continue
        # 区切り行 `|---|---|` はスキップ
        if all(re.fullmatch(r'[-:\s]+', c) for c in cols if c):
            continue

        # ヘッダ行判定: 1行内で api/label キーワードの両方を含む
        cols_lower = [c.lower() for c in cols]
        has_api   = any(c in _API_HDR   for c in cols_lower)
        has_label = any(c in _LABEL_HDR for c in cols_lower)
        if has_api and has_label:
            for i, c in enumerate(cols_lower):
                if c in _API_HDR:   api_col   = i
                elif c in _LABEL_HDR: label_col = i
                elif c in _TYPE_HDR:  type_col  = i
            header_found = True
            in_table = True
            continue
        if not in_table:
            # ヘッダ前の行や表外は一旦スキップ
            continue

        api   = cols[api_col]   if api_col   < len(cols) else ""
        label = cols[label_col] if label_col < len(cols) else ""
        type_ = cols[type_col]  if type_col  < len(cols) else ""
        # API名は英数+アンダースコア。日本語が入っていたら列順の誤認 → スキップ
        if not api or not re.fullmatch(r'[A-Za-z][\w]*', api):
            continue
        objs.append({"api": api, "label": label, "type": type_})

    # ヘッダ行が一度も見つからない旧仕様カタログへのフォールバック
    if not header_found:
        _SKIP = {"オブジェクト", "API名", "---", ""}
        for line in t.splitlines():
            stripped = line.strip()
            if not stripped.startswith("|"): continue
            cols = [_clean_cell(c) for c in stripped.strip("|").split("|")]
            if len(cols) < 2: continue
            label = cols[0]
            api   = cols[1]
            if not api or label in _SKIP or re.fullmatch(r'[-\s:]+', label):
                continue
            if not re.fullmatch(r'[A-Za-z][\w]*', api):
                continue
            objs.append({"api": api, "label": label,
                         "type": cols[2] if len(cols) > 2 else ""})

    # 重複除去（api をキーに、最初のエントリ優先）
    seen = set()
    unique: list[dict] = []
    for o in objs:
        if o["api"] in seen: continue
        seen.add(o["api"])
        unique.append(o)
    return unique[:100]


def parse_data_model(path: Path) -> list[dict]:
    """
    _data-model.md から関連定義をパースする。
    Mermaid ERD 形式（||--o{ など）を優先し、なければ親/子テーブル形式を試みる。
    """
    if not path.exists(): return []
    t = path.read_text(encoding="utf-8")
    rels = []

    # ── Mermaid ERD 形式: "ObjectA ||--o{ ObjectB : "label"" ──
    # 実線（--）・点線（..）のどちらも受理。右端は `o{` `|{` `}o` `o|` `||` 等の多様な表記に対応。
    mermaid_pat = re.compile(
        r'^\s*([\w]+)\s+[|}o][|o][-.]{1,2}[|o][|o{]\s+([\w]+)\s*:\s*"([^"]*)"',
        re.MULTILINE,
    )
    for m in mermaid_pat.finditer(t):
        parent, child, label = m.group(1), m.group(2), m.group(3)
        rel_type = "master-detail" if ("MD" in label or "主従" in label) else "1-N"
        rels.append({"parent": parent, "rel": rel_type, "child": child, "field": label})
    if rels:
        return rels[:30]

    # ── フォールバック: 親オブジェクト/子オブジェクト列のテーブル ──
    in_rel_table = False
    col_parent = col_child = col_rel = col_field = -1
    for line in t.splitlines():
        stripped = line.strip()
        if not stripped.startswith("|"):
            in_rel_table = False
            continue
        cols = [c.strip() for c in stripped.strip("|").split("|")]
        if "親オブジェクト" in cols and "子オブジェクト" in cols:
            col_parent = cols.index("親オブジェクト")
            col_child  = cols.index("子オブジェクト")
            col_rel    = next((i for i, c in enumerate(cols)
                               if c in ("関係", "リレーション", "種別", "rel")), -1)
            col_field  = next((i for i, c in enumerate(cols)
                               if "項目" in c or "field" in c.lower()), -1)
            in_rel_table = True
            continue
        if not in_rel_table:
            continue
        if all(re.fullmatch(r'[-: ]+', c) for c in cols if c):
            continue
        parent = cols[col_parent] if col_parent < len(cols) else ""
        child  = cols[col_child]  if col_child  < len(cols) else ""
        if parent and child and not parent.isdigit() and not child.isdigit():
            rels.append({"parent": parent, "rel": "", "child": child, "field": ""})

    return rels[:20]


def _pick_flows(swimlanes: dict) -> tuple[dict | None, dict | None]:
    """swimlanes.json から As-Is / To-Be フローを抽出する。

    優先順位:
      1. flow_type が `asis` / `tobe` で明示宣言されたフロー
      2. title に As-Is / To-Be 相当のキーワードを含むフロー
      3. overall フロー（全体俯瞰）のフォールバック
         - As-Is が未設定 → overall を As-Is に流用
         - As-Is が明示登録されている & To-Be が未設定 → overall を To-Be に流用
           （as-is=過去、overall=現行=to-be 相当 の意味論に基づく）
    """
    flows = swimlanes.get("flows", [])
    asis = tobe = overall = None
    for f in flows:
        ftype = (f.get("flow_type") or "").lower()
        title = (f.get("title") or "").lower()
        is_asis = (ftype == "asis" or "as-is" in title or "asis" in title or "現状" in title or "導入前" in title)
        is_tobe = (ftype == "tobe" or "to-be" in title or "tobe" in title or "目指す姿" in title or "導入後" in title)
        if is_asis and asis is None:
            asis = f
        elif is_tobe and tobe is None:
            tobe = f
        elif ftype == "overall" and overall is None:
            overall = f
    # As-Is が見つからない場合のみ overall を As-Is 枠へ流用
    if asis is None and overall is not None:
        asis = overall
    # As-Is は明示登録されているが To-Be だけ未登録の場合、overall を To-Be に流用
    # （As-Is に既に overall を当てている場合は重複流用しない）
    elif tobe is None and overall is not None and asis is not overall:
        tobe = overall
    return asis, tobe


# ── シート 1: 表紙 ─────────────────────────────────────────────
def _build_cover(ws, org: dict, req: dict, author: str,
                 history: list | None = None):
    _setup_grid(ws)
    r = 2  # row 1 は余白
    _set_h(ws, 1, 8)
    r = _title_row(ws, r, "プロジェクト概要書")
    r = _margin(ws, r)
    r = _section_row(ws, r, "プロジェクト基本情報")
    # 値がある行のみ表示（空白行は省略）
    _REQUIRED = {"project_name", "system_name", "background", "scope_in"}
    for label, src, key in [
        ("プロジェクト名",     org, "project_name"),
        ("システム名",        org, "system_name"),
        ("導入目的・背景",     req, "background"),
        ("スコープ（対象）",   req, "scope_in"),
        ("スコープ（対象外）", req, "scope_out"),
        ("開始日",          org, "start_date"),
        ("終了予定日",       org, "end_date"),
        ("本番公開日",       org, "go_live_date"),
    ]:
        val = src.get(key, "")
        if not val and key not in _REQUIRED:
            continue  # 値がない任意項目はスキップ
        r = _meta_row(ws, r, label, val)

    r = _margin(ws, r)
    r = _section_row(ws, r, "体制")
    # データにある列のみ表示（空の担当領域・備考は省略）
    stk = org.get("stakeholders", [])
    has_area = any(s.get("area") for s in stk)
    has_note = any(s.get("note") for s in stk)
    if has_area and has_note:
        r = _hdr_row(ws, r, [(2,6,"役割"),(7,16,"氏名 / 組織"),(17,22,"担当領域"),(23,31,"備考")])
        for s in stk:
            r = _data_row(ws, r, [(2,6,s["role"]),(7,16,s["name"]),(17,22,s.get("area","")),
                                   (23,31,s.get("note",""))])
        if len(stk) < 6:
            r = _empty_rows(ws, r, 6 - len(stk), [(2,6),(7,16),(17,22),(23,31)])
    elif has_area:
        r = _hdr_row(ws, r, [(2,6,"役割"),(7,19,"氏名 / 組織"),(20,31,"担当領域")])
        for s in stk:
            r = _data_row(ws, r, [(2,6,s["role"]),(7,19,s["name"]),(20,31,s.get("area",""))])
        if len(stk) < 6:
            r = _empty_rows(ws, r, 6 - len(stk), [(2,6),(7,19),(20,31)])
    else:
        r = _hdr_row(ws, r, [(2,8,"役割"),(9,31,"氏名 / 組織")])
        for s in stk:
            r = _data_row(ws, r, [(2,8,s["role"]),(9,31,s["name"])])
        if len(stk) < 6:
            r = _empty_rows(ws, r, 6 - len(stk), [(2,8),(9,31)])

    r = _margin(ws, r)
    r = _section_row(ws, r, "改版履歴")
    r = _hdr_row(ws, r, [(2,3,"版"),(4,7,"改版日"),(8,12,"改版者"),(13,31,"改版内容")])
    today = date.today().strftime("%Y-%m-%d")
    if not history:
        history = [{"version": "1.0", "date": today, "author": author,
                    "content": "初版作成"}]
    for entry in history:
        r = _data_row(ws, r, [
            (2, 3,  entry.get("version", "")),
            (4, 7,  entry.get("date", "")),
            (8, 12, entry.get("author", "")),
            (13, 31, entry.get("content", "")),
        ])
    remaining = max(0, 10 - len(history))
    if remaining:
        r = _empty_rows(ws, r, remaining, [(2,3),(4,7),(8,12),(13,31)])


# ── シート 2: システム概要 ────────────────────────────────────────
def _build_system_overview(ws, req: dict, system: dict, sys_img_path: str | None):
    _setup_grid(ws)
    r = 2
    _set_h(ws, 1, 8)
    r = _title_row(ws, r, "システム概要")
    r = _margin(ws, r)

    # 導入背景・課題
    r = _section_row(ws, r, "導入背景・解決する課題")
    bg_text = req.get("background", "")
    n_bg_rows = max(5, len(bg_text) // 60 + 2)
    r = _text_area(ws, r, n_bg_rows, bg_text)
    r = _margin(ws, r)

    # 外部連携先一覧（表を先に、図は後）
    r = _section_row(ws, r, "外部連携先一覧")
    r = _hdr_row(ws, r, [(2,8,"連携先システム"),(9,12,"方向"),(13,18,"方式"),
                          (19,22,"頻度"),(23,31,"目的・概要")])
    for ex in system.get("external_systems", [])[:10]:
        r = _data_row(ws, r, [
            (2,8,ex.get("name","")), (9,12,ex.get("direction","")),
            (13,18,ex.get("protocol","")), (19,22,ex.get("frequency","")),
            (23,31,ex.get("purpose","")),
        ])
    if len(system.get("external_systems", [])) < 5:
        r = _empty_rows(ws, r, 5 - len(system.get("external_systems", [])),
                        [(2,8),(9,12),(13,18),(19,22),(23,31)])
    r = _margin(ws, r)

    # システム全体構成図
    r = _section_row(ws, r, "システム全体構成")
    if sys_img_path:
        n_img_rows = dg.embed_image_in_sheet(ws, sys_img_path, anchor_row=r,
                                             anchor_col=GRID_LEFT, max_width_px=1800)
        r += n_img_rows
    else:
        r = _text_area(ws, r, 15, "（system.json が見つかりません）")


# ── シート 3: 業務フロー図 ────────────────────────────────────────
def _build_flow_sheet(ws, asis_flow: dict | None, tobe_flow: dict | None,
                      asis_img: str | None, tobe_img: str | None):
    _setup_grid(ws)
    r = 2
    _set_h(ws, 1, 8)
    r = _title_row(ws, r, "業務フロー図")
    r = _margin(ws, r)

    for label, flow, img_path in [
        ("As-Is 業務フロー（現状）",             asis_flow, asis_img),
        ("To-Be 業務フロー（Salesforce導入後）", tobe_flow, tobe_img),
    ]:
        r = _section_row(ws, r, label)

        # 手順テーブルを先に（表が図に埋もれないよう）
        sub = label.split("（")[0]
        r = _sub_section_row(ws, r, f"{sub} 手順")
        r = _hdr_row(ws, r, [(2,3,"No"),(4,8,"担当"),(9,22,"操作・処理内容"),(23,31,"分岐条件")])
        steps = (flow or {}).get("steps", [])[:20]
        # ステップIDごとの遷移条件を収集（分岐が発生する場合に備考欄へ表示）
        step_conds: dict[str, list[str]] = {}
        for tr in (flow or {}).get("transitions", []):
            src = str(tr.get("from", ""))
            cond = tr.get("condition", "")
            if cond and src:
                step_conds.setdefault(src, []).append(cond)
        for i, s in enumerate(steps):
            sid = str(s.get("id",""))
            cond_text = " / ".join(step_conds.get(sid, []))
            r = _data_row(ws, r, [
                (2,3,sid),
                (4,8,str(s.get("lane",""))),
                (9,22,str(s.get("label","") or s.get("title",""))),
                (23,31,cond_text),
            ], row_h=18)
        if len(steps) < 5:
            r = _empty_rows(ws, r, 5 - len(steps), [(2,3),(4,8),(9,22),(23,31)], row_h=18)
        r = _margin(ws, r)

        # フロー図（表の後）
        r = _sub_section_row(ws, r, f"{sub} フロー図")
        if img_path:
            n_img_rows = dg.embed_image_in_sheet(ws, img_path, anchor_row=r,
                                                 anchor_col=GRID_LEFT, max_width_px=1100)
            r += n_img_rows
        else:
            r = _text_area(ws, r, 12, "（フローデータなし）")
        r = _margin(ws, r, 12)


# ── シート 4: ER図 ──────────────────────────────────────────────
def _build_er_sheet(ws, objects: list, relations: list, er_img: str | None):
    _setup_grid(ws)
    r = 2
    _set_h(ws, 1, 8)
    r = _title_row(ws, r, "ER図（オブジェクト関連図）")
    r = _margin(ws, r)

    # 関連定義表を先に（表が図に埋もれないよう）
    r = _section_row(ws, r, "関連定義表")
    r = _hdr_row(ws, r, [(2,8,"親オブジェクト"),(9,11,"種別"),
                          (12,18,"子オブジェクト"),(19,31,"関係・用途")])
    for rel in relations[:25]:
        r = _data_row(ws, r, [
            (2,8,rel["parent"]), (9,11,rel["rel"]),
            (12,18,rel["child"]), (19,31,rel.get("field","")),
        ])
    if len(relations) < 5:
        r = _empty_rows(ws, r, 5 - len(relations),
                        [(2,8),(9,11),(12,18),(19,31)])
    r = _margin(ws, r)

    # ER図（表の後、max_width 拡大で可読性向上）
    r = _section_row(ws, r, "オブジェクト関連図")
    if er_img:
        n_img_rows = dg.embed_image_in_sheet(ws, er_img, anchor_row=r,
                                             anchor_col=GRID_LEFT, max_width_px=1800)
        r += n_img_rows
    else:
        r = _text_area(ws, r, 18, "（カタログデータなし）")


# ── シート 5: 用語集 ────────────────────────────────────────────
def _build_glossary_sheet(ws, glossary: list):
    _setup_grid(ws)
    r = 2
    _set_h(ws, 1, 8)
    r = _title_row(ws, r, "用語集")
    r = _margin(ws, r)
    r = _section_row(ws, r, "業務用語・Salesforce用語 対照表")
    r = _hdr_row(ws, r, [(2,3,"No"),(4,10,"業務用語"),
                          (11,18,"Salesforce用語 / オブジェクト名"),(19,31,"説明")])
    for i, t in enumerate(glossary):
        r = _data_row(ws, r, [(2,3,str(i+1)),(4,10,t["biz"]),
                               (11,18,t["sf"]),(19,31,t["desc"])], row_h=18)
    if len(glossary) < 10:
        r = _empty_rows(ws, r, 10 - len(glossary), [(2,3),(4,10),(11,18),(19,31)], row_h=18)


# ── メイン ──────────────────────────────────────────────────────
def generate(docs_dir: Path, output: Path, author: str, project_name: str,
             source_file: str = "", version_increment: str = "minor"):
    org      = parse_org(docs_dir / "overview" / "org-profile.md")
    req      = parse_requirements(docs_dir / "requirements" / "requirements.md")
    system     = _normalize_system_json(
                     parse_system_json(docs_dir / "architecture" / "system.json"))
    swim_raw   = parse_swimlanes(docs_dir / "flow" / "swimlanes.json")
    swim       = ({**swim_raw,
                   "flows": [_normalize_flow(f) for f in swim_raw.get("flows", [])]}
                  if swim_raw else {})
    objects  = parse_catalog_index(docs_dir / "catalog" / "_index.md")
    relations = parse_data_model(docs_dir / "catalog" / "_data-model.md")

    # ── 改版履歴: source-file から prev_meta を読み込み、source_hash 比較で差分判定 ──
    today = date.today().strftime("%Y-%m-%d")
    src_paths = [
        docs_dir / "overview"     / "org-profile.md",
        docs_dir / "requirements" / "requirements.md",
        docs_dir / "architecture" / "system.json",
        docs_dir / "catalog"      / "_index.md",
        docs_dir / "catalog"      / "_data-model.md",
        docs_dir / "flow"         / "swimlanes.json",
    ]
    new_hash = compute_source_hash([str(p) for p in src_paths])
    prev_meta = read_meta(source_file) if source_file else None

    if prev_meta:
        prev_history = list(prev_meta.get("history", []))
        prev_version = prev_meta.get("version", "1.0")
        prev_hash    = prev_meta.get("source_hash", "")

        if prev_hash == new_hash and version_increment == "minor":
            current_version = prev_version
            history         = prev_history
            print(f"  [INFO] 差分なし（source_hash 一致）: v{prev_version} 据え置き")
        else:
            # 改版履歴 20 行制限: 既存履歴が 20 以上なら minor 指定でも major 強制昇格
            forced_major = False
            if len(prev_history) >= 20 and version_increment == "minor":
                print(f"  [WARN] 改版履歴が {len(prev_history)} 件に達しているため "
                      f"minor → major に強制昇格し履歴をリセットします")
                version_increment = "major"
                forced_major = True
            is_major = (version_increment == "major")
            current_version = increment_version(prev_version, version_increment)
            history = [] if is_major else prev_history
            next_no = (max((h.get("no", 0) for h in history), default=0) + 1)
            content = ("メジャーバージョンアップ（履歴リセット）"
                       if is_major else "概要情報更新")
            history.append({
                "no":      next_no,
                "version": current_version,
                "date":    today,
                "sheet":   "全シート",
                "content": content,
                "author":  author,
            })
            print(f"  [INFO] 更新: v{prev_version} → v{current_version}"
                  + ("（メジャー昇格・履歴リセット）" if forced_major else
                     "（メジャー）" if is_major else ""))
    else:
        current_version = "1.0"
        history = [{
            "no": 1, "version": "1.0", "date": today,
            "sheet": "全シート", "content": "新規作成", "author": author,
        }]
        print("  [INFO] 新規作成モード: v1.0")

    if project_name:
        org["project_name"] = project_name

    # requirements.mdに背景がなければorg-profile.mdの背景を使う
    if not req.get("background") and org.get("background"):
        req["background"] = org["background"]

    asis_flow, tobe_flow = _pick_flows(swim)

    # 図を一時ディレクトリに生成
    set_project_tmp_dir(output)
    with tempfile.TemporaryDirectory(dir=get_project_tmp_dir()) as tmpdir:
        tmp = Path(tmpdir)

        sys_img = er_img = asis_img = tobe_img = None

        if system:
            try:
                dg.render_system_diagram(system, str(tmp / "system.png"))
                sys_img = str(tmp / "system.png")
                print("  [OK] システム構成図")
            except Exception as e:
                print(f"  [WARN] システム構成図: {e}")

        if objects and relations:
            try:
                dg.render_er_diagram(objects, relations, str(tmp / "er.png"))
                er_img = str(tmp / "er.png")
                print("  [OK] ER図")
            except Exception as e:
                print(f"  [WARN] ER図: {e}")

        if asis_flow:
            try:
                dg.render_swimlane(asis_flow, str(tmp / "asis.png"))
                asis_img = str(tmp / "asis.png")
                print("  [OK] As-Is フロー")
            except Exception as e:
                print(f"  [WARN] As-Is フロー: {e}")

        if tobe_flow:
            try:
                dg.render_swimlane(tobe_flow, str(tmp / "tobe.png"))
                tobe_img = str(tmp / "tobe.png")
                print("  [OK] To-Be フロー")
            except Exception as e:
                print(f"  [WARN] To-Be フロー: {e}")

        wb = Workbook()
        wb.remove(wb.active)

        ws1 = wb.create_sheet("表紙")
        ws2 = wb.create_sheet("システム概要")
        ws3 = wb.create_sheet("業務フロー図")
        ws4 = wb.create_sheet("ER図")
        ws5 = wb.create_sheet("用語集")

        _build_cover(ws1, org, req, author, history=history)
        _build_system_overview(ws2, req, system, sys_img)
        _build_flow_sheet(ws3, asis_flow, tobe_flow, asis_img, tobe_img)
        _build_er_sheet(ws4, objects, relations, er_img)
        _build_glossary_sheet(ws5, org.get("glossary", []))

        write_meta(wb, {
            "version":     current_version,
            "history":     history,
            "source_hash": new_hash,
        })

        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output))

    print(f"saved: {output}")


def main():
    ap = argparse.ArgumentParser(description="プロジェクト概要書.xlsx 生成")
    ap.add_argument("--docs-dir",     required=True)
    ap.add_argument("--output",       required=True)
    ap.add_argument("--author",       default="")
    ap.add_argument("--project-name", default="")
    ap.add_argument("--source-file",  default="",
                    help="更新時: 既存のプロジェクト概要書.xlsx のパス（_meta から履歴引き継ぎ）")
    ap.add_argument("--version-increment", default="minor",
                    choices=["minor", "major"],
                    help="minor: x.y+1 / major: x+1.0（履歴リセット）")
    args = ap.parse_args()
    try:
        generate(Path(args.docs_dir), Path(args.output),
                 args.author, args.project_name,
                 args.source_file, args.version_increment)
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
